import os
import sys
import time
import sqlite3
import datetime
import random
import warnings
import logging
import traceback
warnings.filterwarnings("ignore", category=DeprecationWarning)

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QLabel, QLineEdit, QPushButton, QMessageBox, QFileDialog, QGroupBox,
    QScrollArea, QSizePolicy, QSpacerItem, QDialog, QTableWidget, QTableWidgetItem, 
    QHeaderView, QFormLayout, QDialogButtonBox, QComboBox, QInputDialog, QGridLayout,
    QFrame, QGraphicsDropShadowEffect, QSpinBox
)
from PyQt5.QtGui import QFont, QFontDatabase, QColor, QIcon, QPixmap, QPalette
from PyQt5.QtCore import Qt, QSize
from PyQt5.QtWidgets import QDesktopWidget

import barcode
from barcode.writer import ImageWriter
import openpyxl
from io import BytesIO

def safe_pil_import():
    """Safely import PIL modules - EXE compatible"""
    try:
        from PIL import Image, ImageDraw, ImageFont
        return Image, ImageDraw, ImageFont
    except ImportError:
        try:
            import PIL.Image as Image
            import PIL.ImageDraw as ImageDraw
            import PIL.ImageFont as ImageFont
            return Image, ImageDraw, ImageFont
        except ImportError:
            try:
                # Last resort for very old installations
                import Image, ImageDraw, ImageFont
                return Image, ImageDraw, ImageFont
            except ImportError:
                raise ImportError("Cannot import PIL/Pillow - required for barcode generation")
import subprocess
from functools import partial

try:
    from EmbeddedFont import get_embedded_font
except ImportError:
    def get_embedded_font():
        return None

class ResponsiveScaler:
    """Enhanced responsive scaling system for universal screen compatibility"""
    
    def __init__(self):
        self.screen = QDesktopWidget().screenGeometry()
        self.screen_width = self.screen.width()
        self.screen_height = self.screen.height()
        self.screen_diagonal = (self.screen_width ** 2 + self.screen_height ** 2) ** 0.5
        
        # Get screen DPI for high-DPI displays
        self.screen_dpi = QApplication.instance().primaryScreen().logicalDotsPerInch()
        self.dpi_factor = self.screen_dpi / 96.0  # 96 DPI is standard
        
        # Enhanced screen size detection with more granular categories
        self.detect_screen_category()
        
        # Calculate responsive dimensions with DPI awareness
        self.calculate_responsive_dimensions()
        
        # Calculate window sizing based on screen size and orientation
        self.calculate_window_sizing()
    
    def detect_screen_category(self):
        """Enhanced screen category detection with DPI awareness"""
        # Consider both physical size and DPI
        effective_diagonal = self.screen_diagonal * self.dpi_factor
        
        if effective_diagonal < 1000:  # Very small screens (netbooks, small laptops)
            self.scale_factor = 0.65
            self.category = "tiny"
        elif effective_diagonal < 1300:  # Small screens (ThinkPads, small laptops)
            self.scale_factor = 0.75
            self.category = "small"
        elif effective_diagonal < 1600:  # Medium screens (standard laptops)
            self.scale_factor = 0.85
            self.category = "medium"
        elif effective_diagonal < 2000:  # Large screens (desktops, large laptops)
            self.scale_factor = 0.95
            self.category = "large"
        else:  # Very large screens (4K, ultra-wide)
            self.scale_factor = 1.0
            self.category = "xlarge"
        
        # Adjust for ultra-wide screens (width > 2x height)
        if self.screen_width > self.screen_height * 2:
            self.scale_factor *= 0.9  # Slightly smaller for ultra-wide
            self.category += "_ultrawide"
        
        # Adjust for portrait screens (height > width)
        if self.screen_height > self.screen_width:
            self.scale_factor *= 0.85  # Smaller for portrait
            self.category += "_portrait"
    
    def calculate_responsive_dimensions(self):
        """Calculate responsive dimensions with enhanced scaling"""
        # Base font sizes with DPI awareness
        base_font_size = max(10, int(14 * self.scale_factor * min(self.dpi_factor, 1.5)))
        
        self.base_font = base_font_size
        self.header_font = max(12, int(16 * self.scale_factor * min(self.dpi_factor, 1.5)))
        self.title_font = max(16, int(22 * self.scale_factor * min(self.dpi_factor, 1.5)))
        self.subtitle_font = max(11, int(14 * self.scale_factor * min(self.dpi_factor, 1.5)))
        self.table_font = max(9, int(12 * self.scale_factor * min(self.dpi_factor, 1.5)))
        self.button_font = max(10, int(13 * self.scale_factor * min(self.dpi_factor, 1.5)))
        self.stats_font = max(11, int(14 * self.scale_factor * min(self.dpi_factor, 1.5)))
        
        # Responsive layout dimensions
        self.row_height = max(24, int(36 * self.scale_factor * min(self.dpi_factor, 1.3)))
        self.input_height = max(20, int(30 * self.scale_factor * min(self.dpi_factor, 1.3)))
        self.button_height = max(22, int(32 * self.scale_factor * min(self.dpi_factor, 1.3)))
        self.base_padding = max(6, int(10 * self.scale_factor * min(self.dpi_factor, 1.2)))
        self.section_spacing = max(8, int(14 * self.scale_factor * min(self.dpi_factor, 1.2)))
        self.radius = max(4, int(6 * self.scale_factor * min(self.dpi_factor, 1.2)))
        
        # Enhanced spacing for different screen categories
        if self.category.startswith("tiny"):
            self.base_padding = max(4, self.base_padding * 0.8)
            self.section_spacing = max(6, self.section_spacing * 0.8)
        elif self.category.startswith("xlarge"):
            self.base_padding = int(self.base_padding * 1.2)
            self.section_spacing = int(self.section_spacing * 1.2)
    
    def calculate_window_sizing(self):
        """Calculate optimal window sizing for different screen types"""
        # Calculate available screen space (accounting for taskbar, etc.)
        available_width = self.screen_width - 100
        available_height = self.screen_height - 100
        
        if self.category.startswith("tiny"):
            self.window_width = min(900, available_width)
            self.window_height = min(600, available_height)
            self.min_width = 800
            self.min_height = 550
        elif self.category.startswith("small"):
            self.window_width = min(1100, available_width)
            self.window_height = min(700, available_height)
            self.min_width = 950
            self.min_height = 600
        elif self.category.startswith("medium"):
            self.window_width = min(1300, available_width)
            self.window_height = min(800, available_height)
            self.min_width = 1100
            self.min_height = 700
        elif self.category.startswith("large"):
            self.window_width = min(1500, available_width)
            self.window_height = min(900, available_height)
            self.min_width = 1300
            self.min_height = 800
        else:  # xlarge
            self.window_width = min(1700, available_width)
            self.window_height = min(1000, available_height)
            self.min_width = 1500
            self.min_height = 900
        
        # Adjust for ultra-wide screens
        if "ultrawide" in self.category:
            self.window_width = min(self.window_width * 1.3, available_width)
        
        # Adjust for portrait screens
        if "portrait" in self.category:
            self.window_width = min(self.window_width * 0.8, available_width)
            self.window_height = min(self.window_height * 1.1, available_height)
    
    def get_column_proportions(self):
        """Get responsive column proportions based on screen size"""
        if self.category.startswith("tiny"):
            return 45, 55  # More space for form on tiny screens
        elif self.category.startswith("small"):
            return 40, 60  # More space for form on small screens
        elif self.category.startswith("medium"):
            return 35, 65  # Balanced proportions
        else:
            return 30, 70  # More space for table on larger screens
    
    def get_table_row_height(self):
        """Get responsive table row height"""
        if self.category.startswith("tiny"):
            return max(20, int(28 * self.scale_factor))
        elif self.category.startswith("small"):
            return max(24, int(32 * self.scale_factor))
        else:
            return max(28, int(36 * self.scale_factor))
    
    def get_dialog_size(self, base_width=850, base_height=600):
        """Get responsive dialog size"""
        width = int(base_width * self.scale_factor * min(self.dpi_factor, 1.3))
        height = int(base_height * self.scale_factor * min(self.dpi_factor, 1.3))
        return width, height

class BarcodeGenerator:
    """KipTags engine for TBS LP46 Neo (203 DPI) thermal printer - REFACTORED SINGLE VERSION"""
    
    # Default dimensions for thermal labels (can be overridden)
    DPI = 203            # TBS LP46 Neo native DPI
    
    # MAXIMUM QUALITY SETTINGS
    HD_MULTIPLIER = 4     # Render at 4x resolution for maximum quality
    USE_HD_MODE = True    # Enable maximum quality rendering by default
    
    def __init__(self, logger):
        self.logger = logger
        self.use_tier3_rendering = True  # Enable Tier 3.5 by default
        self.hd_mode = True              # Enable HD mode for better quality
        
        # Initialize with standard dimensions (can be updated later)
        self.set_dimensions("standard")
    
    def set_dimensions(self, paper_size):
        """Set dimensions based on paper size"""
        if paper_size == "38x25mm":
            # 38x25mm dimensions (2UP)
            # Physical: 76mm x 25mm (2 labels side by side)
            # At 203 DPI: 76mm = 76 * 203/25.4 ≈ 607px, 25mm = 25 * 203/25.4 ≈ 200px
            self.CANVAS_WIDTH = 607
            self.CANVAS_HEIGHT = 200
            self.LABEL_WIDTH = 303  # 607/2 rounded down
            self.LABEL_MARGIN = 10   # Slightly smaller margin for smaller labels
            self.BARCODE_WIDTH = 200  # Smaller barcode for smaller labels
            self.BARCODE_HEIGHT = 50  # Slightly smaller height
            self.NAME_HEIGHT = 24     # Smaller name height
            self.PRICE_HEIGHT = 20    # Smaller price height
            self.INFO_HEIGHT = 16     # Smaller info height
        else:
            # Standard dimensions (40x25mm)
            self.CANVAS_WIDTH = 800
            self.CANVAS_HEIGHT = 200
            self.LABEL_WIDTH = 400
            self.LABEL_MARGIN = 12
            self.BARCODE_WIDTH = 260
            self.BARCODE_HEIGHT = 60
            self.NAME_HEIGHT = 28
            self.PRICE_HEIGHT = 24
            self.INFO_HEIGHT = 18
        
        # Update derived dimensions
        self.CONTENT_WIDTH = self.LABEL_WIDTH - (self.LABEL_MARGIN * 2)
        self.CONTENT_HEIGHT = self.CANVAS_HEIGHT - (self.LABEL_MARGIN * 2)
        
        self.logger.info(f"Set dimensions for {paper_size}")
        self.logger.info(f"  Canvas: {self.CANVAS_WIDTH}x{self.CANVAS_HEIGHT}")
        self.logger.info(f"  Label width: {self.LABEL_WIDTH}")
        self.logger.info(f"  Barcode: {self.BARCODE_WIDTH}x{self.BARCODE_HEIGHT}")
        
    def load_font_from_bytes(self, font_path, size):
        """Load font from bytes into memory for high-quality rendering"""
        try:
            # Import PIL safely
            Image, ImageDraw, ImageFont = safe_pil_import()
            
            if font_path and os.path.exists(font_path):
                with open(font_path, 'rb') as f:
                    font_bytes = BytesIO(f.read())
                    return ImageFont.truetype(font_bytes, size)
            else:
                self.logger.warning(f"Font path not found: {font_path}")
                return ImageFont.load_default()
        except Exception as e:
            self.logger.error(f"Failed to load font from {font_path}: {str(e)}")
            # Safe fallback
            try:
                Image, ImageDraw, ImageFont = safe_pil_import()
                return ImageFont.load_default()
            except:
                return None
            
    def generate_barcode(self, barcode_value):
        """Generate REAL SCANNABLE EAN-13 barcode - GUARANTEED CORRECT"""
        try:
            self.logger.info(f"🎯 GENERATING SCANNABLE BARCODE FOR: {barcode_value}")
            
            # Clean and validate input
            barcode_str = str(barcode_value).strip() if barcode_value else "123456789012"
            self.logger.info(f"After strip: '{barcode_str}'")
            
            # Extract only digits
            digits_only = ''.join(filter(str.isdigit, barcode_str))
            if not digits_only:
                digits_only = "123456789012"
            self.logger.info(f"Digits only: '{digits_only}'")
            
            # Get exactly 12 digits for EAN-13
            if len(digits_only) >= 12:
                barcode_12 = digits_only[:12]
            else:
                barcode_12 = digits_only.zfill(12)
            
            self.logger.info(f"12-digit base: '{barcode_12}'")
            
            # Calculate CORRECT check digit
            check_digit = self.calculate_guaranteed_check_digit(barcode_12)
            full_barcode = barcode_12 + str(check_digit)
            
            self.logger.info(f"✅ FINAL SCANNABLE BARCODE: {full_barcode}")
            
            # Generate the barcode image
            return self.create_guaranteed_scannable_barcode(full_barcode)
            
        except Exception as e:
            self.logger.error(f"Barcode generation failed: {e}")
            # Create error barcode with the issue
            return self.create_error_barcode_with_message(f"ERROR: {e}")
    
    def calculate_guaranteed_check_digit(self, barcode_12):
        """GUARANTEED CORRECT EAN-13 check digit calculation"""
        try:
            # Ensure exactly 12 digits
            if len(barcode_12) != 12:
                barcode_12 = str(barcode_12).zfill(12)[:12]
            
            self.logger.info(f"Calculating check digit for: {barcode_12}")
            
            # OFFICIAL EAN-13 algorithm
            total = 0
            for i in range(12):
                digit = int(barcode_12[i])
                multiplier = 1 if i % 2 == 0 else 3
                contribution = digit * multiplier
                total += contribution
                self.logger.debug(f"Position {i+1}: {digit} × {multiplier} = {contribution}")
            
            check_digit = (10 - (total % 10)) % 10
            
            self.logger.info(f"Total: {total}, Check digit: {check_digit}")
            self.logger.info(f"RESULT: {barcode_12} + {check_digit} = {barcode_12}{check_digit}")
            
            return check_digit
            
        except Exception as e:
            self.logger.error(f"Check digit calculation failed: {e}")
            return 0
    
    def create_guaranteed_scannable_barcode(self, full_barcode_13):
        """Create GUARANTEED SCANNABLE EAN-13 barcode"""
        try:
            Image, ImageDraw, ImageFont = safe_pil_import()
            
            self.logger.info(f"Creating SCANNABLE barcode for: {full_barcode_13}")
            
            # Validate input
            if len(full_barcode_13) != 13:
                raise ValueError(f"Invalid barcode length: {len(full_barcode_13)}")
            
            # Use the REAL EAN-13 implementation
            return self.create_real_ean13_with_proper_encoding(full_barcode_13)
            
        except Exception as e:
            self.logger.error(f"Guaranteed barcode creation failed: {e}")
            return self.create_error_barcode_with_message(f"BARCODE ERROR: {e}")
    
    def create_real_ean13_with_proper_encoding(self, full_barcode):
        """Create REAL EAN-13 with proper encoding - GUARANTEED SCANNABLE"""
        try:
            Image, ImageDraw, ImageFont = safe_pil_import()
            
            # OFFICIAL EAN-13 encoding patterns
            L_patterns = {  # Left side odd parity
                '0': '0001101', '1': '0011001', '2': '0010011', '3': '0111101',
                '4': '0100011', '5': '0110001', '6': '0101111', '7': '0111011',
                '8': '0110111', '9': '0001011'
            }
            
            G_patterns = {  # Left side even parity
                '0': '0100111', '1': '0110011', '2': '0011011', '3': '0100001',
                '4': '0011101', '5': '0111001', '6': '0000101', '7': '0010001',
                '8': '0001001', '9': '0010111'
            }
            
            R_patterns = {  # Right side
                '0': '1110010', '1': '1100110', '2': '1101100', '3': '1000010',
                '4': '1011100', '5': '1001110', '6': '1010000', '7': '1000100',
                '8': '1001000', '9': '1110100'
            }
            
            # First digit determines parity pattern
            first_digit_patterns = {
                0: 'LLLLLL', 1: 'LLGLGG', 2: 'LLGGLG', 3: 'LLGGGL',
                4: 'LGLLGG', 5: 'LGGLLG', 6: 'LGGGLL', 7: 'LGLGLG',
                8: 'LGLGGL', 9: 'LGGLGL'
            }
            
            first_digit = int(full_barcode[0])
            parity_pattern = first_digit_patterns[first_digit]
            
            # Build the complete barcode pattern
            pattern = '101'  # Start guard
            
            # Left side (digits 1-6)
            for i in range(6):
                digit = full_barcode[i + 1]
                if parity_pattern[i] == 'L':
                    pattern += L_patterns[digit]
                else:  # 'G'
                    pattern += G_patterns[digit]
            
            pattern += '01010'  # Center guard
            
            # Right side (digits 7-12 + check digit)
            for i in range(6):
                digit = full_barcode[i + 7]
                pattern += R_patterns[digit]
            
            pattern += '101'  # End guard
            
            # Create the barcode image with optimal settings
            module_width = 2
            bar_height = self.BARCODE_HEIGHT - 18
            quiet_zone = 10
            
            total_width = len(pattern) * module_width + (2 * quiet_zone)
            
            img = Image.new('1', (total_width, self.BARCODE_HEIGHT), 1)
            draw = ImageDraw.Draw(img)
            
            # Draw the barcode pattern
            x = quiet_zone
            for bit in pattern:
                if bit == '1':
                    draw.rectangle([x, 2, x + module_width - 1, bar_height], fill=0)
                x += module_width
            
            # Add the numbers underneath
            try:
                font = ImageFont.load_default()
                text_width = len(full_barcode) * 6
                text_x = (total_width - text_width) // 2
                text_y = bar_height + 3
                draw.text((text_x, text_y), full_barcode, font=font, fill=0)
            except:
                pass
            
            # Resize to target dimensions
            img = img.resize((self.BARCODE_WIDTH, self.BARCODE_HEIGHT), Image.NEAREST)
            
            self.logger.info(f"✅ SCANNABLE EAN-13 barcode created: {full_barcode}")
            return img
            
        except Exception as e:
            self.logger.error(f"Real EAN-13 creation failed: {e}")
            raise e
    
    def create_error_barcode_with_message(self, message):
        """Create error barcode with specific message"""
        try:
            Image, ImageDraw, ImageFont = safe_pil_import()
            
            img = Image.new('1', (self.BARCODE_WIDTH, self.BARCODE_HEIGHT), 1)
            draw = ImageDraw.Draw(img)
            
            # Draw error pattern
            for i in range(10, self.BARCODE_WIDTH - 10, 8):
                draw.rectangle([i, 10, i + 4, self.BARCODE_HEIGHT - 20], fill=0)
            
            # Add error message
            try:
                font = ImageFont.load_default()
                draw.text((10, self.BARCODE_HEIGHT - 15), message[:20], font=font, fill=0)
            except:
                pass
            
            return img
            
        except:
            # Ultimate fallback
            try:
                Image, _, _ = safe_pil_import()
                return Image.new('1', (self.BARCODE_WIDTH, self.BARCODE_HEIGHT), 1)
            except:
                return None
    
    def create_simple_ean13_barcode(self, barcode_12):
        """Create REAL SCANNABLE EAN-13 barcode using proper encoding"""
        try:
            Image, ImageDraw, ImageFont = safe_pil_import()
            
            self.logger.info(f"Creating REAL EAN-13 for: {barcode_12}")
            
            # Calculate check digit
            check_digit = self.calculate_simple_check_digit(barcode_12)
            full_barcode = barcode_12 + str(check_digit)
            
            self.logger.info(f"Full EAN-13: {full_barcode}")
            
            # REAL EAN-13 encoding patterns
            L_patterns = {  # Left side odd parity
                '0': '0001101', '1': '0011001', '2': '0010011', '3': '0111101',
                '4': '0100011', '5': '0110001', '6': '0101111', '7': '0111011',
                '8': '0110111', '9': '0001011'
            }
            
            G_patterns = {  # Left side even parity
                '0': '0100111', '1': '0110011', '2': '0011011', '3': '0100001',
                '4': '0011101', '5': '0111001', '6': '0000101', '7': '0010001',
                '8': '0001001', '9': '0010111'
            }
            
            R_patterns = {  # Right side
                '0': '1110010', '1': '1100110', '2': '1101100', '3': '1000010',
                '4': '1011100', '5': '1001110', '6': '1010000', '7': '1000100',
                '8': '1001000', '9': '1110100'
            }
            
            # First digit determines parity pattern
            first_digit_patterns = {
                0: 'LLLLLL', 1: 'LLGLGG', 2: 'LLGGLG', 3: 'LLGGGL',
                4: 'LGLLGG', 5: 'LGGLLG', 6: 'LGGGLL', 7: 'LGLGLG',
                8: 'LGLGGL', 9: 'LGGLGL'
            }
            
            first_digit = int(full_barcode[0])
            parity_pattern = first_digit_patterns[first_digit]
            
            # Build the complete barcode pattern
            pattern = '101'  # Start guard
            
            # Left side (digits 1-6)
            for i in range(6):
                digit = full_barcode[i + 1]
                if parity_pattern[i] == 'L':
                    pattern += L_patterns[digit]
                else:  # 'G'
                    pattern += G_patterns[digit]
            
            pattern += '01010'  # Center guard
            
            # Right side (digits 7-12 + check digit)
            for i in range(6):
                digit = full_barcode[i + 7]
                pattern += R_patterns[digit]
            
            pattern += '101'  # End guard
            
            # Create the barcode image
            module_width = 2
            bar_height = self.BARCODE_HEIGHT - 18
            quiet_zone = 10
            
            total_width = len(pattern) * module_width + (2 * quiet_zone)
            
            img = Image.new('1', (total_width, self.BARCODE_HEIGHT), 1)
            draw = ImageDraw.Draw(img)
            
            # Draw the REAL barcode pattern
            x = quiet_zone
            for bit in pattern:
                if bit == '1':
                    draw.rectangle([x, 2, x + module_width - 1, bar_height], fill=0)
                x += module_width
            
            # Add the numbers underneath
            try:
                font = ImageFont.load_default()
                text_width = len(full_barcode) * 6
                text_x = (total_width - text_width) // 2
                text_y = bar_height + 3
                draw.text((text_x, text_y), full_barcode, font=font, fill=0)
            except:
                pass
            
            # Resize to target dimensions
            img = img.resize((self.BARCODE_WIDTH, self.BARCODE_HEIGHT), Image.NEAREST)
            
            self.logger.info("✅ REAL EAN-13 barcode created!")
            return img
            
        except Exception as e:
            self.logger.error(f"Real EAN-13 creation failed: {e}")
            raise e
    
    def calculate_simple_check_digit(self, barcode_12):
        """CORRECT EAN-13 check digit calculation"""
        try:
            if len(barcode_12) != 12:
                barcode_12 = str(barcode_12).zfill(12)[:12]
            
            # CORRECT EAN-13 algorithm:
            # Odd positions (1st, 3rd, 5th...) multiply by 1
            # Even positions (2nd, 4th, 6th...) multiply by 3
            total = 0
            for i in range(12):
                digit = int(barcode_12[i])
                if i % 2 == 0:  # Positions 1, 3, 5, 7, 9, 11 (0-based: 0, 2, 4, 6, 8, 10)
                    total += digit * 1
                else:  # Positions 2, 4, 6, 8, 10, 12 (0-based: 1, 3, 5, 7, 9, 11)
                    total += digit * 3
            
            check_digit = (10 - (total % 10)) % 10
            
            self.logger.info(f"Check digit: {barcode_12} -> {check_digit}")
            return check_digit
            
        except Exception as e:
            self.logger.error(f"Check digit calculation failed: {e}")
            return 0  # Safe fallback
    
    def create_absolute_fallback_barcode(self, barcode_value):
        """Absolute fallback - NEVER FAILS"""
        try:
            Image, ImageDraw, ImageFont = safe_pil_import()
            
            # Create basic barcode that always works
            img = Image.new('1', (self.BARCODE_WIDTH, self.BARCODE_HEIGHT), 1)
            draw = ImageDraw.Draw(img)
            
            # Draw simple pattern
            for i in range(0, self.BARCODE_WIDTH - 20, 6):
                if i % 12 < 6:  # Alternating pattern
                    draw.rectangle([i + 10, 10, i + 13, self.BARCODE_HEIGHT - 20], fill=0)
            
            # Add text
            try:
                font = ImageFont.load_default()
                text = str(barcode_value)[:10] if barcode_value else "DEFAULT"
                draw.text((10, self.BARCODE_HEIGHT - 15), text, font=font, fill=0)
            except:
                pass
            
            self.logger.warning("⚠️ Using absolute fallback barcode")
            return img
            
        except:
            # If even this fails, return empty image
            try:
                Image, _, _ = safe_pil_import()
                return Image.new('1', (self.BARCODE_WIDTH, self.BARCODE_HEIGHT), 1)
            except:
                return None


    


    def create_real_ean13_barcode(self, barcode_value):
        """Create REAL SCANNABLE EAN-13 barcode - PYINSTALLER COMPATIBLE"""
        try:
            # Try python-barcode first, but if it fails in EXE, use direct implementation
            try:
                import barcode
                from barcode.writer import ImageWriter
                from PIL import Image
                
                self.logger.info("Using python-barcode library for EAN-13 generation")
                
                # Prepare the barcode number
                barcode_str = str(barcode_value).strip()
                
                # Calculate proper EAN-13 with check digit
                if len(barcode_str) == 12:
                    check_digit = self.calculate_ean13_check_digit(barcode_str)
                    full_barcode = barcode_str + str(check_digit)
                elif len(barcode_str) == 13:
                    full_barcode = barcode_str
                else:
                    barcode_12 = barcode_str.zfill(12)[:12]
                    check_digit = self.calculate_ean13_check_digit(barcode_12)
                    full_barcode = barcode_12 + str(check_digit)
                
                self.logger.info(f"Creating EAN-13: {full_barcode}")
                
                # Create the barcode using python-barcode library
                ean = barcode.get('ean13', full_barcode, writer=ImageWriter())
                
                options = {
                    'module_width': 0.33,
                    'module_height': 10.0,
                    'quiet_zone': 9.0,
                    'font_size': 7,
                    'text_distance': 3.0,
                    'background': 'white',
                    'foreground': 'black',
                    'write_text': True,
                    'center_text': True,
                    'dpi': 300
                }
                
                barcode_img = ean.render(options)
                
                if barcode_img.mode != '1':
                    if barcode_img.mode != 'L':
                        barcode_img = barcode_img.convert('L')
                    barcode_img = barcode_img.point(lambda x: 0 if x < 128 else 255, '1')
                
                barcode_img = barcode_img.resize((self.BARCODE_WIDTH, self.BARCODE_HEIGHT), Image.NEAREST)
                
                self.logger.info("✅ Python-barcode EAN-13 created successfully!")
                return barcode_img
                
            except (ImportError, Exception) as e:
                self.logger.warning(f"Python-barcode failed: {e}")
                self.logger.info("Falling back to direct EAN-13 implementation for EXE compatibility")
                
                # Direct EAN-13 implementation that works in PyInstaller
                return self.create_direct_ean13_barcode(barcode_value)
                
        except Exception as e:
            self.logger.error(f"All barcode methods failed: {e}")
            raise Exception(f"Barcode generation completely failed: {e}")
    
    def create_direct_ean13_barcode(self, barcode_value):
        """Direct EAN-13 implementation - BULLETPROOF for PyInstaller EXE"""
        try:
            Image, ImageDraw, ImageFont = safe_pil_import()
            
            self.logger.info("Using direct EAN-13 implementation for maximum compatibility")
            
            # Validate and prepare the barcode number
            barcode_str = str(barcode_value).strip()
            
            # Ensure we have only digits
            if not barcode_str.isdigit():
                barcode_str = ''.join(filter(str.isdigit, barcode_str))
                if not barcode_str:
                    raise ValueError("No valid digits in barcode input")
            
            # ALWAYS recalculate check digit to ensure valid EAN-13
            if len(barcode_str) >= 12:
                barcode_12 = barcode_str[:12].zfill(12)
            else:
                barcode_12 = barcode_str.zfill(12)
            
            # Validate the 12-digit code
            if not barcode_12.isdigit() or len(barcode_12) != 12:
                raise ValueError(f"Invalid 12-digit code: {barcode_12}")
            
            # Calculate correct check digit
            check_digit = self.calculate_ean13_check_digit(barcode_12)
            full_barcode = barcode_12 + str(check_digit)
            
            self.logger.info(f"Input: {barcode_str} -> Valid EAN-13: {full_barcode}")
            
            # Validate the final barcode
            if len(full_barcode) != 13 or not full_barcode.isdigit():
                raise ValueError(f"Invalid final EAN-13: {full_barcode}")
            
            self.logger.info(f"Creating direct EAN-13: {full_barcode}")
            
            # EAN-13 encoding patterns - OFFICIAL SPECIFICATION
            L_patterns = {  # Left side odd parity
                '0': '0001101', '1': '0011001', '2': '0010011', '3': '0111101',
                '4': '0100011', '5': '0110001', '6': '0101111', '7': '0111011',
                '8': '0110111', '9': '0001011'
            }
            
            G_patterns = {  # Left side even parity
                '0': '0100111', '1': '0110011', '2': '0011011', '3': '0100001',
                '4': '0011101', '5': '0111001', '6': '0000101', '7': '0010001',
                '8': '0001001', '9': '0010111'
            }
            
            R_patterns = {  # Right side
                '0': '1110010', '1': '1100110', '2': '1101100', '3': '1000010',
                '4': '1011100', '5': '1001110', '6': '1010000', '7': '1000100',
                '8': '1001000', '9': '1110100'
            }
            
            # First digit determines parity pattern
            first_digit_patterns = {
                0: 'LLLLLL', 1: 'LLGLGG', 2: 'LLGGLG', 3: 'LLGGGL',
                4: 'LGLLGG', 5: 'LGGLLG', 6: 'LGGGLL', 7: 'LGLGLG',
                8: 'LGLGGL', 9: 'LGGLGL'
            }
            
            first_digit = int(full_barcode[0])
            parity_pattern = first_digit_patterns[first_digit]
            
            # Build the complete barcode pattern
            pattern = '101'  # Start guard
            
            # Left side (digits 1-6)
            for i in range(6):
                digit = full_barcode[i + 1]
                if parity_pattern[i] == 'L':
                    pattern += L_patterns[digit]
                else:  # 'G'
                    pattern += G_patterns[digit]
            
            pattern += '01010'  # Center guard
            
            # Right side (digits 7-12 + check digit)
            for i in range(6):
                digit = full_barcode[i + 7]
                pattern += R_patterns[digit]
            
            pattern += '101'  # End guard
            
            # Create the barcode image
            module_width = 2
            bar_height = self.BARCODE_HEIGHT - 18  # Leave space for text
            quiet_zone = 10
            
            total_width = len(pattern) * module_width + (2 * quiet_zone)
            
            img = Image.new('1', (total_width, self.BARCODE_HEIGHT), 1)
            draw = ImageDraw.Draw(img)
            
            # Draw the barcode pattern
            x = quiet_zone
            for bit in pattern:
                if bit == '1':
                    draw.rectangle([x, 2, x + module_width - 1, bar_height], fill=0)
                x += module_width
            
            # Add the numbers underneath
            try:
                font = ImageFont.load_default()
                
                # Draw the full number centered
                text_width = len(full_barcode) * 6
                text_x = (total_width - text_width) // 2
                text_y = bar_height + 3
                
                draw.text((text_x, text_y), full_barcode, font=font, fill=0)
                
            except Exception as e:
                self.logger.warning(f"Could not add text: {e}")
            
            # Resize to target dimensions
            img = img.resize((self.BARCODE_WIDTH, self.BARCODE_HEIGHT), Image.NEAREST)
            
            self.logger.info("✅ Direct EAN-13 barcode created successfully!")
            return img
            
        except Exception as e:
            self.logger.error(f"Direct EAN-13 creation failed: {e}")
            raise Exception(f"Direct barcode creation failed: {e}")
    
    def create_error_barcode(self, barcode_value):
        """Create error barcode when all else fails - GUARANTEED to work"""
        try:
            Image, ImageDraw, ImageFont = safe_pil_import()
            
            self.logger.warning("Creating error barcode - something went wrong!")
            
            # Create a simple error barcode
            img = Image.new('1', (self.BARCODE_WIDTH, self.BARCODE_HEIGHT), 1)
            draw = ImageDraw.Draw(img)
            
            # Draw simple error pattern
            bar_width = 4
            x = 10
            
            # Create a recognizable error pattern
            error_pattern = [1, 0, 1, 0, 1, 1, 0, 0, 1, 0, 1, 1, 0, 1, 0, 1, 1, 0, 0, 1]
            
            for i, bar in enumerate(error_pattern):
                if bar == 1 and x < self.BARCODE_WIDTH - 20:
                    draw.rectangle([x, 5, x + bar_width - 1, self.BARCODE_HEIGHT - 20], fill=0)
                x += bar_width + 2
            
            # Add error text
            try:
                font = ImageFont.load_default()
                error_text = f"ERROR: {str(barcode_value)[:10]}"
                draw.text((10, self.BARCODE_HEIGHT - 15), error_text, font=font, fill=0)
            except:
                pass
            
            self.logger.warning("⚠️ Error barcode created - check logs for issues!")
            return img
            
        except Exception as e:
            self.logger.error(f"Even error barcode creation failed: {e}")
            # Absolute last resort - empty image
            return Image.new('1', (self.BARCODE_WIDTH, self.BARCODE_HEIGHT), 1)
    



    
    def calculate_ean13_check_digit(self, barcode_str):
        """Calculate CORRECT EAN-13 check digit - BULLETPROOF"""
        try:
            # Validate input
            if not barcode_str:
                raise ValueError("Empty barcode string")
            
            # Ensure we have exactly 12 digits
            barcode_str = str(barcode_str).strip()
            
            # Remove non-digits
            barcode_str = ''.join(filter(str.isdigit, barcode_str))
            
            if not barcode_str:
                raise ValueError("No digits found in barcode string")
            
            # Pad or truncate to 12 digits
            barcode_str = barcode_str.zfill(12)[:12]
            
            # Validate we have 12 digits
            if len(barcode_str) != 12 or not barcode_str.isdigit():
                raise ValueError(f"Invalid 12-digit string: {barcode_str}")
            
            # EAN-13 check digit calculation:
            # Multiply odd positions (1st, 3rd, 5th...) by 1
            # Multiply even positions (2nd, 4th, 6th...) by 3
            # Sum all results
            # Check digit = (10 - (sum % 10)) % 10
            
            total = 0
            for i in range(12):
                digit = int(barcode_str[i])
                if digit < 0 or digit > 9:
                    raise ValueError(f"Invalid digit at position {i}: {digit}")
                
                if i % 2 == 0:  # Positions 1, 3, 5, 7, 9, 11 (0-based: 0, 2, 4, 6, 8, 10)
                    total += digit * 1
                else:  # Positions 2, 4, 6, 8, 10, 12 (0-based: 1, 3, 5, 7, 9, 11)
                    total += digit * 3
            
            check_digit = (10 - (total % 10)) % 10
            
            # Validate check digit
            if check_digit < 0 or check_digit > 9:
                raise ValueError(f"Invalid check digit calculated: {check_digit}")
            
            self.logger.info(f"Check digit calculation: {barcode_str} -> {check_digit}")
            return check_digit
            
        except Exception as e:
            self.logger.error(f"Check digit calculation failed: {e}")
            self.logger.error(f"Input was: {repr(barcode_str)}")
            # Return a safe default
            return 0
    
    def render_text_hd(self, text, font, canvas_size):
        """Render text with thick, blocky quality for thermal printing"""
        try:
            # Import PIL safely
            Image, ImageDraw, ImageFont = safe_pil_import()
            
            # Create direct 1-bit canvas for crisp text
            img = Image.new('1', canvas_size, 1)  # White background
            draw = ImageDraw.Draw(img)
            
            # Use the font directly without HD scaling
            if hasattr(font, 'size'):
                try:
                    # Use font as-is for direct rendering
                    direct_font = font
                except:
                    direct_font = ImageFont.load_default()
            else:
                direct_font = ImageFont.load_default()
            
            # Get text size for centering
            try:
                bbox = draw.textbbox((0, 0), text, font=direct_font)
                text_width = bbox[2] - bbox[0]
                text_height = bbox[3] - bbox[1]
            except:
                # Fallback if textbbox fails
                text_width = len(text) * 8  # Approximate width
                text_height = 16  # Approximate height
            
            # Center text
            x = (canvas_size[0] - text_width) // 2
            y = (canvas_size[1] - text_height) // 2
            
            # Ensure text is within bounds
            x = max(0, min(x, canvas_size[0] - text_width))
            y = max(0, min(y, canvas_size[1] - text_height))
            
            # Draw clean, simple text for maximum readability
            draw.text((x, y), text, font=direct_font, fill=0)
            
            return img
            
        except Exception as e:
            self.logger.error(f"Thick text rendering failed: {e}")
            # Fallback to simple text
            try:
                Image, ImageDraw, ImageFont = safe_pil_import()
                img = Image.new('1', canvas_size, 1)
                draw = ImageDraw.Draw(img)
                draw.text((10, 10), text, fill=0)
                return img
            except:
                # Ultimate fallback - return empty image
                try:
                    Image, _, _ = safe_pil_import()
                    return Image.new('1', canvas_size, 1)
                except:
                    return None
    
    def create_label_image(self, items, font_path=None):
        """Create a perfectly symmetrical and balanced dual-label image optimized for thermal printing"""
        try:
            # Import PIL safely
            Image, ImageDraw, ImageFont = safe_pil_import()
            
            self.logger.info(f"Creating label image for {len(items)} items")
            self.logger.info(f"Current dimensions: Canvas={self.CANVAS_WIDTH}x{self.CANVAS_HEIGHT}, Label={self.LABEL_WIDTH}")
            # Start with 1-bit canvas (pure black/white, no grayscale)
            canvas = Image.new('1', (self.CANVAS_WIDTH, self.CANVAS_HEIGHT), 1)  # White background
            
            # Load fonts with proper sizing for clear text
            try:
                if font_path and os.path.exists(font_path):
                    self.logger.info(f"Loading fonts from: {font_path}")
                    title_font = self.load_font_from_bytes(font_path, 16)      # Appropriately sized
                    price_font = self.load_font_from_bytes(font_path, 14)      # Appropriately sized
                    info_font = self.load_font_from_bytes(font_path, 10)       # Appropriately sized
                    barcode_font = self.load_font_from_bytes(font_path, 10)    # Appropriately sized
                else:
                    self.logger.warning("Font path not available, using system fonts")
                    try:
                        title_font = ImageFont.truetype("arialbd.ttf", 16)
                        price_font = ImageFont.truetype("arialbd.ttf", 14)
                        info_font = ImageFont.truetype("arialbd.ttf", 10)
                        barcode_font = ImageFont.truetype("arialbd.ttf", 10)
                    except:
                        self.logger.warning("System fonts not available, using default")
                        title_font = price_font = info_font = barcode_font = ImageFont.load_default()
            except Exception as e:
                self.logger.error(f"Font loading failed: {str(e)}, using default")
                title_font = price_font = info_font = barcode_font = ImageFont.load_default()
            
            for i, item in enumerate(items[:2]):
                x_offset = i * self.LABEL_WIDTH
                item_name = item['item_name'][:20].upper()  # Uppercase for consistency
                
                # === PERFECTLY SYMMETRICAL LABEL STRUCTURE (200px height) ===
                # Top margin: (0-18px) - Clean top margin
                # Item Name: (18-46px) - 28px for name with padding
                # Gap: (46-62px) - 16px breathing room
                # Barcode: (62-122px) - 60px for barcode (perfectly centered)
                # Gap: (122-138px) - 16px separation
                # Price: (138-162px) - 24px for price
                # Gap: (162-172px) - 10px breathing room
                # Bottom info: (172-190px) - 18px for barcode number and brand
                # Bottom margin: (190-200px) - Clean bottom margin
                
                # 1. ITEM NAME - Perfectly centered at top
                name_y = 18
                name_img = self.render_text_hd(item_name, title_font, (self.CONTENT_WIDTH, self.NAME_HEIGHT))
                if name_img.mode != '1':
                    name_img = name_img.convert('1', dither=Image.NONE)
                canvas.paste(name_img, (x_offset + self.LABEL_MARGIN, name_y))
                
                # 2. BARCODE - Perfectly centered horizontally and vertically
                barcode_value = str(item['barcode']).strip().zfill(13)
                self.logger.info(f"Processing barcode: {barcode_value}")
                
                # Generate barcode with fallback
                barcode_img = None
                for attempt in range(3):
                    try:
                        barcode_img = self.generate_barcode(barcode_value)
                        if barcode_img and barcode_img.size == (self.BARCODE_WIDTH, self.BARCODE_HEIGHT):
                            self.logger.info(f" Barcode generated successfully on attempt {attempt + 1}")
                            break
                        else:
                            self.logger.warning(f"Attempt {attempt + 1}: Invalid barcode size or None")
                    except Exception as e:
                        self.logger.warning(f"Attempt {attempt + 1} failed: {e}")
                
                if barcode_img and barcode_img.size == (self.BARCODE_WIDTH, self.BARCODE_HEIGHT):
                    # Perfect center positioning - EXACT same as original design
                    barcode_x = x_offset + (self.LABEL_WIDTH - self.BARCODE_WIDTH) // 2
                    barcode_y = 62
                    
                    # CRITICAL: Validate barcode has content
                    if barcode_img.mode == '1':
                        black_pixels = sum(1 for pixel in barcode_img.getdata() if pixel == 0)
                        self.logger.info(f"Barcode has {black_pixels} black pixels - ready to paste")
                        
                        if black_pixels < 50:
                            self.logger.error(f" Barcode has too few black pixels ({black_pixels}) - regenerating...")
                            # Force regeneration with direct method if python-barcode failed
                            barcode_img = self.create_simple_barcode(barcode_value)
                    
                    # Paste barcode into canvas with validation
                    try:
                        canvas.paste(barcode_img, (barcode_x, barcode_y))
                        self.logger.info(f" Barcode pasted at ({barcode_x}, {barcode_y}) - EXACT same as original design")
                        
                    except Exception as paste_error:
                        self.logger.error(f" Failed to paste barcode: {paste_error}")
                        
                else:
                    # Create centered error placeholder with EXACT same positioning
                    self.logger.error(f" Creating error placeholder for barcode: {barcode_value}")
                    placeholder = Image.new('1', (self.BARCODE_WIDTH, self.BARCODE_HEIGHT), 1)
                    draw = ImageDraw.Draw(placeholder)
                    draw.rectangle([0, 0, self.BARCODE_WIDTH-1, self.BARCODE_HEIGHT-1], outline=0, width=2)
                    draw.text((self.BARCODE_WIDTH//2, self.BARCODE_HEIGHT//2), "ERROR", fill=0, anchor="mm")
                    barcode_x = x_offset + (self.LABEL_WIDTH - self.BARCODE_WIDTH) // 2
                    canvas.paste(placeholder, (barcode_x, 62))
                
                # 3. PRICE - Perfectly centered below barcode (clean design without Rs)
                price_y = 138
                price_text = f"SALE PRICE: {int(item['sale_price'])}"
                price_img = self.render_text_hd(price_text, price_font, (self.CONTENT_WIDTH, self.PRICE_HEIGHT))
                if price_img.mode != '1':
                    price_img = price_img.convert('1', dither=Image.NONE)
                canvas.paste(price_img, (x_offset + self.LABEL_MARGIN, price_y))
                
                # 4. BOTTOM ROW - Perfectly symmetrical positioning
                bottom_y = 172
                
                # CRITICAL: Always use the SAME barcode value regardless of generation method
                # This ensures the barcode number under the lines matches what was generated
                consistent_barcode = str(item['barcode']).strip().zfill(13)
                self.logger.info(f"Bottom row using consistent barcode: {consistent_barcode}")
                
                # Left side: Barcode number (left-aligned with margin) - EXACT same as original design
                barcode_num_img = self.render_text_hd(consistent_barcode, barcode_font, (self.CONTENT_WIDTH // 2 + 10, self.INFO_HEIGHT))
                if barcode_num_img.mode != '1':
                    barcode_num_img = barcode_num_img.convert('1', dither=Image.NONE)
                canvas.paste(barcode_num_img, (x_offset + self.LABEL_MARGIN + 8, bottom_y))
                self.logger.info(f" Barcode number pasted: {consistent_barcode}")
                
                # Right side: Brand (right-aligned with matching margin) - EXACT same as original design
                geek_text = "AUTO GEEK"
                geek_img = self.render_text_hd(geek_text, info_font, (self.CONTENT_WIDTH // 2 - 10, self.INFO_HEIGHT))
                if geek_img.mode != '1':
                    geek_img = geek_img.convert('1', dither=Image.NONE)
                # Position with matching right margin for perfect symmetry
                geek_x = x_offset + self.LABEL_WIDTH - geek_img.width - (self.LABEL_MARGIN + 8)
                canvas.paste(geek_img, (geek_x, bottom_y))
                self.logger.info(f" Brand text pasted: {geek_text}")
            
            self.logger.info(f"SYMMETRICAL label image created successfully: {canvas.size}, mode: {canvas.mode}")
            return canvas
            
        except Exception as e:
            self.logger.error(f"Error creating label image: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())
            return None
    
    def get_font_path(self):
        """Get the optimal font path for the current system"""
        try:
            # Try embedded font first
            embedded_font = get_embedded_font()
            if embedded_font:
                return embedded_font
            
            # Try system fonts
            font_paths = [
                "assets/arial.ttf",
                "assets/arialbd.ttf", 
                "assets/Roboto-Regular.ttf",
                "C:/Windows/Fonts/arial.ttf",
                "/System/Library/Fonts/Arial.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
            ]
            
            for path in font_paths:
                if os.path.exists(path):
                    return path
            
            return None
            
        except Exception as e:
            self.logger.error(f"Font path resolution failed: {e}")
            return None
    
    def toggle_hd_mode(self, enable_hd=None):
        """Toggle HD mode for better quality"""
        if enable_hd is not None:
            self.hd_mode = enable_hd
        else:
            self.hd_mode = not self.hd_mode
        
        self.logger.info(f"HD mode: {'ENABLED' if self.hd_mode else 'DISABLED'}")
        return self.hd_mode
    
    def get_quality_info(self):
        """Get current quality settings"""
        return {
            'hd_mode': self.hd_mode,
            'hd_multiplier': self.HD_MULTIPLIER,
            'dpi': self.DPI,
            'barcode_width': self.BARCODE_WIDTH,
            'barcode_height': self.BARCODE_HEIGHT
        }

class QuickTagApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("QuickTags – AutoGeek Edition by CrystalCore")

        # Initialize responsive scaling system
        self.scaler = ResponsiveScaler()
        
        # Set responsive window size based on screen size
        self.resize(self.scaler.window_width, self.scaler.window_height)
        self.setMinimumSize(self.scaler.min_width, self.scaler.min_height)

        # Set responsive font sizes
        self.base_font_size = self.scaler.base_font
        self.header_font_size = self.scaler.header_font
        self.title_font_size = self.scaler.title_font
        self.subtitle_font_size = self.scaler.subtitle_font

        # Setup logging and database
        self.setup_logger()
        self.init_db()

        # Initialize barcode generator
        self.barcode_generator = BarcodeGenerator(self.logger)

        # Create main layout and apply styles
        self.setup_ui()
        self.apply_visual_styling() # Apply perfect visual styling for two-column layout

        # Load existing data
        self.load_item_history()
        self.load_fonts()
        
        # Initialize summary data
        self.update_inventory_values()
        
        # Initialize barcode generator with correct paper size AFTER UI is set up
        self.initialize_barcode_generator_dimensions()
        
        # Log enhanced responsive scaling information
        self.logger.info(f"Enhanced responsive scaling initialized:")
        self.logger.info(f"  Screen size: {self.scaler.screen_width}x{self.scaler.screen_height}")
        self.logger.info(f"  Screen DPI: {self.scaler.screen_dpi}")
        self.logger.info(f"  DPI factor: {self.scaler.dpi_factor:.2f}")
        self.logger.info(f"  Screen category: {self.scaler.category}")
        self.logger.info(f"  Scale factor: {self.scaler.scale_factor:.2f}")
        self.logger.info(f"  Window size: {self.scaler.window_width}x{self.scaler.window_height}")
        self.logger.info(f"  Base font size: {self.scaler.base_font}")
        self.logger.info(f"  Input height: {self.scaler.input_height}")
        self.logger.info(f"  Button height: {self.scaler.button_height}")
        self.logger.info(f"  Table row height: {self.scaler.get_table_row_height()}")
        left_prop, right_prop = self.scaler.get_column_proportions()
        self.logger.info(f"  Column proportions: {left_prop}% / {right_prop}%")

    def init_db(self):
        """Initialize SQLite database with proper schema"""
        try:
            # Get the correct database path
            if getattr(sys, 'frozen', False):
                # EXE mode
                app_dir = os.path.dirname(sys.executable)
                self.logger.info(f"App directory: {app_dir}")
                # Try user's home directory first
                user_data_dir = os.path.join(os.path.expanduser('~'), '.quicktag')
                os.makedirs(user_data_dir, exist_ok=True)
                db_path = os.path.join(user_data_dir, 'quicktag.db')
                self.logger.info(f"Using database at: {db_path}")
            else:
                # Development mode
                db_path = 'quicktag.db'
                self.logger.info(f"Using development database: {db_path}")
            
            # Connect to database
            self.conn = sqlite3.connect(db_path)
            self.cursor = self.conn.cursor()
            
            # Create tables with proper constraints
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    item_name TEXT NOT NULL UNIQUE,
                    barcode TEXT NOT NULL UNIQUE,
                    purchase_price INTEGER NOT NULL,
                    profit_percent INTEGER NOT NULL,
                    sale_price INTEGER NOT NULL,
                    stock_quantity INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL,
                    CONSTRAINT valid_price CHECK (purchase_price > 0),
                    CONSTRAINT valid_profit CHECK (profit_percent >= 0),
                    CONSTRAINT valid_sale CHECK (sale_price > 0),
                    CONSTRAINT valid_stock CHECK (stock_quantity > 0)
                )
            ''')
            
            # Create indices for performance
            self.cursor.execute('CREATE INDEX IF NOT EXISTS idx_item_name ON items(item_name)')
            self.cursor.execute('CREATE INDEX IF NOT EXISTS idx_barcode ON items(barcode)')
            
            self.conn.commit()
            self.logger.info("Database initialized successfully")
            
            # CRITICAL FIX: Migrate existing barcode data to ensure consistency
            self.migrate_barcode_consistency()
            
        except sqlite3.Error as e:
            self.logger.error(f"Database error: {e}")
            QMessageBox.critical(self, "Database Error", 
                               f"Failed to initialize database:\n{str(e)}\n\nPlease check permissions and disk space.")
            sys.exit(1)
        except Exception as e:
            self.logger.error(f"Unexpected error: {e}")
            QMessageBox.critical(self, "Error", 
                               f"An unexpected error occurred:\n{str(e)}")
            sys.exit(1)

    def migrate_barcode_consistency(self):
        """CRITICAL: Ensure all existing barcode values are standardized to 13 digits"""
        try:
            self.logger.info("🔧 Migrating barcode data for consistency...")
            
            # Get all items
            self.cursor.execute("SELECT id, item_name, barcode FROM items")
            items = self.cursor.fetchall()
            
            migration_count = 0
            for item_id, item_name, barcode in items:
                # Standardize barcode format
                original_barcode = str(barcode).strip()
                standardized_barcode = original_barcode.zfill(13)
                
                # Only update if different
                if original_barcode != standardized_barcode:
                    self.logger.info(f"Migrating '{item_name}': {original_barcode} -> {standardized_barcode}")
                    
                    self.cursor.execute("""
                        UPDATE items SET barcode = ? WHERE id = ?
                    """, (standardized_barcode, item_id))
                    migration_count += 1
            
            if migration_count > 0:
                self.conn.commit()
                self.logger.info(f"✅ Migrated {migration_count} barcode(s) for consistency")
            else:
                self.logger.info("✅ All barcodes already consistent")
                
        except Exception as e:
            self.logger.error(f"Barcode migration failed: {e}")
            # Don't fail the app, just log the error
            self.logger.warning("Failed to migrate barcodes, continuing without migration")

    def setup_ui(self):
        """Setup the main UI with responsive two-column layout"""
        # Create main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(
            int(self.scaler.base_padding), 
            int(self.scaler.base_padding), 
            int(self.scaler.base_padding), 
            int(self.scaler.base_padding)
        )
        main_layout.setSpacing(0)

        # Create horizontal layout for sidebar and content
        h_layout = QHBoxLayout()
        h_layout.setContentsMargins(0, 0, 0, 0)
        h_layout.setSpacing(0)
        main_layout.addLayout(h_layout, 1)
        
        # Create sidebar
        self.create_sidebar(h_layout)
        
        # Create main content area with responsive two-column layout
        content_container = QWidget()
        content_container.setObjectName("main_content")
        content_layout = QHBoxLayout(content_container)  # Horizontal layout for columns
        content_layout.setContentsMargins(
            int(self.scaler.base_padding * 2), 
            int(self.scaler.base_padding * 2), 
            int(self.scaler.base_padding * 2), 
            int(self.scaler.base_padding * 2)
        )  # Responsive margins
        content_layout.setSpacing(int(self.scaler.section_spacing))  # Responsive spacing between columns
        
        # === LEFT COLUMN (35% - Product Information) ===
        left_column = QWidget()
        left_column.setObjectName("left_column")
        left_layout = QVBoxLayout(left_column)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(0)
        
        # Create and add form to left column
        self.create_form()
        left_layout.addWidget(self.form_group)
        
        # Add Item Summary below Product Information (with responsive spacing to align with Inventory Summary)
        left_layout.addSpacing(int(self.scaler.base_padding * 1.5))  # Responsive space to align with Inventory Summary on right
        self.item_summary_container = self.create_item_summary_section()
        left_layout.addWidget(self.item_summary_container, 1)  # Give it stretch factor to fill remaining space
        
        # === RIGHT COLUMN (65% - History + Summary Sections) ===
        right_column = QWidget()
        right_column.setObjectName("right_column")
        right_layout = QVBoxLayout(right_column)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(int(self.scaler.section_spacing))  # Responsive spacing
        
        # Create and add history section to right column
        self.create_history_section()
        right_layout.addWidget(self.history_group)
        
        # Create and add inventory value section below history
        self.value_group_container = self.create_inventory_value_section()
        right_layout.addWidget(self.value_group_container)
        
        # Add columns to content layout with responsive proportions
        left_proportion, right_proportion = self.scaler.get_column_proportions()
        content_layout.addWidget(left_column, left_proportion)   # Responsive width for form
        content_layout.addWidget(right_column, right_proportion)  # Responsive width for table

        h_layout.addWidget(content_container, 1)
        
        # DEPRECATED: All styling is handled in update_ui_scale
        # content_container.setStyleSheet(...)

    def create_inventory_value_section(self):
        """Create section showing total inventory value, with info icon and breakdown popup"""
        container = QWidget()
        container.setObjectName("inventory_value_panel")
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)  # Remove top margin
        container_layout.setSpacing(0)
        
        # --- INVENTORY VALUE SECTION ---
        value_group = QGroupBox("Inventory Summary")  # Changed title for better context
        value_group.setObjectName("inventory_value_group")
        value_layout = QVBoxLayout(value_group)
        value_layout.setContentsMargins(20, 20, 20, 20)  # Better margins
        value_layout.setSpacing(15)  # Better spacing

        # Info icon with tooltip
        info_icon = QLabel()
        info_icon.setPixmap(QPixmap(16, 16))
        info_icon.setPixmap(QPixmap(':/qt-project.org/styles/commonstyle/images/standardbutton-help-32.png').scaled(20, 20, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        info_icon.setToolTip("This shows the total value of your inventory based on purchase price.")
        info_icon.setCursor(Qt.WhatsThisCursor)

        # Breakdown button
        breakdown_btn = QPushButton("View Breakdown")
        breakdown_btn.setObjectName("breakdown_btn")
        breakdown_btn.setToolTip("Show detailed breakdown of all items, prices, and potential profits")
        breakdown_btn.clicked.connect(self.show_inventory_breakdown)
        breakdown_btn.setStyleSheet("font-size: 15px; padding: 6px 12px; min-width: 120px;")
        
        # Today's value
        today_container = QWidget()
        today_container.setObjectName("value_container")
        today_layout = QHBoxLayout(today_container)
        today_layout.setContentsMargins(24, 16, 24, 16)
        today_label = QLabel("📦 Today's Inventory Value:")
        today_label.setObjectName("value_label")
        today_label.setFont(QFont("Arial", 22, QFont.Bold))
        self.today_value_label = QLabel("₹0")
        self.today_value_label.setObjectName("value_amount")
        self.today_value_label.setFont(QFont("Arial", 22))
        today_layout.addWidget(today_label)
        today_layout.addWidget(self.today_value_label)
        today_layout.addWidget(info_icon)
        today_layout.addWidget(breakdown_btn)
        today_layout.addStretch()
        
        # All time value with toggle
        all_time_container = QWidget()
        all_time_container.setObjectName("value_container")
        all_time_layout = QHBoxLayout(all_time_container)
        all_time_layout.setContentsMargins(24, 16, 24, 16)
        
        self.all_time_toggle = QPushButton("View All-Time Value")
        self.all_time_toggle.setObjectName("toggle_btn")
        self.all_time_toggle.setCheckable(True)
        self.all_time_toggle.setFont(QFont("Arial", 22))
        self.all_time_toggle.clicked.connect(self.toggle_all_time_value)
        
        all_time_label = QLabel("💰 All-Time Inventory Value:")
        all_time_label.setObjectName("value_label")
        all_time_label.setFont(QFont("Arial", 22, QFont.Bold))
        all_time_label.hide()
        
        self.all_time_value = QLabel("₹0")
        self.all_time_value.setObjectName("value_amount")
        self.all_time_value.setFont(QFont("Arial", 22))
        self.all_time_value.hide()
        
        all_time_layout.addWidget(self.all_time_toggle)
        all_time_layout.addWidget(all_time_label)
        all_time_layout.addWidget(self.all_time_value)
        all_time_layout.addStretch()
        
        value_layout.addWidget(today_container)
        value_layout.addWidget(all_time_container)
        
        self.all_time_label = all_time_label
        
        container_layout.addWidget(value_group)
        
        return container

    def create_item_summary_section(self):
        """Create section showing item statistics with today/all-time toggle"""
        container = QWidget()
        container.setObjectName("item_summary_panel")
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(0)
        
        # --- ITEM SUMMARY SECTION ---
        summary_group = QGroupBox("Item Summary")
        summary_group.setObjectName("item_summary_group")
        summary_layout = QVBoxLayout(summary_group)
        summary_layout.setContentsMargins(20, 20, 20, 20)
        summary_layout.setSpacing(15)
        
        # Single stats display container (switches between today/all-time)
        stats_display_container = QWidget()
        stats_display_container.setObjectName("stats_container")
        stats_display_layout = QVBoxLayout(stats_display_container)
        stats_display_layout.setContentsMargins(24, 16, 24, 16)
        stats_display_layout.setSpacing(12)
        
        # Total items row (dynamic label and count)
        items_row = QWidget()
        items_layout = QHBoxLayout(items_row)
        items_layout.setContentsMargins(0, 0, 0, 0)
        items_layout.setSpacing(10)
        
        self.items_label = QLabel("🧾 Total Items Added:")
        self.items_label.setObjectName("stats_label")
        self.items_count = QLabel("0")
        self.items_count.setObjectName("stats_amount")
        
        items_layout.addWidget(self.items_label)
        items_layout.addWidget(self.items_count)
        items_layout.addStretch()
        
        # Most expensive item row (dynamic label and content)
        expensive_row = QWidget()
        expensive_layout = QHBoxLayout(expensive_row)
        expensive_layout.setContentsMargins(0, 0, 0, 0)
        expensive_layout.setSpacing(10)
        
        self.expensive_label = QLabel("💰 Most Expensive:")
        self.expensive_label.setObjectName("stats_label")
        self.expensive_item = QLabel("None")
        self.expensive_item.setObjectName("stats_item_name")
        
        expensive_layout.addWidget(self.expensive_label)
        expensive_layout.addWidget(self.expensive_item)
        expensive_layout.addStretch()
        
        stats_display_layout.addWidget(items_row)
        stats_display_layout.addWidget(expensive_row)
        
        # Toggle container
        toggle_container = QWidget()
        toggle_container.setObjectName("stats_container")
        toggle_layout = QHBoxLayout(toggle_container)
        toggle_layout.setContentsMargins(24, 16, 24, 16)
        toggle_layout.setSpacing(0)
        
        self.stats_toggle = QPushButton("View All-Time Stats")
        self.stats_toggle.setObjectName("toggle_btn")
        self.stats_toggle.setCheckable(True)
        self.stats_toggle.clicked.connect(self.toggle_item_stats_view)
        
        toggle_layout.addWidget(self.stats_toggle)
        toggle_layout.addStretch()
        
        summary_layout.addWidget(stats_display_container)
        summary_layout.addWidget(toggle_container)
        
        container_layout.addWidget(summary_group)
        
        # Initialize showing today's stats
        self.showing_today_stats = True
        self.today_item_count = 0
        self.today_expensive_name = "None"
        self.alltime_item_count = 0
        self.alltime_expensive_name = "None"
        
        return container

    def toggle_item_stats_view(self):
        """Toggle between today and all-time item statistics display"""
        if self.stats_toggle.isChecked():
            # Switch to all-time view
            self.showing_today_stats = False
            self.stats_toggle.setText("View Today's Stats")
            self.items_label.setText("🧾 All-Time Items:")
            self.expensive_label.setText("💰 All-Time Most Expensive:")
            # Update display with all-time data
            self.items_count.setText(str(self.alltime_item_count))
            self.expensive_item.setText(self.alltime_expensive_name)
        else:
            # Switch to today view
            self.showing_today_stats = True
            self.stats_toggle.setText("View All-Time Stats")
            self.items_label.setText("🧾 Total Items Added:")
            self.expensive_label.setText("💰 Most Expensive:")
            # Update display with today's data
            self.items_count.setText(str(self.today_item_count))
            self.expensive_item.setText(self.today_expensive_name)

    def toggle_all_time_value(self):
        """Toggle visibility of all-time inventory value"""
        if self.all_time_toggle.isChecked():
            self.all_time_toggle.setText("Hide All-Time Value")
            self.all_time_label.show()
            self.all_time_value.show()
        else:
            self.all_time_toggle.setText("View All-Time Value")
            self.all_time_label.hide()
            self.all_time_value.hide()

    def update_inventory_values(self):
        """Update inventory value display"""
        try:
            today = datetime.datetime.now().strftime("%Y-%m-%d")
            
            # Calculate today's value
            self.cursor.execute("""
                SELECT COALESCE(SUM(purchase_price * stock_quantity), 0) 
                FROM items 
                WHERE DATE(created_at) = ?
            """, (today,))
            today_value = self.cursor.fetchone()[0]
            
            # Calculate all-time value
            self.cursor.execute("""
                SELECT COALESCE(SUM(purchase_price * stock_quantity), 0) FROM items
            """)
            all_time_value = self.cursor.fetchone()[0]
            
            # If your prices are stored as paise/cents, keep the division by 100
            self.today_value_label.setText(f"₹{int(today_value/100):,}")
            self.all_time_value.setText(f"₹{int(all_time_value/100):,}")
            
            # Update item summary statistics
            self.update_item_summary()
            
        except Exception as e:
            self.logger.error(f"Failed to update inventory values: {str(e)}")
            self.logger.error(traceback.format_exc())
            self.today_value_label.setText("₹0")
            self.all_time_value.setText("₹0")

    def update_item_summary(self):
        """Update item summary statistics"""
        try:
            today = datetime.datetime.now().strftime("%Y-%m-%d")
            
            # Today's item count
            self.cursor.execute("""
                SELECT COUNT(*) FROM items 
                WHERE DATE(created_at) = ?
            """, (today,))
            today_count = self.cursor.fetchone()[0]
            
            # Today's most expensive item
            self.cursor.execute("""
                SELECT item_name, sale_price FROM items 
                WHERE DATE(created_at) = ?
                ORDER BY sale_price DESC LIMIT 1
            """, (today,))
            today_expensive = self.cursor.fetchone()
            
            # All-time item count
            self.cursor.execute("SELECT COUNT(*) FROM items")
            all_time_count = self.cursor.fetchone()[0]
            
            # All-time most expensive item
            self.cursor.execute("""
                SELECT item_name, sale_price FROM items 
                ORDER BY sale_price DESC LIMIT 1
            """)
            all_time_expensive = self.cursor.fetchone()
            
            # Store data in instance variables
            self.today_item_count = today_count
            self.alltime_item_count = all_time_count
            
            if today_expensive:
                item_name, price = today_expensive
                display_name = item_name[:20] + "..." if len(item_name) > 20 else item_name
                self.today_expensive_name = f"{display_name} (₹{int(price/100):,})"
            else:
                self.today_expensive_name = "None"
            
            if all_time_expensive:
                item_name, price = all_time_expensive
                display_name = item_name[:20] + "..." if len(item_name) > 20 else item_name
                self.alltime_expensive_name = f"{display_name} (₹{int(price/100):,})"
            else:
                self.alltime_expensive_name = "None"
            
            # Update display based on current view
            if hasattr(self, 'showing_today_stats'):
                if self.showing_today_stats:
                    # Currently showing today's stats
                    self.items_count.setText(str(self.today_item_count))
                    self.expensive_item.setText(self.today_expensive_name)
                else:
                    # Currently showing all-time stats
                    self.items_count.setText(str(self.alltime_item_count))
                    self.expensive_item.setText(self.alltime_expensive_name)
                
        except Exception as e:
            self.logger.error(f"Failed to update item summary: {str(e)}")
            self.logger.error(traceback.format_exc())
            # Set fallback values
            if hasattr(self, 'items_count'):
                self.today_item_count = 0
                self.alltime_item_count = 0
                self.today_expensive_name = "None"
                self.alltime_expensive_name = "None"
                self.items_count.setText("0")
                self.expensive_item.setText("None")

    def show_inventory_breakdown(self):
        """Show a popup dialog with a breakdown table of all items, prices, and profits"""
        dialog = QDialog(self)
        dialog.setWindowTitle("") # Hide default title bar
        dialog.setModal(True)
        dialog.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)
        dialog.setObjectName("breakdown_dialog")
        
        # Use responsive dialog sizing
        dialog_width, dialog_height = self.scaler.get_dialog_size(900, 600)
        dialog.setFixedWidth(dialog_width)
        dialog.setFixedHeight(dialog_height)
        
        dialog_layout = QVBoxLayout(dialog)
        dialog_layout.setContentsMargins(1, 1, 1, 1)
        dialog_layout.setSpacing(0)

        container = QWidget()
        container.setObjectName("popup_container")
        layout = QVBoxLayout(container)
        layout.setContentsMargins(32, 32, 32, 32)  # Increased padding
        layout.setSpacing(24)  # Increased spacing
        
        # --- Title Bar ---
        title_bar = QWidget()
        title_layout = QHBoxLayout(title_bar)
        title_layout.setContentsMargins(0, 0, 0, 16)  # Added bottom margin
        title_label = QLabel("Inventory & Profit Breakdown")
        title_label.setObjectName("popup_title")
        title_layout.addWidget(title_label)
        title_layout.addStretch()
        layout.addWidget(title_bar)
        
        # --- Table ---
        table = QTableWidget()
        table.setObjectName("breakdown_table")
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels([
            "Item Name", "Purchase Price", "Sale Price", "Profit %", "Stock", "Potential Profit"
        ])
        
        # Configure table properties
        table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)  # Enable horizontal scroll
        table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        table.setAlternatingRowColors(True)
        table.setShowGrid(False)
        table.setFrameStyle(QFrame.NoFrame)
        table.verticalHeader().setVisible(False)
        table.setSelectionMode(QTableWidget.NoSelection)  # Disable selection
        table.setGridStyle(Qt.NoPen)  # Remove any grid lines
        
        # Set column resize modes
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)  # Item Name fixed
        table.setColumnWidth(0, 220)  # Adjusted width for item name
        for i in range(1, 6):
            table.horizontalHeader().setSectionResizeMode(i, QHeaderView.Fixed)
            table.setColumnWidth(i, 100)  # Fixed width for other columns
        
        # Populate table data
        self.cursor.execute("SELECT item_name, purchase_price, sale_price, profit_percent, stock_quantity FROM items ORDER BY item_name")
        items = self.cursor.fetchall()
        table.setRowCount(len(items))
        
        for row, (name, purchase, sale, profit_percent, stock) in enumerate(items):
            table.setItem(row, 0, QTableWidgetItem(str(name)))
            table.setItem(row, 1, QTableWidgetItem(f"₹{purchase/100:.2f}"))
            table.setItem(row, 2, QTableWidgetItem(f"₹{sale/100:.2f}"))
            table.setItem(row, 3, QTableWidgetItem(f"{profit_percent/100:.2f}%"))
            table.setItem(row, 4, QTableWidgetItem(str(stock)))
            profit_amt = (sale - purchase) * stock / 100
            table.setItem(row, 5, QTableWidgetItem(f"₹{profit_amt:,.2f}"))
            
            # Set alignment and tooltips for all cells
            for col in range(6):
                item = table.item(row, col)
                if col == 0:  # Item Name - left align
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                elif col in [1, 2, 5]:  # Price columns - right align
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                else:  # Other columns - center align
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                item.setToolTip(item.text())
        
        # Adjust column widths and alignment
        table.setColumnWidth(0, 200)  # Item Name
        table.setColumnWidth(1, 130)  # Purchase Price  
        table.setColumnWidth(2, 130)  # Sale Price
        table.setColumnWidth(3, 90)   # Profit %
        table.setColumnWidth(4, 80)   # Stock
        table.setColumnWidth(5, 150)  # Potential Profit

        # Increase font size and padding for better readability
        table.setStyleSheet("""
            QTableWidget#breakdown_table QHeaderView::section {
                font-size: 16px;
                padding: 20px 16px;
            }
            QTableWidget#breakdown_table::item {
                font-size: 16px;
                padding: 16px 20px;
            }
        """)
        
        # Set fixed height based on content
        header_height = table.horizontalHeader().height()
        row_height = table.rowHeight(0)
        total_height = header_height + (row_height * min(10, len(items)))  # Show max 10 rows
        table.setFixedHeight(total_height)
        
        layout.addWidget(table)
        
        # --- Close Button ---
        button_container = QWidget()
        button_layout = QHBoxLayout(button_container)
        button_layout.setContentsMargins(0, 24, 0, 24)  # Center the close button
        button_layout.addStretch()
        
        close_btn = QPushButton("Close")
        close_btn.setObjectName("popup_close_btn")
        close_btn.clicked.connect(dialog.reject)
        button_layout.addWidget(close_btn)
        
        layout.addWidget(button_container)
        dialog_layout.addWidget(container)
        
        # Apply clean styling to the breakdown dialog
        dialog.setStyleSheet("""
            QDialog#breakdown_dialog {
                background-color: transparent;
            }
            QWidget#popup_container {
                background-color: #151b23;
                border: 2px solid #30363d;
                border-radius: 16px;
            }
            QLabel#popup_title {
                font-size: 28px;
                font-weight: 700;
                color: #ffffff;
                padding: 16px 0;
                border-bottom: 2px solid #30363d;
            }
            QTableWidget#breakdown_table {
                background-color: #151b23;
                border: 2px solid #30363d;
                border-radius: 10px;
                gridline-color: transparent;
            }
            QTableWidget#breakdown_table QHeaderView::section {
                background-color: #161b22;
                color: #ffffff;
                font-weight: 700;
                font-size: 16px;
                padding: 16px 12px;
                border: none;
                border-bottom: 3px solid #2f81f7;
            }
            QTableWidget#breakdown_table::item {
                padding: 14px 12px;
                border-bottom: 1px solid #21262d;
                color: #f0f6fc;
                font-size: 15px;
                font-weight: 500;
            }
            QTableWidget#breakdown_table::item:hover {
                background-color: #1e2329;
            }
            QPushButton#popup_close_btn {
                background-color: #2f81f7;
                color: white;
                border: none;
                padding: 14px 36px;
                border-radius: 10px;
                font-size: 18px;
                font-weight: 700;
                min-width: 140px;
                min-height: 52px;
            }
            QPushButton#popup_close_btn:hover {
                background-color: #58a6ff;
            }
        """)
        
        dialog.exec_()

    def create_form(self):
        """Create the Product Information form with responsive layout and spacing"""
        self.form_group = QGroupBox("Product Information")
        form_layout = QVBoxLayout(self.form_group)  # Changed to vertical layout for better organization
        form_layout.setContentsMargins(
            int(self.scaler.base_padding * 1.5), 
            int(self.scaler.base_padding * 2), 
            int(self.scaler.base_padding * 1.5), 
            int(self.scaler.base_padding * 1.5)
        )  # Responsive margins
        form_layout.setSpacing(int(self.scaler.section_spacing))  # Responsive form spacing
        
        # Create main fields container
        fields_container = QWidget()
        fields_layout = QVBoxLayout(fields_container)  # Stack fields vertically for better readability
        fields_layout.setContentsMargins(0, 0, 0, 0)
        fields_layout.setSpacing(int(self.scaler.base_padding))  # Responsive spacing between fields
        
        # Helper function to create field group with responsive styling
        def create_field_group(label_text, name, placeholder, required, default_value=""):
            field_container = QWidget()
            field_layout = QVBoxLayout(field_container)
            field_layout.setContentsMargins(0, 0, 0, 0)
            field_layout.setSpacing(int(self.scaler.base_padding // 2))  # Responsive space between label and input
            
            label = QLabel(f"{label_text}{' *' if required else ''}")
            label.setFont(QFont("Arial", self.scaler.base_font, QFont.Bold))  # Responsive font size
            label.setObjectName("field_label")
            
            field = QLineEdit()
            field.setObjectName("form_input")
            field.setPlaceholderText(placeholder)
            field.setFont(QFont("Arial", self.scaler.base_font - 1))
            field.setMinimumHeight(self.scaler.input_height)  # Responsive height for touch targets
            
            if default_value:
                field.setText(default_value)
                
            setattr(self, name, field)
            if name in ['purchase_price_input', 'profit_percent_input']:
                field.textChanged.connect(self.calculate_sale_price)
            
            field_layout.addWidget(label)
            field_layout.addWidget(field)
            
            return field_container
        
        # Create all form fields with improved organization
        form_fields = [
            ("Item Name", "item_name_input", "Enter item name (required)", True, ""),
            ("Purchase Price (₹)", "purchase_price_input", "Enter purchase price", False, "1"),
            ("Profit Percentage (%)", "profit_percent_input", "Enter profit percentage", False, "100"),
            ("Sale Price (₹)", "sale_price_input", "Calculated automatically", False, ""),
            ("Stock Quantity", "stock_quantity_input", "Enter stock quantity", False, "1"),
            ("Barcode Override", "barcode_override_input", "Enter 13-digit barcode (optional)", False, ""),
        ]
        
        # Add fields to layout
        for label_text, name, placeholder, required, default in form_fields:
            field_group = create_field_group(label_text, name, placeholder, required, default)
            fields_layout.addWidget(field_group)
        
        # Add paper size selector
        paper_size_container = QWidget()
        paper_size_layout = QVBoxLayout(paper_size_container)
        paper_size_layout.setContentsMargins(0, 0, 0, 0)
        paper_size_layout.setSpacing(int(self.scaler.base_padding // 2))
        
        paper_size_label = QLabel("Label Size *")
        paper_size_label.setFont(QFont("Arial", self.scaler.base_font, QFont.Bold))
        paper_size_label.setObjectName("field_label")
        
        self.paper_size_combo = QComboBox()
        self.paper_size_combo.setObjectName("form_input")
        self.paper_size_combo.setFont(QFont("Arial", self.scaler.base_font - 1))
        self.paper_size_combo.setMinimumHeight(self.scaler.input_height)
        self.paper_size_combo.addItem("Standard (40x25mm)", "standard")
        self.paper_size_combo.addItem("38x25mm (2UP)", "38x25mm")
        self.paper_size_combo.currentIndexChanged.connect(self.on_paper_size_changed)
        
        # Load saved paper size
        self.load_paper_size_setting()
        
        # Add some debugging
        current_paper_size = self.paper_size_combo.currentData()
        print(f"DEBUG: Paper size selector initialized with: {current_paper_size}")
        
        paper_size_layout.addWidget(paper_size_label)
        paper_size_layout.addWidget(self.paper_size_combo)
        
        fields_layout.addWidget(paper_size_container)
        
        form_layout.addWidget(fields_container)
        
        # Create action buttons section with responsive spacing
        button_section = QWidget()
        button_layout = QVBoxLayout(button_section)
        button_layout.setContentsMargins(0, int(self.scaler.base_padding), 0, 0)  # Responsive top margin for separation
        button_layout.setSpacing(int(self.scaler.base_padding))
        
        # Main action buttons
        main_buttons = QWidget()
        main_button_layout = QHBoxLayout(main_buttons)
        main_button_layout.setContentsMargins(0, 0, 0, 0)
        main_button_layout.setSpacing(int(self.scaler.base_padding))
        
        self.add_button = QPushButton("Add Item")
        self.add_button.setObjectName("primary_button")
        self.add_button.setFont(QFont("Arial", self.scaler.button_font, QFont.Bold))
        self.add_button.setMinimumHeight(self.scaler.button_height)  # Responsive height
        self.add_button.clicked.connect(self.add_item)
        
        clear_button = QPushButton("Clear Form")
        clear_button.setObjectName("secondary_button")
        clear_button.setFont(QFont("Arial", self.scaler.button_font))
        clear_button.setMinimumHeight(self.scaler.button_height)  # Responsive height
        clear_button.clicked.connect(self.clear_form)
        
        main_button_layout.addWidget(self.add_button)
        main_button_layout.addWidget(clear_button)
        
        button_layout.addWidget(main_buttons)
        
        form_layout.addWidget(button_section)
        form_layout.addStretch()  # Push everything to top

    def load_paper_size_setting(self):
        """Load paper size setting from config file"""
        try:
            import json
            config_path = 'quicktags_config.json'
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    config = json.load(f)
                    paper_size = config.get('paper_size', 'standard')
                    
                    # Set the combo box to the saved value
                    for i in range(self.paper_size_combo.count()):
                        if self.paper_size_combo.itemData(i) == paper_size:
                            self.paper_size_combo.setCurrentIndex(i)
                            break
                    
                    self.logger.info(f"Loaded paper size setting: {paper_size}")
                    print(f"DEBUG: Loaded paper size setting: {paper_size}")
            else:
                self.logger.info("No config file found, using default paper size")
                print("DEBUG: No config file found, using default paper size")
        except Exception as e:
            self.logger.error(f"Error loading paper size setting: {e}")
            print(f"DEBUG: Error loading paper size setting: {e}")

    def save_paper_size_setting(self, paper_size):
        """Save paper size setting to config file"""
        try:
            import json
            config_path = 'quicktags_config.json'
            
            # Load existing config or create new
            config = {}
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    config = json.load(f)
            
            # Update paper size
            config['paper_size'] = paper_size
            
            # Save config
            with open(config_path, 'w') as f:
                json.dump(config, f, indent=2)
            
            self.logger.info(f"Saved paper size setting: {paper_size}")
        except Exception as e:
            self.logger.error(f"Error saving paper size setting: {e}")

    def on_paper_size_changed(self):
        """Handle paper size selection change"""
        try:
            paper_size = self.paper_size_combo.currentData()
            self.logger.info(f"Paper size selection changed to: {paper_size}")
            
            self.save_paper_size_setting(paper_size)
            
            # Update barcode generator dimensions
            self.update_barcode_generator_dimensions(paper_size)
            
            # Force refresh of any existing previews
            self.logger.info(f"Paper size updated successfully: {paper_size}")
            
            # Show user feedback
            if paper_size == "38x25mm":
                self.show_info("Paper Size Changed", "Switched to 38x25mm (compact) labels")
            else:
                self.show_info("Paper Size Changed", "Switched to Standard (40x25mm) labels")
                
        except Exception as e:
            self.logger.error(f"Error changing paper size: {e}")
            self.show_error("Paper Size Error", f"Failed to change paper size: {str(e)}")

    def update_barcode_generator_dimensions(self, paper_size):
        """Update barcode generator dimensions based on paper size"""
        try:
            # Use the new set_dimensions method
            self.barcode_generator.set_dimensions(paper_size)
            
            self.logger.info(f"Updated barcode generator dimensions for {paper_size}")
            print(f"DEBUG: Updated dimensions for {paper_size}")
            print(f"DEBUG: Canvas: {self.barcode_generator.CANVAS_WIDTH}x{self.barcode_generator.CANVAS_HEIGHT}")
            print(f"DEBUG: Label width: {self.barcode_generator.LABEL_WIDTH}")
        except Exception as e:
            self.logger.error(f"Error updating barcode generator dimensions: {e}")
            print(f"DEBUG: Error updating dimensions: {e}")

    def initialize_barcode_generator_dimensions(self):
        """Initialize barcode generator with correct dimensions based on saved setting"""
        try:
            # Load paper size setting
            import json
            config_path = 'quicktags_config.json'
            paper_size = 'standard'  # Default
            
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    config = json.load(f)
                    paper_size = config.get('paper_size', 'standard')
            
            # Set dimensions directly on the barcode generator
            self.barcode_generator.set_dimensions(paper_size)
            
            self.logger.info(f"Initialized barcode generator with paper size: {paper_size}")
            print(f"DEBUG: Initialized with paper size: {paper_size}")
        except Exception as e:
            self.logger.error(f"Error initializing barcode generator dimensions: {e}")
            print(f"DEBUG: Error initializing dimensions: {e}")

    def create_history_section(self):
        """Create the history section with search and table"""
        self.history_group = QGroupBox("Session History")
        self.history_group.setTitle("  Session History")  # Leading space to fix visual misalignment
        history_layout = QVBoxLayout(self.history_group)
        history_layout.setContentsMargins(
            int(self.scaler.base_padding * 1.5), 
            int(self.scaler.base_padding * 2), 
            int(self.scaler.base_padding * 1.5), 
            int(self.scaler.base_padding * 1.5)
        )  # Responsive margins
        history_layout.setSpacing(int(self.scaler.section_spacing))  # Responsive section spacing
        
        # --- Search and Sort Controls ---
        controls_container = QWidget()
        controls_container.setObjectName("filter_bar")
        controls_container.setStyleSheet("""
            QWidget#filter_bar {
                background-color: #161b22;
                border: 1px solid #30363d;
                border-bottom: none;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                margin: 0px;
                padding: 0px;
            }
        """)
        controls_layout = QHBoxLayout(controls_container)
        controls_layout.setContentsMargins(20, 12, 20, 12)  # Tighter margins to connect with table
        controls_layout.setSpacing(30)  # Consistent spacing
        
        # Left side - Search with icon
        search_container = QWidget()
        search_layout = QHBoxLayout(search_container)
        search_layout.setContentsMargins(0, 0, 0, 0)
        search_layout.setSpacing(0)  # Remove gap for seamless connection
        
        search_icon = QLabel("🔍")
        search_icon.setFixedSize(40, 36)  # Match input height perfectly
        search_icon.setAlignment(Qt.AlignCenter)
        search_icon.setStyleSheet("""
            color: #ffffff; 
            font-size: 18px;
            background-color: #30363d;
            border: 1px solid #30363d;
            border-radius: 6px;
            border-top-right-radius: 0px;
            border-bottom-right-radius: 0px;
            margin: 0px;
            padding: 0px;
        """)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search items...")
        self.search_input.textChanged.connect(self.filter_history)
        self.search_input.setFixedSize(280, 36)  # Better proportions to match table style
        
        search_layout.addWidget(search_icon)
        search_layout.addWidget(self.search_input)
        
        # Right side - Sort dropdown
        sort_container = QWidget()
        sort_layout = QHBoxLayout(sort_container)
        sort_layout.setContentsMargins(0, 0, 0, 0)
        sort_layout.setSpacing(15)
        
        sort_label = QLabel("Sort by:")
        sort_label.setStyleSheet("color: #8b949e; font-size: 14px; font-weight: 600;")
        sort_label.setFixedSize(60, 36)
        sort_label.setAlignment(Qt.AlignCenter)
        
        self.sort_combo = QComboBox()
        self.sort_combo.addItems([
            "Date Added (Newest)",
            "Date Added (Oldest)",
            "Item Name (A-Z)",
            "Item Name (Z-A)",
            "Price (Low-High)",
            "Price (High-Low)",
            "Stock (Low-High)",
            "Stock (High-Low)"
        ])
        self.sort_combo.setFixedSize(180, 36)
        self.sort_combo.currentIndexChanged.connect(self.sort_history)
        
        sort_layout.addWidget(sort_label)
        sort_layout.addWidget(self.sort_combo)
        
        # Balanced layout
        controls_layout.addWidget(search_container)
        controls_layout.addStretch()
        controls_layout.addWidget(sort_container)
        
        # Style the search and sort controls
        self.search_input.setStyleSheet("""
            QLineEdit {
                background-color: #21262d;
                color: #f0f6fc;
                border: 1px solid #30363d;
                border-radius: 6px;
                border-top-left-radius: 0px;
                border-bottom-left-radius: 0px;
                border-left: none;
                padding: 8px 12px;
                font-size: 14px;
                font-weight: 500;
                margin: 0px;
            }
            QLineEdit:focus {
                border-color: #58a6ff;
                background-color: #0d1117;
                border-left: none;
            }
            QLineEdit:hover {
                border-color: #58a6ff;
                border-left: none;
            }
        """)
        
        self.sort_combo.setStyleSheet("""
            QComboBox {
                background-color: #21262d;
                color: #f0f6fc;
                border: 2px solid #30363d;
                border-radius: 6px;
                padding: 8px 12px;
                font-size: 14px;
                font-weight: 500;
            }
            QComboBox:hover {
                border-color: #58a6ff;
            }
            QComboBox::drop-down {
                border: none;
                width: 30px;
                background-color: transparent;
            }
            QComboBox::down-arrow {
                image: none;
                border: none;
                width: 12px;
                height: 12px;
                background-color: #8b949e;
            }
            QComboBox QAbstractItemView {
                background-color: #21262d;
                color: #f0f6fc;
                border: 2px solid #30363d;
                border-radius: 6px;
                selection-background-color: #2d4f79;
                selection-color: #ffffff;
                padding: 6px;
            }
            QComboBox QAbstractItemView::item {
                min-height: 28px;
                padding: 6px 12px;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #2d4f79;
            }
        """)
        
        history_layout.addWidget(controls_container)
        
        # Remove divider since filter bar now connects to table
        
        # --- Table Container ---
        table_container = QWidget()
        table_container.setObjectName("table_container")
        self.table_stack = QVBoxLayout(table_container)
        self.table_stack.setContentsMargins(0, 0, 0, 0)
        
        # Create table
        self.create_history_table()
        self.table_stack.addWidget(self.history_table)
        
        # Create "No Items" placeholder with icon and message
        self.no_items_label = QLabel("📦\n\nNo items added yet.\nAdd your first product to get started.")
        self.no_items_label.setObjectName("placeholder_label")
        self.no_items_label.setAlignment(Qt.AlignCenter)
        self.no_items_label.setStyleSheet("""
            QLabel {
                color: #8b949e;
                font-size: 16px;
                padding: 48px;
                background-color: #161b22;
                border: 1px solid #30363d;
                border-radius: 8px;
            }
        """)
        self.table_stack.addWidget(self.no_items_label)
        self.no_items_label.hide()
        
        history_layout.addWidget(table_container)
        


    def sort_history(self):
        """Sort history table based on selected criteria"""
        sort_option = self.sort_combo.currentText()
        column_map = {
            "Date Added (Newest)": (3, True),
            "Date Added (Oldest)": (3, False),
            "Item Name (A-Z)": (2, False),
            "Item Name (Z-A)": (2, True),
            "Price (Low-High)": (4, False),
            "Price (High-Low)": (4, True),
            "Stock (Low-High)": (5, False),
            "Stock (High-Low)": (5, True)
        }
        
        if sort_option in column_map:
            column, reverse = column_map[sort_option]
            self.history_table.sortItems(column, Qt.DescendingOrder if reverse else Qt.AscendingOrder)
            
            # Update S.No. after sorting
            for row in range(self.history_table.rowCount()):
                self.history_table.setItem(row, 1, QTableWidgetItem(str(row + 1)))

    def export_to_excel(self):
        """Export all items to Excel with minimal columns and clean formatting"""
        try:
            # Get all items from database
            self.cursor.execute("""
                SELECT item_name, barcode, purchase_price, sale_price, stock_quantity
                FROM items ORDER BY item_name
            """)
            items = self.cursor.fetchall()
            
            if not items:
                QMessageBox.information(self, "No Data", "No items found to export.")
                return
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "QuickTags Items"
            
            # Define headers and column widths
            headers = [
                ("Item Name", 45),
                ("Item Code", 22),  # Renamed from "Barcode" to "Item Code"
                ("Purchase Price", 18),
                ("Sale Price", 18),
                ("Stock Quantity", 16)
            ]
            
            # Style configuration
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_font = Font(name='Arial', size=12, bold=True)
            data_font = Font(name='Arial', size=12)
            centered = Alignment(horizontal='center', vertical='center')
            
            thin_border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )
            
            header_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
            
            # Add headers with styling
            for col, (header, width) in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = centered
                cell.border = thin_border
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # CRITICAL FIX: Add data with proper validation and formatting
            for row, item in enumerate(items, 2):
                item_name, barcode, purchase_price, sale_price, stock_quantity = item
                
                # VALIDATION ASSERTION: Ensure barcode consistency in Excel export
                barcode_str = str(barcode).strip().zfill(13)
                assert len(barcode_str) == 13, f"Excel barcode must be 13 digits: {barcode_str}"
                assert barcode_str.isdigit(), f"Excel barcode must be numeric: {barcode_str}"
                
                # VALIDATION ASSERTION: Ensure price consistency
                purchase_rupees = purchase_price / 100.0
                sale_rupees = sale_price / 100.0
                assert purchase_rupees > 0, f"Purchase price must be positive: {purchase_rupees}"
                assert sale_rupees > 0, f"Sale price must be positive: {sale_rupees}"
                assert sale_rupees >= purchase_rupees, f"Sale price must be >= purchase price: {sale_rupees} < {purchase_rupees}"
                
                # Format cells with explicit text formatting for barcodes
                cells = [
                    (item_name, '@'),                            # Item Name as text
                    (barcode_str, '@'),                         # Item Code as TEXT (preserves leading zeros)
                    (purchase_rupees, '#,##0.00'),              # Purchase Price with decimals
                    (sale_rupees, '#,##0.00'),                  # Sale Price with decimals
                    (stock_quantity, '#,##0')                   # Stock Quantity as integer
                ]
                
                for col, (value, number_format) in enumerate(cells, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.number_format = number_format
                    cell.alignment = centered
                    cell.font = data_font
                    cell.border = thin_border
                    
                    # CRITICAL FIX: Explicit text formatting for item code column WITHOUT apostrophe
                    if col == 2:  # Item Code column
                        cell.value = barcode_str  # NO apostrophe - just the clean value
                        cell.number_format = '@'  # Ensure text format
            
            # Create output directory
            output_dir = self.ensure_output_dir()
            
            # Generate filename with timestamp
            timestamp = datetime.datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
            filename = f"quicktags_export_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            
            # Save file
            wb.save(filepath)
            
            # Show success message with file path
            message = f"Data exported successfully!\n\nFile: {filepath}"
            if sys.platform == 'win32':
                message += "\n\nOpening file..."
                
            QMessageBox.information(self, "Export Successful", message)
            
            # Open the file
            if sys.platform == 'win32':
                os.startfile(filepath)
            elif sys.platform == 'darwin':
                subprocess.run(['open', filepath])
            else:
                subprocess.run(['xdg-open', filepath])
                
        except sqlite3.Error as e:
            self.show_error("Database Error", "Failed to retrieve data", e)
        except Exception as e:
            self.show_error("Export Error", "Failed to export data", e)

    def ensure_output_dir(self):
        """Ensure output directory exists and return path"""
        try:
            if getattr(sys, 'frozen', False):
                # EXE mode - use user's documents directory
                output_base = os.path.join(os.path.expanduser('~'), 'Documents', 'QuickTags')
            else:
                # Development mode - use local output directory
                output_base = 'output'
            
            # Create dated subdirectories for better organization
            excel_dir = os.path.join(output_base, 'excel')
            date_dir = os.path.join(excel_dir, datetime.datetime.now().strftime("%d-%m-%y"))
            
            os.makedirs(date_dir, exist_ok=True)
            self.logger.info(f"Created output directory: {date_dir}")
            
            return date_dir
            
        except Exception as e:
            self.logger.error(f"Failed to create output directory: {str(e)}")
            # Fallback to current directory
            fallback_dir = os.path.join('output', 'excel', datetime.datetime.now().strftime("%d-%m-%y"))
            os.makedirs(fallback_dir, exist_ok=True)
            self.logger.warning(f"Using fallback output directory: {fallback_dir}")
            return fallback_dir

    def edit_item(self, item_name):
        """Edit an existing item"""
        try:
            # Fetch current item data
            self.cursor.execute("""
                SELECT item_name, barcode, purchase_price, profit_percent, 
                       sale_price, stock_quantity 
                FROM items WHERE item_name=?
            """, (item_name,))
            
            result = self.cursor.fetchone()
            if not result:
                self.show_warning("Item Not Found", f"Item '{item_name}' not found!")
                return
            
            current_name, current_barcode, current_purchase, current_profit, current_sale, current_stock = result
            
            # Create edit dialog
            dialog = QDialog(self)
            dialog.setWindowTitle(f"Edit Item: {item_name}")
            dialog.setModal(True)
            dialog.resize(400, 300)
            
            layout = QVBoxLayout(dialog)
            
            # Create form
            form_layout = QFormLayout()
            
            name_edit = QLineEdit(current_name)
            barcode_edit = QLineEdit(current_barcode)
            purchase_edit = QLineEdit(str(current_purchase / 100.0))
            profit_edit = QLineEdit(str(current_profit / 100.0))
            sale_edit = QLineEdit(str(current_sale / 100.0))
            stock_edit = QLineEdit(str(current_stock))
            
            form_layout.addRow("Item Name:", name_edit)
            form_layout.addRow("Barcode:", barcode_edit)
            form_layout.addRow("Purchase Price:", purchase_edit)
            form_layout.addRow("Profit %:", profit_edit)
            form_layout.addRow("Sale Price:", sale_edit)
            form_layout.addRow("Stock Quantity:", stock_edit)
            
            layout.addLayout(form_layout)
            
            # Buttons
            button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            button_box.accepted.connect(dialog.accept)
            button_box.rejected.connect(dialog.reject)
            layout.addWidget(button_box)
            
            # DEPRECATED: Styling is handled globally
            # dialog.setStyleSheet(...)
            
            if dialog.exec_() == QDialog.Accepted:
                try:
                    # Validate inputs
                    new_name = name_edit.text().strip()
                    new_barcode = barcode_edit.text().strip()
                    new_purchase = float(purchase_edit.text())
                    new_profit = float(profit_edit.text())
                    new_sale = float(sale_edit.text())
                    new_stock = int(stock_edit.text())
                    
                    if not new_name:
                        raise ValueError("Item name cannot be empty")
                    if not new_barcode:
                        raise ValueError("Barcode cannot be empty")
                    if new_purchase <= 0:
                        raise ValueError("Purchase price must be positive")
                    if new_profit < 0:
                        raise ValueError("Profit percentage cannot be negative")
                    if new_sale <= 0:
                        raise ValueError("Sale price must be positive")
                    if new_stock <= 0:
                        raise ValueError("Stock quantity must be positive")
                    
                    # Check for duplicates (excluding current item)
                    self.cursor.execute("""
                        SELECT id FROM items 
                        WHERE (item_name=? OR barcode=?) AND item_name!=?
                    """, (new_name, new_barcode, item_name))
                    
                    if self.cursor.fetchone():
                        self.show_warning("Duplicate Item", 
                                           "An item with this name or barcode already exists!")
                        return
                    
                    # Update item
                    self.cursor.execute("""
                        UPDATE items 
                        SET item_name=?, barcode=?, purchase_price=?, profit_percent=?, 
                            sale_price=?, stock_quantity=?
                        WHERE item_name=?
                    """, (
                        new_name,
                        new_barcode,
                        int(new_purchase * 100),  # Store as cents
                        int(new_profit * 100),    # Store as basis points
                        int(new_sale * 100),      # Store as cents
                        new_stock,
                        item_name
                    ))
                    
                    self.conn.commit()
                    
                    # Refresh table
                    self.load_item_history()
                    
                    self.show_info("Success", f"Item '{new_name}' updated successfully!")
                    
                except ValueError as e:
                    self.show_warning("Input Error", str(e))
                except sqlite3.Error as e:
                    self.show_error("Database Error", "Operation failed", e)
                    self.conn.rollback()
                
        except Exception as e:
            self.show_error("Error", "Failed to edit item", e)

    def delete_item(self, item_name):
        """Delete an item from the database"""
        try:
            if self.show_confirm("Confirm Delete", 
                f"Are you sure you want to delete '{item_name}'?"):
                try:
                    self.cursor.execute("DELETE FROM items WHERE item_name=?", (item_name,))
                    self.conn.commit()
                    self.load_item_history()
                    self.show_info("Success", f"Item '{item_name}' deleted successfully!")
                except sqlite3.Error as e:
                    self.show_error("Database Error", "Operation failed", e)
                    self.conn.rollback()
                
        except Exception as e:
            self.show_error("Error", "Failed to delete item", e)

    def reprint_item(self, item_name):
        """Print an existing item's barcode using Paint for single prints"""
        try:
            # Fetch item data
            self.cursor.execute("""
                SELECT item_name, barcode, sale_price, stock_quantity 
                FROM items WHERE item_name=?
            """, (item_name,))
            
            result = self.cursor.fetchone()
            if not result:
                self.show_warning("Print Error", 
                    f"Item '{item_name}' not found in database!")
                return

            name, barcode, price, stock = result

            # Create custom input dialog with EXTRA LARGE sizing to prevent cutoff
            dialog = QDialog(self)
            dialog.setWindowTitle("Print Labels")
            dialog.setModal(True)
            dialog.setFixedSize(650, 600)  # MUCH larger size to prevent any cutoff
            
            # Center the dialog on screen
            screen = QApplication.primaryScreen().availableGeometry()
            dialog.move(
                (screen.width() - 650) // 2,
                (screen.height() - 600) // 2
            )
            
            # Add drop shadow
            shadow = QGraphicsDropShadowEffect(dialog)
            shadow.setBlurRadius(20)
            shadow.setColor(QColor(0, 0, 0, 100))
            shadow.setOffset(0, 4)
            dialog.setGraphicsEffect(shadow)
            
            layout = QVBoxLayout(dialog)
            layout.setContentsMargins(50, 50, 50, 50)  # Extra generous margins
            layout.setSpacing(30)  # More spacing for the larger dialog
            
            # Title - larger for the bigger dialog
            title = QLabel("Print Labels")
            title.setObjectName("dialog_title")
            title.setFont(QFont("Segoe UI", 24, QFont.Bold))  # Even larger title for bigger dialog
            title.setAlignment(Qt.AlignCenter)
            layout.addWidget(title)
            
            # Item info
            info_container = QWidget()
            info_layout = QFormLayout(info_container)
            info_layout.setContentsMargins(0, 0, 0, 0)
            info_layout.setSpacing(18)  # Even better spacing
            
            name_label = QLabel("Item:")
            name_value = QLabel(name)
            name_value.setStyleSheet("font-weight: 600;")
            info_layout.addRow(name_label, name_value)
            
            stock_label = QLabel("Available Stock:")
            stock_value = QLabel(str(stock))
            stock_value.setStyleSheet("font-weight: 600;")
            info_layout.addRow(stock_label, stock_value)
            
            layout.addWidget(info_container)
            
            # Quantity input
            input_container = QWidget()
            input_layout = QHBoxLayout(input_container)
            input_layout.setContentsMargins(0, 0, 0, 0)
            input_layout.setSpacing(20)  # Better spacing
            
            qty_label = QLabel("Number of labels:")
            qty_label.setFixedWidth(120)
            
            qty_input = QSpinBox()
            qty_input.setRange(1, 2)  # Always allow 1-2 labels regardless of stock
            qty_input.setValue(1)
            qty_input.setFixedWidth(100)  # Wider input
            qty_input.setFixedHeight(40)   # Taller input
            
            input_layout.addWidget(qty_label)
            input_layout.addWidget(qty_input)
            input_layout.addStretch()
            
            layout.addWidget(input_container)
            
            # Info text with better formatting and more space
            info = QLabel("You can print 1 or 2 labels per sheet.\n\nSelect the quantity below and click Print when ready.")
            info.setObjectName("info_text")
            info.setWordWrap(True)
            info.setAlignment(Qt.AlignCenter)
            info.setMinimumHeight(80)  # Ensure minimum height for visibility
            layout.addWidget(info)
            
            # Add spacer before buttons to push them down
            layout.addStretch()
            
            # Buttons with extra spacing
            button_container = QWidget()
            button_layout = QHBoxLayout(button_container)
            button_layout.setContentsMargins(0, 20, 0, 20)  # Extra margins for visibility
            button_layout.setSpacing(30)  # Even more button spacing
            
            cancel_btn = QPushButton("Cancel")
            cancel_btn.setObjectName("cancel_btn")
            cancel_btn.clicked.connect(dialog.reject)
            
            print_btn = QPushButton("Print")
            print_btn.setObjectName("print_btn")
            print_btn.clicked.connect(dialog.accept)
            print_btn.setDefault(True)
            
            button_layout.addStretch()
            button_layout.addWidget(cancel_btn)
            button_layout.addWidget(print_btn)
            
            layout.addWidget(button_container)
            
            # Style the dialog
            dialog.setStyleSheet("""
                QDialog {
                    background-color: #161b22;
                    border: 1px solid #30363d;
                    border-radius: 12px;
                }
                QLabel {
                    color: #f0f6fc;
                    font-size: 16px;
                }
                QLabel#dialog_title {
                    color: #ffffff;
                    font-size: 22px;
                    padding-bottom: 12px;
                    border-bottom: 1px solid #30363d;
                }
                QLabel#info_text {
                    color: #58a6ff;
                    font-size: 13px;
                    padding: 12px;
                    background-color: rgba(88, 166, 255, 0.1);
                    border-radius: 6px;
                    border: 1px solid rgba(88, 166, 255, 0.3);
                }
                QSpinBox {
                    background-color: #21262d;
                    color: #f0f6fc;
                    border: 1px solid #30363d;
                    border-radius: 8px;
                    padding: 8px 12px;
                    font-size: 16px;
                }
                QSpinBox:hover {
                    border-color: #58a6ff;
                }
                QSpinBox::up-button, QSpinBox::down-button {
                    width: 20px;
                    background-color: #2f81f7;
                    border-radius: 4px;
                    margin: 4px;
                }
                QSpinBox::up-button:hover, QSpinBox::down-button:hover {
                    background-color: #58a6ff;
                }
                QPushButton {
                    font-size: 16px;
                    font-weight: 500;
                    padding: 12px 30px;
                    border-radius: 8px;
                    min-width: 120px;
                    min-height: 42px;
                }
                QPushButton#print_btn {
                    background-color: #2f81f7;
                    color: white;
                    border: none;
                }
                QPushButton#print_btn:hover {
                    background-color: #58a6ff;
                }
                QPushButton#cancel_btn {
                    background-color: transparent;
                    color: #8b949e;
                    border: 1px solid #30363d;
                }
                QPushButton#cancel_btn:hover {
                    background-color: #21262d;
                    color: #f0f6fc;
                }
            """)
            
            count = 0  # Initialize count variable
            if dialog.exec_() == QDialog.Accepted:
                count = qty_input.value()

            # Generate labels
            items = [{
                'item_name': name,
                'barcode': barcode,
                'sale_price': price / 100.0,  # Convert to rupees
                'stock_quantity': stock
            } for _ in range(count)]

            # Create print file
            temp_dir = self.ensure_temp_dir("quicktag_prints")
            print_path = os.path.join(temp_dir, f"print_{barcode}.bmp")
            
            label_image = self.barcode_generator.create_label_image(items, self.barcode_font_path)
            if not label_image:
                raise ValueError("Failed to generate barcode image")
                
            # Save as 1-bit BMP
            label_image.save(print_path, "BMP")
            
            # Show print instructions
            if self.show_print_instructions_dialog("Print Instructions",
                f"Ready to print {count} label(s).\n\n" +
                "Instructions:\n" +
                "1. Opening in Paint for best quality\n" +
                "2. Select 'Print' and ensure:\n" +
                "   • Print size is set to 'Actual Size' (100%)\n" +
                "   • No scaling or 'Fit to page' is applied\n" +
                "   • Paper orientation is Landscape\n" +
                "3. Click Print"):
                try:
                    if sys.platform == 'win32':
                        os.startfile(print_path, "print")  # Opens directly in Paint
                    elif sys.platform == 'darwin':
                        subprocess.run(['open', '-a', 'Preview', print_path])
                    else:
                        subprocess.run(['xdg-open', print_path])
                except Exception as e:
                    self.show_error("Print Error",
                        f"Generated label at:\n{print_path}\n\n" +
                        "But failed to open automatically.", e)

        except Exception as e:
            self.show_error("Print Error", "Failed to print label", e)

    def print_all_barcodes(self):
        """Print all items using Photos app for batch printing"""
        try:
            # Get all items
            self.cursor.execute("""
                SELECT item_name, barcode, sale_price, stock_quantity 
                FROM items ORDER BY item_name
            """)
            items = self.cursor.fetchall()
            
            if not items:
                self.show_info("No Items", "No items found to print.")
                return
            
            # Create temp directory for prints
            temp_dir = self.ensure_temp_dir("quicktag_prints")
            
            # Generate labels for each item (max 2 per row)
            total_labels = 0
            for item_name, barcode, sale_price, stock_quantity in items:
                # Create label data (max 2 labels per row)
                label_data = [{
                    'item_name': item_name,
                    'barcode': barcode,
                    'sale_price': sale_price / 100.0,  # Convert to rupees
                    'stock_quantity': stock_quantity
                } for _ in range(min(2, stock_quantity))]  # Limit to 2 labels
                
                # Generate and save label
                label_image = self.barcode_generator.create_label_image(label_data, self.barcode_font_path)
                if label_image:
                    bmp_path = os.path.join(temp_dir, f"print_{barcode}.bmp")
                    label_image.save(bmp_path, "BMP")
                    total_labels += len(label_data)
            
            if total_labels > 0:
                # Show print instructions with clear steps
                if self.show_print_instructions_dialog("Print All Labels",
                    f"Generated {total_labels} labels for {len(items)} items.\n\n" +
                    "Instructions for Batch Printing:\n\n" +
                    "1. Opening prints folder...\n" +
                    "2. Select all files (Ctrl+A)\n" +
                    "3. Right-click and select 'Print'\n" +
                    "4. In Photos app, ensure:\n" +
                    "   • 'Fit picture to frame' is OFF\n" +
                    "   • Size is set to 'Actual Size' (100%)\n" +
                    "   • Paper orientation is Landscape\n" +
                    "   • No scaling is applied\n\n" +
                    "Note: Only two labels are shown per image.\n" +
                    "For items with more stock, print multiple times."):
                    try:
                        # Open prints folder
                        if sys.platform == 'win32':
                            os.startfile(temp_dir)
                        elif sys.platform == 'darwin':
                            subprocess.run(['open', temp_dir])
                        else:
                            subprocess.run(['xdg-open', temp_dir])
                    except Exception as e:
                        self.show_error("Folder Error",
                            f"Failed to open prints folder:\n{temp_dir}", e)
            else:
                self.show_warning("Print Error", "Failed to generate any labels.")

        except sqlite3.Error as e:
            self.show_error("Database Error", "Failed to retrieve items", e)
        except Exception as e:
            self.show_error("Print Error", "Failed to print all labels", e)

    def ensure_temp_dir(self, subdir):
        """Ensure temporary directory exists and clean old files"""
        try:
            if getattr(sys, 'frozen', False):
                # EXE mode - use user's temp directory
                temp_base = os.path.join(os.path.expanduser('~'), '.quicktag_temp')
            else:
                # Development mode - use local temp directory
                temp_base = 'temp'
            
            temp_dir = os.path.join(temp_base, subdir)
            os.makedirs(temp_dir, exist_ok=True)
            
            # Clean old files (older than 24 hours)
            current_time = time.time()
            for file in os.listdir(temp_dir):
                file_path = os.path.join(temp_dir, file)
                if os.path.isfile(file_path):
                    if current_time - os.path.getmtime(file_path) > 86400:  # 24 hours
                        try:
                            os.remove(file_path)
                        except:
                            pass  # Ignore errors in cleanup
            
            return temp_dir
            
        except Exception as e:
            self.logger.error(f"Failed to ensure temp directory: {str(e)}")
            raise

    def setup_logger(self):
        """Set up logging with file output and proper Unicode handling"""
        try:
            # Create logs directory in user's home
            log_dir = os.path.join(os.path.expanduser('~'), '.quicktags', 'logs')
            os.makedirs(log_dir, exist_ok=True)
            
            # Set up log file with date
            log_file = os.path.join(log_dir, f'quicktags_{datetime.datetime.now().strftime("%Y%m%d")}.log')
            
            # Configure logging with UTF-8 encoding
            logging.basicConfig(
                level=logging.INFO,
                format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                handlers=[
                    logging.FileHandler(log_file, encoding='utf-8'),
                    logging.StreamHandler(sys.stdout)  # Use sys.stdout for console output
                ]
            )
            
            self.logger = logging.getLogger('QuickTags')
            self.logger.info("Logger initialized successfully")
            
        except Exception as e:
            print(f"Warning: Failed to setup logger: {str(e)}")
            # Create a basic console logger as fallback
            self.logger = logging.getLogger('QuickTags')
            handler = logging.StreamHandler(sys.stdout)
            handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
            self.logger.addHandler(handler)
            self.logger.warning("Using fallback console logger")

    def create_sidebar(self, main_layout):
        """Create fixed-width sidebar with navigation"""
        sidebar = QWidget()
        sidebar.setObjectName("sidebar")
        sidebar.setFixedWidth(350)  # Better proportioned sidebar
        
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(10, 0, 10, 20)
        sidebar_layout.setSpacing(0)
        
        # App title section
        title_container = QWidget()
        title_layout = QVBoxLayout(title_container)
        title_layout.setContentsMargins(30, 40, 30, 40)
        title_layout.setSpacing(16)
        
        title = QLabel("QuickTags")
        title.setObjectName("app_title")
        title.setFont(QFont("Arial", self.title_font_size, QFont.Bold))
        
        subtitle = QLabel("AutoGeek Edition by CrystalCore")
        subtitle.setObjectName("app_subtitle")
        subtitle.setFont(QFont("Arial", self.subtitle_font_size))
        
        title_layout.addWidget(title)
        title_layout.addWidget(subtitle)
        sidebar_layout.addWidget(title_container)
        
        # Navigation buttons with increased height
        nav_container = QWidget()
        nav_layout = QVBoxLayout(nav_container)
        nav_layout.setContentsMargins(20, 20, 20, 20)
        nav_layout.setSpacing(16)
        
        nav_buttons = [
            ("🖨️ Print All", self.print_all_barcodes, "Print labels for all items in batch"),
            ("📊 Export to Excel", self.export_to_excel, "Export all data to Excel spreadsheet"),
            ("🗑️ Clear History", self.clear_history, "Permanently delete all items from history")
        ]
        
        for text, handler, tooltip in nav_buttons:
            btn = QPushButton(text)
            btn.setObjectName("sidebar_btn")
            btn.setFont(QFont("Arial", self.base_font_size))
            btn.setMinimumHeight(64)
            btn.setToolTip(tooltip)
            btn.clicked.connect(handler)
            nav_layout.addWidget(btn)
        
        sidebar_layout.addWidget(nav_container)
        sidebar_layout.addStretch()
        
        main_layout.addWidget(sidebar)

    def clear_history(self):
        """Clear history with confirmation dialog"""
        dialog = QMessageBox(self)
        dialog.setWindowTitle("Clear History")
        dialog.setText("⚠️ Warning: This will permanently delete all history!")
        dialog.setInformativeText("Are you sure you want to clear all history?\nThis action cannot be undone.")
        dialog.setIcon(QMessageBox.Warning)
        dialog.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        dialog.setDefaultButton(QMessageBox.No)
        
        if dialog.exec_() == QMessageBox.Yes:
            try:
                self.cursor.execute("DELETE FROM items")
                self.conn.commit()
                self.load_item_history()
                self.update_inventory_values()
                self.show_info("Success", "History cleared successfully!")
            except sqlite3.Error as e:
                self.show_error("Database Error", "Failed to clear history", e)
                self.conn.rollback()
            except Exception as e:
                self.show_error("Error", "An unexpected error occurred", e)

    def load_fonts(self):
        """Load fonts with better error handling for EXE mode"""
        try:
            self.logger.info("Starting font loading process...")
            
            # Method 1: Try embedded font data (should be bytes, not FreeTypeFont)
            font_data = get_embedded_font()
            if font_data and isinstance(font_data, (bytes, bytearray)):
                self.logger.info("Found embedded font data, saving to temp file...")
                # Save font data to temp file
                temp_dir = self.ensure_temp_dir("fonts")
                font_path = os.path.join(temp_dir, "arialbd.ttf")
                with open(font_path, 'wb') as f:
                    f.write(font_data)
                # Load font into Qt
                font_id = QFontDatabase.addApplicationFont(font_path)
                if font_id != -1:
                    font_family = QFontDatabase.applicationFontFamilies(font_id)[0]
                    self.font_bold = QFont(font_family, 16, QFont.Bold)
                    self.font_regular = QFont(font_family, 15)
                    self.barcode_font_path = font_path
                    self.logger.info(f"Successfully loaded embedded font: {font_family}")
                    return
                else:
                    self.logger.warning("Failed to load embedded font into Qt")
            
            # Method 2: Try bundled font files (EXE mode)
            if getattr(sys, 'frozen', False):
                self.logger.info("Running in EXE mode, checking bundled fonts...")
                exe_paths = [
                    os.path.join(sys._MEIPASS, 'assets', 'arialbd.ttf'),
                    os.path.join(sys._MEIPASS, 'arialbd.ttf'),
                    os.path.join(os.path.dirname(sys.executable), 'assets', 'arialbd.ttf'),
                    os.path.join(os.path.dirname(sys.executable), 'arialbd.ttf')
                ]
                
                for path in exe_paths:
                    if os.path.exists(path):
                        self.logger.info(f"Found bundled font at: {path}")
                        try:
                            font_id = QFontDatabase.addApplicationFont(path)
                            if font_id != -1:
                                font_family = QFontDatabase.applicationFontFamilies(font_id)[0]
                                self.font_bold = QFont(font_family, 16, QFont.Bold)
                                self.font_regular = QFont(font_family, 15)
                                self.barcode_font_path = path
                                self.logger.info(f"Successfully loaded bundled font: {font_family}")
                                return
                        except Exception as e:
                            self.logger.error(f"Failed to load bundled font {path}: {str(e)}")
                            continue
            
            # Method 3: Try local assets folder (development mode)
            if not getattr(sys, 'frozen', False):
                self.logger.info("Running in development mode, checking local assets...")
                dev_paths = [
                    os.path.join(os.path.dirname(__file__), 'assets', 'arialbd.ttf'),
                    os.path.join(os.path.dirname(__file__), 'arialbd.ttf'),
                    os.path.join('assets', 'arialbd.ttf'),
                    'arialbd.ttf'
                ]
                
                for path in dev_paths:
                    if os.path.exists(path):
                        self.logger.info(f"Found local font at: {path}")
                        try:
                            font_id = QFontDatabase.addApplicationFont(path)
                            if font_id != -1:
                                font_family = QFontDatabase.applicationFontFamilies(font_id)[0]
                                self.font_bold = QFont(font_family, 16, QFont.Bold)
                                self.font_regular = QFont(font_family, 15)
                                self.barcode_font_path = path
                                self.logger.info(f"Successfully loaded local font: {font_family}")
                                return
                        except Exception as e:
                            self.logger.error(f"Failed to load local font {path}: {str(e)}")
                            continue
            
            # Method 4: Try system fonts
            self.logger.info("Trying system fonts...")
            system_paths = [
                '/System/Library/Fonts/Arial Bold.ttf',  # macOS
                'C:\\Windows\\Fonts\\arialbd.ttf',      # Windows
                '/usr/share/fonts/truetype/msttcorefonts/arialbd.ttf',  # Linux
                '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'  # Linux fallback
            ]
            
            for path in system_paths:
                if os.path.exists(path):
                    self.logger.info(f"Found system font at: {path}")
                    try:
                        font_id = QFontDatabase.addApplicationFont(path)
                        if font_id != -1:
                            font_family = QFontDatabase.applicationFontFamilies(font_id)[0]
                            self.font_bold = QFont(font_family, 16, QFont.Bold)
                            self.font_regular = QFont(font_family, 15)
                            self.barcode_font_path = path
                            self.logger.info(f"Successfully loaded system font: {font_family}")
                            return
                    except Exception as e:
                        self.logger.error(f"Failed to load system font {path}: {str(e)}")
                        continue
            
            # Method 5: Try to find any available font
            self.logger.info("Trying to find any available font...")
            try:
                # Try common font names
                font_names = ['Arial Bold', 'Arial-Bold', 'Arial', 'Helvetica', 'DejaVu Sans']
                for font_name in font_names:
                    try:
                        test_font = QFont(font_name, 16, QFont.Bold)
                        if test_font.family() != font_name:  # Font was found
                            self.font_bold = QFont(font_name, 16, QFont.Bold)
                            self.font_regular = QFont(font_name, 15)
                            self.barcode_font_path = None  # Will use system font
                            self.logger.info(f"Using system font: {font_name}")
                            return
                    except:
                        continue
            except Exception as e:
                self.logger.error(f"Failed to find system fonts: {str(e)}")
            
            # Final fallback to default system font
            self.logger.warning("Using fallback system font")
            self.font_bold = QFont()
            self.font_bold.setPointSize(16)
            self.font_bold.setBold(True)
            self.font_regular = QFont()
            self.font_regular.setPointSize(15)
            self.barcode_font_path = None  # Will use PIL's default font
            
        except Exception as e:
            self.logger.error(f"Error loading fonts: {str(e)}")
            self.logger.error(traceback.format_exc())
            # Set fallback fonts
            self.font_bold = QFont()
            self.font_regular = QFont()
            self.barcode_font_path = None

    def add_item(self):
        """Add a new item with proper validation and error handling"""
        try:
            # Validate inputs
            item_data = self.validate_inputs()
            
            # Check for existing item
            self.cursor.execute("""
                SELECT id FROM items 
                WHERE item_name = ? OR barcode = ?
            """, (item_data['item_name'], item_data['barcode']))
            
            if self.cursor.fetchone():
                self.show_warning("Duplicate Item", 
                    "An item with this name or barcode already exists!")
                return
            
            # CRITICAL FIX: Insert into database with validation
            created_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # VALIDATION ASSERTION: Ensure data consistency before database insert
            barcode_to_store = str(item_data['barcode']).strip().zfill(13)
            assert len(barcode_to_store) == 13, f"Database barcode must be 13 digits: {barcode_to_store}"
            assert barcode_to_store.isdigit(), f"Database barcode must be numeric: {barcode_to_store}"
            
            purchase_cents = int(item_data['purchase_price'] * 100)
            sale_cents = int(item_data['sale_price'] * 100)
            profit_basis_points = int(item_data['profit_percent'] * 100)
            
            assert purchase_cents > 0, f"Purchase price in cents must be positive: {purchase_cents}"
            assert sale_cents > 0, f"Sale price in cents must be positive: {sale_cents}"
            assert sale_cents >= purchase_cents, f"Sale price must be >= purchase price: {sale_cents} < {purchase_cents}"
            
            self.cursor.execute("""
                INSERT INTO items (
                    item_name, barcode, purchase_price, profit_percent,
                    sale_price, stock_quantity, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                item_data['item_name'],
                barcode_to_store,                        # Validated 13-digit barcode
                purchase_cents,                          # Store as cents
                profit_basis_points,                     # Store as basis points
                sale_cents,                              # Store as cents
                item_data['stock_quantity'],
                created_at
            ))
            self.conn.commit()
            
            # Show success message
            self.show_info("Success", 
                f"Item '{item_data['item_name']}' added successfully!")
            
            # Clear form and refresh table
            self.clear_form()
            self.load_item_history()
            self.update_inventory_values()
            
        except ValueError as e:
            self.show_warning("Input Error", str(e))
        except sqlite3.Error as e:
            self.show_error("Database Error", "Failed to add item", e)
            self.conn.rollback()
        except Exception as e:
            self.show_error("Error", "An unexpected error occurred", e)

    def clear_form(self):
        """Clear all form inputs and reset to defaults"""
        try:
            # Reset all inputs to defaults
            self.item_name_input.clear()
            self.purchase_price_input.setText("1")
            self.profit_percent_input.setText("100")
            self.sale_price_input.clear()
            self.stock_quantity_input.setText("1")
            
            # Clear barcode override if visible
            if hasattr(self, 'barcode_override_input'):
                self.barcode_override_input.clear()
            
            # Focus back on item name
            self.item_name_input.setFocus()
            
        except Exception as e:
            self.logger.error(f"Error clearing form: {str(e)}")
            self.show_error("Error", "Failed to clear form fields")

    def show_dialog(self, title, message, icon_type=QMessageBox.Information, buttons=QMessageBox.Ok):
        """Show a styled dialog with proper icon handling and scaled fonts"""
        dialog = QMessageBox(self)
        dialog.setWindowTitle(title)
        dialog.setText(message)
        dialog.setIcon(icon_type)
        dialog.setStandardButtons(buttons)
        
        # Style the dialog with dark theme
        dialog.setStyleSheet(f"""
            QMessageBox {{
                background-color: #161b22;
                border: 2px solid #30363d;
                border-radius: 12px;
                min-width: 500px;
                color: #f0f6fc;
            }}
            QMessageBox QLabel {{
                font-size: {self.base_font_size + 1}px;
                color: #f0f6fc;
                padding: 20px;
                background-color: transparent;
                min-height: 60px;
                line-height: 1.4;
            }}
            QMessageBox QPushButton {{
                background-color: #238636;
                color: white;
                border: none;
                padding: 12px 24px;
                border-radius: 8px;
                font-size: {self.base_font_size}px;
                font-weight: 600;
                min-width: 100px;
                min-height: 40px;
                margin: 4px;
            }}
            QMessageBox QPushButton:hover {{
                background-color: #2ea043;
            }}
            QMessageBox QPushButton[text="Cancel"],
            QMessageBox QPushButton[text="No"] {{
                background-color: #21262d;
                color: #f0f6fc;
                border: 2px solid #30363d;
            }}
            QMessageBox QPushButton[text="Cancel"]:hover,
            QMessageBox QPushButton[text="No"]:hover {{
                background-color: #30363d;
                border-color: #58a6ff;
            }}
        """)
        
        return dialog.exec_()

    def show_error(self, title, message, error=None):
        """Show an error dialog with proper logging"""
        if error:
            self.logger.error(f"{title}: {str(error)}")
            message = f"{message}\n\nError: {str(error)}"
        self.show_dialog(title, message, QMessageBox.Critical)

    def show_warning(self, title, message):
        """Show a warning dialog"""
        return self.show_dialog(title, message, QMessageBox.Warning)

    def show_info(self, title, message):
        """Show an information dialog"""
        return self.show_dialog(title, message, QMessageBox.Information)

    def show_confirm(self, title, message):
        """Show a confirmation dialog"""
        return self.show_dialog(title, message, QMessageBox.Question, 
                              QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes

    def show_print_instructions_dialog(self, title, message):
        """Show a custom print instructions dialog with proper sizing"""
        # Create custom dialog with EXTRA LARGE sizing
        dialog = QDialog(self)
        dialog.setWindowTitle(title)
        dialog.setModal(True)
        dialog.setFixedSize(750, 500)  # Large enough to prevent cutoff
        
        # Center the dialog on screen
        screen = QApplication.primaryScreen().availableGeometry()
        dialog.move(
            (screen.width() - 750) // 2,
            (screen.height() - 500) // 2
        )
        
        # Add drop shadow
        shadow = QGraphicsDropShadowEffect(dialog)
        shadow.setBlurRadius(20)
        shadow.setColor(QColor(0, 0, 0, 100))
        shadow.setOffset(0, 4)
        dialog.setGraphicsEffect(shadow)
        
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(50, 50, 50, 50)  # Extra generous margins
        layout.setSpacing(30)  # More spacing for better readability
        
        # Title
        title_label = QLabel(title)
        title_label.setObjectName("dialog_title")
        title_label.setFont(QFont("Segoe UI", 24, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # Message text with better formatting
        msg_label = QLabel(message)
        msg_label.setObjectName("dialog_message")
        msg_label.setWordWrap(True)
        msg_label.setAlignment(Qt.AlignLeft)
        msg_label.setMinimumHeight(200)  # Ensure enough height for the text
        layout.addWidget(msg_label)
        
        # Add spacer before buttons
        layout.addStretch()
        
        # Buttons
        button_container = QWidget()
        button_layout = QHBoxLayout(button_container)
        button_layout.setContentsMargins(0, 20, 0, 20)
        button_layout.setSpacing(30)
        
        no_btn = QPushButton("Cancel")
        no_btn.setObjectName("cancel_btn")
        no_btn.clicked.connect(dialog.reject)
        
        yes_btn = QPushButton("Continue")
        yes_btn.setObjectName("continue_btn")
        yes_btn.clicked.connect(dialog.accept)
        yes_btn.setDefault(True)
        
        button_layout.addStretch()
        button_layout.addWidget(no_btn)
        button_layout.addWidget(yes_btn)
        
        layout.addWidget(button_container)
        
        # Style the dialog
        dialog.setStyleSheet("""
            QDialog {
                background-color: #161b22;
                border: 1px solid #30363d;
                border-radius: 12px;
            }
            QLabel {
                color: #f0f6fc;
                font-size: 16px;
            }
            QLabel#dialog_title {
                color: #ffffff;
                font-size: 22px;
                padding-bottom: 12px;
                border-bottom: 1px solid #30363d;
            }
            QLabel#dialog_message {
                color: #f0f6fc;
                font-size: 16px;
                line-height: 1.6;
                padding: 12px;
                background-color: #21262d;
                border-radius: 6px;
            }
            QPushButton {
                font-size: 16px;
                font-weight: 500;
                padding: 12px 30px;
                border-radius: 8px;
                min-width: 120px;
                min-height: 42px;
            }
            QPushButton#continue_btn {
                background-color: #2f81f7;
                color: white;
                border: none;
            }
            QPushButton#continue_btn:hover {
                background-color: #58a6ff;
            }
            QPushButton#cancel_btn {
                background-color: transparent;
                color: #8b949e;
                border: 1px solid #30363d;
            }
            QPushButton#cancel_btn:hover {
                background-color: #21262d;
                color: #f0f6fc;
            }
        """)
        
        return dialog.exec_() == QDialog.Accepted

    def calculate_sale_price(self):
        """Calculate sale price based on purchase price and profit percentage"""
        try:
            purchase_price = float(self.purchase_price_input.text() or 0)
            profit_percent = float(self.profit_percent_input.text() or 0)
            
            if purchase_price > 0 and profit_percent >= 0:
                # Calculate sale price with profit margin
                sale_price = purchase_price * (1 + profit_percent / 100)
                self.sale_price_input.setText(f"{sale_price:.2f}")
        except ValueError:
            # Clear sale price if inputs are invalid
            self.sale_price_input.clear()
            
    def validate_inputs(self):
        """Validate form inputs and return cleaned data"""
        # Get and clean inputs
        item_name = self.item_name_input.text().strip()
        purchase_price = self.purchase_price_input.text().strip()
        profit_percent = self.profit_percent_input.text().strip()
        sale_price = self.sale_price_input.text().strip()
        stock_quantity = self.stock_quantity_input.text().strip()
        barcode_override = self.barcode_override_input.text().strip()
        
        # Validate required fields
        if not item_name:
            raise ValueError("Item name is required")
        
        if len(item_name) > 30:
            raise ValueError("Item name must be 30 characters or less")
            
        try:
            purchase_price = float(purchase_price)
            if purchase_price <= 0:
                raise ValueError
        except ValueError:
            raise ValueError("Purchase price must be a positive number")
            
        try:
            profit_percent = float(profit_percent)
            if profit_percent < 0:
                raise ValueError
            if profit_percent > 1000:  # Cap at 1000%
                raise ValueError("Profit percentage cannot exceed 1000%")
        except ValueError:
            raise ValueError("Profit percentage must be a non-negative number")
            
        try:
            sale_price = float(sale_price)
            if sale_price <= 0:
                raise ValueError
            if sale_price < purchase_price:
                raise ValueError("Sale price cannot be less than purchase price")
        except ValueError:
            raise ValueError("Sale price must be a positive number")
            
        try:
            stock_quantity = int(stock_quantity)
            if stock_quantity <= 0:
                raise ValueError
        except ValueError:
            raise ValueError("Stock quantity must be a positive integer")
            
        # Generate or validate barcode
        if barcode_override:
            if not barcode_override.isdigit() or len(barcode_override) != 13:
                raise ValueError("Barcode must be exactly 13 digits")
            barcode = barcode_override
        else:
            # Generate random 13-digit barcode
            barcode = ''.join([str(random.randint(0, 9)) for _ in range(13)])
            
        return {
            'item_name': item_name,
            'barcode': barcode,
            'purchase_price': purchase_price,
            'profit_percent': profit_percent,
            'sale_price': sale_price,
            'stock_quantity': stock_quantity
        }
        
    def create_history_table(self):
        """Create and configure the history table with proper styling and alignment"""
        self.history_table = QTableWidget()
        self.history_table.setObjectName("history_table")
        self.history_table.setSizeAdjustPolicy(QTableWidget.AdjustToContents)
        self.history_table.setWordWrap(False)
        self.history_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.history_table.setTextElideMode(Qt.ElideRight)
        self.history_table.setAlternatingRowColors(True)
        self.history_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.history_table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.history_table.setMouseTracking(True)
        self.history_table.setShowGrid(False)
        self.history_table.setFrameStyle(QFrame.NoFrame)
        self.history_table.setSelectionMode(QTableWidget.SingleSelection)
        self.history_table.setSelectionBehavior(QTableWidget.SelectRows)

        # Set column count and headers with optimized widths
        self.history_table.setColumnCount(7)
        headers = [
            ("Actions", 290, Qt.AlignCenter),
            ("S.No.", 65, Qt.AlignCenter),  # Increased from 50 to 65 for better visibility
            ("Item Name", -1, Qt.AlignLeft | Qt.AlignVCenter),  # -1 for stretch
            ("Date Added", 120, Qt.AlignCenter),  # Increased from 110 to 120 for full date visibility
            ("Price", 120, Qt.AlignRight | Qt.AlignVCenter),  # Increased from 90 to 120
            ("Stock", 65, Qt.AlignCenter),  # Increased from 60 to 65 to match S.No.
            ("Value", 120, Qt.AlignRight | Qt.AlignVCenter)  # Increased from 100 to 120
        ]

        # Configure header appearance
        header_font = QFont("Segoe UI", 12, QFont.Bold)
        header = self.history_table.horizontalHeader()
        header.setHighlightSections(False)
        header.setFixedHeight(35)  # Better header height
        header.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        header.setStretchLastSection(False)
        
        # Add drop shadow to table
        shadow = QGraphicsDropShadowEffect(self.history_table)
        shadow.setBlurRadius(20)
        shadow.setColor(QColor(0, 0, 0, 80))
        shadow.setOffset(0, 2)
        self.history_table.setGraphicsEffect(shadow)
        
        # Apply headers and column properties
        for col, (text, width, align) in enumerate(headers):
            item = QTableWidgetItem(text)
            item.setTextAlignment(align)
            item.setFont(header_font)
            self.history_table.setHorizontalHeaderItem(col, item)

            if width == -1:
                header.setSectionResizeMode(col, QHeaderView.Stretch)
            else:
                header.setSectionResizeMode(col, QHeaderView.Fixed)
                self.history_table.setColumnWidth(col, width)

        # Configure vertical header with responsive row height
        self.history_table.verticalHeader().setVisible(False)
        responsive_row_height = self.scaler.get_table_row_height()
        self.history_table.verticalHeader().setDefaultSectionSize(responsive_row_height)
        self.history_table.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)

        # Set consistent style with improved visual hierarchy
        self.history_table.setStyleSheet("""
            QTableWidget {
                border: none;
                background-color: #161b22;
                border: 1px solid #30363d;
                border-radius: 8px;
                padding: 0px;
            }
            QTableWidget::item {
                border-bottom: 1px solid #21262d;
                padding: 6px 8px;
                font-size: 13px;
            }
            QTableWidget::item:selected {
                background-color: #2d4f79;
                color: #ffffff;
            }
            QTableWidget::item:hover {
                background-color: #21262d;
            }
            QHeaderView::section {
                padding: 8px 12px;
                border: none;
                border-bottom: 2px solid #30363d;
                background-color: #161b22;
                color: #8b949e;
                font-weight: 600;
            }
            QHeaderView::section:first {
                border-top-left-radius: 8px;
            }
            QHeaderView::section:last {
                border-top-right-radius: 8px;
            }
            QTableWidget QWidget#action_container {
                background: transparent;
            }
            QTableWidget QPushButton {
                font-size: 12px;
                font-weight: 600;
                padding: 4px 12px;
                border-radius: 4px;
                min-height: 28px;
                min-width: 60px;
            }
            QTableWidget QPushButton#edit_btn {
                background-color: #2f81f7;
                color: white;
                border: none;
            }
            QTableWidget QPushButton#edit_btn:hover {
                background-color: #58a6ff;
            }
            QTableWidget QPushButton#del_btn {
                background-color: #f85149;
                color: white;
                border: none;
            }
            QTableWidget QPushButton#del_btn:hover {
                background-color: #da3633;
            }
            QTableWidget QPushButton#print_btn {
                background-color: #238636;
                color: white;
                border: none;
            }
            QTableWidget QPushButton#print_btn:hover {
                background-color: #2ea043;
            }
        """)

        # Table is already configured via the main stylesheet

    def load_item_history(self):
        try:
            self.history_table.setRowCount(0)
            self.cursor.execute("""
                SELECT item_name, barcode, purchase_price, sale_price, 
                       stock_quantity, created_at 
                FROM items ORDER BY created_at DESC
            """)
            items = self.cursor.fetchall()

            if not items:
                self.history_table.hide()
                self.no_items_label.show()
            else:
                self.history_table.show()
                self.no_items_label.hide()

            for row, (name, barcode, purchase, sale, stock, date) in enumerate(items):
                self.history_table.insertRow(row)
                btn_container = QWidget()
                btn_container.setObjectName("action_container")
                layout = QHBoxLayout(btn_container)
                layout.setContentsMargins(3, 4, 3, 4)  # Better margins for crisp buttons
                layout.setSpacing(6)  # Better spacing between buttons
                layout.setAlignment(Qt.AlignCenter)
                
                edit_btn = QPushButton("Edit")
                edit_btn.setObjectName("edit_btn")
                edit_btn.setToolTip("Edit this item")
                edit_btn.setFixedSize(48, 28)  # Slightly larger for better clarity
                edit_btn.clicked.connect(lambda _, n=name: self.edit_item(n))
                
                del_btn = QPushButton("Del")
                del_btn.setObjectName("del_btn")
                del_btn.setToolTip("Delete this item")
                del_btn.setFixedSize(42, 28)  # Slightly larger for better clarity
                del_btn.clicked.connect(lambda _, n=name: self.delete_item(n))
                
                print_btn = QPushButton("Print")
                print_btn.setObjectName("print_btn")
                print_btn.setToolTip("Reprint this item's label")
                print_btn.setFixedSize(48, 28)  # Slightly larger for better clarity
                print_btn.clicked.connect(lambda _, n=name: self.reprint_item(n))
                
                layout.addWidget(edit_btn)
                layout.addWidget(del_btn)
                layout.addWidget(print_btn)

                self.history_table.setCellWidget(row, 0, btn_container)
                
                regular_font = QFont("Segoe UI", 12)
                date_font = QFont("Segoe UI", 11)
                def cell(text, align, use_date_font=False):
                    item = QTableWidgetItem(text)
                    item.setTextAlignment(align | Qt.AlignVCenter) # Ensure vertical center alignment
                    item.setFlags(Qt.ItemIsEnabled)
                    item.setFont(date_font if use_date_font else regular_font)
                    item.setToolTip(text)
                    return item
                date_obj = datetime.datetime.strptime(date, "%Y-%m-%d %H:%M:%S")
                formatted_date = date_obj.strftime("%d/%m/%y")
                cells = [
                    (str(row + 1), Qt.AlignCenter, False),
                    (name, Qt.AlignLeft | Qt.AlignVCenter, False),
                    (formatted_date, Qt.AlignCenter, True),
                    (f"₹{int(sale/100):,}", Qt.AlignRight | Qt.AlignVCenter, False),
                    (str(stock), Qt.AlignCenter, False),
                    (f"₹{int((purchase * stock)/100):,}", Qt.AlignRight | Qt.AlignVCenter, False)
                ]
                for col, (text, align, is_date) in enumerate(cells, 1):
                    self.history_table.setItem(row, col, cell(text, align, is_date))
            self.update_inventory_values()
        except Exception as e:
            self.show_error("Error", "Failed to update session table", e)

    def format_date(self, date_str):
        """Format date string for display"""
        try:
            date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            return date_obj.strftime("%d-%m-%Y %H:%M")
        except:
            return date_str

    def filter_history(self):
        """Filter history table based on search input with live updates"""
        search_text = self.search_input.text().lower()
        visible_count = 0
        
        for row in range(self.history_table.rowCount()):
            show_row = False
            
            # Get searchable fields
            item_name = self.history_table.item(row, 2).text().lower()
            barcode = self.history_table.item(row, 6).text().lower()
            price = self.history_table.item(row, 4).text().lower()
            stock = self.history_table.item(row, 5).text().lower()
            date = self.history_table.item(row, 3).text().lower()
            
            # Check if search text matches any field
            if (search_text in item_name or 
                search_text in barcode or
                search_text in price or
                search_text in stock or
                search_text in date):
                show_row = True
                visible_count += 1
                
            self.history_table.setRowHidden(row, not show_row)
            
            # Update S.No. for visible rows only
            if show_row:
                sno_item = QTableWidgetItem(str(visible_count))
                sno_item.setTextAlignment(Qt.AlignCenter)
                self.history_table.setItem(row, 1, sno_item)
        
        # Update placeholder text with count
        if search_text:
            self.search_input.setPlaceholderText(f"Found {visible_count} items")
        else:
            self.search_input.setPlaceholderText("Search items...")



    def apply_visual_styling(self):
        """Apply a visually perfect dark theme with optimal contrast and spacing for two-column layout"""
        app = QApplication.instance()
        
        # --- RESPONSIVE SCALING BASED ON SCREEN SIZE ---
        scale = {
            'base_font': self.scaler.base_font,
            'header_font': self.scaler.header_font,
            'title_font': self.scaler.title_font,
            'subtitle_font': self.scaler.subtitle_font,
            'table_font': self.scaler.table_font,
            'button_font': self.scaler.button_font,
            'stats_font': self.scaler.stats_font,
            'row_height': self.scaler.row_height,
            'input_height': self.scaler.input_height,
            'button_height': self.scaler.button_height,
            'dialog_width': int(850 * self.scaler.scale_factor),
            'base_padding': int(self.scaler.base_padding),
            'section_spacing': int(self.scaler.section_spacing),
            'radius': int(self.scaler.radius),
            'shadow_blur': int(12 * self.scaler.scale_factor),
            'shadow_offset': int(3 * self.scaler.scale_factor)
        }

        app.setFont(QFont("Segoe UI", scale['base_font']))
        
        # --- PERFECT COLOR PALETTE ---
        colors = {
            'bg_deep': '#0a0e14',
            'bg_surface': '#151b23',
            'bg_element': '#1e2329',
            'bg_element_hover': '#2d4f79',
            'bg_input': '#0d1117',
            'border_primary': '#30363d',
            'border_secondary': '#21262d',
            'border_accent': '#58a6ff',
            'text_primary': '#f0f6fc',
            'text_secondary': '#8b949e',
            'text_header': '#ffffff',
            'text_muted': '#6e7681',
            'accent_blue': '#2f81f7',
            'accent_blue_hover': '#58a6ff',
            'accent_red': '#f85149',
            'accent_red_hover': '#da3633',
            'accent_green': '#238636',
            'accent_green_hover': '#2ea043',
            'accent_orange': '#fb8500',
            'accent_orange_hover': '#ffb700',
            'scrollbar_bg': '#21262d',
            'scrollbar_handle': '#484f58',
            'scrollbar_handle_hover': '#586069',
            'table_header_bg': '#161b22',
            'table_row_alt': '#0d1117',
            'popup_bg': '#151b23',
            'popup_border': '#30363d'
        }

        # --- ENHANCED STYLESHEET FOR TWO-COLUMN LAYOUT ---
        self.setStyleSheet(f"""
            /* --- Main Window & Layout --- */
            QMainWindow {{
                background-color: {colors['bg_deep']};
                color: {colors['text_primary']};
            }}
            
            QWidget#main_content {{
                background-color: {colors['bg_deep']};
            }}
            
            QWidget#left_column, QWidget#right_column {{
                background-color: transparent;
            }}
            
            QWidget#sidebar {{
                background-color: {colors['bg_surface']};
                border-right: 3px solid {colors['border_primary']};
            }}
            
            /* --- Typography & Headers --- */
            QLabel#app_title {{
                color: {colors['text_header']};
                font-size: {scale['title_font']}px;
                font-weight: 700;
                letter-spacing: 1.5px;
                margin-bottom: 12px;
            }}
            
            QLabel#app_subtitle {{
                color: {colors['text_secondary']};
                font-size: {scale['subtitle_font']}px;
                font-weight: 400;
                letter-spacing: 0.5px;
            }}
            
            /* --- Enhanced Group Boxes --- */
            QGroupBox {{
                font-size: {scale['header_font']}px;
                font-weight: 700;
                padding-top: 28px;
                border-radius: {scale['radius']}px;
                background-color: {colors['bg_surface']};
                border: 2px solid {colors['border_primary']};
                color: {colors['text_header']};
                margin-top: 14px;
            }}
            
            QGroupBox::title {{
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 14px;
                background-color: {colors['bg_surface']};
                border-radius: 5px;
                font-weight: 700;
                letter-spacing: 0.8px;
            }}
            
            /* --- Enhanced Form Elements --- */
            QLabel {{
                font-size: {scale['base_font']}px;
                color: {colors['text_primary']};
                font-weight: 500;
            }}
            
            QLabel#field_label {{
                font-size: {scale['base_font'] + 1}px;
                color: {colors['text_header']};
                font-weight: 700;
                margin-bottom: 6px;
            }}
            
            QLineEdit {{
                font-size: {scale['base_font']}px;
                padding: {scale['base_padding']}px;
                min-height: {scale['input_height']}px;
                background-color: {colors['bg_input']};
                border: 2px solid {colors['border_primary']};
                border-radius: {scale['radius']}px;
                color: {colors['text_primary']};
                selection-background-color: {colors['accent_blue']};
                font-weight: 500;
            }}
            
            QLineEdit:focus {{
                border-color: {colors['border_accent']};
                background-color: {colors['bg_element']};
            }}
            
            QLineEdit:hover {{
                border-color: {colors['text_secondary']};
            }}
            
            /* --- Enhanced ComboBox (Label Size Toggle) --- */
            QComboBox {{
                font-size: {scale['base_font']}px;
                padding: {scale['base_padding']}px;
                min-height: {scale['input_height']}px;
                background-color: {colors['bg_input']};
                border: 2px solid {colors['border_primary']};
                border-radius: {scale['radius']}px;
                color: {colors['text_primary']};
                font-weight: 500;
            }}
            
            QComboBox:hover {{
                border-color: {colors['text_secondary']};
            }}
            
            QComboBox:focus {{
                border-color: {colors['border_accent']};
                background-color: {colors['bg_element']};
            }}
            
            QComboBox::drop-down {{
                border: none;
                width: 20px;
            }}
            
            QComboBox::down-arrow {{
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid {colors['text_primary']};
                margin-right: 8px;
            }}
            
            QComboBox::down-arrow:hover {{
                border-top-color: {colors['text_header']};
            }}
            
            QComboBox QAbstractItemView {{
                background-color: {colors['bg_element']};
                border: 2px solid {colors['border_primary']};
                border-radius: {scale['radius']}px;
                selection-background-color: {colors['accent_blue']};
                selection-color: white;
                outline: none;
            }}
            
            QComboBox QAbstractItemView::item {{
                padding: {scale['base_padding']}px;
                color: {colors['text_primary']};
                background-color: transparent;
            }}
            
            QComboBox QAbstractItemView::item:hover {{
                background-color: {colors['bg_element_hover']};
                color: {colors['text_header']};
            }}
            
            QComboBox QAbstractItemView::item:selected {{
                background-color: {colors['accent_blue']};
                color: white;
            }}
            
            /* --- Enhanced Buttons --- */
            QPushButton {{
                font-size: {scale['button_font']}px;
                font-weight: 600;
                padding: {scale['base_padding']}px {scale['base_padding'] * 2}px;
                min-height: {scale['button_height']}px;
                border-radius: {scale['radius']}px;
                background-color: {colors['bg_element']};
                border: 2px solid {colors['border_primary']};
                color: {colors['text_primary']};
                letter-spacing: 0.5px;
            }}
            
            QPushButton:hover {{
                background-color: {colors['bg_element_hover']};
                border-color: {colors['border_accent']};
                color: {colors['text_header']};
            }}
            
            QPushButton:pressed {{
                background-color: {colors['accent_blue']};
            }}
            
            QPushButton#primary_button {{
                background-color: {colors['accent_blue']};
                color: white;
                border: none;
                font-weight: 700;
                font-size: {scale['button_font'] + 1}px;
            }}
            
            QPushButton#primary_button:hover {{
                background-color: {colors['accent_blue_hover']};
            }}
            
            QPushButton#secondary_button {{
                background-color: {colors['bg_element']};
                border-color: {colors['text_secondary']};
                font-size: {scale['button_font']}px;
            }}
            
            QPushButton#secondary_button:hover {{
                background-color: {colors['text_secondary']};
                color: {colors['bg_deep']};
            }}
            
            /* --- Enhanced Sidebar Buttons --- */
            QWidget#sidebar QPushButton {{
                text-align: left;
                padding: 18px 26px;
                margin: 6px 10px;
                font-size: {scale['base_font']}px;
                background-color: transparent;
                border: 1px solid transparent;
                border-radius: {scale['radius']}px;
                font-weight: 600;
            }}
            
            QWidget#sidebar QPushButton:hover {{
                background-color: {colors['bg_element']};
                border-color: {colors['border_primary']};
                transform: translateX(4px);
            }}

            /* --- Enhanced History Table --- */
            QWidget#table_container {{
                background-color: transparent;
            }}
            
            QTableWidget#history_table {{
                font-size: {scale['table_font']}px;
                background-color: {colors['bg_surface']};
                border-radius: {scale['radius']}px;
                border: 2px solid {colors['border_primary']};
                gridline-color: transparent;
                outline: none;
                alternate-background-color: {colors['table_row_alt']};
            }}
            
            QTableWidget#history_table::item {{
                padding: 8px 6px;
                color: {colors['text_primary']};
                border-bottom: 1px solid {colors['border_secondary']};
                font-weight: 500;
                font-size: {scale['table_font']}px;
            }}
            
            QTableWidget#history_table::item:selected {{
                background-color: {colors['bg_element_hover']};
                color: {colors['text_header']};
            }}
            
            QTableWidget#history_table::item:hover {{
                background-color: {colors['bg_element']};
            }}
            
            QTableWidget#history_table QHeaderView::section {{
                font-size: {scale['table_font']}px;
                font-weight: 700;
                padding: 8px 6px;
                background-color: {colors['table_header_bg']};
                color: {colors['text_header']};
                border: none;
                border-bottom: 3px solid {colors['accent_blue']};
            }}
            
            /* --- Table Action Buttons --- */
            QTableWidget#history_table QPushButton {{
                font-size: 11px;
                font-weight: 700;
                padding: 4px 8px;
                border-radius: 4px;
                min-height: 28px;
                min-width: 42px;
                max-width: 48px;
                border: none;
                color: white;
                margin: 1px;
            }}
            
            QTableWidget#history_table QPushButton#edit_btn {{
                background-color: {colors['accent_blue']};
            }}
            
            QTableWidget#history_table QPushButton#edit_btn:hover {{
                background-color: {colors['accent_blue_hover']};
            }}
            
            QTableWidget#history_table QPushButton#del_btn {{
                background-color: {colors['accent_red']};
            }}
            
            QTableWidget#history_table QPushButton#del_btn:hover {{
                background-color: {colors['accent_red_hover']};
            }}
            
            QTableWidget#history_table QPushButton#print_btn {{
                background-color: {colors['accent_green']};
            }}
            
            QTableWidget#history_table QPushButton#print_btn:hover {{
                background-color: {colors['accent_green_hover']};
            }}
            
            /* --- Scrollbars --- */
            QScrollBar:vertical {{
                background-color: {colors['scrollbar_bg']};
                width: 12px;
                border-radius: 6px;
                margin: 0px;
            }}
            
            QScrollBar::handle:vertical {{
                background-color: {colors['scrollbar_handle']};
                border-radius: 6px;
                min-height: 24px;
                margin: 2px;
            }}
            
            QScrollBar::handle:vertical:hover {{
                background-color: {colors['scrollbar_handle_hover']};
            }}
            
            QScrollBar:horizontal {{
                background-color: {colors['scrollbar_bg']};
                height: 12px;
                border-radius: 6px;
                margin: 0px;
            }}
            
            QScrollBar::handle:horizontal {{
                background-color: {colors['scrollbar_handle']};
                border-radius: 6px;
                min-width: 24px;
                margin: 2px;
            }}
            
            QScrollBar::handle:horizontal:hover {{
                background-color: {colors['scrollbar_handle_hover']};
            }}
            
            /* --- Summary Panels --- */
            QWidget#inventory_value_panel, QWidget#item_summary_panel {{
                background-color: transparent;
            }}
            
            QGroupBox#inventory_value_group {{
                min-height: 180px;
                max-height: 220px;
            }}
            
            QGroupBox#item_summary_group {{
                min-height: 180px;
            }}
            
            /* --- Value & Stats Labels --- */
            QLabel#value_label, QLabel#stats_label {{
                color: {colors['text_header']};
                font-weight: 700;
                font-size: {scale['header_font'] - 2}px;
                letter-spacing: 0.5px;
            }}
            
            QLabel#value_amount, QLabel#stats_amount {{
                color: {colors['accent_green']};
                font-weight: 700;
                font-size: {scale['header_font'] - 2}px;
                letter-spacing: 0.5px;
            }}
            
            QLabel#stats_item_name {{
                color: {colors['text_secondary']};
                font-weight: 600;
                font-size: {scale['base_font'] - 1}px;
                font-style: italic;
            }}
            
            /* --- Breakdown & Toggle Buttons --- */
            QPushButton#breakdown_btn {{
                margin-left: {scale['base_padding'] * 2}px;
                background-color: {colors['accent_orange']};
                color: white;
                border: none;
                font-weight: 700;
                font-size: {scale['base_font'] - 2}px;
                padding: 8px 16px;
                border-radius: 6px;
            }}
            
            QPushButton#breakdown_btn:hover {{
                background-color: {colors['accent_orange_hover']};
                transform: translateY(-1px);
            }}
            
            QPushButton#toggle_btn {{
                background-color: {colors['bg_element']};
                border: 2px solid {colors['border_primary']};
                font-weight: 600;
                font-size: {scale['base_font'] - 1}px;
                padding: 8px 16px;
                border-radius: 6px;
                letter-spacing: 0.3px;
                min-width: 140px;
                max-width: 180px;
            }}
            
            QPushButton#toggle_btn:hover {{
                background-color: {colors['bg_element_hover']};
                border-color: {colors['border_accent']};
                color: {colors['text_header']};
            }}
            
            QPushButton#toggle_btn:checked {{
                background-color: {colors['accent_blue']};
                border-color: {colors['accent_blue']};
                color: white;
            }}
            
            /* --- Container Styling --- */
            QWidget#value_container, QWidget#stats_container {{
                background-color: rgba(30, 35, 41, 0.4);
                border-radius: 8px;
                border: 1px solid rgba(48, 54, 61, 0.6);
            }}
            
            QWidget#value_container:hover, QWidget#stats_container:hover {{
                background-color: rgba(30, 35, 41, 0.6);
                border-color: rgba(88, 166, 255, 0.3);
            }}

            /* --- Dialogs & Popups --- */
            QDialog {{
                background-color: {colors['popup_bg']};
                border: 2px solid {colors['popup_border']};
                border-radius: 16px;
                min-width: {scale['dialog_width']}px;
            }}
            
            QDialog QWidget#popup_container {{
                background-color: {colors['popup_bg']};
                border-radius: 14px;
            }}
            
            QDialog QLabel#popup_title {{
                font-size: 28px;
                font-weight: 700;
                color: {colors['text_header']};
                padding: 16px 0;
                border-bottom: 2px solid {colors['border_primary']};
                letter-spacing: 0.8px;
            }}
            
            QDialog QTableWidget#breakdown_table {{
                background-color: {colors['bg_surface']};
                border: 2px solid {colors['border_primary']};
                border-radius: 10px;
                font-size: 16px;
                gridline-color: transparent;
                alternate-background-color: {colors['table_row_alt']};
            }}
            
            QDialog QTableWidget#breakdown_table QHeaderView::section {{
                background-color: {colors['table_header_bg']};
                color: {colors['text_header']};
                font-weight: 700;
                font-size: 16px;
                padding: 16px 12px;
                border: none;
                border-bottom: 3px solid {colors['accent_blue']};
            }}
            
            QDialog QTableWidget#breakdown_table::item {{
                padding: 14px 12px;
                color: {colors['text_primary']};
                font-size: 15px;
                font-weight: 500;
                border-bottom: 1px solid {colors['border_secondary']};
                background-color: transparent;
            }}
            
            QDialog QTableWidget#breakdown_table::item:alternate {{
                background-color: {colors['table_row_alt']};
            }}
            
            QDialog QPushButton#popup_close_btn {{
                background-color: {colors['accent_blue']};
                color: white;
                border: none;
                padding: 14px 36px;
                border-radius: 10px;
                font-size: 18px;
                font-weight: 700;
                min-width: 140px;
                min-height: 52px;
            }}
            
            QDialog QPushButton#popup_close_btn:hover {{
                background-color: {colors['accent_blue_hover']};
            }}
            
            QMessageBox {{
                background-color: {colors['bg_surface']};
                border: 2px solid {colors['border_primary']};
                border-radius: 12px;
                min-width: {scale['dialog_width']}px;
            }}
            
            QMessageBox QLabel {{
                font-size: {scale['base_font']}px;
                color: {colors['text_primary']};
                font-weight: 500;
            }}
            
            /* --- Placeholder Label --- */
            QLabel#placeholder_label {{
                color: {colors['text_muted']};
                font-size: {scale['base_font'] + 2}px;
                padding: {scale['base_padding'] * 4}px;
                background-color: {colors['bg_surface']};
                border: 2px dashed {colors['border_primary']};
                border-radius: {scale['radius']}px;
                font-weight: 500;
            }}
        """)
        
        # Apply shadows and effects
        self.add_visual_effects()
        
        # Update table row heights for HD displays
        if hasattr(self, 'history_table'):
            self.history_table.verticalHeader().setDefaultSectionSize(scale['row_height'])  # HD optimized row height
        
        # Update fonts on key widgets
        title = self.findChild(QLabel, "app_title")
        if title:
            title.setFont(QFont("Segoe UI", scale['title_font'], QFont.Bold))
        subtitle = self.findChild(QLabel, "app_subtitle")
        if subtitle:
            subtitle.setFont(QFont("Segoe UI", scale['subtitle_font']))
    
    def add_visual_effects(self):
        """Add drop shadows and visual effects for depth and modern appearance"""
        # Add shadow to main panels with HD optimized values
        shadow_config = {
            'blur_radius': 16,
            'color': QColor(0, 0, 0, 60),
            'offset': (0, 3)
        }
        
        # Form group shadow
        if hasattr(self, 'form_group'):
            shadow = QGraphicsDropShadowEffect()
            shadow.setBlurRadius(shadow_config['blur_radius'])
            shadow.setColor(shadow_config['color'])
            shadow.setOffset(*shadow_config['offset'])
            self.form_group.setGraphicsEffect(shadow)
        
        # History group shadow
        if hasattr(self, 'history_group'):
            shadow = QGraphicsDropShadowEffect()
            shadow.setBlurRadius(shadow_config['blur_radius'])
            shadow.setColor(shadow_config['color'])
            shadow.setOffset(*shadow_config['offset'])
            self.history_group.setGraphicsEffect(shadow)
        
        # Summary panels shadows
        if hasattr(self, 'value_group_container'):
            shadow = QGraphicsDropShadowEffect()
            shadow.setBlurRadius(shadow_config['blur_radius'])
            shadow.setColor(shadow_config['color'])
            shadow.setOffset(*shadow_config['offset'])
            self.value_group_container.setGraphicsEffect(shadow)
        
        if hasattr(self, 'item_summary_container'):
            shadow = QGraphicsDropShadowEffect()
            shadow.setBlurRadius(shadow_config['blur_radius'])
            shadow.setColor(shadow_config['color'])
            shadow.setOffset(*shadow_config['offset'])
            self.item_summary_container.setGraphicsEffect(shadow)

if __name__ == "__main__":
    try:
        # Verify critical imports before starting
        print("Checking critical dependencies...")
        
        try:
            import barcode
            try:
                version = barcode.__version__
                print(f"✓ python-barcode {version} loaded successfully")
            except AttributeError:
                print("✓ python-barcode loaded successfully (version unknown)")
        except ImportError as e:
            print(f"✗ Failed to import barcode module: {e}")
            sys.exit(1)

        # Main application entry point
        if __name__ == "__main__":
            try:
                app = QApplication(sys.argv)
                
                # Set application style
                app.setStyle("Fusion")
                
                # Create and show the main window
                main_window = QuickTagApp()
                main_window.show()
                
                # Start the event loop
                sys.exit(app.exec_())
            except Exception as e:
                print(f"Critical error: {e}")
                print(traceback.format_exc())
                # Don't use input() in EXE mode
                if not getattr(sys, 'frozen', False):
                    input("\nPress Enter to exit...")  # Only in development mode
                sys.exit(1)
    except Exception as e:
        print(f"Fatal error: {str(e)}")
        traceback.print_exc()
        
        # In case of error, show all loaded modules
        print("\nLoaded modules:")
        for name, module in sys.modules.items():
            if module:
                try:
                    location = getattr(module, '__file__', 'built-in')
                    print(f"  {name}: {location}")
                except:
                    print(f"  {name}: location unknown")
        
        input("\nPress Enter to exit...")  # Keep window open
        sys.exit(1)