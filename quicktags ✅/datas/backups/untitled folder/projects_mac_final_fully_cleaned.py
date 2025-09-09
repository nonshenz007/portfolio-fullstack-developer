
import sys
import ctypes
import multiprocessing

def is_already_running():
    mutex = ctypes.windll.kernel32.CreateMutexW(None, 1, "Global\\ProjectsByShenzC_FullSafeApp")
    return ctypes.GetLastError() == 183

#!/usr/local/bin/python3.12
import subprocess
import sys
import platform
import logging
import os
import random
import datetime
import sqlite3

from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QFormLayout, QDialog,
    QPushButton, QMessageBox, QLineEdit, QFileDialog, QLabel,
    QComboBox, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView,
    QCheckBox, QStatusBar
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont
# Add this near the top (after imports)
def get_app_data_path():
    if platform.system() == "Windows":
        base_path = os.path.join(os.environ["APPDATA"], "BarcodeGenerator")
    else:  # Keep macOS compatibility
        base_path = os.path.join(os.path.expanduser("~"), "Library", "Application Support", "BarcodeGenerator")
    os.makedirs(base_path, exist_ok=True)
    return base_path

# Replace these in your code:
DB_PATH = os.path.join(get_app_data_path(), "barcodes.db")
LOG_PATH = os.path.join(get_app_data_path(), "barcode_app.log")

# Update the barcode output directory in generate_barcode_image():
output_dir = os.path.join(get_app_data_path(), "barcodes")

# ------------------------------------------------------
# Auto-install required packages
# ------------------------------------------------------
required = [
    "PyQt5", "python-barcode", "reportlab", "openpyxl", "Pillow"
]
if platform.system() == "Windows":
    required.append("pywin32")

if not getattr(sys, 'frozen', False):  # Only install when not in PyInstaller EXE
    for pkg in required:
        try:
            __import__(pkg.split("-")[0].replace("pyqt5", "PyQt5"))
        except ImportError:
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])

# ------------------------------------------------------
# Logging Setup
# ------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler("barcode_app.log"), logging.StreamHandler()]
)
logger = logging.getLogger(__name__)

TITLE = "Projects \U0001F4C8 by Shenz C (v1 beta)"

LABEL_SIZES = {
    "Small (60x25 mm)": (60, 25),
    "Medium (100x25 mm)": (100, 25),
    "Large (150x50 mm)": (150, 50),
    "A4 Sheet (210x297 mm)": (210, 297),
    "2 Labels (50 * 25mm)": (100, 25),
    "Custom Size": None
}
PRINTERS = [
    "TVS LP46 Neo",
    "Generic PDF Printer",
    "Zebra GK420"
]

# ------------------------------------------------------
# Database
# ------------------------------------------------------
class DatabaseManager:
    def __init__(self, db_path="barcodes.db"):
        self.db_path = db_path
        try:
            self.conn = sqlite3.connect(self.db_path)
            self.create_tables()
            logger.info("Database connected and tables created/verified.")
        except Exception as e:
            logger.error(f"Database connection failed: {e}")
            raise

    def create_tables(self):
        try:
            c = self.conn.cursor()
            c.execute('''
                CREATE TABLE IF NOT EXISTS barcode_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    item_name TEXT,
                    barcode TEXT,
                    sale_price REAL,
                    purchase_price REAL,
                    company_name TEXT,
                    generation_date TEXT,
                    username TEXT DEFAULT 'admin'
                )
            ''')
            c.execute('''
                CREATE TABLE IF NOT EXISTS barcodes (
                    barcode TEXT PRIMARY KEY,
                    item_name TEXT,
                    sale_price REAL,
                    purchase_price REAL,
                    company_name TEXT,
                    quantity INTEGER
                )
            ''')
            c.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    username TEXT PRIMARY KEY,
                    password TEXT,
                    role TEXT
                )
            ''')
            c.execute('SELECT COUNT(*) FROM users')
            if c.fetchone()[0] == 0:
                c.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)',
                          ("admin", "Kungfukenny", "admin"))
            self.conn.commit()
        except Exception as e:
            logger.error(f"Error creating tables: {e}")
            raise

    def add_barcode_history(self, username, item_name, barcode, sale_price, purchase_price, company_name,
                            generation_date):
        try:
            c = self.conn.cursor()
            c.execute('''
                INSERT INTO barcode_history
                (username, item_name, barcode, sale_price, purchase_price, company_name, generation_date)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (username, item_name, barcode, sale_price, purchase_price, company_name, generation_date))
            self.conn.commit()
            logger.info("Barcode history added.")
        except Exception as e:
            logger.error(f"Failed to add barcode history: {e}")
            raise

    def add_or_update_barcode(self, item_name, barcode, sale_price, purchase_price, company_name, quantity):
        try:
            c = self.conn.cursor()
            c.execute('SELECT quantity FROM barcodes WHERE barcode=?', (barcode,))
            result = c.fetchone()
            if result:
                new_quantity = result[0] + quantity
                c.execute('''
                    UPDATE barcodes
                    SET item_name=?, sale_price=?, purchase_price=?, company_name=?, quantity=?
                    WHERE barcode=?
                ''', (item_name, sale_price, purchase_price, company_name, new_quantity, barcode))
            else:
                c.execute('''
                    INSERT INTO barcodes
                    (barcode, item_name, sale_price, purchase_price, company_name, quantity)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (barcode, item_name, sale_price, purchase_price, company_name, quantity))
            self.conn.commit()
            logger.info("Barcode updated/added in barcodes table.")
        except Exception as e:
            logger.error(f"Failed to add or update barcode: {e}")
            raise

    def fetch_history(self, search_filter=""):
        try:
            c = self.conn.cursor()
            query = """
                SELECT
                    id,
                    username,
                    item_name,
                    barcode,
                    sale_price,
                    purchase_price,
                    company_name,
                    generation_date
                FROM barcode_history
            """
            params = ()
            if search_filter:
                query += " WHERE item_name LIKE ? OR barcode LIKE ?"
                params = (f"%{search_filter}%", f"%{search_filter}%")
            query += " ORDER BY id DESC"
            c.execute(query, params)
            history = c.fetchall()
            logger.info("History fetched from database.")
            return history
        except Exception as e:
            logger.error(f"Failed to fetch history: {e}")
            return []

    def clear_history(self):
        try:
            c = self.conn.cursor()
            c.execute('DELETE FROM barcode_history')
            self.conn.commit()
            logger.info("Barcode history cleared.")
        except Exception as e:
            logger.error(f"Failed to clear history: {e}")
            raise

    def add_user(self, username, password, role="user"):
        try:
            c = self.conn.cursor()
            c.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)', (username, password, role))
            self.conn.commit()
            logger.info(f"User   {username} added.")
        except Exception as e:
            logger.error(f"Failed to add user: {e}")
            raise

    def remove_user(self, username):
        try:
            c = self.conn.cursor()
            c.execute('DELETE FROM users WHERE username=?', (username,))
            self.conn.commit()
            logger.info(f"User   {username} removed.")
        except Exception as e:
            logger.error(f"Failed to remove user: {e}")
            raise

    def fetch_users(self):
        try:
            c = self.conn.cursor()
            c.execute('SELECT username, role FROM users ORDER BY username ASC')
            users = c.fetchall()
            logger.info("Fetched users from database.")
            return users
        except Exception as e:
            logger.error(f"Failed to fetch users: {e}")
            return []

    def validate_user(self, username, password):
        try:
            c = self.conn.cursor()
            c.execute('SELECT username, role FROM users WHERE username=? AND password=?', (username, password))
            result = c.fetchone()
            logger.info(f"User   {username} validation attempted.")
            return result
        except Exception as e:
            logger.error(f"User   validation failed: {e}")
            return None

# ------------------------------------------------------
# Barcode Generation with Code128
# ------------------------------------------------------
try:
    from barcode import get_barcode_class
    from barcode.writer import ImageWriter
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "python-barcode"])
    from barcode import get_barcode_class
    from barcode.writer import ImageWriter

def sanitize_barcode_input(barcode_str):
    """
    Code128 typically supports ASCII 32-126.
    We'll remove any out-of-range chars, to avoid errors.
    """
    valid_str = "".join(ch for ch in barcode_str if 32 <= ord(ch) <= 126)
    return valid_str

def generate_barcode_image(barcode_str):
    """
    Generate a Code128 barcode as a PNG using python-barcode.
    Raises an Exception if the PNG is not actually created.
    """
    # 1) Sanitize for Code128
    safe_barcode = sanitize_barcode_input(barcode_str)
    if not safe_barcode:
        raise ValueError("Barcode has no valid Code128 characters after sanitizing.")

    try:
        Code128 = get_barcode_class('code128')
        output_dir = os.path.join(os.path.expanduser("~"), "Desktop", "barcodes")
        os.makedirs(output_dir, exist_ok=True)

        base_filename = os.path.join(output_dir, safe_barcode)
        code128_obj = Code128(safe_barcode, writer=ImageWriter())
        saved_path = code128_obj.save(base_filename)

        if not os.path.isfile(saved_path):
            raise OSError(f"Barcode file not created at: {saved_path}")

        logger.info("Barcode image generated at %s", saved_path)
        return saved_path

    except Exception as e:
        logger.error(f"Barcode generation failed: {e}")
        raise

# ------------------------------------------------------
# PDF Generation (ReportLab)
# ------------------------------------------------------
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm

def draw_single_label(c, x_offset_mm, y_offset_mm, w_label_mm, h_label_mm,
                      barcode_path, details):
    item = details.get("item", "")
    sale = details.get("sale", "")
    purchase = details.get("purchase", "")
    company = details.get("company", "")

    c.rect(x_offset_mm * mm, y_offset_mm * mm, w_label_mm * mm, h_label_mm * mm)
    c.setFont("Helvetica", 8)

    c.drawCentredString((x_offset_mm + w_label_mm/2) * mm,
                        (y_offset_mm + h_label_mm - 5) * mm,
                        item)
    c.drawImage(barcode_path,
                (x_offset_mm + w_label_mm/2 - 20) * mm,
                (y_offset_mm + h_label_mm - 16) * mm,
                40 * mm, 8 * mm)

    c.drawCentredString((x_offset_mm + w_label_mm/2) * mm,
                        (y_offset_mm + h_label_mm - 20) * mm,
                        f"Sale: {sale} | {company}")

def generate_label_pdf(pdf_path, w_mm, h_mm, barcode_path, details):
    c = canvas.Canvas(pdf_path, pagesize=(w_mm*mm, h_mm*mm))
    count = details.get("count", 1)

    if (w_mm, h_mm) == (100, 25):
        i = 0
        while i < count:
            draw_single_label(c, 0, 0, 50, 25, barcode_path, details)
            i += 1
            if i < count:
                draw_single_label(c, 50, 0, 50, 25, barcode_path, details)
                i += 1
            c.showPage()
        c.save()
        logger.info("PDF generated (2-up) at %s", pdf_path)
        return

    for _ in range(count):
        draw_single_label(c, 0, 0, w_mm, h_mm, barcode_path, details)
        c.showPage()
    c.save()
    logger.info("PDF generated at %s", pdf_path)

# ------------------------------------------------------
# Excel Update (Append or Update)
# ------------------------------------------------------
def update_excel(item_code, item_name, sale_price, purchase_price, company_name, generation_date):
    from openpyxl import Workbook, load_workbook

    file_name = "Items.xlsx"
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Items"
        ws.append(["Item Code", "Item Name", "Sale Price", "Purchase Price", "Company", "Generation Date"])
        wb.save(file_name)

    wb = load_workbook(file_name)
    ws = wb.active
    found = False

    # check if item_code already exists
    for row in ws.iter_rows(min_row=2, values_only=False):
        code_cell = row[0]
        if code_cell.value == item_code:
            row[1].value = item_name
            row[2].value = sale_price
            row[3].value = purchase_price
            row[4].value = company_name
            row[5].value = generation_date
            found = True
            break

    if not found:
        ws.append([item_code, item_name, sale_price, purchase_price, company_name, generation_date])

    # -------------- Concurrency Fix: Retry on lock --------------
    for attempt in range(3):
        try:
            wb.save(file_name)
            logger.info("Excel updated: item_code=%s (found=%s)", item_code, found)
            break
        except Exception as e:
            logger.error(f"Excel save failed: {e}")
            ret = QMessageBox.question(None, "Excel Locked",
                f"Unable to save {file_name} (maybe open/locked?).\n"
                 "Close the file and click 'Retry', or 'Cancel' to skip updating.",
                 QMessageBox.Retry | QMessageBox.Cancel)
            if ret == QMessageBox.Retry:
                continue
            else:
                # user canceled, skip
                break

# ------------------------------------------------------
# Windows PDF Print
# ------------------------------------------------------
def print_pdf_windows(pdf_path):
    if platform.system() != "Windows":
        logger.warning("print_pdf_windows called on non-Windows system.")
        return
    try:
        os.startfile(pdf_path, "print")
        logger.info(f"Sent PDF to default printer: {pdf_path}")
    except Exception as e:
        logger.error(f"Windows PDF print failed: {e}")

# ------------------------------------------------------
# Login Dialog
# ------------------------------------------------------
class LoginDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(TITLE)
        self.setFixedSize(320, 220)
        self.setStyleSheet("""
            QDialog {
                background-color: #1c1c1c;
                color: #e0e0e0;
                font-family: 'Segoe UI', sans-serif;
            }
            QLabel {
                font-size: 16px;
                font-weight: 500;
                color: #ffffff;
            }
            QLineEdit {
                background-color: #2a2a2a;
                border: 1px solid #444444;
                border-radius: 6px;
                padding: 6px;
                color: #ffffff;
            }
            QPushButton {
                background-color: #333333;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                color: #ffffff;
                font-weight: bold;
                font-size: 14px;
            }
        """)
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(10)

        title_label = QLabel(TITLE)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 18px; font-weight: 600; margin-bottom: 8px;")
        layout.addWidget(title_label)

        form_layout = QFormLayout()
        self.usernameField = QLineEdit()
        form_layout.addRow("Username:", self.usernameField)
        self.passwdField = QLineEdit()
        self.passwdField.setEchoMode(QLineEdit.Password)
        form_layout.addRow("Password:", self.passwdField)
        layout.addLayout(form_layout)

        btn = QPushButton("Login")
        btn.clicked.connect(self.check_credentials)
        layout.addWidget(btn, alignment=Qt.AlignCenter)
        self.setLayout(layout)
        self.user = None
        self.role = None

    def check_credentials(self):
        username = self.usernameField.text().strip()
        password = self.passwdField.text().strip()
        if not username or not password:
            QMessageBox.warning(self, "Access Denied", "Enter both username and password")
            return
        try:
            db = DatabaseManager()
            result = db.validate_user(username, password)
            if result:
                self.user, self.role = result
                logger.info(f"User   {username} logged in successfully.")
                self.accept()
            else:
                QMessageBox.warning(self, "Access Denied", "Incorrect username or password")
                self.passwdField.clear()
        except Exception as e:
            QMessageBox.warning(self, "Error", "An error occurred during login.")
            logger.error(f"Login error: {e}")

# ------------------------------------------------------
# Main Application
# ------------------------------------------------------
class BarcodeGeneratorUI(QWidget):
    def __init__(self, current_user, current_role):
        super().__init__()
        self.current_user = current_user
        self.current_role = current_role

        try:
            self.db = DatabaseManager()
        except Exception as e:
            QMessageBox.critical(self, "Error", "Failed to initialize the database.")
            sys.exit(1)
        self.initUI()

    def initUI(self):
        self.setWindowTitle(TITLE)
        self.setGeometry(250, 250, 1100, 600)
        self.setStyleSheet("""
            QWidget {
                background-color: #1c1c1c;
                color: #e0e0e0;
                font-family: 'Segoe UI', sans-serif;
                font-size: 14px;
            }
            QLineEdit, QComboBox {
                background-color: #2a2a2a;
                border: 1px solid #444444;
                border-radius: 6px;
                padding: 6px;
                color: #ffffff;
            }
            QPushButton {
                background-color: #333333;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                color: #ffffff;
                font-weight: bold;
            }
            QLabel {
                color: #ffffff;
                font-size: 14px;
            }
            QTableWidget {
                background-color: #2a2a2a;
                border: 1px solid #444444;
                border-radius: 6px;
            }
        """)
        main_layout = QVBoxLayout()

        heading_label = QLabel(TITLE)
        heading_label.setAlignment(Qt.AlignCenter)
        heading_label.setStyleSheet("font-size: 18px; font-weight: 600; margin-bottom: 8px;")
        main_layout.addWidget(heading_label)

        # Row 1: item, purchase, profit, sale
        row1_layout = QHBoxLayout()
        row1_layout.addWidget(QLabel("Item Name:"))
        self.itemField = QLineEdit()
        row1_layout.addWidget(self.itemField)

        row1_layout.addWidget(QLabel("Purchase Price:"))
        self.purchaseField = QLineEdit()
        row1_layout.addWidget(self.purchaseField)

        row1_layout.addWidget(QLabel("Profit (%):"))
        self.profitField = QLineEdit()
        self.profitField.setPlaceholderText("optional")
        row1_layout.addWidget(self.profitField)

        row1_layout.addWidget(QLabel("Sale Price:"))
        self.saleField = QLineEdit()
        row1_layout.addWidget(self.saleField)
        calcBtn = QPushButton("Calc")
        calcBtn.clicked.connect(self.apply_profit_calculator)
        row1_layout.addWidget(calcBtn)
        main_layout.addLayout(row1_layout)

        # Row 2: company, count, barcode
        row2_layout = QHBoxLayout()
        row2_layout.addWidget(QLabel("Company:"))
        self.companyField = QLineEdit()
        row2_layout.addWidget(self.companyField)

        row2_layout.addWidget(QLabel("No. of Labels:"))
        self.countField = QLineEdit()
        row2_layout.addWidget(self.countField)

        row2_layout.addWidget(QLabel("Barcode (optional):"))
        self.barcodeField = QLineEdit()
        self.barcodeField.setPlaceholderText("If empty, generate random 12-digit code")
        row2_layout.addWidget(self.barcodeField)

        main_layout.addLayout(row2_layout)

        # Row 3: label size
        row3_layout = QHBoxLayout()
        row3_layout.addWidget(QLabel("Label Size:"))
        self.sizeBox = QComboBox()
        for label in LABEL_SIZES:
            self.sizeBox.addItem(label)
        row3_layout.addWidget(self.sizeBox)

        row3_layout.addWidget(QLabel("Custom W x H (mm):"))
        self.customWidth = QLineEdit()
        self.customHeight = QLineEdit()
        self.customWidth.setVisible(False)
        self.customHeight.setVisible(False)
        row3_layout.addWidget(self.customWidth)
        row3_layout.addWidget(self.customHeight)

        row3_layout.addWidget(QLabel("Printer Profile:"))
        self.printerBox = QComboBox()
        for p in PRINTERS:
            self.printerBox.addItem(p)
        row3_layout.addWidget(self.printerBox)

        self.continuousRollCheck = QCheckBox("Continuous Roll")
        row3_layout.addWidget(self.continuousRollCheck)
        main_layout.addLayout(row3_layout)

        # Buttons
        btn_layout = QHBoxLayout()
        genBtn = QPushButton("Generate Labels PDF")
        printNowBtn = QPushButton("Print (Win10)")
        reprintBtn = QPushButton("Reprint Selected")
        clearHistBtn = QPushButton("Clear History")
        analyticsBtn = QPushButton("Analytics Dashboard")
        openExcelBtn = QPushButton("Open Items Excel")

        genBtn.clicked.connect(self.generate_pdf)
        printNowBtn.clicked.connect(self.print_labels_win)
        reprintBtn.clicked.connect(self.reprint_selected_barcode)
        clearHistBtn.clicked.connect(self.clear_history)
        analyticsBtn.clicked.connect(self.show_analytics)
        openExcelBtn.clicked.connect(self.open_excel_file)

        btn_layout.addWidget(genBtn)
        btn_layout.addWidget(printNowBtn)
        btn_layout.addWidget(reprintBtn)
        btn_layout.addWidget(clearHistBtn)
        btn_layout.addWidget(analyticsBtn)
        btn_layout.addWidget(openExcelBtn)

        if self.current_role == "admin":
            manageUsersBtn = QPushButton("Manage Users")
            manageUsersBtn.clicked.connect(self.manage_users)
            btn_layout.addWidget(manageUsersBtn)

        main_layout.addLayout(btn_layout)

        # Search + History
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Search History:"))
        self.searchField = QLineEdit()
        self.searchField.textChanged.connect(self.refresh_history)
        search_layout.addWidget(self.searchField)
        main_layout.addLayout(search_layout)

        self.historyTable = QTableWidget()
        self.historyTable.setColumnCount(8)
        self.historyTable.setHorizontalHeaderLabels(
            ["ID", "User", "Item Name", "Barcode", "Sale Price", "Purchase Price", "Company", "Date"]
        )
        self.historyTable.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        main_layout.addWidget(QLabel("Barcode Generation History:"))
        main_layout.addWidget(self.historyTable)

        self.statusBar = QStatusBar()
        main_layout.addWidget(self.statusBar)
        self.setLayout(main_layout)

        self.sizeBox.currentTextChanged.connect(self.toggleCustomSize)
        self.refresh_history()

    def toggleCustomSize(self):
        is_custom = (self.sizeBox.currentText() == "Custom Size")
        self.customWidth.setVisible(is_custom)
        self.customHeight.setVisible(is_custom)

    def apply_profit_calculator(self):
        # Replace commas with periods to handle locale
        p_txt = self.purchaseField.text().strip().replace(",", ".")
        profit_txt = self.profitField.text().strip().replace(",", ".")

        try:
            purchase_val = float(p_txt)
        except ValueError:
            QMessageBox.warning(self, "Error", "Enter a valid numeric Purchase Price.")
            return

        if not profit_txt:
            QMessageBox.warning(self, "Error", "Enter a valid numeric Profit (%) or type 0.")
            return

        try:
            profit_percent = float(profit_txt)
        except ValueError:
            QMessageBox.warning(self, "Error", "Profit (%) must be numeric.")
            return

        calc_sale = purchase_val + (purchase_val * (profit_percent / 100.0))
        self.saleField.setText(f"{calc_sale:.2f}")

    def get_label_size(self):
        label = self.sizeBox.currentText()
        if label == "Custom Size":
            w_txt = self.customWidth.text().strip().replace(",", ".")
            h_txt = self.customHeight.text().strip().replace(",", ".")
            try:
                w = float(w_txt)
                h = float(h_txt)
                return w, h
            except ValueError:
                QMessageBox.warning(self, "Error", "Enter valid numeric custom dimensions.")
                logger.error("Invalid custom size.")
                return None
        return LABEL_SIZES[label]

    def generate_pdf(self):
        # Gather user input
        item_name = self.itemField.text().strip()
        purchase_txt = self.purchaseField.text().strip().replace(",", ".")
        sale_txt = self.saleField.text().strip().replace(",", ".")
        company = self.companyField.text().strip()
        count_txt = self.countField.text().strip()
        barcode_in = self.barcodeField.text().strip()

        # Basic validations
        if not item_name or not purchase_txt or not sale_txt or not company or not count_txt:
            QMessageBox.warning(self, "Error", "All fields except Barcode/Profit must be filled out.")
            return

        try:
            purchase_val = float(purchase_txt)
            sale_val = float(sale_txt)
            count_val = int(count_txt)
        except ValueError:
            QMessageBox.warning(self, "Error", "Purchase, Sale, and No. of Labels must be numeric.")
            return

        size = self.get_label_size()
        if not size:
            return
        w_mm, h_mm = size

        # If no user-specified barcode, generate random
        if barcode_in:
            barcode_str = barcode_in
        else:
            barcode_str = "".join(random.choices("0123456789", k=12))

        # Generate barcode image
        try:
            barcode_path = generate_barcode_image(barcode_str)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Barcode generation failed:\n{e}")
            return

        # Choose folder for PDF
        folder = QFileDialog.getExistingDirectory(self, "Select folder to save PDF")
        if not folder:
            return
        pdf_path = os.path.join(folder, f"labels_{barcode_str}.pdf")

        details = {
            "item": item_name,
            "sale": sale_txt,
            "purchase": purchase_txt,
            "company": company,
            "count": count_val
        }

        # Generate PDF
        try:
            generate_label_pdf(pdf_path, w_mm, h_mm, barcode_path, details)
            QMessageBox.information(self, "Done", f"PDF saved:\n{pdf_path}")
            self.statusBar.showMessage("PDF generated successfully.", 5000)
        except Exception as ex:
            QMessageBox.warning(self, "Error", f"PDF generation failed:\n{ex}")
            logger.error(f"PDF generation error: {ex}")
            return

        # Update DB
        gen_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            self.db.add_barcode_history(self.current_user, item_name, barcode_str,
                                        sale_val, purchase_val, company, gen_date)
            self.db.add_or_update_barcode(item_name, barcode_str, sale_val, purchase_val, company, count_val)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Database update failed:\n{e}")
            logger.error(f"Database update error: {e}")

        # Excel update
        try:
            update_excel(
                item_code=barcode_str,
                item_name=item_name,
                sale_price=sale_txt,
                purchase_price=purchase_txt,
                company_name=company,
                generation_date=gen_date
            )
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Excel update failed:\n{e}")
            logger.error(f"Excel update error: {e}")

        self.refresh_history()

    def print_labels_win(self):
        if platform.system() != "Windows":
            QMessageBox.warning(self, "Not Windows", "This print function is only for Windows.")
            return
    file_dialog = QFileDialog()
        file_dialog.setNameFilter("PDF Files (*.pdf)")
    sys.exit(app.exec_())
    pdf_files = file_dialog.selectedFiles()
    if pdf_files:
                pdf_path = pdf_files[0]
                print_pdf_windows(pdf_path)
                QMessageBox.information(self, "Printing", f"Sent to default printer:\n{pdf_path}")

    def reprint_selected_barcode(self):
        pass

    def refresh_history(self):
        search_text = self.searchField.text().strip()
        history = self.db.fetch_history(search_text)
        self.historyTable.setRowCount(len(history))

        header = self.historyTable.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)

        for i, row in enumerate(history):
            for j, value in enumerate(row):
                cell_item = QTableWidgetItem(str(value))
                cell_item.setTextAlignment(Qt.AlignCenter)
                self.historyTable.setItem(i, j, cell_item)
        self.historyTable.resizeRowsToContents()

    def clear_history(self):
        reply = QMessageBox.question(
            self,
            "Clear History",
            "Are you sure you want to clear the entire barcode generation history?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            try:
                self.db.clear_history()
                self.refresh_history()
                QMessageBox.information(self, "Clear History", "Barcode generation history cleared.")
                self.statusBar.showMessage("History cleared.", 5000)
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to clear history: {e}")
                logger.error(f"Clear history error: {e}")

    def open_excel_file(self):
        excel_file = "Items.xlsx"
        if not os.path.exists(excel_file):
            QMessageBox.information(self, "Excel Not Found", "Items.xlsx does not exist yet.")
            return

        try:
            if platform.system() == "Windows":
                os.startfile(excel_file)
            elif platform.system() == "Darwin":
                subprocess.run(["open", excel_file])
            else:
                subprocess.run(["xdg-open", excel_file])
        except Exception as e:
            logger.error(f"Failed to open {excel_file}: {e}")
            QMessageBox.warning(self, "Error", f"Could not open {excel_file}:\n{e}")

    def show_analytics(self):
        pass

    def manage_users(self):
        pass

# ------------------------------------------------------
# Main
# ------------------------------------------------------

def main():


if __name__ == '__main__':


import win32print
import win32con

def print_to_tvs_lp46(text="Hello World", width_mm=100, height_mm=50, copies=1):
    printer_name = "TVS LP46 Neo"  # Adjust if printer shows differently in Control Panel
    try:
        # Convert mm to dots (203 DPI = 8 dots/mm)
        width_dots = int(width_mm * 8)
        height_dots = int(height_mm * 8)

        # TSPL/BPLE command example
        commands = f"""
SIZE {width_mm} mm,{height_mm} mm
GAP 2 mm,0 mm
DIRECTION 1
CLS
TEXT 100,100,"3",0,1,1,"{text}"
PRINT {copies}
"""
        # Open printer and send raw data
        handle = win32print.OpenPrinter(printer_name)
        job = win32print.StartDocPrinter(handle, 1, ("LabelJob", None, "RAW"))
        win32print.StartPagePrinter(handle)
        win32print.WritePrinter(handle, commands.encode("utf-8"))
        win32print.EndPagePrinter(handle)
        win32print.EndDocPrinter(handle)
        win32print.ClosePrinter(handle)
        print("✅ Label sent to printer.")

    except Exception as e:
        print(f"❌ Printer error: {e}")

if __name__ == "__main__":
    if is_already_running():
        print("App already running.")
        sys.exit(0)
    app = QApplication(sys.argv)
    font = QFont("Segoe UI", 10)
    app.setFont(font)
    login = LoginDialog()
        window = BarcodeGeneratorUI(login.user, login.role)
        window.show()
    else:
        sys.exit(0)
    # Replace with actual startup call
    pass
    multiprocessing.freeze_support()
    if is_already_running():
        print("Another instance is already running. Exiting.")
        sys.exit(0)
    main()