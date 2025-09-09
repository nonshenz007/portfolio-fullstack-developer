import os
import sys
import time
import sqlite3
import datetime
import random
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)

from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                            QLabel, QLineEdit, QPushButton, QMessageBox, QFileDialog, QGroupBox,
                            QScrollArea, QSizePolicy, QSpacerItem, QDialog, QTableWidget, QTableWidgetItem, QHeaderView)
from PyQt5.QtGui import QPixmap, QImage, QPainter, QFont, QFontDatabase, QIcon
from PyQt5.QtCore import Qt, QSize, QSizeF, QRectF
import barcode
from barcode.writer import ImageWriter
import openpyxl
from PIL import Image, ImageDraw, ImageFont
from io import BytesIO
import subprocess

class QuickTagApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Quicktags – AutoGeek Edition by Shenz")
        self.setMinimumSize(600, 400)
   
        # Initialize database first
        self.init_db()
        
        # Then create central widget and layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        self.layout = QVBoxLayout(central_widget)
        
        # Now you can set layout properties
        self.layout.setSpacing(8)
        self.layout.setContentsMargins(15, 15, 15, 15)
        
        # Add header title
        self.add_header()
        
        # Create form fields
        self.create_form()
        
        # Create buttons
        self.create_buttons()
        
        # Create history table
        self.create_history_table()
        
        # Load existing items from database
        self.load_item_history()
        
        # Load fonts
        self.load_fonts()

    def create_form(self):
        form_group = QGroupBox("Product Information")
        form_group.setMinimumWidth(400)
        form_layout = QVBoxLayout()
        form_layout.setSpacing(8)
        
        # Item Name
        self.item_name_label = QLabel("Item Name (required):")
        self.item_name_input = QLineEdit()
        form_layout.addWidget(self.item_name_label)
        form_layout.addWidget(self.item_name_input)
        
        # Purchase Price
        self.purchase_price_label = QLabel("Purchase Price (default: 1):")
        self.purchase_price_input = QLineEdit("1")
        form_layout.addWidget(self.purchase_price_label)
        form_layout.addWidget(self.purchase_price_input)
        
        # Profit %
        self.profit_percent_label = QLabel("Profit % (default: 100):")
        self.profit_percent_input = QLineEdit("100")
        form_layout.addWidget(self.profit_percent_label)
        form_layout.addWidget(self.profit_percent_input)
        
        # Sale Price
        self.sale_price_label = QLabel("Sale Price (optional, overrides %Profit):")
        self.sale_price_input = QLineEdit()
        form_layout.addWidget(self.sale_price_label)
        form_layout.addWidget(self.sale_price_input)
        
        # Stock Quantity
        self.stock_quantity_label = QLabel("Stock Quantity (default: 1):")
        self.stock_quantity_input = QLineEdit("1")
        form_layout.addWidget(self.stock_quantity_label)
        form_layout.addWidget(self.stock_quantity_input)
        
        # Barcode (hidden override)
        self.barcode_override_label = QLabel("Barcode Override:")
        self.barcode_override_input = QLineEdit()
        self.barcode_override_label.hide()
        self.barcode_override_input.hide()
        form_layout.addWidget(self.barcode_override_label)
        form_layout.addWidget(self.barcode_override_input)

        form_group.setLayout(form_layout)
        self.layout.addWidget(form_group)

        

        
    def add_header(self):
        header = QLabel("Quicktags – AutoGeek Edition by Shenz")
        header.setFont(QFont("Arial", 18, QFont.Bold))  # Changed from 14 to 18
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("padding: 5px;")
        self.layout.addWidget(header)
        
    def init_db(self):
        try:
            # Get the correct database path for both development and EXE modes
            if getattr(sys, 'frozen', False):
                # Running in EXE mode
                db_path = os.path.join(os.path.dirname(sys.executable), 'quicktag.db')
            else:
                # Running in development mode
                db_path = 'quicktag.db'
                
            self.conn = sqlite3.connect(db_path)
            self.cursor = self.conn.cursor()
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    item_name TEXT UNIQUE,
                    barcode TEXT UNIQUE,
                    purchase_price INTEGER,
                    profit_percent INTEGER,
                    sale_price INTEGER,
                    stock_quantity INTEGER,
                    created_at TEXT
                )
            ''')
            self.conn.commit()
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"Failed to initialize database: {str(e)}")
            sys.exit(1)
        
    def load_fonts(self):
        # Try to load from bundled resources first
        try:
            self.font_bold = QFont("Arial", 16, QFont.Bold)
            self.font_regular = QFont("Arial", 15)
            # Store font sizes as attributes
            self.font_bold_size = 16
            self.font_regular_size = 15
        except:
            # Fallback to system fonts
            font_names = ["Arial", "Helvetica", "San Francisco", ".SF NS Text"]
            self.font_bold = QFont()
            self.font_regular = QFont()
            
            for name in font_names:
                if self.font_bold.exactMatch(): break
                self.font_bold = QFont(name, 16, QFont.Bold)
                self.font_bold_size = 16
                
            for name in font_names:
                if self.font_regular.exactMatch(): break
                self.font_regular = QFont(name, 15)
                self.font_regular_size = 15
        
   

    def create_buttons(self):
        button_container = QWidget()
        button_layout = QHBoxLayout(button_container)
        button_layout.setSpacing(8)
        button_layout.setContentsMargins(0, 0, 0, 0)
        
        button_style = """
            QPushButton {
                font-size: 14px;
                font-weight: 600;
                padding: 8px 12px;
                border-radius: 6px;
                border: 1px solid #ddd;
                background: white;
            }
            QPushButton:hover {
                background: #f0f0f0;
            }
            #add_item_button {
                background: #28a745;
                color: white;
                border-color: #218838;
            }
            #preview_button {
                background: #17a2b8;
                color: white;
                border-color: #117a8b;
            }
            #print_button, #print_all_button {
                background: #007bff;
                color: white;
                border-color: #0069d9;
            }
            #export_excel_button {
                background: #6c757d;
                color: white;
                border-color: #5a6268;
            }
            #clear_button {
                background: #dc3545;
                color: white;
                border-color: #c82333;
            }
        """
        
        # Remove button styling
        # button_container.setStyleSheet(button_style)
        
        # Add Item button
        self.add_item_button = QPushButton("➕ Add Item")
        self.add_item_button.setObjectName("add_item_button")
        self.add_item_button.clicked.connect(self.add_item)
        button_layout.addWidget(self.add_item_button)  # Changed from self.button_layout
        
        # Preview button
        self.preview_button = QPushButton("👁️ Preview")
        self.preview_button.clicked.connect(self.preview_label)
        button_layout.addWidget(self.preview_button)  # Changed from self.button_layout
        
        # Print button
        self.print_button = QPushButton("🖨️ Print")
        self.print_button.clicked.connect(self.print_label)
        button_layout.addWidget(self.print_button)  # Changed from self.button_layout
        
        # Print All button
        self.print_all_button = QPushButton("🧾 Print All")
        self.print_all_button.clicked.connect(self.print_all_labels)
        button_layout.addWidget(self.print_all_button)  # Changed from self.button_layout
        
        # Export Excel button
        self.export_excel_button = QPushButton("📁 Export Excel")
        self.export_excel_button.clicked.connect(self.export_to_excel)
        button_layout.addWidget(self.export_excel_button)  # Changed from self.button_layout
        
        # Add Clear button
        self.clear_button = QPushButton("🗑️ Clear")
        self.clear_button.clicked.connect(self.clear_items)
        button_layout.addWidget(self.clear_button)
        
        # Add Barcode Toggle button
        self.toggle_barcode_button = QPushButton("🔢 Toggle Barcode")
        self.toggle_barcode_button.clicked.connect(self.toggle_barcode_override)
        button_layout.addWidget(self.toggle_barcode_button)
        
        self.layout.addWidget(button_container)  # Changed from self.central_layout
        
    def create_history_table(self):
        table_group = QGroupBox("Session History")
        table_layout = QVBoxLayout(table_group)
        
        self.history_table = QTableWidget()
        self.history_table.setColumnCount(4)
        self.history_table.setHorizontalHeaderLabels(["Item Name", "Date", "Sale Price", "Stock Qty"])
        
        # Revert to original table settings
        self.history_table.verticalHeader().setVisible(False)
        self.history_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.history_table.setSelectionMode(QTableWidget.NoSelection)
        self.history_table.setShowGrid(False)
        self.history_table.setAlternatingRowColors(True)
        
        # Revert to original row height
        self.history_table.verticalHeader().setDefaultSectionSize(40)
        
        # Configure fonts
        # Update font sizes (reduced by ~5%)
        header_font = QFont("Arial", 15, QFont.Bold)  # Reduced from 16 to 15
        cell_font = QFont("Arial", 13)  # Reduced from 14 to 13
        self.history_table.horizontalHeader().setFont(header_font)
        self.history_table.setFont(cell_font)
        
        # Configure column behavior
        self.history_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        table_layout.addWidget(self.history_table)
        self.layout.addWidget(table_group)

    def add_item_to_table(self, item_name, created_at, price, quantity):
        row = self.history_table.rowCount()
        self.history_table.insertRow(row)
        
        try:
            # First try parsing with the expected format
            formatted_date = datetime.datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y")
        except ValueError:
            try:
                # If that fails, try parsing as a barcode number (fallback)
                formatted_date = created_at
            except:
                # If all fails, use current date
                formatted_date = datetime.datetime.now().strftime("%d-%m-%Y")
        
        for i, value in enumerate([item_name, formatted_date, f"₹{price}", str(quantity)]):
            item = QTableWidgetItem(value)
            item.setTextAlignment(Qt.AlignCenter)
            self.history_table.setItem(row, i, item)

    def toggle_barcode_override(self):
        """Toggle visibility of barcode override fields"""
        if self.barcode_override_label.isVisible():
            self.barcode_override_label.hide()
            self.barcode_override_input.hide()
        else:
            self.barcode_override_label.show()
            self.barcode_override_input.show()
            
    def add_item(self):
        item_name = self.item_name_input.text().strip()
        if not item_name:
            QMessageBox.warning(self, "Error", "Item Name is required!")
            return
            
        # Check for duplicate item name
        self.cursor.execute("SELECT COUNT(*) FROM items WHERE item_name=?", (item_name,))
        if self.cursor.fetchone()[0] > 0:
            QMessageBox.warning(self, "Error", "Item Name already exists!")
            return
            
        # Generate or use override barcode
        barcode_value = self.barcode_override_input.text().strip() or self.generate_barcode()
        
        # Check for duplicate barcode
        self.cursor.execute("SELECT COUNT(*) FROM items WHERE barcode=?", (barcode_value,))
        if self.cursor.fetchone()[0] > 0:
            QMessageBox.warning(self, "Error", "Barcode already exists!")
            return
            
        # Get other values
        try:
            purchase_price = int(float(self.purchase_price_input.text() or 1))  # Convert to int
            if purchase_price <= 0:
                raise ValueError("Purchase price must be positive")
                    
            profit_percent = int(float(self.profit_percent_input.text() or 100))  # Convert to int
            if profit_percent < 0:
                raise ValueError("Profit percent cannot be negative")
                    
            stock_quantity = int(self.stock_quantity_input.text() or 1)
            if stock_quantity <= 0:
                raise ValueError("Stock quantity must be positive")
            
            # Calculate sale price
            sale_price_input = self.sale_price_input.text().strip()
            if sale_price_input:
                sale_price = int(float(sale_price_input))  # Convert to int
            else:
                sale_price = int(purchase_price + (purchase_price * profit_percent / 100))  # Convert to int
                
            # Add to database
            created_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.cursor.execute('''
                INSERT INTO items 
                (item_name, barcode, purchase_price, profit_percent, sale_price, stock_quantity, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (item_name, barcode_value, purchase_price, profit_percent, sale_price, stock_quantity, created_at))
            self.conn.commit()
            
            # Add to history table
            self.add_item_to_table(item_name, barcode_value, sale_price, stock_quantity)
            
            # Remove session_items logic and just refresh from DB
            self.load_item_history()
            
            QMessageBox.information(self, "Success", "Item added successfully!")
            
        except ValueError as e:
            QMessageBox.warning(self, "Error", f"Invalid input: {str(e)}")
            
    def generate_barcode(self):
        while True:
            barcode_value = str(random.randint(100000000000, 999999999999))
            self.cursor.execute("SELECT COUNT(*) FROM items WHERE barcode=?", (barcode_value,))
            if self.cursor.fetchone()[0] == 0:
                return barcode_value
                
    def generate_barcode_image(self, barcode_value):
        """Generate high-quality 1-bit barcode image following KipTag specs"""
        try:
            # Configure barcode writer for KipTag compatibility
            writer = ImageWriter()
            writer.set_options({
                'module_width': 0.25,  # Thinner bars for KipTag
                'module_height': 25,    # Standard KipTag height
                'font_size': 0,         # No text
                'text_distance': 0,     # Remove text completely
                'quiet_zone': 5,        # KipTag standard quiet zone
                'background': 'white',  # Ensure white background
                'foreground': 'black'   # Ensure black bars
            })
            
            # Generate to memory buffer
            buffer = BytesIO()
            ean = barcode.get('ean13', barcode_value, writer=writer)
            ean.write(buffer)
            buffer.seek(0)
            
            # Process image for KipTag compatibility
            barcode_img = Image.open(buffer)
            barcode_img = barcode_img.resize((280, 55), Image.BOX)  # KipTag standard size
            return barcode_img.convert('1')  # Already returns 1-bit image
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to generate barcode: {str(e)}")
            return Image.new('1', (280, 55), 1)  # Return blank KipTag-sized image
    
    def generate_row_label_image(self, items, output_path):
        """Generate multi-row label image with 2 labels per row"""
        try:
            from math import ceil
            # Calculate required rows and canvas height
            rows = ceil(len(items) / 2)
            height = rows * 200
            
            # Create blank image (800xheight @ 203 DPI)
            img = Image.new('RGB', (800, height), 'white')
            draw = ImageDraw.Draw(img)
            
            # Load KipTag standard fonts
            try:
                title_font = ImageFont.truetype("arialbd.ttf", 16)
                price_font = ImageFont.truetype("arialbd.ttf", 20)
                brand_font = ImageFont.truetype("arialbd.ttf", 14)
            except:
                title_font = ImageFont.load_default()
                price_font = ImageFont.load_default()
                brand_font = ImageFont.load_default()
            
            # Process each item
            for index, item in enumerate(items):
                row = index // 2
                col = index % 2
                x_offset = col * 400
                y_offset = row * 200
                
                # Draw item name (centered in 400px block)
                text_width = draw.textlength(item['item_name'], font=title_font)
                draw.text(
                    (x_offset + (400 - text_width)/2, y_offset + 15),
                    item['item_name'],
                    font=title_font,
                    fill='black'
                )
                
                # Draw barcode (fixed position within 400px block)
                barcode_img = self.generate_barcode_image(item['barcode'])
                img.paste(barcode_img, (x_offset + 60, y_offset + 40))
                
                # Draw price (centered in 400px block)
                price_text = f"₹{item['sale_price']}"
                text_width = draw.textlength(price_text, font=price_font)
                draw.text(
                    (x_offset + (400 - text_width)/2, y_offset + 110),
                    price_text,
                    font=price_font,
                    fill='black'
                )
                
                # Draw brand name (centered in 400px block)
                text_width = draw.textlength("AUTO GEEK", font=brand_font)
                draw.text(
                    (x_offset + (400 - text_width)/2, y_offset + 150),
                    "AUTO GEEK",
                    font=brand_font,
                    fill='black'
                )
            
            img.save(output_path, "BMP", dpi=(203, 203))
            return True
            
        except Exception as e:
            QMessageBox.warning(self, "Image Error", f"Failed to generate labels: {str(e)}")
            return False




        
    
        
    def save_label_files(self, items):
        try:
            # Create date-based folders
            today = datetime.datetime.now().strftime("%d-%m-%y")
            folders = {
                'bmp': os.path.join("barcodes", today, "BMP"),
                'png': os.path.join("barcodes", today, "PNG"),
                'pdf': os.path.join("barcodes", today, "PDF")
            }
            
            for folder in folders.values():
                os.makedirs(folder, exist_ok=True)
                
            # Generate all labels considering stock quantities
            all_labels = []
            for item in items:
                all_labels.extend([item] * item['stock_quantity'])
                
            temp_path = os.path.join("temp_print", datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
            os.makedirs(temp_path, exist_ok=True)
            
            # Process labels in pairs
            for i in range(0, len(all_labels), 2):
                item_pair = all_labels[i:i+2]
                # Save files with counter for duplicates
                base_name = f"{item_pair[0]['item_name']}_{i//2}"
                bmp_path = os.path.join(folders['bmp'], f"{base_name}.bmp")
                png_path = os.path.join(folders['png'], f"{base_name}.png")
                pdf_path = os.path.join(folders['pdf'], f"{base_name}.pdf")
                
                self.generate_row_label_image(item_pair, bmp_path)
                
                # Just save PNG and PDF versions
                image = Image.open(bmp_path)
                image.save(png_path, "PNG", dpi=(203, 203))
                image.save(pdf_path, "PDF", resolution=203.0)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to save label files: {str(e)}")
            return False
        return True
        




    def get_save_location(self):
        config_file = os.path.expanduser('~/.quicktag_config')
        try:
            with open(config_file, 'r') as f:
                return f.read().strip()
        except:
            # First time - prompt user and save location
            save_dir = QFileDialog.getExistingDirectory(
                self, "Select Default Save Location", 
                os.path.expanduser("~/Documents")
            )
            if save_dir:
                with open(config_file, 'w') as f:
                    f.write(save_dir)
                return save_dir
            return os.path.expanduser("~/Documents")  # Fallback
    def export_to_excel(self):
        try:
            today = datetime.datetime.now().strftime("%d-%m-%y")
            excel_folder = os.path.join("excel", today)
            os.makedirs(excel_folder, exist_ok=True)
            
            excel_path = os.path.join(excel_folder, f"Stock {today}.xlsx")
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Item Name", "Barcode", "Purchase Price", "Sale Price", "Opening Stock"])
            
            self.cursor.execute("SELECT item_name, barcode, purchase_price, sale_price, stock_quantity FROM items")
            for row in self.cursor.fetchall():
                sheet.append([
                    row[0],  # item_name
                    row[1],  # barcode
                    row[2],  # purchase_price
                    row[3],  # sale_price
                    row[4]   # stock_quantity
                ])
                
            workbook.save(excel_path)
            QMessageBox.information(self, "Success", f"Excel exported to:\n{excel_path}")
            self.load_item_history()  # Refresh history after export
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to export Excel: {str(e)}")
            
    def clear_items(self):
        reply = QMessageBox.question(
            self, 
            "Confirm Clear", 
            "This will permanently delete ALL items from the database. Continue?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                self.cursor.execute("DELETE FROM items")
                self.conn.commit()
                self.history_table.setRowCount(0)
                self.load_item_history()  # Explicit refresh
                QMessageBox.information(self, "Cleared", "All items have been cleared from the database.")
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to clear items: {str(e)}")

    def closeEvent(self, event):
        if hasattr(self, 'conn'):
            self.conn.close()
        if hasattr(self, 'preview_window'):
            self.preview_window.close()
        event.accept()

    def load_item_history(self):
        self.history_table.setRowCount(0)
        self.cursor.execute("SELECT item_name, created_at, sale_price, stock_quantity FROM items")
        for item in self.cursor.fetchall():
            self.add_item_to_table(*item)

    def preview_label(self):
        """Preview the label that will be printed"""
        try:
            import copy
            item_name = self.item_name_input.text().strip()
            if not item_name:
                QMessageBox.warning(self, "Error", "Item Name is required!")
                return

            # Get current item details
            barcode_value = self.barcode_override_input.text().strip() or self.generate_barcode()
            sale_price = self.calculate_current_price()
            stock_qty = int(self.stock_quantity_input.text() or 1)

            # Create full item list based on stock
            item = {
                'item_name': item_name,
                'barcode': barcode_value,
                'sale_price': sale_price
            }
            items = [copy.deepcopy(item) for _ in range(stock_qty)]

            # Create temp directory
            temp_dir = os.path.join(os.path.expanduser("~"), "quicktag_temp")
            os.makedirs(temp_dir, exist_ok=True)
            temp_bmp = os.path.join(temp_dir, 'quicktag_preview.bmp')

            # Generate row image with proper layout
            if self.generate_row_label_image(items, temp_bmp):
                if sys.platform == 'win32':
                    os.startfile(temp_bmp)
                else:
                    subprocess.run(['open', temp_bmp])
                    
        except Exception as e:
            QMessageBox.warning(self, "Preview Error", f"Failed to generate preview: {str(e)}")

    def generate_dual_label_image(self, items, output_path=None):
        """Generate a single 800x200 image with two labels side by side"""
        try:
            # Create blank image (800x200 @ 203 DPI)
            img = Image.new('RGB', (800, 200), 'white')
            draw = ImageDraw.Draw(img)
            
            # Load fonts
            try:
                title_font = ImageFont.truetype("arialbd.ttf", 16)
                price_font = ImageFont.truetype("arialbd.ttf", 20)
                brand_font = ImageFont.truetype("arialbd.ttf", 14)
            except:
                title_font = ImageFont.load_default()
                price_font = ImageFont.load_default()
                brand_font = ImageFont.load_default()
            
            # Process each item in the pair
            for i, item in enumerate(items[:2]):  # Only process up to 2 items
                x_offset = 0 if i == 0 else 400
                
                # Draw item name (centered in 400px block)
                text_width = draw.textlength(item['item_name'], font=title_font)
                draw.text(
                    (x_offset + (400 - text_width)/2, 15),
                    item['item_name'],
                    font=title_font,
                    fill='black'
                )
                
                # Draw barcode (fixed position within 400px block)
                barcode_img = self.generate_barcode_image(item['barcode'])
                img.paste(barcode_img, (x_offset + 60, 40))
                
                # Draw price (centered in 400px block)
                price_text = f"₹{item['sale_price']}"
                text_width = draw.textlength(price_text, font=price_font)
                draw.text(
                    (x_offset + (400 - text_width)/2, 110),
                    price_text,
                    font=price_font,
                    fill='black'
                )
                
                # Draw brand name (centered in 400px block)
                text_width = draw.textlength("AUTO GEEK", font=brand_font)
                draw.text(
                    (x_offset + (400 - text_width)/2, 150),
                    "AUTO GEEK",
                    font=brand_font,
                    fill='black'
                )
            
            if output_path:
                img.save(output_path, "BMP", dpi=(203, 203))
                return True
            return img
            
        except Exception as e:
            QMessageBox.warning(self, "Image Error", f"Failed to generate dual labels: {str(e)}")
            return False




        
    
        
    def save_label_files(self, items):
        try:
            # Create date-based folders
            today = datetime.datetime.now().strftime("%d-%m-%y")
            folders = {
                'bmp': os.path.join("barcodes", today, "BMP"),
                'png': os.path.join("barcodes", today, "PNG"),
                'pdf': os.path.join("barcodes", today, "PDF")
            }
            
            for folder in folders.values():
                os.makedirs(folder, exist_ok=True)
                
            # Generate all labels considering stock quantities
            all_labels = []
            for item in items:
                all_labels.extend([item] * item['stock_quantity'])
                
            temp_path = os.path.join("temp_print", datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
            os.makedirs(temp_path, exist_ok=True)
            
            # Process labels in pairs
            for i in range(0, len(all_labels), 2):
                item_pair = all_labels[i:i+2]
                # Save files with counter for duplicates
                base_name = f"{item_pair[0]['item_name']}_{i//2}"
                bmp_path = os.path.join(folders['bmp'], f"{base_name}.bmp")
                png_path = os.path.join(folders['png'], f"{base_name}.png")
                pdf_path = os.path.join(folders['pdf'], f"{base_name}.pdf")
                
                self.generate_row_label_image(item_pair, bmp_path)
                
                # Just save PNG and PDF versions
                image = Image.open(bmp_path)
                image.save(png_path, "PNG", dpi=(203, 203))
                image.save(pdf_path, "PDF", resolution=203.0)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to save label files: {str(e)}")
            return False
        return True
        
    def print_label(self):
        """Print current label"""
        try:
            item_data = {
                'item_name': self.item_name_input.text().strip(),
                'barcode': self.barcode_override_input.text().strip() or self.generate_barcode(),
                'sale_price': self.calculate_current_price()
            }
            quantity = int(self.stock_quantity_input.text() or 1)
            items = [item_data.copy() for _ in range(quantity)]

            temp_bmp = os.path.join(os.environ['TEMP'], 'quicktag_print.bmp')
            os.makedirs(os.path.dirname(temp_bmp), exist_ok=True)

            if self.generate_row_label_image(items, temp_bmp):
                if sys.platform == 'win32':
                    os.startfile(temp_bmp, "print")
                else:
                    subprocess.run(['lp', temp_bmp])
                QMessageBox.information(self, "Success", "Label printed successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Print Error", f"Failed to print label: {str(e)}")



    def get_save_location(self):
        config_file = os.path.expanduser('~/.quicktag_config')
        try:
            with open(config_file, 'r') as f:
                return f.read().strip()
        except:
            # First time - prompt user and save location
            save_dir = QFileDialog.getExistingDirectory(
                self, "Select Default Save Location", 
                os.path.expanduser("~/Documents")
            )
            if save_dir:
                with open(config_file, 'w') as f:
                    f.write(save_dir)
                return save_dir
            return os.path.expanduser("~/Documents")  # Fallback
    def export_to_excel(self):
        try:
            today = datetime.datetime.now().strftime("%d-%m-%y")
            excel_folder = os.path.join("excel", today)
            os.makedirs(excel_folder, exist_ok=True)
            
            excel_path = os.path.join(excel_folder, f"Stock {today}.xlsx")
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Item Name", "Barcode", "Purchase Price", "Sale Price", "Opening Stock"])
            
            self.cursor.execute("SELECT item_name, barcode, purchase_price, sale_price, stock_quantity FROM items")
            for row in self.cursor.fetchall():
                sheet.append([
                    row[0],  # item_name
                    row[1],  # barcode
                    row[2],  # purchase_price
                    row[3],  # sale_price
                    row[4]   # stock_quantity
                ])
                
            workbook.save(excel_path)
            QMessageBox.information(self, "Success", f"Excel exported to:\n{excel_path}")
            self.load_item_history()  # Refresh history after export
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to export Excel: {str(e)}")
            
    def clear_items(self):
        reply = QMessageBox.question(
            self, 
            "Confirm Clear", 
            "This will permanently delete ALL items from the database. Continue?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                self.cursor.execute("DELETE FROM items")
                self.conn.commit()
                self.history_table.setRowCount(0)
                self.load_item_history()  # Explicit refresh
                QMessageBox.information(self, "Cleared", "All items have been cleared from the database.")
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to clear items: {str(e)}")

    def closeEvent(self, event):
        if hasattr(self, 'conn'):
            self.conn.close()
        if hasattr(self, 'preview_window'):
            self.preview_window.close()
        event.accept()

    def load_item_history(self):
        self.history_table.setRowCount(0)
        self.cursor.execute("SELECT item_name, created_at, sale_price, stock_quantity FROM items")
        for item in self.cursor.fetchall():
            self.add_item_to_table(*item)

    def preview_label(self):
        """Preview the label that will be printed"""
        try:
            import copy
            item_name = self.item_name_input.text().strip()
            if not item_name:
                QMessageBox.warning(self, "Error", "Item Name is required!")
                return

            # Get current item details
            barcode_value = self.barcode_override_input.text().strip() or self.generate_barcode()
            sale_price = self.calculate_current_price()
            stock_qty = int(self.stock_quantity_input.text() or 1)

            # Create full item list based on stock
            item = {
                'item_name': item_name,
                'barcode': barcode_value,
                'sale_price': sale_price
            }
            items = [copy.deepcopy(item) for _ in range(stock_qty)]

            # Create temp directory
            temp_dir = os.path.join(os.path.expanduser("~"), "quicktag_temp")
            os.makedirs(temp_dir, exist_ok=True)
            temp_bmp = os.path.join(temp_dir, 'quicktag_preview.bmp')

            # Generate row image with proper layout
            if self.generate_row_label_image(items, temp_bmp):
                if sys.platform == 'win32':
                    os.startfile(temp_bmp)
                else:
                    subprocess.run(['open', temp_bmp])
                    
        except Exception as e:
            QMessageBox.warning(self, "Preview Error", f"Failed to generate preview: {str(e)}")

    def generate_dual_label_image(self, items, output_path=None):
        """Generate a single 800x200 image with two labels side by side"""
        try:
            # Create blank image (800x200 @ 203 DPI)
            img = Image.new('RGB', (800, 200), 'white')
            draw = ImageDraw.Draw(img)
            
            # Load fonts
            try:
                title_font = ImageFont.truetype("arialbd.ttf", 16)
                price_font = ImageFont.truetype("arialbd.ttf", 20)
                brand_font = ImageFont.truetype("arialbd.ttf", 14)
            except:
                title_font = ImageFont.load_default()
                price_font = ImageFont.load_default()
                brand_font = ImageFont.load_default()
            
            # Process each item in the pair
            for i, item in enumerate(items[:2]):  # Only process up to 2 items
                x_offset = 0 if i == 0 else 400
                
                # Draw item name (centered in 400px block)
                text_width = draw.textlength(item['item_name'], font=title_font)
                draw.text(
                    (x_offset + (400 - text_width)/2, 15),
                    item['item_name'],
                    font=title_font,
                    fill='black'
                )
                
                # Draw barcode (fixed position within 400px block)
                barcode_img = self.generate_barcode_image(item['barcode'])
                img.paste(barcode_img, (x_offset + 60, 40))
                
                # Draw price (centered in 400px block)
                price_text = f"₹{item['sale_price']}"
                text_width = draw.textlength(price_text, font=price_font)
                draw.text(
                    (x_offset + (400 - text_width)/2, 110),
                    price_text,
                    font=price_font,
                    fill='black'
                )
                
                # Draw brand name (centered in 400px block)
                text_width = draw.textlength("AUTO GEEK", font=brand_font)
                draw.text(
                    (x_offset + (400 - text_width)/2, 150),
                    "AUTO GEEK",
                    font=brand_font,
                    fill='black'
                )
            
            if output_path:
                img.save(output_path, "BMP", dpi=(203, 203))
                return True
            return img
            
        except Exception as e:
            QMessageBox.warning(self, "Image Error", f"Failed to generate dual labels: {str(e)}")
            return False

      
    def save_label_files(self, items):
        try:
            # Create date-based folders
            today = datetime.datetime.now().strftime("%d-%m-%y")
            folders = {
                'bmp': os.path.join("barcodes", today, "BMP"),
                'png': os.path.join("barcodes", today, "PNG"), 
                'pdf': os.path.join("barcodes", today, "PDF")
            }
            
            for folder in folders.values():
                os.makedirs(folder, exist_ok=True)
                
            # Generate all labels considering stock quantities
            all_labels = []
            for item in items:
                all_labels.extend([item] * item['stock_quantity'])
                
            temp_path = os.path.join("temp_print", datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
            os.makedirs(temp_path, exist_ok=True)
            
            # Process labels in pairs
            for i in range(0, len(all_labels), 2):
                item_pair = all_labels[i:i+2]
                base_name = f"{item_pair[0]['item_name']}_{i//2}"
                bmp_path = os.path.join(folders['bmp'], f"{base_name}.bmp")
                png_path = os.path.join(folders['png'], f"{base_name}.png")
                pdf_path = os.path.join(folders['pdf'], f"{base_name}.pdf")
                
                self.generate_row_label_image(item_pair, bmp_path)
                
                image = Image.open(bmp_path)
                image.save(png_path, "PNG", dpi=(203, 203))
                image.save(pdf_path, "PDF", resolution=203.0)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to save label files: {str(e)}")
            return False
        return True

    def print_label(self):
        """Print current label with smart handling of stock quantities"""
        try:
            item = {
                'item_name': self.item_name_input.text().strip(),
                'barcode': self.barcode_override_input.text().strip() or self.generate_barcode(),
                'sale_price': self.calculate_current_price()
            }
            stock_qty = int(self.stock_quantity_input.text() or 1)
            items = [item.copy() for _ in range(stock_qty)]

            output_dir = os.path.join(os.path.expanduser("~"), "quicktag_single")
            os.makedirs(output_dir, exist_ok=True)
            label_path = os.path.join(output_dir, "label_preview.bmp")

            if stock_qty <= 2:
                self.generate_row_label_image(items, label_path)
            else:
                # Create stacked image for this single item
                height = (stock_qty // 2 + (1 if stock_qty % 2 else 0)) * 200
                img = Image.new('RGB', (800, height), 'white')
                draw = ImageDraw.Draw(img)
                
                # Load fonts
                try:
                    title_font = ImageFont.truetype("arialbd.ttf", 16)
                    price_font = ImageFont.truetype("arialbd.ttf", 20)
                    brand_font = ImageFont.truetype("arialbd.ttf", 14)
                except:
                    title_font = ImageFont.load_default()
                    price_font = ImageFont.load_default()
                    brand_font = ImageFont.load_default()
                
                # Process each copy of the item
                for index in range(stock_qty):
                    row = index // 2
                    col = index % 2
                    x_offset = col * 400
                    y_offset = row * 200
                    
                    # Draw item name
                    text_width = draw.textlength(item['item_name'], font=title_font)
                    draw.text(
                        (x_offset + (400 - text_width)/2, y_offset + 15),
                        item['item_name'],
                        font=title_font,
                        fill='black'
                    )
                    
                    # Draw barcode
                    barcode_img = self.generate_barcode_image(item['barcode'])
                    img.paste(barcode_img, (x_offset + 60, y_offset + 40))
                    
                    # Draw price
                    price_text = f"₹{item['sale_price']}"
                    text_width = draw.textlength(price_text, font=price_font)
                    draw.text(
                        (x_offset + (400 - text_width)/2, y_offset + 110),
                        price_text,
                        font=price_font,
                        fill='black'
                    )
                    
                    # Draw brand name
                    text_width = draw.textlength("AUTO GEEK", font=brand_font)
                    draw.text(
                        (x_offset + (400 - text_width)/2, y_offset + 150),
                        "AUTO GEEK",
                        font=brand_font,
                        fill='black'
                    )
                
                img.save(label_path, "BMP", dpi=(203, 203))

            if sys.platform == 'win32':
                os.startfile(label_path, "open")
            else:
                subprocess.run(['open', label_path])
            QMessageBox.information(self, "Ready", "Label ready in default viewer.\nClick Print manually from there.")
        except Exception as e:
            QMessageBox.critical(self, "Print Error", f"Print failed: {str(e)}")
            
    def print_all_labels(self):
        """Print all labels with smart handling of stock quantities and item separation"""
        try:
            output_dir = os.path.join(os.path.expanduser("~"), "quicktag_batch")
            os.makedirs(output_dir, exist_ok=True)

            self.cursor.execute("SELECT item_name, barcode, sale_price, stock_quantity FROM items")
            records = self.cursor.fetchall()

            if not records:
                QMessageBox.information(self, "No Items", "No items found in database.")
                return

            # Generate one image per item
            label_paths = []
            for idx, (item_name, barcode, price, qty) in enumerate(records):
                item = {
                    'item_name': item_name,
                    'barcode': barcode,
                    'sale_price': price
                }
                items = [item.copy() for _ in range(qty)]

                output_path = os.path.join(output_dir, f"label_{idx}.bmp")
                if qty <= 2:
                    self.generate_row_label_image(items, output_path)
                else:
                    # Create stacked image for this single item
                    height = (qty // 2 + (1 if qty % 2 else 0)) * 200
                    img = Image.new('RGB', (800, height), 'white')
                    draw = ImageDraw.Draw(img)
                    
                    # Load fonts
                    try:
                        title_font = ImageFont.truetype("arialbd.ttf", 16)
                        price_font = ImageFont.truetype("arialbd.ttf", 20)
                        brand_font = ImageFont.truetype("arialbd.ttf", 14)
                    except:
                        title_font = ImageFont.load_default()
                        price_font = ImageFont.load_default()
                        brand_font = ImageFont.load_default()
                    
                    # Process each copy of the item
                    for index in range(qty):
                        row = index // 2
                        col = index % 2
                        x_offset = col * 400
                        y_offset = row * 200
                        
                        # Draw item name
                        text_width = draw.textlength(item['item_name'], font=title_font)
                        draw.text(
                            (x_offset + (400 - text_width)/2, y_offset + 15),
                            item['item_name'],
                            font=title_font,
                            fill='black'
                        )
                        
                        # Draw barcode
                        barcode_img = self.generate_barcode_image(item['barcode'])
                        img.paste(barcode_img, (x_offset + 60, y_offset + 40))
                        
                        # Draw price
                        price_text = f"₹{item['sale_price']}"
                        text_width = draw.textlength(price_text, font=price_font)
                        draw.text(
                            (x_offset + (400 - text_width)/2, y_offset + 110),
                            price_text,
                            font=price_font,
                            fill='black'
                        )
                        
                        # Draw brand name
                        text_width = draw.textlength("AUTO GEEK", font=brand_font)
                        draw.text(
                            (x_offset + (400 - text_width)/2, y_offset + 150),
                            "AUTO GEEK",
                            font=brand_font,
                            fill='black'
                        )
                    
                    img.save(output_path, "BMP", dpi=(203, 203))

                label_paths.append(output_path)

            # Combine all item BMPs into one tall final BMP
            images = [Image.open(path) for path in label_paths]
            total_height = sum(im.height for im in images)
            stacked = Image.new('RGB', (800, total_height), 'white')

            y = 0
            for im in images:
                stacked.paste(im, (0, y))
                y += im.height

            final_path = os.path.join(output_dir, "stacked_labels.bmp")
            stacked.save(final_path, "BMP", dpi=(203, 203))
            
            if sys.platform == 'win32':
                os.startfile(final_path, "open")
            else:
                subprocess.run(['open', final_path])
            QMessageBox.information(self, "Batch Ready", "All labels combined.\nNow open in default viewer.\nClick Print.")
        except Exception as e:
            QMessageBox.critical(self, "Batch Print Error", f"Failed to prepare labels: {str(e)}")
    def get_save_location(self):
        config_file = os.path.expanduser('~/.quicktag_config')
        try:
            with open(config_file, 'r') as f:
                return f.read().strip()
        except:
            # First time - prompt user and save location
            save_dir = QFileDialog.getExistingDirectory(
                self, "Select Default Save Location", 
                os.path.expanduser("~/Documents")
            )
            if save_dir:
                with open(config_file, 'w') as f:
                    f.write(save_dir)
                return save_dir
            return os.path.expanduser("~/Documents")  # Fallback
    def export_to_excel(self):
        try:
            today = datetime.datetime.now().strftime("%d-%m-%y")
            excel_folder = os.path.join("excel", today)
            os.makedirs(excel_folder, exist_ok=True)
            
            excel_path = os.path.join(excel_folder, f"Stock {today}.xlsx")
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Item Name", "Barcode", "Purchase Price", "Sale Price", "Opening Stock"])
            
            self.cursor.execute("SELECT item_name, barcode, purchase_price, sale_price, stock_quantity FROM items")
            for row in self.cursor.fetchall():
                sheet.append([
                    row[0],  # item_name
                    row[1],  # barcode
                    row[2],  # purchase_price
                    row[3],  # sale_price
                    row[4]   # stock_quantity
                ])
                
            workbook.save(excel_path)
            QMessageBox.information(self, "Success", f"Excel exported to:\n{excel_path}")
            self.load_item_history()  # Refresh history after export
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to export Excel: {str(e)}")
            
    def clear_items(self):
        reply = QMessageBox.question(
            self, 
            "Confirm Clear", 
            "This will permanently delete ALL items from the database. Continue?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                self.cursor.execute("DELETE FROM items")
                self.conn.commit()
                self.history_table.setRowCount(0)
                self.load_item_history()  # Explicit refresh
                QMessageBox.information(self, "Cleared", "All items have been cleared from the database.")
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to clear items: {str(e)}")

    def closeEvent(self, event):
        if hasattr(self, 'conn'):
            self.conn.close()
        if hasattr(self, 'preview_window'):
            self.preview_window.close()
        event.accept()

    def load_item_history(self):
        self.history_table.setRowCount(0)
        self.cursor.execute("SELECT item_name, created_at, sale_price, stock_quantity FROM items")
        for item in self.cursor.fetchall():
            self.add_item_to_table(*item)

    def preview_label(self):
        """Preview the label that will be printed"""
        try:
            import copy
            item_name = self.item_name_input.text().strip()
            if not item_name:
                QMessageBox.warning(self, "Error", "Item Name is required!")
                return

            # Get current item details
            barcode_value = self.barcode_override_input.text().strip() or self.generate_barcode()
            sale_price = self.calculate_current_price()
            stock_qty = int(self.stock_quantity_input.text() or 1)

            # Create full item list based on stock
            item = {
                'item_name': item_name,
                'barcode': barcode_value,
                'sale_price': sale_price
            }
            items = [copy.deepcopy(item) for _ in range(stock_qty)]

            # Create temp directory
            temp_dir = os.path.join(os.path.expanduser("~"), "quicktag_temp")
            os.makedirs(temp_dir, exist_ok=True)
            temp_bmp = os.path.join(temp_dir, 'quicktag_preview.bmp')

            # Generate row image with proper layout
            if self.generate_row_label_image(items, temp_bmp):
                if sys.platform == 'win32':
                    os.startfile(temp_bmp)
                else:
                    subprocess.run(['open', temp_bmp])
                    
        except Exception as e:
            QMessageBox.warning(self, "Preview Error", f"Failed to generate preview: {str(e)}")

    def calculate_current_price(self):
        """Calculate the current price based on inputs"""
        try:
            purchase_price = int(float(self.purchase_price_input.text() or 1))
            profit_percent = int(float(self.profit_percent_input.text() or 100))
            sale_price_input = self.sale_price_input.text().strip()
            
            if sale_price_input:
                return int(float(sale_price_input))
            else:
                return int(purchase_price + (purchase_price * profit_percent / 100))
        except:
            return 0


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = QuickTagApp()
    window.show()
    sys.exit(app.exec_())

