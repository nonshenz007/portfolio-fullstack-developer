# ------------------------------
# Barcode Generator v1.2.5 - Full Implementation
# ------------------------------
import sys
import ctypes
import multiprocessing
import subprocess
import platform
import logging
import os
import random
import datetime
import sqlite3
import configparser
from pathlib import Path

from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QFormLayout, QDialog,
    QPushButton, QMessageBox, QLineEdit, QFileDialog, QLabel,
    QComboBox, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView,
    QCheckBox, QStatusBar, QSizePolicy
)
from PyQt5.QtCore import Qt, QSizeF, QRectF
from PyQt5.QtGui import QImage, QPainter, QFont
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog


# ------------------------------
# Configuration Manager
# ------------------------------
class ConfigManager:
    def __init__(self):
        self.config_path = os.path.join(get_app_data_path(), "config.ini")
        self.config = configparser.ConfigParser()
        self.load_config()

    def load_config(self):
        if os.path.exists(self.config_path):
            self.config.read(self.config_path)
        else:
            self.config['DEFAULT'] = {
                'output_dir': os.path.join(get_app_data_path(), "barcodes"),
                'default_label_size': 'Medium (100x25 mm)',
                'excel_path': os.path.join(get_app_data_path(), "Excel"),
                'last_used_printer': '',
                'default_company': '',
                'continuous_roll': 'False'
            }
            self.save_config()

    def save_config(self):
        with open(self.config_path, 'w') as configfile:
            self.config.write(configfile)

    def get(self, section, option, fallback=None):
        try:
            return self.config.get(section, option, fallback=fallback)
        except (configparser.NoSectionError, configparser.NoOptionError):
            return fallback

    def set(self, section, option, value):
        if not self.config.has_section(section):
            self.config.add_section(section)
        self.config.set(section, option, str(value))
        self.save_config()


# Initialize config manager early
config_manager = ConfigManager()


# ------------------------------
# Single Instance Check (Windows Only)
# ------------------------------
def is_already_running():
    if platform.system() != "Windows":
        return False
    mutex_name = "Global\\ProjectsByShenzC_FullSafeApp_v1.2.5"
    mutex = ctypes.windll.kernel32.CreateMutexW(None, False, mutex_name)
    last_error = ctypes.windll.kernel32.GetLastError()

    if last_error == 183:  # ERROR_ALREADY_EXISTS
        return True
    return False


# ------------------------------
# App Data Directory and Global Paths
# ------------------------------
def get_app_data_path():
    base_path = os.path.join(os.environ["APPDATA"], "BarcodeGenerator")
    os.makedirs(base_path, exist_ok=True)
    return base_path


DB_PATH = os.path.join(get_app_data_path(), "barcodes.db")
LOG_PATH = os.path.join(get_app_data_path(), "barcode_app.log")
output_dir = config_manager.get('DEFAULT', 'output_dir',
                                os.path.join(get_app_data_path(), "barcodes"))
os.makedirs(output_dir, exist_ok=True)

# Create required subdirectories
for subdir in ['Excel', 'Labels', 'Backups']:
    os.makedirs(os.path.join(get_app_data_path(), subdir), exist_ok=True)

# ------------------------------
# Logging Setup
# ------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH),
        logging.StreamHandler(),
        logging.handlers.RotatingFileHandler(
            LOG_PATH, maxBytes=5 * 1024 * 1024, backupCount=3
        )
    ]
)
logger = logging.getLogger(__name__)

TITLE = "Projects \U0001F4C8 by Shenz C (v1.2.5)"

LABEL_SIZES = {
    "Small (60x25 mm)": (60, 25),
    "Medium (100x25 mm)": (100, 25),
    "Large (150x50 mm)": (150, 50),
    "A4 Sheet (210x297 mm)": (210, 297),
    "2 Labels (50x25mm)": (100, 25),
    "Custom Size": None
}


# ------------------------------
# Database Manager (Enhanced)
# ------------------------------
class DatabaseManager:
    def __init__(self, db_path=DB_PATH):
        self.db_path = db_path
        self.conn = None
        self.connect()

    def connect(self):
        try:
            self.conn = sqlite3.connect(self.db_path)
            self.conn.execute("PRAGMA foreign_keys = ON")
            self.conn.execute("PRAGMA journal_mode = WAL")
            self.create_tables()
            logger.info("Database connected and tables verified.")
        except Exception as e:
            logger.error(f"Database connection failed: {e}")
            raise

    def backup_database(self):
        try:
            backup_dir = os.path.join(get_app_data_path(), "Backups")
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(backup_dir, f"barcodes_backup_{timestamp}.db")

            # Use SQLite backup API
            backup_conn = sqlite3.connect(backup_path)
            self.conn.backup(backup_conn)
            backup_conn.close()
            logger.info(f"Database backup created at {backup_path}")
            return backup_path
        except Exception as e:
            logger.error(f"Database backup failed: {e}")
            return None

    def create_tables(self):
        try:
            c = self.conn.cursor()

            # Barcode history with opening_stock
            c.execute('''
                CREATE TABLE IF NOT EXISTS barcode_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    item_name TEXT NOT NULL,
                    barcode TEXT NOT NULL,
                    sale_price REAL NOT NULL,
                    purchase_price REAL NOT NULL,
                    company_name TEXT,
                    generation_date TEXT NOT NULL,
                    username TEXT NOT NULL DEFAULT 'admin',
                    opening_stock INTEGER DEFAULT 0,
                    label_size TEXT,
                    print_count INTEGER DEFAULT 1,
                    CONSTRAINT unique_barcode UNIQUE (barcode, generation_date)
                )
            ''')

            # Barcodes table with opening_stock
            c.execute('''
                CREATE TABLE IF NOT EXISTS barcodes (
                    barcode TEXT PRIMARY KEY,
                    item_name TEXT NOT NULL,
                    sale_price REAL NOT NULL,
                    purchase_price REAL NOT NULL,
                    company_name TEXT,
                    quantity INTEGER DEFAULT 1,
                    opening_stock INTEGER DEFAULT 0,
                    last_updated TEXT NOT NULL,
                    label_size TEXT
                )
            ''')

            # Users table with enhanced security
            c.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    username TEXT PRIMARY KEY,
                    password TEXT NOT NULL,
                    role TEXT NOT NULL DEFAULT 'user',
                    last_login TEXT,
                    login_count INTEGER DEFAULT 0,
                    created_date TEXT NOT NULL,
                    CONSTRAINT valid_role CHECK (role IN ('admin', 'manager', 'user'))
                )
            ''')

            # Settings table for application preferences
            c.execute('''
                CREATE TABLE IF NOT EXISTS settings (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL,
                    description TEXT,
                    last_modified TEXT NOT NULL
                )
            ''')

            # Insert default admin if none exists
            c.execute('SELECT COUNT(*) FROM users WHERE username = "admin"')
            if c.fetchone()[0] == 0:
                default_pass = "Kungfukenny"  # In production, use proper hashing
                created_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                c.execute('''
                    INSERT INTO users 
                    (username, password, role, created_date) 
                    VALUES (?, ?, ?, ?)
                ''', ("admin", default_pass, "admin", created_date))

            self.conn.commit()
        except Exception as e:
            logger.error(f"Error creating tables: {e}")
            self.conn.rollback()
            raise

    def add_barcode_history(self, username, item_name, barcode, sale_price,
                            purchase_price, company_name, generation_date,
                            opening_stock, label_size, print_count=1):
        try:
            c = self.conn.cursor()
            c.execute('''
                INSERT INTO barcode_history
                (username, item_name, barcode, sale_price, purchase_price, 
                 company_name, generation_date, opening_stock, label_size, print_count)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (username, item_name, barcode, sale_price, purchase_price,
                  company_name, generation_date, opening_stock, label_size, print_count))
            self.conn.commit()
            logger.info("Barcode history added successfully.")
            return True
        except sqlite3.IntegrityError as e:
            logger.warning(f"Duplicate barcode entry detected: {e}")
            return False
        except Exception as e:
            logger.error(f"Failed to add barcode history: {e}")
            self.conn.rollback()
            raise

    def add_or_update_barcode(self, item_name, barcode, sale_price, purchase_price,
                              company_name, quantity, opening_stock, label_size):
        try:
            c = self.conn.cursor()
            last_updated = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            c.execute('''
                INSERT INTO barcodes
                (barcode, item_name, sale_price, purchase_price, company_name, 
                 quantity, opening_stock, last_updated, label_size)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(barcode) DO UPDATE SET
                    item_name = excluded.item_name,
                    sale_price = excluded.sale_price,
                    purchase_price = excluded.purchase_price,
                    company_name = excluded.company_name,
                    quantity = quantity + excluded.quantity,
                    opening_stock = excluded.opening_stock,
                    last_updated = excluded.last_updated,
                    label_size = excluded.label_size
            ''', (barcode, item_name, sale_price, purchase_price, company_name,
                  quantity, opening_stock, last_updated, label_size))

            self.conn.commit()
            logger.info("Barcode updated/added in barcodes table.")
            return True
        except Exception as e:
            logger.error(f"Failed to add or update barcode: {e}")
            self.conn.rollback()
            raise

    def fetch_history(self, search_filter="", limit=500):
        try:
            c = self.conn.cursor()
            query = """
                SELECT id, username, item_name, barcode, sale_price, 
                       purchase_price, company_name, opening_stock, 
                       generation_date, label_size, print_count
                FROM barcode_history
            """
            params = ()

            if search_filter:
                query += """
                    WHERE item_name LIKE ? 
                    OR barcode LIKE ? 
                    OR company_name LIKE ?
                """
                search_term = f"%{search_filter}%"
                params = (search_term, search_term, search_term)

            query += " ORDER BY generation_date DESC LIMIT ?"
            params += (limit,)

            c.execute(query, params)
            history = c.fetchall()
            logger.info(f"Fetched {len(history)} history records.")
            return history
        except Exception as e:
            logger.error(f"Failed to fetch history: {e}")
            return []

    def clear_history(self, older_than_days=None):
        try:
            c = self.conn.cursor()
            if older_than_days:
                cutoff_date = (datetime.datetime.now() - datetime.timedelta(days=older_than_days)).strftime("%Y-%m-%d")
                c.execute('''
                    DELETE FROM barcode_history 
                    WHERE date(generation_date) < date(?)
                ''', (cutoff_date,))
                rows_affected = c.rowcount
            else:
                c.execute('DELETE FROM barcode_history')
                rows_affected = c.rowcount

            self.conn.commit()
            logger.info(f"Cleared {rows_affected} history records.")
            return rows_affected
        except Exception as e:
            logger.error(f"Failed to clear history: {e}")
            self.conn.rollback()
            raise

    def add_user(self, username, password, role="user"):
        try:
            if role not in ('admin', 'manager', 'user'):
                raise ValueError("Invalid role specified")

            created_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            c = self.conn.cursor()
            c.execute('''
                INSERT INTO users 
                (username, password, role, created_date) 
                VALUES (?, ?, ?, ?)
            ''', (username, password, role, created_date))

            self.conn.commit()
            logger.info(f"User {username} added successfully.")
            return True
        except sqlite3.IntegrityError:
            logger.warning(f"User {username} already exists.")
            return False
        except Exception as e:
            logger.error(f"Failed to add user: {e}")
            self.conn.rollback()
            raise

    def remove_user(self, username):
        try:
            if username == "admin":
                raise ValueError("Cannot remove default admin user")

            c = self.conn.cursor()
            c.execute('DELETE FROM users WHERE username = ?', (username,))
            rows_affected = c.rowcount

            if rows_affected == 0:
                logger.warning(f"User {username} not found.")
                return False

            self.conn.commit()
            logger.info(f"User {username} removed successfully.")
            return True
        except Exception as e:
            logger.error(f"Failed to remove user: {e}")
            self.conn.rollback()
            raise

    def fetch_users(self):
        try:
            c = self.conn.cursor()
            c.execute('''
                SELECT username, role, last_login, login_count 
                FROM users 
                ORDER BY username ASC
            ''')
            users = c.fetchall()
            logger.info(f"Fetched {len(users)} users from database.")
            return users
        except Exception as e:
            logger.error(f"Failed to fetch users: {e}")
            return []

    def validate_user(self, username, password):
        try:
            c = self.conn.cursor()
            c.execute('''
                SELECT username, role 
                FROM users 
                WHERE username = ? AND password = ?
            ''', (username, password))

            result = c.fetchone()
            if result:
                # Update last login info
                login_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                c.execute('''
                    UPDATE users 
                    SET last_login = ?, login_count = login_count + 1 
                    WHERE username = ?
                ''', (login_time, username))
                self.conn.commit()
                logger.info(f"User {username} logged in successfully.")
            else:
                logger.warning(f"Failed login attempt for user {username}")

            return result
        except Exception as e:
            logger.error(f"User validation failed: {e}")
            return None

    def update_user_password(self, username, new_password):
        try:
            c = self.conn.cursor()
            c.execute('''
                UPDATE users 
                SET password = ? 
                WHERE username = ?
            ''', (new_password, username))

            rows_affected = c.rowcount
            self.conn.commit()

            if rows_affected > 0:
                logger.info(f"Password updated for user {username}")
                return True
            return False
        except Exception as e:
            logger.error(f"Failed to update password: {e}")
            self.conn.rollback()
            raise

    def get_setting(self, key, default=None):
        try:
            c = self.conn.cursor()
            c.execute('''
                SELECT value 
                FROM settings 
                WHERE key = ?
            ''', (key,))

            result = c.fetchone()
            return result[0] if result else default
        except Exception as e:
            logger.error(f"Failed to get setting {key}: {e}")
            return default

    def set_setting(self, key, value, description=None):
        try:
            last_modified = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            c = self.conn.cursor()
            c.execute('''
                INSERT INTO settings 
                (key, value, description, last_modified) 
                VALUES (?, ?, ?, ?)
                ON CONFLICT(key) DO UPDATE SET
                    value = excluded.value,
                    description = excluded.description,
                    last_modified = excluded.last_modified
            ''', (key, value, description, last_modified))

            self.conn.commit()
            logger.info(f"Setting {key} updated to {value}")
            return True
        except Exception as e:
            logger.error(f"Failed to set setting {key}: {e}")
            self.conn.rollback()
            raise


# ------------------------------
# Barcode Generation (Enhanced Code128)
# ------------------------------
try:
    from barcode import get_barcode_class
    from barcode.writer import ImageWriter
    import barcode
except ImportError as e:
    logger.error("Missing dependency 'python-barcode'. Please install it.")
    raise e


class CustomBarcodeWriter(ImageWriter):
    """Custom writer with no text and optimized settings"""

    def __init__(self):
        super().__init__()
        self.text = ''  # Remove text under barcode
        self.font_size = 0  # No text means no need for font
        self.dpi = 300  # Higher DPI for better print quality
        self.module_width = 0.2  # Make bars slightly thinner
        self.module_height = 15.0  # Make barcode taller
        self.quiet_zone = 3.0  # Add more quiet zone
        self.background = 'white'  # White background
        self.foreground = 'black'  # Black bars
        self.write_text = False  # Don't write text
        self.text_distance = 0  # No text distance


def sanitize_barcode_input(barcode_str):
    """Ensure barcode only contains valid Code128 characters"""
    return "".join(ch for ch in barcode_str if 32 <= ord(ch) <= 126)


def generate_barcode_image(barcode_str, output_dir=None):
    """Generate barcode image with enhanced options"""
    safe_barcode = sanitize_barcode_input(barcode_str)
    if not safe_barcode:
        raise ValueError("No valid Code128 characters after sanitizing.")

    try:
        Code128 = get_barcode_class('code128')
        output_dir = output_dir or config_manager.get('DEFAULT', 'output_dir')
        os.makedirs(output_dir, exist_ok=True)

        # Generate filename with timestamp
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base_filename = os.path.join(output_dir, f"barcode_{safe_barcode}_{timestamp}")

        # Create barcode with custom writer
        writer = CustomBarcodeWriter()
        code128_obj = Code128(safe_barcode, writer=writer)

        # Save options
        options = {
            'module_width': writer.module_width,
            'module_height': writer.module_height,
            'quiet_zone': writer.quiet_zone,
            'font_size': writer.font_size,
            'text_distance': writer.text_distance,
            'background': writer.background,
            'foreground': writer.foreground,
            'write_text': writer.write_text
        }

        # Save in multiple formats
        png_path = code128_obj.save(base_filename, options)
        bmp_path = f"{base_filename}.bmp"

        # Convert to BMP format
        img = QImage(png_path)
        if img.isNull():
            raise IOError(f"Failed to load generated barcode image at {png_path}")
        img.save(bmp_path, 'BMP')

        logger.info(f"Barcode images generated: {png_path}, {bmp_path}")
        return {
            'png': png_path,
            'bmp': bmp_path,
            'barcode': safe_barcode
        }
    except barcode.errors.BarcodeError as e:
        logger.error(f"Invalid barcode data: {e}")
        raise ValueError(f"Invalid barcode data: {e}")
    except Exception as e:
        logger.error(f"Barcode generation failed: {e}")
        raise


# ------------------------------
# PDF Generation (Enhanced)
# ------------------------------
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm, inch
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, Frame, KeepInFrame
from reportlab.lib.enums import TA_CENTER, TA_LEFT


def mm_to_points(mm_val):
    """Convert millimeters to points (1mm = 2.83465 points)"""
    return mm_val * 2.83465


def create_label_styles():
    """Create styles for label text"""
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name='LabelTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        alignment=TA_CENTER,
        leading=12,
        spaceAfter=2
    ))

    styles.add(ParagraphStyle(
        name='LabelInfo',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        alignment=TA_CENTER,
        leading=10,
        spaceBefore=2
    ))

    return styles


def draw_label(c, x_offset_mm, y_offset_mm, w_mm, h_mm, barcode_path, details, styles):
    """Draw a single label with proper layout"""
    # Convert mm to points
    x = mm_to_points(x_offset_mm)
    y = mm_to_points(y_offset_mm)
    width = mm_to_points(w_mm)
    height = mm_to_points(h_mm)

    # Label border (optional)
    if config_manager.get('DEFAULT', 'draw_borders', 'False').lower() == 'true':
        c.rect(x, y, width, height)

    # Get label content
    item_name = details.get("item", "").strip()
    sale_price = details.get("sale", "").strip()
    company = details.get("company", "").strip()
    barcode_value = details.get("barcode", "").strip()

    # Calculate positions
    title_y = y + height - mm_to_points(5)  # 5mm from top
    barcode_y = y + height / 2 - mm_to_points(8) / 2  # Centered vertically
    info_y = y + mm_to_points(3)  # 3mm from bottom

    # Draw item name (top)
    title = Paragraph(f"<b>{item_name}</b>", styles['LabelTitle'])
    title_width, title_height = title.wrap(width - mm_to_points(2), mm_to_points(5))
    title.drawOn(c, x + (width - title_width) / 2, title_y - title_height)

    # Draw barcode (center)
    barcode_width = min(width - mm_to_points(10), mm_to_points(40))  # Max 40mm wide
    barcode_height = mm_to_points(8)  # 8mm tall
    c.drawImage(
        barcode_path,
        x + (width - barcode_width) / 2,
        barcode_y,
        width=barcode_width,
        height=barcode_height,
        preserveAspectRatio=True,
        mask='auto'
    )

    # Draw info line (bottom)
    info_text = f"Sale: {sale_price} | {company}"
    info = Paragraph(info_text, styles['LabelInfo'])
    info_width, info_height = info.wrap(width - mm_to_points(2), mm_to_points(5))
    info.drawOn(c, x + (width - info_width) / 2, info_y)

    # Optional: Draw small barcode number (very small at bottom)
    if config_manager.get('DEFAULT', 'show_barcode_number', 'False').lower() == 'true':
        barcode_num_style = ParagraphStyle(
            name='BarcodeNumber',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=6,
            alignment=TA_CENTER,
            leading=6,
            textColor='gray'
        )
        barcode_num = Paragraph(barcode_value, barcode_num_style)
        barcode_num.drawOn(c, x + (width - mm_to_points(20)) / 2, y + mm_to_points(1))


def generate_label_pdf(pdf_path, label_size_mm, barcode_path, details):
    """Generate PDF with labels using enhanced layout"""
    w_mm, h_mm = label_size_mm
    count = details.get("count", 1)
    barcode_value = details.get("barcode", "")

    # Create PDF canvas
    if label_size_mm == (210, 297):  # A4
        c = canvas.Canvas(pdf_path, pagesize=A4)
        page_width, page_height = A4
    else:
        # Custom size in points
        page_width = mm_to_points(w_mm)
        page_height = mm_to_points(h_mm)
        c = canvas.Canvas(pdf_path, pagesize=(page_width, page_height))

    # Create styles
    styles = create_label_styles()

    # Handle different label layouts
    if label_size_mm == (100, 25):  # Dual-label layout
        labels_per_page = 2
        single_label_w = 50  # mm
        pages_needed = (count + labels_per_page - 1) // labels_per_page

        for page in range(pages_needed):
            for i in range(labels_per_page):
                label_num = page * labels_per_page + i
                if label_num >= count:
                    break

                x_offset = i * single_label_w
                draw_label(c, x_offset, 0, single_label_w, h_mm,
                           barcode_path, details, styles)

            c.showPage()
    else:  # Single label per page
        for _ in range(count):
            draw_label(c, 0, 0, w_mm, h_mm, barcode_path, details, styles)
            c.showPage()

    c.save()
    logger.info(f"PDF generated at {pdf_path} with {count} labels")
    return pdf_path


# ------------------------------
# Excel Export (Enhanced)
# ------------------------------
def update_excel(item_code, item_name, sale_price, purchase_price,
                 company_name, generation_date, opening_stock):
    """Update Excel with enhanced features"""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Get Excel directory from config
        excel_dir = config_manager.get('DEFAULT', 'excel_path',
                                       os.path.join(get_app_data_path(), "Excel"))
        os.makedirs(excel_dir, exist_ok=True)

        # Create filename with current month/year
        today = datetime.datetime.now()
        month_year = today.strftime("%m-%Y")
        file_name = os.path.join(excel_dir, f"Stock_{month_year}.xlsx")

        # Create new workbook if needed
        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "Items"

            # Create header row with styling
            headers = [
                "Item Code", "Item Name", "Sale Price",
                "Purchase Price", "Company", "Opening Stock",
                "Last Updated"
            ]

            ws.append(headers)

            # Style header row
            bold_font = Font(bold=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = bold_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            # Set column widths
            column_widths = [15, 30, 12, 12, 25, 12, 20]
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = width

            wb.save(file_name)

        # Open existing workbook
        wb = load_workbook(file_name)
        ws = wb.active
        found = False

        # Search for existing item
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
            if row[0].value == item_code:
                # Update existing item
                row[0].offset(column=1).value = item_name  # Item Name
                row[0].offset(column=2).value = float(sale_price)  # Sale Price
                row[0].offset(column=3).value = float(purchase_price)  # Purchase Price
                row[0].offset(column=4).value = company_name  # Company
                row[0].offset(column=5).value = int(opening_stock)  # Opening Stock
                row[0].offset(column=6).value = generation_date  # Last Updated
                found = True
                break

        if not found:
            # Add new item
            ws.append([
                item_code,
                item_name,
                float(sale_price),
                float(purchase_price),
                company_name,
                int(opening_stock),
                generation_date
            ])

        # Save with retry logic
        for attempt in range(3):
            try:
                wb.save(file_name)
                logger.info(f"Excel updated: item_code={item_code} (found={found})")
                return True
            except PermissionError as e:
                logger.warning(f"Excel file locked (attempt {attempt + 1}/3): {e}")
                if attempt < 2:
                    time.sleep(1)  # Wait before retrying
                    continue

                QMessageBox.warning(
                    None,
                    "Excel Locked",
                    f"Could not save to {file_name}. File may be open in another program.",
                    QMessageBox.Ok
                )
                return False
            except Exception as e:
                logger.error(f"Excel save error: {e}")
                QMessageBox.warning(
                    None,
                    "Excel Error",
                    f"Failed to update Excel file:\n{e}",
                    QMessageBox.Ok
                )
                return False

        return True
    except ImportError:
        logger.warning("openpyxl not available - Excel export disabled")
        return False
    except Exception as e:
        logger.error(f"Excel update failed: {e}")
        QMessageBox.warning(
            None,
            "Excel Error",
            f"Failed to update Excel file:\n{e}",
            QMessageBox.Ok
        )
        return False


# ------------------------------
# Printing Functions (Enhanced)
# ------------------------------
def print_image_via_dialog(parent, image_path, label_size_mm=(50, 25)):
    """Print label with proper sizing and options"""
    try:
        printer = QPrinter(QPrinter.HighResolution)

        # Set printer properties
        printer.setFullPage(True)
        printer.setColorMode(QPrinter.Color)
        printer.setPageMargins(0, 0, 0, 0, QPrinter.Millimeter)

        # Set paper size based on label dimensions
        width_mm, height_mm = label_size_mm
        printer.setPaperSize(QSizeF(width_mm, height_mm), QPrinter.Millimeter)

        # Show print dialog
        dialog = QPrintDialog(printer, parent)
        dialog.setWindowTitle("Print Barcode Label")

        if dialog.exec_() == QDialog.Accepted:
            painter = QPainter(printer)
            img = QImage(image_path)

            if img.isNull():
                raise IOError(f"Failed to load image: {image_path}")

            # Calculate aspect ratio preserving dimensions
            img_ratio = img.width() / img.height()
            page_ratio = printer.pageRect().width() / printer.pageRect().height()

            if img_ratio > page_ratio:
                # Image is wider than page
                new_width = printer.pageRect().width()
                new_height = new_width / img_ratio
            else:
                # Image is taller than page
                new_height = printer.pageRect().height()
                new_width = new_height * img_ratio

            # Center the image on the page
            x_offset = (printer.pageRect().width() - new_width) / 2
            y_offset = (printer.pageRect().height() - new_height) / 2

            # Draw image
            painter.drawImage(QRectF(x_offset, y_offset, new_width, new_height), img)
            painter.end()

            # Update print count in database
            try:
                barcode = os.path.basename(image_path).split('_')[1]  # Extract barcode from filename
                db = DatabaseManager()

                # Get the most recent history entry for this barcode
                c = db.conn.cursor()
                c.execute('''
                    UPDATE barcode_history 
                    SET print_count = print_count + 1 
                    WHERE barcode = ? 
                    ORDER BY generation_date DESC 
                    LIMIT 1
                ''', (barcode,))
                db.conn.commit()
                logger.info(f"Updated print count for barcode {barcode}")
            except Exception as db_error:
                logger.error(f"Failed to update print count: {db_error}")

            QMessageBox.information(
                parent,
                "Printing Complete",
                "Label successfully sent to printer.",
                QMessageBox.Ok
            )
        else:
            QMessageBox.information(
                parent,
                "Printing Canceled",
                "Print job was canceled.",
                QMessageBox.Ok
            )
    except Exception as ex:
        logger.error(f"Printing error: {ex}")
        QMessageBox.warning(
            parent,
            "Print Error",
            f"Failed to print label:\n{ex}",
            QMessageBox.Ok
        )


# ------------------------------
# User Management Dialog (Enhanced)
# ------------------------------
class UserManagementDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("User Management")
        self.setMinimumSize(600, 400)
        self.db = DatabaseManager()
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()

        # Title
        title = QLabel("User Management")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 18px; font-weight: bold;")
        layout.addWidget(title)

        # User table
        self.user_table = QTableWidget()
        self.user_table.setColumnCount(5)
        self.user_table.setHorizontalHeaderLabels([
            "Username", "Role", "Last Login", "Login Count", "Created"
        ])
        self.user_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.user_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.user_table.setSelectionMode(QTableWidget.SingleSelection)
        layout.addWidget(self.user_table)

        # Form for adding/editing users
        form_layout = QFormLayout()

        self.username_edit = QLineEdit()
        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.Password)
        self.role_combo = QComboBox()
        self.role_combo.addItems(["admin", "manager", "user"])

        form_layout.addRow("Username:", self.username_edit)
        form_layout.addRow("Password:", self.password_edit)
        form_layout.addRow("Role:", self.role_combo)

        # Buttons
        btn_layout = QHBoxLayout()

        self.add_btn = QPushButton("Add User")
        self.add_btn.clicked.connect(self.add_user)

        self.update_btn = QPushButton("Update User")
        self.update_btn.clicked.connect(self.update_user)
        self.update_btn.setEnabled(False)

        self.remove_btn = QPushButton("Remove User")
        self.remove_btn.clicked.connect(self.remove_user)
        self.remove_btn.setEnabled(False)

        btn_layout.addWidget(self.add_btn)
        btn_layout.addWidget(self.update_btn)
        btn_layout.addWidget(self.remove_btn)

        # Password change section
        pw_change_layout = QFormLayout()

        self.current_pw_edit = QLineEdit()
        self.current_pw_edit.setEchoMode(QLineEdit.Password)
        self.new_pw_edit = QLineEdit()
        self.new_pw_edit.setEchoMode(QLineEdit.Password)
        self.confirm_pw_edit = QLineEdit()
        self.confirm_pw_edit.setEchoMode(QLineEdit.Password)

        pw_change_layout.addRow("Current Password:", self.current_pw_edit)
        pw_change_layout.addRow("New Password:", self.new_pw_edit)
        pw_change_layout.addRow("Confirm Password:", self.confirm_pw_edit)

        self.change_pw_btn = QPushButton("Change Password")
        self.change_pw_btn.clicked.connect(self.change_password)

        # Add to main layout
        layout.addLayout(form_layout)
        layout.addLayout(btn_layout)
        layout.addWidget(QLabel("Change Password:"))
        layout.addLayout(pw_change_layout)
        layout.addWidget(self.change_pw_btn)

        self.setLayout(layout)

        # Connect table selection
        self.user_table.itemSelectionChanged.connect(self.user_selected)

        # Load users
        self.load_users()

    def load_users(self):
        users = self.db.fetch_users()
        self.user_table.setRowCount(len(users))

        for row, user in enumerate(users):
            for col, value in enumerate(user):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignCenter)

                # Make username and role editable
                if col in (0, 1):
                    item.setFlags(item.flags() | Qt.ItemIsEditable)

                self.user_table.setItem(row, col, item)

    def user_selected(self):
        selected = self.user_table.selectedItems()

        if selected:
            username = selected[0].text()
            self.username_edit.setText(username)
            self.password_edit.clear()

            # Enable/disable buttons based on selection
            self.update_btn.setEnabled(True)
            self.remove_btn.setEnabled(username != "admin")
            self.change_pw_btn.setEnabled(True)
        else:
            self.username_edit.clear()
            self.password_edit.clear()
            self.update_btn.setEnabled(False)
            self.remove_btn.setEnabled(False)
            self.change_pw_btn.setEnabled(False)

    def add_user(self):
        username = self.username_edit.text().strip()
        password = self.password_edit.text().strip()
        role = self.role_combo.currentText()

        if not username or not password:
            QMessageBox.warning(self, "Error", "Username and password are required.")
            return

        try:
            if self.db.add_user(username, password, role):
                QMessageBox.information(self, "Success", "User added successfully.")
                self.load_users()
            else:
                QMessageBox.warning(self, "Error", "User already exists.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to add user:\n{e}")

    def update_user(self):
        selected = self.user_table.selectedItems()
        if not selected:
            return

        old_username = selected[0].text()
        new_username = self.username_edit.text().strip()
        password = self.password_edit.text().strip()
        role = self.role_combo.currentText()

        if not new_username:
            QMessageBox.warning(self, "Error", "Username cannot be empty.")
            return

        try:
            # Update username if changed
            if old_username != new_username:
                # In a real app, we'd need to update all references to this user
                QMessageBox.warning(self, "Info", "Username change not implemented fully.")
                return

            # Update role in the table
            selected[1].setText(role)

            # Update password if provided
            if password:
                if not self.db.update_user_password(old_username, password):
                    QMessageBox.warning(self, "Error", "Failed to update password.")
                    return

            QMessageBox.information(self, "Success", "User updated successfully.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to update user:\n{e}")

    def remove_user(self):
        selected = self.user_table.selectedItems()
        if not selected:
            return

        username = selected[0].text()

        if username == "admin":
            QMessageBox.warning(self, "Error", "Cannot delete the admin user.")
            return

        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Are you sure you want to delete user '{username}'?",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                if self.db.remove_user(username):
                    QMessageBox.information(self, "Success", "User deleted successfully.")
                    self.load_users()
                else:
                    QMessageBox.warning(self, "Error", "User not found.")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete user:\n{e}")

    def change_password(self):
        selected = self.user_table.selectedItems()
        if not selected:
            return

        username = selected[0].text()
        current_pw = self.current_pw_edit.text()
        new_pw = self.new_pw_edit.text()
        confirm_pw = self.confirm_pw_edit.text()

        if not current_pw or not new_pw or not confirm_pw:
            QMessageBox.warning(self, "Error", "All password fields are required.")
            return

        if new_pw != confirm_pw:
            QMessageBox.warning(self, "Error", "New passwords do not match.")
            return

        try:
            # Verify current password
            if not self.db.validate_user(username, current_pw):
                QMessageBox.warning(self, "Error", "Current password is incorrect.")
                return

            # Update password
            if self.db.update_user_password(username, new_pw):
                QMessageBox.information(self, "Success", "Password changed successfully.")
                self.current_pw_edit.clear()
                self.new_pw_edit.clear()
                self.confirm_pw_edit.clear()
            else:
                QMessageBox.warning(self, "Error", "Failed to update password.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to change password:\n{e}")


# ------------------------------
# Login Dialog (Enhanced)
# ------------------------------
class LoginDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(TITLE)
        self.setFixedSize(500, 300)
        self.setStyleSheet("""
            QDialog { 
                background-color: #1c1c1c; 
                color: #e0e0e0; 
                font-family: 'Segoe UI', sans-serif; 
            }
            QLabel { 
                font-size: 16px; 
                font-weight: 500; 
                color: #ffffff; 
            }
            QLineEdit { 
                background-color: #2a2a2a; 
                border: 1px solid #444444; 
                border-radius: 6px; 
                padding: 6px; 
                color: #ffffff; 
            }
            QPushButton { 
                background-color: #333333; 
                border: none; 
                border-radius: 6px; 
                padding: 8px 16px; 
                color: #ffffff; 
                font-weight: bold; 
                font-size: 14px; 
            }
        """)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(10)

        # Title
        title_label = QLabel(TITLE)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("""
            font-size: 18px; 
            font-weight: 600; 
            margin-bottom: 8px;
        """)
        layout.addWidget(title_label)

        # Form
        form_layout = QFormLayout()

        self.usernameField = QLineEdit()
        self.usernameField.setPlaceholderText("Enter username")

        self.passwdField = QLineEdit()
        self.passwdField.setPlaceholderText("Enter password")
        self.passwdField.setEchoMode(QLineEdit.Password)

        form_layout.addRow("Username:", self.usernameField)
        form_layout.addRow("Password:", self.passwdField)

        # Remember me checkbox
        self.rememberCheck = QCheckBox("Remember username")
        self.rememberCheck.setChecked(True)

        # Buttons
        btn_layout = QHBoxLayout()

        login_btn = QPushButton("Login")
        login_btn.clicked.connect(self.check_credentials)

        exit_btn = QPushButton("Exit")
        exit_btn.clicked.connect(self.reject)

        btn_layout.addWidget(login_btn)
        btn_layout.addWidget(exit_btn)

        # Add to main layout
        layout.addLayout(form_layout)
        layout.addWidget(self.rememberCheck)
        layout.addLayout(btn_layout)

        self.setLayout(layout)

        # Load saved username if available
        saved_username = config_manager.get('LOGIN', 'last_username', '')
        if saved_username:
            self.usernameField.setText(saved_username)
            self.passwdField.setFocus()

        self.user = None
        self.role = None

    def check_credentials(self):
        username = self.usernameField.text().strip()
        password = self.passwdField.text().strip()

        if not username or not password:
            QMessageBox.warning(self, "Error", "Please enter both username and password.")
            return

        try:
            db = DatabaseManager()
            result = db.validate_user(username, password)

            if result:
                self.user, self.role = result
                logger.info(f"User {username} logged in successfully.")

                # Save username if "Remember me" is checked
                if self.rememberCheck.isChecked():
                    config_manager.set('LOGIN', 'last_username', username)

                self.accept()
            else:
                QMessageBox.warning(self, "Error", "Invalid username or password.")
                self.passwdField.clear()
                self.passwdField.setFocus()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Login failed:\n{e}")
            logger.error(f"Login error: {e}")


# ------------------------------
# Settings Dialog
# ------------------------------
class SettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Application Settings")
        self.setMinimumSize(500, 400)
        self.db = DatabaseManager()
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()

        # Tab widget for different settings categories
        self.tabs = QTabWidget()

        # General Settings Tab
        general_tab = QWidget()
        general_layout = QFormLayout()

        # Output directory
        self.output_dir_edit = QLineEdit()
        self.output_dir_edit.setText(config_manager.get('DEFAULT', 'output_dir'))
        self.output_dir_btn = QPushButton("Browse...")
        self.output_dir_btn.clicked.connect(self.select_output_dir)

        dir_layout = QHBoxLayout()
        dir_layout.addWidget(self.output_dir_edit)
        dir_layout.addWidget(self.output_dir_btn)

        general_layout.addRow("Output Directory:", dir_layout)

        # Default label size
        self.label_size_combo = QComboBox()
        self.label_size_combo.addItems(LABEL_SIZES.keys())

        current_size = config_manager.get('DEFAULT', 'default_label_size', 'Medium (100x25 mm)')
        index = self.label_size_combo.findText(current_size)
        if index >= 0:
            self.label_size_combo.setCurrentIndex(index)

        general_layout.addRow("Default Label Size:", self.label_size_combo)

        # Show barcode numbers
        self.show_numbers_check = QCheckBox("Show barcode numbers on labels")
        show_numbers = config_manager.get('DEFAULT', 'show_barcode_number', 'False')
        self.show_numbers_check.setChecked(show_numbers.lower() == 'true')
        general_layout.addRow(self.show_numbers_check)

        # Draw label borders
        self.draw_borders_check = QCheckBox("Draw borders around labels")
        draw_borders = config_manager.get('DEFAULT', 'draw_borders', 'False')
        self.draw_borders_check.setChecked(draw_borders.lower() == 'true')
        general_layout.addRow(self.draw_borders_check)

        general_tab.setLayout(general_layout)

        # Database Tab
        db_tab = QWidget()
        db_layout = QVBoxLayout()

        # Backup button
        backup_btn = QPushButton("Create Database Backup")
        backup_btn.clicked.connect(self.create_backup)

        # Clear history
        clear_history_btn = QPushButton("Clear Old History")
        clear_history_btn.clicked.connect(self.clear_history)

        db_layout.addWidget(backup_btn)
        db_layout.addWidget(clear_history_btn)
        db_layout.addStretch()

        db_tab.setLayout(db_layout)

        # Add tabs
        self.tabs.addTab(general_tab, "General")
        self.tabs.addTab(db_tab, "Database")

        # Buttons
        btn_layout = QHBoxLayout()

        save_btn = QPushButton("Save")
        save_btn.clicked.connect(self.save_settings)

        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)

        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(cancel_btn)

        # Add to main layout
        layout.addWidget(self.tabs)
        layout.addLayout(btn_layout)

        self.setLayout(layout)

    def select_output_dir(self):
        dir_path = QFileDialog.getExistingDirectory(
            self,
            "Select Output Directory",
            self.output_dir_edit.text()
        )

        if dir_path:
            self.output_dir_edit.setText(dir_path)

    def create_backup(self):
        try:
            backup_path = self.db.backup_database()
            if backup_path:
                QMessageBox.information(
                    self,
                    "Backup Created",
                    f"Database backup created at:\n{backup_path}",
                    QMessageBox.Ok
                )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Backup Failed",
                f"Failed to create backup:\n{e}",
                QMessageBox.Ok
            )

    def clear_history(self):
        try:
            days, ok = QInputDialog.getInt(
                self,
                "Clear Old History",
                "Delete records older than (days):",
                30, 1, 365, 1
            )

            if ok:
                count = self.db.clear_history(days)
                QMessageBox.information(
                    self,
                    "History Cleared",
                    f"Deleted {count} old records.",
                    QMessageBox.Ok
                )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Clear Failed",
                f"Failed to clear history:\n{e}",
                QMessageBox.Ok
            )

    def save_settings(self):
        try:
            # Save general settings
            config_manager.set('DEFAULT', 'output_dir', self.output_dir_edit.text())
            config_manager.set('DEFAULT', 'default_label_size', self.label_size_combo.currentText())
            config_manager.set('DEFAULT', 'show_barcode_number', str(self.show_numbers_check.isChecked()))
            config_manager.set('DEFAULT', 'draw_borders', str(self.draw_borders_check.isChecked()))

            QMessageBox.information(
                self,
                "Settings Saved",
                "Settings have been saved successfully.",
                QMessageBox.Ok
            )

            self.accept()
        except Exception as e:
            QMessageBox.critical(
                self,
                "Save Failed",
                f"Failed to save settings:\n{e}",
                QMessageBox.Ok
            )


# ------------------------------
# Main Application UI (Enhanced)
# ------------------------------
class BarcodeGeneratorUI(QWidget):
    def __init__(self, current_user, current_role):
        super().__init__()
        self.current_user = current_user
        self.current_role = current_role
        self.db = DatabaseManager()
        self.initUI()
        self.last_barcode_data = None
        self.last_label_size = None

    def initUI(self):
        self.setWindowTitle(TITLE)
        self.setGeometry(250, 250, 1100, 600)
        self.setStyleSheet("""
            QWidget { 
                background-color: #1c1c1c; 
                color: #e0e0e0; 
                font-family: 'Segoe UI', sans-serif; 
                font-size: 14px; 
            }
            QLineEdit, QComboBox { 
                background-color: #2a2a2a; 
                border: 1px solid #444444; 
                border-radius: 6px; 
                padding: 6px; 
                color: #ffffff; 
            }
            QPushButton { 
                background-color: #333333; 
                border: none; 
                border-radius: 6px; 
                padding: 10px 20px; 
                color: #ffffff; 
                font-weight: bold; 
            }
            QLabel { 
                color: #ffffff; 
                font-size: 14px; 
            }
            QTableWidget { 
                background-color: #2a2a2a; 
                border: 1px solid #444444; 
                border-radius: 6px; 
            }
        """)

        main_layout = QVBoxLayout()

        # Header
        header_layout = QHBoxLayout()

        title_label = QLabel(TITLE)
        title_label.setStyleSheet("font-size: 18px; font-weight: 600;")

        user_label = QLabel(f"User: {self.current_user} ({self.current_role})")
        user_label.setStyleSheet("font-size: 12px; color: #aaaaaa;")

        header_layout.addWidget(title_label)
        header_layout.addStretch()
        header_layout.addWidget(user_label)

        main_layout.addLayout(header_layout)

        # Form Rows
        self.create_form_rows(main_layout)
        self.create_button_row(main_layout)
        self.create_history_section(main_layout)

        self.setLayout(main_layout)

        # Load default values
        self.load_defaults()

    def create_form_rows(self, main_layout):
        # Row 1: Item Name, Purchase Price, Profit %, Sale Price, Calculator
        row1_layout = QHBoxLayout()

        row1_layout.addWidget(QLabel("Item Name:"))
        self.itemField = QLineEdit()
        self.itemField.setPlaceholderText("Product name")
        row1_layout.addWidget(self.itemField)

        row1_layout.addWidget(QLabel("Purchase Price:"))
        self.purchaseField = QLineEdit()
        self.purchaseField.setPlaceholderText("0.00")
        row1_layout.addWidget(self.purchaseField)

        row1_layout.addWidget(QLabel("Profit (%):"))
        self.profitField = QLineEdit()
        self.profitField.setPlaceholderText("optional")
        row1_layout.addWidget(self.profitField)

        row1_layout.addWidget(QLabel("Sale Price:"))
        self.saleField = QLineEdit()
        self.saleField.setPlaceholderText("0.00")
        row1_layout.addWidget(self.saleField)

        calcBtn = QPushButton("Calc")
        calcBtn.setToolTip("Calculate sale price from purchase price and profit %")
        calcBtn.clicked.connect(self.apply_profit_calculator)
        row1_layout.addWidget(calcBtn)

        main_layout.addLayout(row1_layout)

        # Row 2: Company, No. of Labels, Opening Stock, Barcode
        row2_layout = QHBoxLayout()

        row2_layout.addWidget(QLabel("Company:"))
        self.companyField = QLineEdit()
        self.companyField.setPlaceholderText("Supplier name")
        row2_layout.addWidget(self.companyField)

        row2_layout.addWidget(QLabel("No. of Labels:"))
        self.countField = QLineEdit()
        self.countField.setPlaceholderText("1")
        self.countField.setText("1")
        row2_layout.addWidget(self.countField)

        row2_layout.addWidget(QLabel("Opening Stock:"))
        self.stockField = QLineEdit()
        self.stockField.setPlaceholderText("0")
        self.stockField.setText("0")
        row2_layout.addWidget(self.stockField)

        row2_layout.addWidget(QLabel("Barcode (optional):"))
        self.barcodeField = QLineEdit()
        self.barcodeField.setPlaceholderText("Auto-generate if empty")
        row2_layout.addWidget(self.barcodeField)

        main_layout.addLayout(row2_layout)

        # Row 3: Label Size, Custom Dimensions, Continuous Roll
        row3_layout = QHBoxLayout()

        row3_layout.addWidget(QLabel("Label Size:"))
        self.sizeBox = QComboBox()
        self.sizeBox.addItems(LABEL_SIZES.keys())
        row3_layout.addWidget(self.sizeBox)

        row3_layout.addWidget(QLabel("Custom W x H (mm):"))
        self.customWidth = QLineEdit()
        self.customWidth.setPlaceholderText("Width")
        self.customWidth.setVisible(False)
        row3_layout.addWidget(self.customWidth)

        self.customHeight = QLineEdit()
        self.customHeight.setPlaceholderText("Height")
        self.customHeight.setVisible(False)
        row3_layout.addWidget(self.customHeight)

        self.continuousRollCheck = QCheckBox("Continuous Roll")
        self.continuousRollCheck.setToolTip("For continuous roll printers")
        row3_layout.addWidget(self.continuousRollCheck)

        main_layout.addLayout(row3_layout)

        # Connect signals
        self.sizeBox.currentTextChanged.connect(self.toggleCustomSize)

    def create_button_row(self, main_layout):
        btn_layout = QHBoxLayout()

        # Generate PDF Button
        genBtn = QPushButton("Generate Labels PDF")
        genBtn.setToolTip("Generate PDF with labels")
        genBtn.clicked.connect(self.generate_pdf)

        # Print Button
        printNowBtn = QPushButton("Print Label")
        printNowBtn.setToolTip("Print the last generated label")
        printNowBtn.clicked.connect(self.print_last_label)

        # Reprint Selected Button
        reprintBtn = QPushButton("Reprint Selected")
        reprintBtn.setToolTip("Reprint selected barcode from history")
        reprintBtn.clicked.connect(self.reprint_selected_barcode)

        # Clear History Button
        clearHistBtn = QPushButton("Clear History")
        clearHistBtn.setToolTip("Clear barcode generation history")
        clearHistBtn.clicked.connect(self.clear_history)

        # Open Excel Button
        openExcelBtn = QPushButton("Open Excel")
        openExcelBtn.setToolTip("Open the current Excel stock file")
        openExcelBtn.clicked.connect(self.open_excel_file)

        # Settings Button
        settingsBtn = QPushButton("Settings")
        settingsBtn.setToolTip("Application settings")
        settingsBtn.clicked.connect(self.open_settings)

        # Add buttons to layout
        btn_layout.addWidget(genBtn)
        btn_layout.addWidget(printNowBtn)
        btn_layout.addWidget(reprintBtn)
        btn_layout.addWidget(clearHistBtn)
        btn_layout.addWidget(openExcelBtn)
        btn_layout.addWidget(settingsBtn)

        # Add user management button for admins
        if self.current_role == "admin":
            manageUsersBtn = QPushButton("Manage Users")
            manageUsersBtn.setToolTip("Add, edit or remove users")
            manageUsersBtn.clicked.connect(self.manage_users)
            btn_layout.addWidget(manageUsersBtn)

        main_layout.addLayout(btn_layout)

    def create_history_section(self, main_layout):
        # Search bar
        search_layout = QHBoxLayout()

        search_layout.addWidget(QLabel("Search History:"))
        self.searchField = QLineEdit()
        self.searchField.setPlaceholderText("Search by item, barcode or company")
        self.searchField.textChanged.connect(self.refresh_history)
        search_layout.addWidget(self.searchField)

        main_layout.addLayout(search_layout)

        # History table
        self.historyTable = QTableWidget()
        self.historyTable.setColumnCount(11)
        self.historyTable.setHorizontalHeaderLabels([
            "ID", "User", "Item Name", "Barcode", "Sale Price",
            "Purchase Price", "Company", "Opening Stock", "Date",
            "Label Size", "Print Count"
        ])

        # Set column widths
        self.historyTable.setColumnWidth(0, 50)  # ID
        self.historyTable.setColumnWidth(1, 80)  # User
        self.historyTable.setColumnWidth(2, 150)  # Item Name
        self.historyTable.setColumnWidth(3, 120)  # Barcode
        self.historyTable.setColumnWidth(4, 80)  # Sale Price
        self.historyTable.setColumnWidth(5, 80)  # Purchase Price
        self.historyTable.setColumnWidth(6, 120)  # Company
        self.historyTable.setColumnWidth(7, 80)  # Opening Stock
        self.historyTable.setColumnWidth(8, 120)  # Date
        self.historyTable.setColumnWidth(9, 100)  # Label Size
        self.historyTable.setColumnWidth(10, 70)  # Print Count

        self.historyTable.setSelectionBehavior(QTableWidget.SelectRows)
        self.historyTable.setSelectionMode(QTableWidget.SingleSelection)

        main_layout.addWidget(QLabel("Barcode Generation History:"))
        main_layout.addWidget(self.historyTable)

        # Status bar
        self.statusBar = QStatusBar()
        main_layout.addWidget(self.statusBar)

    def load_defaults(self):
        """Load default values from config"""
        # Set default label size
        default_size = config_manager.get('DEFAULT', 'default_label_size', 'Medium (100x25 mm)')
        index = self.sizeBox.findText(default_size)
        if index >= 0:
            self.sizeBox.setCurrentIndex(index)

        # Set default company if available
        default_company = config_manager.get('DEFAULT', 'default_company', '')
        if default_company:
            self.companyField.setText(default_company)

        # Load history
        self.refresh_history()

    def toggleCustomSize(self):
        """Show/hide custom size fields based on selection"""
        is_custom = (self.sizeBox.currentText() == "Custom Size")
        self.customWidth.setVisible(is_custom)
        self.customHeight.setVisible(is_custom)

        if is_custom:
            self.customWidth.setFocus()

    def apply_profit_calculator(self):
        """Calculate sale price based on purchase price and profit %"""
        p_txt = self.purchaseField.text().strip().replace(",", ".")
        profit_txt = self.profitField.text().strip().replace(",", ".")

        try:
            purchase_val = float(p_txt)
        except ValueError:
            QMessageBox.warning(self, "Error", "Enter a valid numeric Purchase Price.")
            return

        if not profit_txt:
            QMessageBox.warning(self, "Error", "Enter a Profit % or type 0.")
            return

        try:
            profit_percent = float(profit_txt)
            calc_sale = round(purchase_val * (1 + profit_percent / 100), 2)
            self.saleField.setText(f"{calc_sale:.2f}")
        except ValueError:
            QMessageBox.warning(self, "Error", "Profit % must be numeric.")
            return

    def get_label_size(self):
        """Get current label size in mm (width, height)"""
        label = self.sizeBox.currentText()

        if label == "Custom Size":
            try:
                w = float(self.customWidth.text().strip().replace(",", "."))
                h = float(self.customHeight.text().strip().replace(",", "."))

                if w <= 0 or h <= 0:
                    raise ValueError("Dimensions must be positive")

                return (w, h)
            except ValueError:
                QMessageBox.warning(self, "Error", "Enter valid numeric dimensions.")
                return None

        return LABEL_SIZES[label]

    def validate_inputs(self):
        """Validate all form inputs before generation"""
        item_name = self.itemField.text().strip()
        purchase_txt = self.purchaseField.text().strip().replace(",", ".")
        sale_txt = self.saleField.text().strip().replace(",", ".")
        company = self.companyField.text().strip()
        count_txt = self.countField.text().strip()
        stock_txt = self.stockField.text().strip()

        # Required fields
        if not all([item_name, purchase_txt, sale_txt, company, count_txt, stock_txt]):
            QMessageBox.warning(self, "Error", "All fields except Barcode/Profit must be filled.")
            return False

        # Numeric validation
        try:
            purchase_val = float(purchase_txt)
            sale_val = float(sale_txt)
            count_val = int(count_txt)
            stock_val = int(stock_txt)

            if purchase_val < 0 or sale_val < 0 or count_val <= 0 or stock_val < 0:
                raise ValueError("Negative values not allowed")
        except ValueError:
            QMessageBox.warning(self, "Error",
                                "Purchase, Sale, Label Count and Stock must be positive numbers.")
            return False

        # Label size validation
        label_size = self.get_label_size()
        if not label_size:
            return False

        return True

    def generate_pdf(self):
        """Generate PDF with barcode labels"""
        if not self.validate_inputs():
            return

        # Get all input values
        item_name = self.itemField.text().strip()
        purchase_txt = self.purchaseField.text().strip().replace(",", ".")
        sale_txt = self.saleField.text().strip().replace(",", ".")
        company = self.companyField.text().strip()
        count = int(self.countField.text().strip())
        stock = int(self.stockField.text().strip())
        barcode_in = self.barcodeField.text().strip()
        label_size = self.get_label_size()

        # Generate or use provided barcode
        barcode_str = barcode_in if barcode_in else "".join(random.choices("0123456789", k=12))

        try:
            # Generate barcode image
            barcode_data = generate_barcode_image(barcode_str)
            self.last_barcode_data = barcode_data
            self.last_label_size = label_size

            # Prepare details for label
            details = {
                "item": item_name,
                "sale": sale_txt,
                "purchase": purchase_txt,
                "company": company,
                "count": count,
                "barcode": barcode_str
            }

            # Get output directory
            output_dir = config_manager.get('DEFAULT', 'output_dir')

            # Generate PDF filename
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            pdf_filename = f"labels_{item_name}_{timestamp}.pdf"
            pdf_path = os.path.join(output_dir, pdf_filename)

            # Generate the PDF
            generate_label_pdf(pdf_path, label_size, barcode_data['png'], details)

            # Show success message
            QMessageBox.information(
                self,
                "Success",
                f"PDF with {count} labels generated:\n{pdf_path}",
                QMessageBox.Ok
            )

            # Record in database
            gen_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            label_size_str = self.sizeBox.currentText()

            self.db.add_barcode_history(
                self.current_user, item_name, barcode_str,
                float(sale_txt), float(purchase_txt), company,
                gen_date, stock, label_size_str
            )

            self.db.add_or_update_barcode(
                item_name, barcode_str, float(sale_txt),
                float(purchase_txt), company, count, stock,
                label_size_str
            )

            # Update Excel
            update_excel(
                barcode_str, item_name, sale_txt, purchase_txt,
                company, gen_date, stock
            )

            # Refresh history
            self.refresh_history()

            # Update status bar
            self.statusBar.showMessage(
                f"Generated {count} labels for {item_name}",
                5000
            )

        except ValueError as e:
            QMessageBox.warning(
                self,
                "Barcode Error",
                f"Invalid barcode data:\n{e}",
                QMessageBox.Ok
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Generation Error",
                f"Failed to generate labels:\n{e}",
                QMessageBox.Ok
            )
            logger.error(f"Label generation failed: {e}")

    def print_last_label(self):
        """Print the last generated label"""
        if not self.last_barcode_data:
            QMessageBox.warning(
                self,
                "Print Error",
                "No label has been generated yet.",
                QMessageBox.Ok
            )
            return

        try:
            print_image_via_dialog(
                self,
                self.last_barcode_data['png'],
                self.last_label_size
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Print Error",
                f"Failed to print label:\n{e}",
                QMessageBox.Ok
            )
            logger.error(f"Print failed: {e}")

    def reprint_selected_barcode(self):
        """Reprint a barcode selected from history"""
        selected = self.historyTable.selectedItems()
        if not selected:
            QMessageBox.warning(
                self,
                "Reprint",
                "Please select a barcode record to reprint.",
                QMessageBox.Ok
            )
            return

        try:
            barcode_value = selected[3].text()  # Barcode is in column 3
            label_size_str = selected[9].text()  # Label size is in column 9

            # Get label size in mm
            if label_size_str in LABEL_SIZES:
                label_size = LABEL_SIZES[label_size_str]
            else:
                label_size = (50, 25)  # Default size

            # Generate barcode image
            barcode_data = generate_barcode_image(barcode_value)

            # Print it
            print_image_via_dialog(
                self,
                barcode_data['png'],
                label_size
            )

            # Update status
            self.statusBar.showMessage(
                f"Reprinted barcode: {barcode_value}",
                3000
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Reprint Error",
                f"Failed to reprint barcode:\n{e}",
                QMessageBox.Ok
            )
            logger.error(f"Reprint failed: {e}")

    def refresh_history(self):
        """Refresh the history table with current filter"""
        search_text = self.searchField.text().strip()
        history = self.db.fetch_history(search_text)

        self.historyTable.setRowCount(len(history))

        for row_idx, record in enumerate(history):
            for col_idx, value in enumerate(record):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignCenter)

                # Make certain columns non-editable
                if col_idx not in (2, 3, 6):  # Item, Barcode, Company
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)

                self.historyTable.setItem(row_idx, col_idx, item)

        self.historyTable.resizeRowsToContents()

    def clear_history(self):
        """Clear the barcode generation history"""
        reply = QMessageBox.question(
            self,
            "Clear History",
            "Are you sure you want to clear the entire barcode history?\n"
            "This action cannot be undone.",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                rows_affected = self.db.clear_history()

                QMessageBox.information(
                    self,
                    "History Cleared",
                    f"Cleared {rows_affected} history records.",
                    QMessageBox.Ok
                )

                self.refresh_history()
                self.statusBar.showMessage(
                    "History cleared successfully.",
                    5000
                )
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Clear Error",
                    f"Failed to clear history:\n{e}",
                    QMessageBox.Ok
                )
                logger.error(f"History clear failed: {e}")

    def open_excel_file(self):
        """Open the current Excel stock file"""
        try:
            excel_dir = config_manager.get('DEFAULT', 'excel_path')
            today = datetime.datetime.now()
            month_year = today.strftime("%m-%Y")
            excel_file = os.path.join(excel_dir, f"Stock_{month_year}.xlsx")

            if not os.path.exists(excel_file):
                QMessageBox.information(
                    self,
                    "Excel Not Found",
                    "No Excel file exists for the current month yet.",
                    QMessageBox.Ok
                )
                return

            if platform.system() == "Windows":
                os.startfile(excel_file)
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", excel_file])
            else:  # Linux
                subprocess.run(["xdg-open", excel_file])

            self.statusBar.showMessage(
                f"Opened Excel file: {excel_file}",
                3000
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Open Error",
                f"Failed to open Excel file:\n{e}",
                QMessageBox.Ok
            )
            logger.error(f"Excel open failed: {e}")

    def manage_users(self):
        """Open the user management dialog"""
        if self.current_role != "admin":
            QMessageBox.warning(
                self,
                "Access Denied",
                "Only admin users can manage users.",
                QMessageBox.Ok
            )
            return

        dialog = UserManagementDialog(self)
        dialog.exec_()

    def open_settings(self):
        """Open the settings dialog"""
        dialog = SettingsDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            # Reload any settings that might have changed
            self.load_defaults()

            self.statusBar.showMessage(
                "Settings updated successfully.",
                3000
            )


# ------------------------------
# Application Entry Point
# ------------------------------
if __name__ == "__main__":
    # Required for PyInstaller and multiprocessing on Windows
    multiprocessing.freeze_support()

    # Check if another instance is already running
    if is_already_running():
        print("Another instance is already running. Exiting.")
        sys.exit(0)

    # Windows-specific checks
    if platform.system() != "Windows":
        print("This application is optimized for Windows. Some features may not work properly.")

    # Enable High DPI scaling
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    # Create application instance
    app = QApplication(sys.argv)
    app.setStyle('Fusion')

    # Show login dialog
    login_dialog = LoginDialog()

    if login_dialog.exec_() == QDialog.Accepted:
        # Login successful - create and show main window
        main_window = BarcodeGeneratorUI(login_dialog.user, login_dialog.role)
        main_window.show()

        # Start application event loop
        sys.exit(app.exec_())
    else:
        # Login canceled or failed
        sys.exit(0)