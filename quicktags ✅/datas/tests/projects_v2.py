import sys
import ctypes
import multiprocessing
import subprocess
import platform
import logging
import os
import random
import datetime
import sqlite3
import csv
from typing import Dict, List, Tuple, Optional
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QFormLayout, QDialog, QDialogButtonBox,
    QPushButton, QMessageBox, QLineEdit, QFileDialog, QLabel,
    QComboBox, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView,
    QCheckBox, QStatusBar, QGroupBox, QScrollArea, QTextEdit, QSpinBox,
    QDoubleSpinBox, QTabWidget, QSizePolicy, QGraphicsView, QGraphicsScene,
    QGraphicsPixmapItem, QSplitter, QToolButton, QMenu, QAction, QInputDialog
)
from PyQt5.QtCore import Qt, QSize, QRectF, pyqtSignal
from PyQt5.QtGui import (
    QFont, QImage, QPainter, QPixmap, QColor, QIcon,
    QTextCursor, QFontDatabase
)
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
# ------------------------------
# Constants and Configurations
# ------------------------------
TITLE = "Projects \U0001F4C8 by Shenz C (v2 AutoGeek Edition)"
LOGO_PATH = "logo.ico"  # Assumed to be in the same directory
DEFAULT_FONT = "Arial"
DEFAULT_FONT_SIZE = 8
DPI_OPTIONS = [150, 200, 300, 600]
DEFAULT_DPI = 300
# Label sizes in mm (width, height)
LABEL_SIZES = {
    "Small (60x25 mm)": (60, 25),
    "Medium (100x25 mm)": (100, 25),
    "Large (150x50 mm)": (150, 50),
    "A4 Sheet (210x297 mm)": (210, 297),
    "2 Labels (50x25 mm)": (100, 25),  # Will be split into two 50x25 labels
    "Custom Size": None
}
# ------------------------------
# Single Instance Check (Windows Only)
# ------------------------------
def is_already_running():
    if platform.system() != "Windows":
        return False
    mutex = ctypes.windll.kernel32.CreateMutexW(None, 1, "Global\\ProjectsByShenzC_FullSafeApp_V2")
    return ctypes.GetLastError() == 183
# ------------------------------
# App Data Directory and Global Paths
# ------------------------------
def get_app_data_path():
    base_path = os.path.join(os.environ["APPDATA"], "BarcodeGeneratorV2")
    os.makedirs(base_path, exist_ok=True)
    return base_path
def get_today_folder(base_path: str) -> str:
    today = datetime.datetime.now().strftime("%d-%m-%y")
    today_path = os.path.join(base_path, today)
    os.makedirs(today_path, exist_ok=True)
    return today_path
def setup_output_folders(base_path: str) -> Dict[str, str]:
    today_path = get_today_folder(base_path)
    folders = {
        "pdf": os.path.join(today_path, "PDF"),
        "png": os.path.join(today_path, "PNG"),
        "bmp": os.path.join(today_path, "BMP"),
        "excel": os.path.join(today_path, "Excel")
    }
    for folder in folders.values():
        os.makedirs(folder, exist_ok=True)
    return folders
APP_DATA_PATH = get_app_data_path()
DB_PATH = os.path.join(APP_DATA_PATH, "barcodes_v2.db")
LOG_PATH = os.path.join(APP_DATA_PATH, "barcode_app_v2.log")
TEMPLATES_PATH = os.path.join(APP_DATA_PATH, "templates")
os.makedirs(TEMPLATES_PATH, exist_ok=True)
# ------------------------------
# Logging Setup
# ------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)
def show_error(message: str, parent=None):
    """Show error message both in UI and log it."""
    logger.error(message)
    QMessageBox.critical(parent, "Error", message)
def show_warning(message: str, parent=None):
    """Show warning message both in UI and log it."""
    logger.warning(message)
    QMessageBox.warning(parent, "Warning", message)
def show_info(message: str, parent=None):
    """Show info message both in UI and log it."""
    logger.info(message)
    QMessageBox.information(parent, "Information", message)
# ------------------------------
# Database Manager (Upgraded for V2)
# ------------------------------
class DatabaseManager:
    def __init__(self, db_path=DB_PATH):
        self.db_path = db_path
        try:
            self.conn = sqlite3.connect(self.db_path)
            self.create_tables()
            self.create_default_templates()
            logger.info("Database connected and tables created/verified.")
        except Exception as e:
            show_error(f"Database connection failed: {e}")
            raise
    def create_tables(self):
        try:
            c = self.conn.cursor()
            # Barcode history with stock_count
            c.execute('''
                CREATE TABLE IF NOT EXISTS barcode_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    item_name TEXT,
                    barcode TEXT,
                    sale_price REAL,
                    purchase_price REAL,
                    company_name TEXT,
                    generation_date TEXT,
                    username TEXT DEFAULT 'admin',
                    opening_stock INTEGER DEFAULT 0,
                    template_used TEXT
                )
            ''')
            # Barcodes table with stock
            c.execute('''
                CREATE TABLE IF NOT EXISTS barcodes (
                    barcode TEXT PRIMARY KEY,
                    item_name TEXT,
                    sale_price REAL,
                    purchase_price REAL,
                    company_name TEXT,
                    quantity INTEGER,
                    opening_stock INTEGER DEFAULT 0,
                    template_used TEXT
                )
            ''')
            # Users table
            c.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    username TEXT PRIMARY KEY,
                    password TEXT,
                    role TEXT
                )
            ''')
            # Templates table
            c.execute('''
                CREATE TABLE IF NOT EXISTS templates (
                    name TEXT PRIMARY KEY,
                    font_family TEXT,
                    font_size INTEGER,
                    barcode_scale REAL,
                    show_price BOOLEAN,
                    show_company BOOLEAN,
                    show_stock BOOLEAN,
                    created_by TEXT,
                    created_date TEXT
                )
            ''')
            # Insert default admin if none exist
            c.execute('SELECT COUNT(*) FROM users')
            if c.fetchone()[0] == 0:
                c.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)',
                          ("admin", "Kungfukenny", "admin"))
            self.conn.commit()
        except Exception as e:
            show_error(f"Error creating tables: {e}")
            raise
    def create_default_templates(self):
        try:
            c = self.conn.cursor()
            c.execute('SELECT COUNT(*) FROM templates')
            if c.fetchone()[0] == 0:
                # Create a default template
                default_template = (
                    "Default", "Arial", 8, 1.0, True, True, False,
                    "system", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                )
                c.execute('''
                    INSERT INTO templates 
                    (name, font_family, font_size, barcode_scale, show_price, show_company, show_stock, created_by, created_date)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', default_template)
                self.conn.commit()
        except Exception as e:
            show_error(f"Error creating default templates: {e}")
    def add_barcode_history(self, username: str, item_name: str, barcode: str, sale_price: float,
                            purchase_price: float, company_name: str, generation_date: str,
                            opening_stock: int, template_used: str):
        try:
            c = self.conn.cursor()
            c.execute('''
                INSERT INTO barcode_history
                (username, item_name, barcode, sale_price, purchase_price, company_name, 
                 generation_date, opening_stock, template_used)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (username, item_name, barcode, sale_price, purchase_price,
                  company_name, generation_date, opening_stock, template_used))
            self.conn.commit()
            logger.info("Barcode history added.")
        except Exception as e:
            show_error(f"Failed to add barcode history: {e}")
            raise
    def add_or_update_barcode(self, item_name: str, barcode: str, sale_price: float,
                              purchase_price: float, company_name: str, quantity: int,
                              opening_stock: int, template_used: str):
        try:
            c = self.conn.cursor()
            c.execute('SELECT quantity FROM barcodes WHERE barcode=?', (barcode,))
            result = c.fetchone()
            if result:
                new_quantity = result[0] + quantity
                c.execute('''
                    UPDATE barcodes
                    SET item_name=?, sale_price=?, purchase_price=?, company_name=?, 
                        quantity=?, opening_stock=?, template_used=?
                    WHERE barcode=?
                ''', (item_name, sale_price, purchase_price, company_name,
                      new_quantity, opening_stock, template_used, barcode))
            else:
                c.execute('''
                    INSERT INTO barcodes
                    (barcode, item_name, sale_price, purchase_price, company_name, 
                     quantity, opening_stock, template_used)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (barcode, item_name, sale_price, purchase_price, company_name,
                      quantity, opening_stock, template_used))
            self.conn.commit()
            logger.info("Barcode updated/added in barcodes table.")
        except Exception as e:
            show_error(f"Failed to add or update barcode: {e}")
            raise
    def fetch_history(self, search_filter: str = "") -> List[Tuple]:
        try:
            c = self.conn.cursor()
            query = """
                SELECT id, username, item_name, barcode, sale_price, purchase_price, 
                       company_name, opening_stock, generation_date, template_used
                FROM barcode_history
            """
            params = ()
            if search_filter:
                query += " WHERE item_name LIKE ? OR barcode LIKE ?"
                params = (f"%{search_filter}%", f"%{search_filter}%")
            query += " ORDER BY id DESC"
            c.execute(query, params)
            history = c.fetchall()
            logger.info("History fetched from database.")
            return history
        except Exception as e:
            show_error(f"Failed to fetch history: {e}")
            return []
    def clear_history(self):
        try:
            c = self.conn.cursor()
            c.execute('DELETE FROM barcode_history')
            self.conn.commit()
            logger.info("Barcode history cleared.")
        except Exception as e:
            show_error(f"Failed to clear history: {e}")
            raise
    def add_user(self, username: str, password: str, role: str = "user"):
        try:
            c = self.conn.cursor()
            c.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)',
                      (username, password, role))
            self.conn.commit()
            logger.info(f"User {username} added.")
        except Exception as e:
            show_error(f"Failed to add user: {e}")
            raise
    def remove_user(self, username: str):
        try:
            c = self.conn.cursor()
            c.execute('DELETE FROM users WHERE username=?', (username,))
            self.conn.commit()
            logger.info(f"User {username} removed.")
        except Exception as e:
            show_error(f"Failed to remove user: {e}")
            raise
    def fetch_users(self) -> List[Tuple]:
        try:
            c = self.conn.cursor()
            c.execute('SELECT username, role FROM users ORDER BY username ASC')
            users = c.fetchall()
            logger.info("Fetched users from database.")
            return users
        except Exception as e:
            show_error(f"Failed to fetch users: {e}")
            return []
    def validate_user(self, username: str, password: str) -> Optional[Tuple]:
        try:
            c = self.conn.cursor()
            c.execute('SELECT username, role FROM users WHERE username=? AND password=?',
                      (username, password))
            result = c.fetchone()
            logger.info(f"User {username} validation attempted.")
            return result
        except Exception as e:
            show_error(f"User validation failed: {e}")
            return None
    def add_template(self, name: str, font_family: str, font_size: int, barcode_scale: float,
                     show_price: bool, show_company: bool, show_stock: bool, created_by: str):
        try:
            c = self.conn.cursor()
            created_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            c.execute('''
                INSERT INTO templates 
                (name, font_family, font_size, barcode_scale, show_price, show_company, show_stock, created_by, created_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (name, font_family, font_size, barcode_scale, show_price, show_company, show_stock, created_by,
                  created_date))
            self.conn.commit()
            logger.info(f"Template {name} added by {created_by}.")
            return True
        except sqlite3.IntegrityError:
            show_warning(f"Template '{name}' already exists.")
            return False
        except Exception as e:
            show_error(f"Failed to add template: {e}")
            return False
    def update_template(self, name: str, font_family: str, font_size: int, barcode_scale: float,
                        show_price: bool, show_company: bool, show_stock: bool):
        try:
            c = self.conn.cursor()
            c.execute('''
                UPDATE templates 
                SET font_family=?, font_size=?, barcode_scale=?, show_price=?, show_company=?, show_stock=?
                WHERE name=?
            ''', (font_family, font_size, barcode_scale, show_price, show_company, show_stock, name))
            self.conn.commit()
            logger.info(f"Template {name} updated.")
            return True
        except Exception as e:
            show_error(f"Failed to update template: {e}")
            return False
    def delete_template(self, name: str):
        try:
            c = self.conn.cursor()
            c.execute('DELETE FROM templates WHERE name=?', (name,))
            self.conn.commit()
            logger.info(f"Template {name} deleted.")
            return True
        except Exception as e:
            show_error(f"Failed to delete template: {e}")
            return False
    def get_template(self, name: str) -> Optional[Dict]:
        try:
            c = self.conn.cursor()
            c.execute('''
                SELECT name, font_family, font_size, barcode_scale, show_price, show_company, show_stock
                FROM templates WHERE name=?
            ''', (name,))
            result = c.fetchone()
            if result:
                return {
                    "name": result[0],
                    "font_family": result[1],
                    "font_size": result[2],
                    "barcode_scale": result[3],
                    "show_price": bool(result[4]),
                    "show_company": bool(result[5]),
                    "show_stock": bool(result[6])
                }
            return None
        except Exception as e:
            show_error(f"Failed to get template: {e}")
            return None
    def get_all_templates(self) -> List[Tuple]:
        try:
            c = self.conn.cursor()
            c.execute('SELECT name, created_by, created_date FROM templates ORDER BY name ASC')
            return c.fetchall()
        except Exception as e:
            show_error(f"Failed to get templates: {e}")
            return []


# ------------------------------
# Barcode Generation (Code128)
# ------------------------------
try:
    from barcode import get_barcode_class
    from barcode.writer import ImageWriter
except ImportError as e:
    show_error("Missing dependency 'python-barcode'. Please install it before running this application.")
    raise e


def sanitize_barcode_input(barcode_str: str) -> str:
    """Ensure barcode only contains valid Code128 characters."""
    return "".join(ch for ch in barcode_str if 32 <= ord(ch) <= 126)


def generate_barcode_image(barcode_str: str, scale: float = 1.0, dpi: int = 300,
                           font_path: str = None) -> str:
    """
    Generate a barcode image with the given specifications.
    Returns path to the generated image.
    """
    safe_barcode = sanitize_barcode_input(barcode_str)
    if not safe_barcode:
        raise ValueError("No valid Code128 characters after sanitizing.")

    try:
        Code128 = get_barcode_class('code128')
        output_folders = setup_output_folders(APP_DATA_PATH)
        base_filename = os.path.join(output_folders["png"], safe_barcode)

        # Custom writer with specified DPI and scale
        class CustomWriter(ImageWriter):
            def __init__(self):
                super().__init__()
                self.dpi = dpi
                self.set_options({"module_width": 0.2 * scale, "module_height": 15.0 * scale})
                if font_path:
                    self.set_options({"font_path": font_path})

        writer = CustomWriter()
        code128_obj = Code128(safe_barcode, writer=writer)

        # Save to PNG first (highest quality)
        png_path = code128_obj.save(base_filename)

        # Convert to BMP if needed (for thermal printers)
        img = QImage(png_path)
        bmp_path = os.path.join(output_folders["bmp"], f"{safe_barcode}.bmp")
        img.save(bmp_path, "BMP")

        logger.info(f"Barcode images generated at {png_path} and {bmp_path}")
        return png_path
    except Exception as e:
        show_error(f"Barcode generation failed: {e}")
        raise


# ------------------------------
# PDF Generation for Labels (using ReportLab)
# ------------------------------
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
except ImportError as e:
    show_error("Missing dependency 'reportlab'. Please install it before running this application.")
    raise e


def draw_label(c: canvas.Canvas, x_offset: float, y_offset: float,
               width: float, height: float, barcode_path: str,
               details: Dict, template: Dict, dpi: int = 300):
    """
    Draw a single label with the given specifications.
    """
    # Draw border
    c.rect(x_offset * mm, y_offset * mm, width * mm, height * mm)

    # Register font if not already registered
    font_name = template["font_family"]
    try:
        pdfmetrics.registerFont(TTFont(font_name, f"{font_name}.ttf"))
    except:
        # Fallback to Helvetica if custom font not found
        font_name = "Helvetica"

    # Set font for text
    c.setFont(font_name, template["font_size"])

    # Calculate positions
    center_x = x_offset + width / 2
    top_y = y_offset + height - 2  # 2mm from top

    # Item name (top center)
    c.drawCentredString(center_x * mm, top_y * mm, details["item"])

    # Barcode (center)
    barcode_y = y_offset + height - 15  # 15mm from top
    barcode_width = 40 * template["barcode_scale"]
    c.drawImage(barcode_path, (center_x - barcode_width / 2) * mm,
                barcode_y * mm, barcode_width * mm, 10 * mm, preserveAspectRatio=True)

    # Additional info (bottom)
    info_lines = []
    if template["show_price"]:
        info_lines.append(f"Sale: {details['sale']}")
    if template["show_company"]:
        info_lines.append(details["company"])
    if template["show_stock"]:
        info_lines.append(f"Stock: {details['opening_stock']}")

    info_text = " | ".join(info_lines)
    if info_text:
        c.drawCentredString(center_x * mm, (y_offset + 5) * mm, info_text)

    # Add footer branding
    c.setFont("Helvetica", 6)
    c.drawCentredString(center_x * mm, (y_offset + 1) * mm, "Projects by Shenz C - AutoGeek Edition")


def generate_label_pdf(pdf_path: str, width_mm: float, height_mm: float,
                       barcode_path: str, details: Dict, template: Dict,
                       dpi: int = 300) -> bool:
    """
    Generate a PDF with labels according to specifications.
    Returns True if successful, False otherwise.
    """
    try:
        count = details.get("count", 1)
        is_dual = (width_mm, height_mm) == (100, 25)  # Special case for 2 labels per row

        # Create canvas with appropriate page size
        if is_dual:
            c = canvas.Canvas(pdf_path, pagesize=(width_mm * mm, height_mm * mm))
        else:
            c = canvas.Canvas(pdf_path,
                              pagesize=A4 if (width_mm, height_mm) == (210, 297) else (width_mm * mm, height_mm * mm))

        # For dual labels (2 per row)
        if is_dual:
            labels_per_page = 2
            labels_remaining = count

            while labels_remaining > 0:
                # First label in row
                draw_label(c, 0, 0, 50, 25, barcode_path, details, template, dpi)
                labels_remaining -= 1

                # Second label in row if needed
                if labels_remaining > 0:
                    draw_label(c, 50, 0, 50, 25, barcode_path, details, template, dpi)
                    labels_remaining -= 1

                c.showPage()
        else:
            # Single label per page
            for _ in range(count):
                draw_label(c, 0, 0, width_mm, height_mm, barcode_path, details, template, dpi)
                c.showPage()

        c.save()
        logger.info(f"PDF generated successfully at {pdf_path}")
        return True
    except Exception as e:
        show_error(f"PDF generation failed: {e}")
        return False


# ------------------------------
# Excel Export with Stock Management
# ------------------------------
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment
except ImportError as e:
    show_error("Missing dependency 'openpyxl'. Please install it before running this application.")
    raise e


def update_excel(item_code: str, item_name: str, sale_price: str,
                 purchase_price: str, company_name: str, generation_date: str,
                 opening_stock: int) -> bool:
    """
    Update or create Excel file with item data.
    Returns True if successful, False otherwise.
    """
    try:
        output_folders = setup_output_folders(APP_DATA_PATH)
        today = datetime.datetime.now().strftime("%d-%m-%y")
        file_name = os.path.join(output_folders["excel"], f"Stock {today}.xlsx")

        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "Items"
            # Header row with styling
            headers = ["Item Code", "Item Name", "Sale Price", "Purchase Price",
                       "Company", "Opening Stock", "Generation Date"]
            ws.append(headers)

            # Style header row
            bold_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = bold_font
                cell.alignment = Alignment(horizontal='center')

            wb.save(file_name)

        wb = load_workbook(file_name)
        ws = wb.active
        found = False

        # Check if item already exists
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
            if row[0].value == item_code:
                # Update existing row
                row[1].value = item_name
                row[2].value = sale_price
                row[3].value = purchase_price
                row[4].value = company_name
                row[5].value = opening_stock
                row[6].value = generation_date
                found = True
                break

        if not found:
            # Add new row
            ws.append([item_code, item_name, sale_price, purchase_price,
                       company_name, opening_stock, generation_date])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save with retry logic
        for attempt in range(3):
            try:
                wb.save(file_name)
                logger.info(f"Excel updated: item_code={item_code} (found={found})")
                return True
            except Exception as e:
                if attempt == 2:
                    show_error(f"Failed to save Excel after 3 attempts: {e}")
                    return False

                ret = QMessageBox.question(None, "Excel Locked",
                                           f"Unable to save {file_name} (maybe open/locked?).\n"
                                           "Close the file and click 'Retry' or 'Cancel' to skip updating.",
                                           QMessageBox.Retry | QMessageBox.Cancel)
                if ret == QMessageBox.Retry:
                    continue
                else:
                    logger.warning("User canceled Excel update.")
                    return False
    except Exception as e:
        show_error(f"Excel update failed: {e}")
        return False


# ------------------------------
# Preview Widget for Label
# ------------------------------
class LabelPreview(QGraphicsView):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.scene = QGraphicsScene(self)
        self.setScene(self.scene)
        self.setRenderHint(QPainter.Antialiasing)
        self.setRenderHint(QPainter.SmoothPixmapTransform)
        self.setRenderHint(QPainter.TextAntialiasing)
        self.setBackgroundBrush(QColor(240, 240, 240))
        self.setMinimumSize(200, 100)
        
        # Add placeholder text
        self.placeholder = QGraphicsTextItem("Preview will appear here after generation")
        self.placeholder.setDefaultTextColor(QColor(150, 150, 150))
        self.scene.addItem(self.placeholder)

    def update_preview(self, pixmap: QPixmap):
        """Update the preview with a new pixmap."""
        self.scene.clear()
        if pixmap.isNull():
            # Show placeholder if no valid pixmap
            self.scene.addItem(self.placeholder)
            return

        # Scale the pixmap to fit the view while maintaining aspect ratio
        scaled_pixmap = pixmap.scaled(self.width(), self.height(),
                                    Qt.KeepAspectRatio, Qt.SmoothTransformation)
        item = QGraphicsPixmapItem(scaled_pixmap)
        self.scene.addItem(item)
        self.fitInView(item, Qt.KeepAspectRatio)
# ------------------------------
# Template Management Dialog
# ------------------------------
class TemplateDialog(QDialog):
    def __init__(self, db: DatabaseManager, current_user: str, parent=None):
        super().__init__(parent)
        self.db = db
        self.current_user = current_user
        self.current_template = None
        self.setWindowTitle("Manage Label Templates")
        self.setMinimumSize(600, 500)

        self.init_ui()
        self.load_templates()

    def init_ui(self):
        layout = QVBoxLayout()

        # Template selection
        self.template_combo = QComboBox()
        self.template_combo.currentTextChanged.connect(self.template_selected)
        self.load_template_btn = QPushButton("Load Template")
        self.load_template_btn.clicked.connect(self.load_template)

        # Template controls
        btn_layout = QHBoxLayout()
        self.new_template_btn = QPushButton("New Template")
        self.new_template_btn.clicked.connect(self.create_template)
        self.save_template_btn = QPushButton("Save Template")
        self.save_template_btn.clicked.connect(self.save_template)
        self.delete_template_btn = QPushButton("Delete Template")
        self.delete_template_btn.clicked.connect(self.delete_template)
        btn_layout.addWidget(self.new_template_btn)
        btn_layout.addWidget(self.save_template_btn)
        btn_layout.addWidget(self.delete_template_btn)

        # Template settings
        settings_group = QGroupBox("Template Settings")
        form_layout = QFormLayout()

        self.font_combo = QComboBox()
        self.font_combo.addItems(sorted(QFontDatabase().families()))
        self.font_combo.setCurrentText(DEFAULT_FONT)

        self.font_size = QSpinBox()
        self.font_size.setRange(6, 24)
        self.font_size.setValue(DEFAULT_FONT_SIZE)

        self.barcode_scale = QDoubleSpinBox()
        self.barcode_scale.setRange(0.5, 2.0)
        self.barcode_scale.setSingleStep(0.1)
        self.barcode_scale.setValue(1.0)

        self.show_price = QCheckBox()
        self.show_price.setChecked(True)

        self.show_company = QCheckBox()
        self.show_company.setChecked(True)

        self.show_stock = QCheckBox()
        self.show_stock.setChecked(False)

        form_layout.addRow("Font Family:", self.font_combo)
        form_layout.addRow("Font Size:", self.font_size)
        form_layout.addRow("Barcode Scale:", self.barcode_scale)
        form_layout.addRow("Show Price:", self.show_price)
        form_layout.addRow("Show Company:", self.show_company)
        form_layout.addRow("Show Stock:", self.show_stock)

        settings_group.setLayout(form_layout)

        # Preview area
        self.preview = LabelPreview()

        # Add widgets to main layout
        layout.addWidget(QLabel("Select Template:"))
        layout.addWidget(self.template_combo)
        layout.addWidget(self.load_template_btn)
        layout.addLayout(btn_layout)
        layout.addWidget(settings_group)
        layout.addWidget(QLabel("Preview:"))
        layout.addWidget(self.preview)

        self.setLayout(layout)

    def load_templates(self):
        """Load available templates from database."""
        self.template_combo.clear()
        templates = self.db.get_all_templates()
        for name, created_by, _ in templates:
            self.template_combo.addItem(f"{name} (by {created_by})", name)

    def template_selected(self, text):
        """Handle template selection change."""
        if not text:
            return

        template_name = self.template_combo.currentData()
        self.current_template = self.db.get_template(template_name)

        if self.current_template:
            self.font_combo.setCurrentText(self.current_template["font_family"])
            self.font_size.setValue(self.current_template["font_size"])
            self.barcode_scale.setValue(self.current_template["barcode_scale"])
            self.show_price.setChecked(self.current_template["show_price"])
            self.show_company.setChecked(self.current_template["show_company"])
            self.show_stock.setChecked(self.current_template["show_stock"])

            # Generate preview
            self.generate_preview()

    def load_template(self):
        """Load the selected template and close the dialog."""
        if self.current_template:
            self.accept()

    def create_template(self):
        """Create a new template."""
        name, ok = QInputDialog.getText(self, "New Template", "Template Name:")
        if not ok or not name:
            return

        # Check if template already exists
        if self.db.get_template(name):
            show_warning(f"Template '{name}' already exists.", self)
            return

        # Create new template with current settings
        template_data = {
            "font_family": self.font_combo.currentText(),
            "font_size": self.font_size.value(),
            "barcode_scale": self.barcode_scale.value(),
            "show_price": self.show_price.isChecked(),
            "show_company": self.show_company.isChecked(),
            "show_stock": self.show_stock.isChecked()
        }

        if self.db.add_template(name, **template_data, created_by=self.current_user):
            self.load_templates()
            self.template_combo.setCurrentText(f"{name} (by {self.current_user})")
            show_info(f"Template '{name}' created successfully.", self)

    def save_template(self):
        """Save changes to the current template."""
        if not self.current_template:
            show_warning("No template selected to save.", self)
            return

        template_name = self.current_template["name"]

        # Update template with current settings
        success = self.db.update_template(
            name=template_name,
            font_family=self.font_combo.currentText(),
            font_size=self.font_size.value(),
            barcode_scale=self.barcode_scale.value(),
            show_price=self.show_price.isChecked(),
            show_company=self.show_company.isChecked(),
            show_stock=self.show_stock.isChecked()
        )

        if success:
            show_info(f"Template '{template_name}' updated successfully.", self)
            self.load_templates()
            self.template_combo.setCurrentText(f"{template_name} (by {self.current_user})")

    def delete_template(self):
        """Delete the current template."""
        if not self.current_template:
            show_warning("No template selected to delete.", self)
            return

        template_name = self.current_template["name"]

        if template_name == "Default":
            show_warning("Cannot delete the default template.", self)
            return

        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete template '{template_name}'?",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            if self.db.delete_template(template_name):
                show_info(f"Template '{template_name}' deleted.", self)
                self.load_templates()
                self.current_template = None

    def generate_preview(self):
        """Generate a preview of the current template settings."""
        if not self.current_template:
            return

        # Create a temporary QPixmap to preview
        pixmap = QPixmap(300, 150)
        pixmap.fill(Qt.white)

        painter = QPainter(pixmap)
        try:
            # Set font
            font = QFont(self.font_combo.currentText(), self.font_size.value())
            painter.setFont(font)

            # Draw sample text
            painter.drawText(10, 20, "Sample Item Name")

            # Draw sample barcode (just a rectangle for preview)
            barcode_height = 40 * self.barcode_scale.value()
            painter.drawRect(100, 40, 100, barcode_height)
            painter.drawText(100, 40 + barcode_height + 15, "Barcode Preview")

            # Draw sample info
            info_lines = []
            if self.show_price.isChecked():
                info_lines.append("Sale: $10.00")
            if self.show_company.isChecked():
                info_lines.append("Sample Co.")
            if self.show_stock.isChecked():
                info_lines.append("Stock: 50")

            info_text = " | ".join(info_lines)
            if info_text:
                painter.drawText(10, 130, info_text)

            # Draw footer
            small_font = QFont("Arial", 6)
            painter.setFont(small_font)
            painter.drawText(10, 145, "Projects by Shenz C - AutoGeek Edition")

        finally:
            painter.end()

        self.preview.update_preview(pixmap)

    def get_selected_template(self) -> Optional[Dict]:
        """Return the currently selected template."""
        return self.current_template


# ------------------------------
# Bulk Entry Dialog
# ------------------------------
class BulkEntryDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Bulk Barcode Entry")
        self.setMinimumSize(800, 600)

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        # Tab widget for different entry methods
        self.tabs = QTabWidget()

        # Manual Entry Tab
        self.manual_tab = QWidget()
        manual_layout = QVBoxLayout()

        self.manual_table = QTableWidget()
        self.manual_table.setColumnCount(6)
        self.manual_table.setHorizontalHeaderLabels([
            "Item Name", "Purchase Price", "Profit %", "Sale Price",
            "Company", "Opening Stock"
        ])
        self.manual_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        add_row_btn = QPushButton("Add Row")
        add_row_btn.clicked.connect(self.add_manual_row)
        remove_row_btn = QPushButton("Remove Selected")
        remove_row_btn.clicked.connect(self.remove_manual_row)

        btn_layout = QHBoxLayout()
        btn_layout.addWidget(add_row_btn)
        btn_layout.addWidget(remove_row_btn)

        manual_layout.addWidget(self.manual_table)
        manual_layout.addLayout(btn_layout)
        self.manual_tab.setLayout(manual_layout)

        # CSV Import Tab
        self.csv_tab = QWidget()
        csv_layout = QVBoxLayout()

        self.csv_text = QTextEdit()
        self.csv_text.setPlaceholderText(
            "Paste CSV data here or use the import button below.\n"
            "Format: Item Name,Purchase Price,Profit %,Company,Opening Stock"
        )

        csv_btn_layout = QHBoxLayout()
        import_csv_btn = QPushButton("Import CSV File")
        import_csv_btn.clicked.connect(self.import_csv)
        parse_csv_btn = QPushButton("Parse CSV Data")
        parse_csv_btn.clicked.connect(self.parse_csv)
        csv_btn_layout.addWidget(import_csv_btn)
        csv_btn_layout.addWidget(parse_csv_btn)

        csv_layout.addWidget(self.csv_text)
        csv_layout.addLayout(csv_btn_layout)
        self.csv_tab.setLayout(csv_layout)

        self.tabs.addTab(self.manual_tab, "Manual Entry")
        self.tabs.addTab(self.csv_tab, "CSV Import")

        # Preview and buttons
        self.preview_table = QTableWidget()
        self.preview_table.setColumnCount(7)
        self.preview_table.setHorizontalHeaderLabels([
            "Item Name", "Purchase", "Profit %", "Sale",
            "Company", "Stock", "Barcode"
        ])
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.validate_and_accept)
        button_box.rejected.connect(self.reject)

        layout.addWidget(self.tabs)
        layout.addWidget(QLabel("Preview:"))
        layout.addWidget(self.preview_table)
        layout.addWidget(button_box)

        self.setLayout(layout)

        # Add initial row to manual entry
        self.add_manual_row()

    def add_manual_row(self):
        row = self.manual_table.rowCount()
        self.manual_table.insertRow(row)

        # Add empty items
        for col in range(6):
            item = QTableWidgetItem()
            if col in (1, 2, 3, 5):  # Numeric columns
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.manual_table.setItem(row, col, item)

    def remove_manual_row(self):
        selected = self.manual_table.selectionModel().selectedRows()
        if not selected:
            show_warning("Please select rows to remove.", self)
            return

        for idx in sorted((row.row() for row in selected), reverse=True):
            self.manual_table.removeRow(idx)

    def import_csv(self):
        file_name, _ = QFileDialog.getOpenFileName(
            self, "Open CSV File", "", "CSV Files (*.csv);;All Files (*)"
        )

        if not file_name:
            return

        try:
            with open(file_name, 'r', encoding='utf-8') as f:
                content = f.read()
                self.csv_text.setPlainText(content)
        except Exception as e:
            show_error(f"Failed to read CSV file: {e}", self)

    def parse_csv(self):
        csv_data = self.csv_text.toPlainText().strip()
        if not csv_data:
            show_warning("No CSV data to parse.", self)
            return

        try:
            # Clear manual table first
            self.manual_table.setRowCount(0)

            # Parse CSV
            reader = csv.reader(csv_data.splitlines())
            for row in reader:
                if len(row) < 5:  # Skip incomplete rows
                    continue

                # Add to manual table
                r = self.manual_table.rowCount()
                self.manual_table.insertRow(r)

                # Item Name
                self.manual_table.setItem(r, 0, QTableWidgetItem(row[0]))

                # Purchase Price
                item = QTableWidgetItem(row[1])
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.manual_table.setItem(r, 1, item)

                # Profit % (default to 20 if empty)
                profit = row[2] if len(row) > 2 and row[2] else "20"
                item = QTableWidgetItem(profit)
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.manual_table.setItem(r, 2, item)

                # Sale Price (calculate from purchase + profit)
                try:
                    purchase = float(row[1])
                    profit_pct = float(profit)
                    sale = purchase * (1 + profit_pct / 100)
                    item = QTableWidgetItem(f"{sale:.2f}")
                except:
                    item = QTableWidgetItem("")
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.manual_table.setItem(r, 3, item)

                # Company
                company = row[3] if len(row) > 3 else ""
                self.manual_table.setItem(r, 4, QTableWidgetItem(company))

                # Opening Stock
                stock = row[4] if len(row) > 4 else "0"
                item = QTableWidgetItem(stock)
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.manual_table.setItem(r, 5, item)

            show_info(f"Imported {self.manual_table.rowCount()} items from CSV.", self)
            self.tabs.setCurrentIndex(0)  # Switch to manual tab

        except Exception as e:
            show_error(f"Failed to parse CSV: {e}", self)

    def validate_and_accept(self):
        """Validate input and prepare preview before accepting."""
        # Clear preview
        self.preview_table.setRowCount(0)

        # Process manual entries
        valid_rows = []
        for row in range(self.manual_table.rowCount()):
            item_name = self.manual_table.item(row, 0).text().strip() if self.manual_table.item(row, 0) else ""
            purchase = self.manual_table.item(row, 1).text().strip() if self.manual_table.item(row, 1) else ""
            profit = self.manual_table.item(row, 2).text().strip() if self.manual_table.item(row, 2) else ""
            sale = self.manual_table.item(row, 3).text().strip() if self.manual_table.item(row, 3) else ""
            company = self.manual_table.item(row, 4).text().strip() if self.manual_table.item(row, 4) else ""
            stock = self.manual_table.item(row, 5).text().strip() if self.manual_table.item(row, 5) else ""

            # Validate required fields
            if not item_name or not purchase or not sale or not stock:
                continue

            # Generate barcode (12 digits)
            barcode = "".join(random.choices("0123456789", k=12))

            valid_rows.append({
                "item_name": item_name,
                "purchase": purchase,
                "profit": profit,
                "sale": sale,
                "company": company,
                "stock": stock,
                "barcode": barcode
            })

        if not valid_rows:
            show_warning("No valid items to process.", self)
            return

        # Populate preview table
        self.preview_table.setRowCount(len(valid_rows))
        for row, data in enumerate(valid_rows):
            self.preview_table.setItem(row, 0, QTableWidgetItem(data["item_name"]))

            item = QTableWidgetItem(data["purchase"])
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.preview_table.setItem(row, 1, item)

            item = QTableWidgetItem(data["profit"])
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.preview_table.setItem(row, 2, item)

            item = QTableWidgetItem(data["sale"])
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.preview_table.setItem(row, 3, item)

            self.preview_table.setItem(row, 4, QTableWidgetItem(data["company"]))

            item = QTableWidgetItem(data["stock"])
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.preview_table.setItem(row, 5, item)

            self.preview_table.setItem(row, 6, QTableWidgetItem(data["barcode"]))

        self.accept()

    def get_items(self) -> List[Dict]:
        """Return the list of items to process."""
        items = []
        for row in range(self.preview_table.rowCount()):
            items.append({
                "item_name": self.preview_table.item(row, 0).text(),
                "purchase_price": self.preview_table.item(row, 1).text(),
                "sale_price": self.preview_table.item(row, 3).text(),
                "company": self.preview_table.item(row, 4).text(),
                "opening_stock": self.preview_table.item(row, 5).text(),
                "barcode": self.preview_table.item(row, 6).text()
            })
        return items


# ------------------------------
# User Management Dialog
# ------------------------------
class UserManagementDialog(QDialog):
    def __init__(self, db: DatabaseManager, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Manage Users")
        self.setFixedSize(500, 400)

        self.init_ui()
        self.refresh_users()

    def init_ui(self):
        layout = QVBoxLayout()

        # User table
        self.user_table = QTableWidget()
        self.user_table.setColumnCount(3)
        self.user_table.setHorizontalHeaderLabels(["Username", "Role", "Actions"])
        self.user_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.user_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.user_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.user_table.verticalHeader().setVisible(False)

        # Add user form
        form_group = QGroupBox("Add New User")
        form_layout = QFormLayout()

        self.new_username = QLineEdit()
        self.new_password = QLineEdit()
        self.new_password.setEchoMode(QLineEdit.Password)
        self.new_role = QComboBox()
        self.new_role.addItems(["user", "admin"])

        form_layout.addRow("Username:", self.new_username)
        form_layout.addRow("Password:", self.new_password)
        form_layout.addRow("Role:", self.new_role)

        add_btn = QPushButton("Add User")
        add_btn.clicked.connect(self.add_user)

        form_layout.addRow(add_btn)
        form_group.setLayout(form_layout)

        layout.addWidget(self.user_table)
        layout.addWidget(form_group)

        self.setLayout(layout)

    def refresh_users(self):
        users = self.db.fetch_users()
        self.user_table.setRowCount(len(users))

        for row, (username, role) in enumerate(users):
            # Username
            self.user_table.setItem(row, 0, QTableWidgetItem(username))

            # Role
            self.user_table.setItem(row, 1, QTableWidgetItem(role))

            # Delete button (except for admin)
            if username == "admin":
                self.user_table.setItem(row, 2, QTableWidgetItem("System"))
            else:
                delete_btn = QPushButton("Delete")
                delete_btn.clicked.connect(lambda _, u=username: self.delete_user(u))
                self.user_table.setCellWidget(row, 2, delete_btn)

    def add_user(self):
        username = self.new_username.text().strip()
        password = self.new_password.text().strip()
        role = self.new_role.currentText()

        if not username or not password:
            show_warning("Please enter both username and password.", self)
            return

        try:
            self.db.add_user(username, password, role)
            show_info(f"User '{username}' added successfully.", self)
            self.refresh_users()
            self.new_username.clear()
            self.new_password.clear()
        except Exception as e:
            show_error(f"Failed to add user: {e}", self)

    def delete_user(self, username: str):
        if username == "admin":
            show_warning("Cannot delete the default admin user.", self)
            return

        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete user '{username}'?",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                self.db.remove_user(username)
                show_info(f"User '{username}' deleted.", self)
                self.refresh_users()
            except Exception as e:
                show_error(f"Failed to delete user: {e}", self)


# ------------------------------
# Login Dialog (UI)
# ------------------------------
class LoginDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(TITLE)
        self.setFixedSize(500, 300)

        # Try to set window icon if available
        if os.path.exists(LOGO_PATH):
            self.setWindowIcon(QIcon(LOGO_PATH))

        self.init_ui()
        self.user = None
        self.role = None

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Title
        title_label = QLabel(TITLE)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("""
            font-size: 22px; 
            font-weight: 600; 
            color: #2c3e50;
            margin-bottom: 20px;
        """)

        # Logo (if available)
        if os.path.exists(LOGO_PATH):
            logo_label = QLabel()
            logo_pixmap = QPixmap(LOGO_PATH).scaled(80, 80, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            logo_label.setPixmap(logo_pixmap)
            logo_label.setAlignment(Qt.AlignCenter)
            layout.addWidget(logo_label)

        layout.addWidget(title_label)

        # Form
        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        self.username_field = QLineEdit()
        self.username_field.setPlaceholderText("Enter username")

        self.password_field = QLineEdit()
        self.password_field.setPlaceholderText("Enter password")
        self.password_field.setEchoMode(QLineEdit.Password)

        form_layout.addRow("Username:", self.username_field)
        form_layout.addRow("Password:", self.password_field)

        # Login button
        login_btn = QPushButton("Login")
        login_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 10px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        login_btn.clicked.connect(self.check_credentials)

        layout.addLayout(form_layout)
        layout.addWidget(login_btn)

        self.setLayout(layout)

    def check_credentials(self):
        username = self.username_field.text().strip()
        password = self.password_field.text().strip()

        if not username or not password:
            show_warning("Please enter both username and password.", self)
            return

        try:
            db = DatabaseManager()
            result = db.validate_user(username, password)

            if result:
                self.user, self.role = result
                logger.info(f"User {username} logged in successfully.")
                self.accept()
            else:
                show_warning("Incorrect username or password.", self)
                self.password_field.clear()
        except Exception as e:
            show_error("An error occurred during login.", self)
            logger.error(f"Login error: {e}")


# ------------------------------
# Printing via Print Dialog
# ------------------------------
def print_image_via_dialog(parent, image_path: str):
    try:
        printer = QPrinter(QPrinter.HighResolution)
        printer.setFullPage(True)

        dialog = QPrintDialog(printer, parent)
        if dialog.exec_() == QDialog.Accepted:
            painter = QPainter(printer)
            img = QImage(image_path)

            if img.isNull():
                raise IOError("Failed to open image for printing.")

            # Calculate aspect ratio preserving dimensions
            rect = painter.viewport()
            size = img.size()
            size.scale(rect.size(), Qt.KeepAspectRatio)

            # Center the image
            x = (rect.width() - size.width()) / 2
            y = (rect.height() - size.height()) / 2

            painter.setViewport(x, y, size.width(), size.height())
            painter.setWindow(img.rect())
            painter.drawImage(0, 0, img)
            painter.end()

            show_info("Label sent to printer successfully.", parent)
        else:
            show_info("Print job canceled.", parent)
    except Exception as ex:
        show_error(f"Printing error: {ex}", parent)
        logger.error(f"Printing failed: {ex}")


# ------------------------------
# Main Application UI
# ------------------------------
class BarcodeGeneratorUI(QWidget):
    def __init__(self, current_user: str, current_role: str):
        super().__init__()
        self.current_user = current_user
        self.current_role = current_role
        self.current_template = None
        self.last_barcode_png = None
        self.output_folders = setup_output_folders(APP_DATA_PATH)

        try:
            self.db = DatabaseManager()
            # Load default template
            self.current_template = self.db.get_template("Default")
        except Exception as e:
            show_error("Failed to initialize the database.", self)
            sys.exit(1)

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle(TITLE)
        self.setGeometry(100, 100, 1200, 800)

        # Try to set window icon if available
        if os.path.exists(LOGO_PATH):
            self.setWindowIcon(QIcon(LOGO_PATH))

        # Main layout
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)

        # Header
        header_layout = QHBoxLayout()

        title_label = QLabel(TITLE)
        title_label.setStyleSheet("""
            font-size: 18px; 
            font-weight: bold; 
            color: #2c3e50;
        """)

        user_label = QLabel(f"User: {self.current_user} ({self.current_role})")
        user_label.setStyleSheet("font-size: 12px; color: #7f8c8d;")

        header_layout.addWidget(title_label)
        header_layout.addStretch()
        header_layout.addWidget(user_label)

        main_layout.addLayout(header_layout)

        # Splitter for main content
        splitter = QSplitter(Qt.Vertical)

        # Top part - Form and preview
        top_widget = QWidget()
        top_layout = QHBoxLayout()
        top_layout.setContentsMargins(0, 0, 0, 0)

        # Form area
        form_widget = QWidget()
        form_widget.setMaximumWidth(500)
        form_layout = QVBoxLayout()

        # Label size and template
        size_group = QGroupBox("Label Settings")
        size_layout = QFormLayout()

        self.size_combo = QComboBox()
        self.size_combo.addItems(LABEL_SIZES.keys())
        size_layout.addRow("Label Size:", self.size_combo)

        self.custom_width = QLineEdit()
        self.custom_width.setPlaceholderText("Width (mm)")
        self.custom_width.setVisible(False)

        self.custom_height = QLineEdit()
        self.custom_height.setPlaceholderText("Height (mm)")
        self.custom_height.setVisible(False)

        size_layout.addRow("Custom Width:", self.custom_width)
        size_layout.addRow("Custom Height:", self.custom_height)

        self.template_combo = QComboBox()
        self.load_templates()
        size_layout.addRow("Template:", self.template_combo)

        if self.current_role == "admin":
            self.template_btn = QPushButton("Manage Templates")
            self.template_btn.clicked.connect(self.manage_templates)
            size_layout.addRow(self.template_btn)

        self.dpi_combo = QComboBox()
        self.dpi_combo.addItems(map(str, DPI_OPTIONS))
        self.dpi_combo.setCurrentText(str(DEFAULT_DPI))
        size_layout.addRow("DPI:", self.dpi_combo)

        self.continuous_roll = QCheckBox("Continuous Roll Mode")
        size_layout.addRow(self.continuous_roll)

        size_group.setLayout(size_layout)

        # Product info
        product_group = QGroupBox("Product Information")
        product_layout = QFormLayout()

        self.item_name = QLineEdit()
        product_layout.addRow("Item Name:", self.item_name)

        self.company = QLineEdit()
        product_layout.addRow("Company:", self.company)

        self.barcode = QLineEdit()
        self.barcode.setPlaceholderText("Leave empty to generate random barcode")
        product_layout.addRow("Barcode:", self.barcode)

        product_group.setLayout(product_layout)

        # Pricing and stock
        price_group = QGroupBox("Pricing & Stock")
        price_layout = QFormLayout()

        self.purchase_price = QLineEdit()
        self.purchase_price.setPlaceholderText("0.00")
        price_layout.addRow("Purchase Price:", self.purchase_price)

        self.profit_percent = QLineEdit()
        self.profit_percent.setPlaceholderText("20")
        price_layout.addRow("Profit %:", self.profit_percent)

        self.sale_price = QLineEdit()
        self.sale_price.setPlaceholderText("Calculated automatically")
        self.sale_price.setReadOnly(True)
        price_layout.addRow("Sale Price:", self.sale_price)

        self.stock = QLineEdit()
        self.stock.setPlaceholderText("0")
        price_layout.addRow("Opening Stock:", self.stock)

        self.label_count = QSpinBox()
        self.label_count.setMinimum(1)
        self.label_count.setMaximum(1000)
        self.label_count.setValue(1)
        price_layout.addRow("Number of Labels:", self.label_count)

        price_group.setLayout(price_layout)

        # Buttons
        button_layout = QHBoxLayout()

        self.calc_btn = QPushButton("Calculate Price")
        self.calc_btn.clicked.connect(self.calculate_price)

        self.generate_btn = QPushButton("Generate Labels")
        self.generate_btn.clicked.connect(self.generate_labels)
        self.generate_btn.setStyleSheet("background-color: #27ae60; color: white;")

        self.bulk_btn = QPushButton("Bulk Entry")
        self.bulk_btn.clicked.connect(self.bulk_entry)

        button_layout.addWidget(self.calc_btn)
        button_layout.addWidget(self.generate_btn)
        button_layout.addWidget(self.bulk_btn)

        # Add all to form layout
        form_layout.addWidget(size_group)
        form_layout.addWidget(product_group)
        form_layout.addWidget(price_group)
        form_layout.addLayout(button_layout)
        form_layout.addStretch()

        form_widget.setLayout(form_layout)

        # Preview area
        preview_group = QGroupBox("Label Preview")
        preview_layout = QVBoxLayout()

        self.preview = LabelPreview()
        preview_layout.addWidget(self.preview)

        preview_btn_layout = QHBoxLayout()
        self.print_btn = QPushButton("Print Label")
        self.print_btn.clicked.connect(self.print_label)
        self.print_btn.setEnabled(False)

        preview_btn_layout.addWidget(self.print_btn)
        preview_btn_layout.addStretch()

        preview_layout.addLayout(preview_btn_layout)
        preview_group.setLayout(preview_layout)

        top_layout.addWidget(form_widget)
        top_layout.addWidget(preview_group)
        top_widget.setLayout(top_layout)

        # Bottom part - History
        history_group = QGroupBox("Barcode History")
        history_layout = QVBoxLayout()

        search_layout = QHBoxLayout()
        self.search_field = QLineEdit()
        self.search_field.setPlaceholderText("Search by item name or barcode...")
        self.search_field.textChanged.connect(self.refresh_history)
        search_layout.addWidget(self.search_field)

        self.clear_history_btn = QPushButton("Clear History")
        self.clear_history_btn.clicked.connect(self.clear_history)
        if self.current_role != "admin":
            self.clear_history_btn.setEnabled(False)
        search_layout.addWidget(self.clear_history_btn)

        history_layout.addLayout(search_layout)

        self.history_table = QTableWidget()
        self.history_table.setColumnCount(9)
        self.history_table.setHorizontalHeaderLabels([
            "ID", "User", "Item", "Barcode", "Sale", "Purchase",
            "Company", "Stock", "Date"
        ])
        self.history_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.history_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.history_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.history_table.doubleClicked.connect(self.reprint_from_history)

        history_layout.addWidget(self.history_table)
        history_group.setLayout(history_layout)

        # Add to splitter
        splitter.addWidget(top_widget)
        splitter.addWidget(history_group)
        splitter.setSizes([500, 300])

        main_layout.addWidget(splitter)

        # Status bar
        self.status_bar = QStatusBar()
        self.status_bar.showMessage("Ready")
        main_layout.addWidget(self.status_bar)

        # Admin menu
        if self.current_role == "admin":
            admin_menu = QMenu()

            user_action = QAction("Manage Users", self)
            user_action.triggered.connect(self.manage_users)
            admin_menu.addAction(user_action)

            folder_action = QAction("Set Output Folder", self)
            folder_action.triggered.connect(self.set_output_folder)
            admin_menu.addAction(folder_action)

            admin_btn = QToolButton()
            admin_btn.setText("Admin")
            admin_btn.setPopupMode(QToolButton.InstantPopup)
            admin_btn.setMenu(admin_menu)

            header_layout.addWidget(admin_btn)

        self.setLayout(main_layout)

        # Connect signals
        self.size_combo.currentTextChanged.connect(self.toggle_custom_size)
        self.template_combo.currentTextChanged.connect(self.template_changed)
        self.purchase_price.textChanged.connect(self.calculate_price)
        self.profit_percent.textChanged.connect(self.calculate_price)

        # Initial refresh
        self.refresh_history()

    def load_templates(self):
        """Load available templates into the combo box."""
        self.template_combo.clear()
        templates = self.db.get_all_templates()
        for name, created_by, _ in templates:
            self.template_combo.addItem(f"{name} (by {created_by})", name)

        if self.current_template:
            index = self.template_combo.findData(self.current_template["name"])
            if index >= 0:
                self.template_combo.setCurrentIndex(index)

    def toggle_custom_size(self):
        """Show/hide custom size fields based on selection."""
        is_custom = self.size_combo.currentText() == "Custom Size"
        self.custom_width.setVisible(is_custom)
        self.custom_height.setVisible(is_custom)

    def template_changed(self, text):
        """Handle template selection change."""
        if not text:
            return

        template_name = self.template_combo.currentData()
        self.current_template = self.db.get_template(template_name)

        if self.current_template:
            # Update preview with new template
            self.update_preview()

    def calculate_price(self):
        """Calculate sale price based on purchase price and profit percentage."""
        purchase_text = self.purchase_price.text().replace(",", ".").strip()
        profit_text = self.profit_percent.text().replace(",", ".").strip()

        if not purchase_text or not profit_text:
            return

        try:
            purchase = float(purchase_text)
            profit = float(profit_text)
            sale = purchase * (1 + profit / 100)
            self.sale_price.setText(f"{sale:.2f}")
        except ValueError:
            self.sale_price.clear()

    def update_preview(self):
        """Update the label preview with current settings."""
        if not self.current_template:
            return

        # Create a temporary QPixmap to preview
        width, height = 300, 150  # Default preview size

        # Adjust for selected label size if possible
        label_size = self.size_combo.currentText()
        if label_size in LABEL_SIZES and LABEL_SIZES[label_size]:
            w, h = LABEL_SIZES[label_size]
            # Scale to fit preview while maintaining aspect ratio
            if w > h:
                width = 300
                height = int(300 * h / w)
            else:
                height = 150
                width = int(150 * w / h)

        pixmap = QPixmap(width, height)
        pixmap.fill(Qt.white)

        painter = QPainter(pixmap)
        try:
            # Set font
            font = QFont(self.current_template["font_family"], self.current_template["font_size"])
            painter.setFont(font)

            # Draw sample text
            item_name = self.item_name.text() or "Sample Item"
            painter.drawText(10, 20, item_name)

            # Draw sample barcode (just a rectangle for preview)
            barcode_height = 40 * self.current_template["barcode_scale"]
            painter.drawRect(width // 2 - 50, 40, 100, barcode_height)

            # Draw sample info
            info_lines = []
            if self.current_template["show_price"]:
                sale_price = self.sale_price.text() or "10.00"
                info_lines.append(f"Sale: {sale_price}")
            if self.current_template["show_company"]:
                company = self.company.text() or "Sample Co."
                info_lines.append(company)
            if self.current_template["show_stock"]:
                stock = self.stock.text() or "50"
                info_lines.append(f"Stock: {stock}")

            info_text = " | ".join(info_lines)
            if info_text:
                painter.drawText(10, height - 20, info_text)

            # Draw footer
            small_font = QFont("Arial", 6)
            painter.setFont(small_font)
            painter.drawText(10, height - 5, "Projects by Shenz C - AutoGeek Edition")

        finally:
            painter.end()

        self.preview.update_preview(pixmap)

    def validate_inputs(self) -> bool:
        """Validate all input fields before generation."""
        if not self.item_name.text().strip():
            show_warning("Please enter an item name.", self)
            return False

        if not self.purchase_price.text().strip():
            show_warning("Please enter a purchase price.", self)
            return False

        if not self.sale_price.text().strip():
            show_warning("Please calculate a sale price.", self)
            return False

        if not self.company.text().strip():
            show_warning("Please enter a company name.", self)
            return False

        if not self.stock.text().strip():
            show_warning("Please enter an opening stock value.", self)
            return False

        # Validate label size
        label_size = self.size_combo.currentText()
        if label_size == "Custom Size":
            try:
                width = float(self.custom_width.text())
                height = float(self.custom_height.text())
                if width <= 0 or height <= 0:
                    show_warning("Custom dimensions must be positive numbers.", self)
                    return False
            except ValueError:
                show_warning("Please enter valid custom dimensions.", self)
                return False

        return True

    def get_label_size(self) -> Tuple[float, float]:
        """Get the current label size in mm (width, height)."""
        label_size = self.size_combo.currentText()

        if label_size == "Custom Size":
            try:
                return float(self.custom_width.text()), float(self.custom_height.text())
            except ValueError:
                show_warning("Invalid custom dimensions. Using default.", self)
                return 60, 25  # Default to small label

        return LABEL_SIZES.get(label_size, (60, 25))

    def generate_labels(self):
        """Generate barcode labels based on current settings."""
        if not self.validate_inputs():
            return

        if not self.current_template:
            show_warning("No template selected.", self)
            return

        # Get all values
        item_name = self.item_name.text().strip()
        purchase_price = self.purchase_price.text().replace(",", ".").strip()
        sale_price = self.sale_price.text().replace(",", ".").strip()
        company = self.company.text().strip()
        stock = self.stock.text().strip()
        label_count = self.label_count.value()
        barcode = self.barcode.text().strip() or "".join(random.choices("0123456789", k=12))
        width_mm, height_mm = self.get_label_size()
        dpi = int(self.dpi_combo.currentText())
        template_name = self.current_template["name"]

        try:
            # Generate barcode image
            barcode_path = generate_barcode_image(
                barcode,
                scale=self.current_template["barcode_scale"],
                dpi=dpi,
                font_path=None  # Use default font for now
            )
            self.last_barcode_png = barcode_path
            self.print_btn.setEnabled(True)

            # Generate PDF
            pdf_filename = f"{item_name} - {barcode}.pdf"
            pdf_path = os.path.join(self.output_folders["pdf"], pdf_filename)

            details = {
                "item": item_name,
                "sale": sale_price,
                "purchase": purchase_price,
                "company": company,
                "opening_stock": stock,
                "count": label_count
            }

            if generate_label_pdf(pdf_path, width_mm, height_mm, barcode_path, details, self.current_template, dpi):
                show_info(f"Label PDF generated:\n{pdf_path}", self)
                self.status_bar.showMessage("Label generated successfully", 5000)

            # Update database
            gen_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.db.add_barcode_history(
                self.current_user, item_name, barcode, float(sale_price),
                float(purchase_price), company, gen_date, int(stock), template_name
            )
            self.db.add_or_update_barcode(
                item_name, barcode, float(sale_price), float(purchase_price),
                company, label_count, int(stock), template_name
            )

            # Update Excel
            update_excel(
                barcode, item_name, sale_price, purchase_price,
                company, gen_date, stock
            )

            # Refresh history
            self.refresh_history()

            # Update preview with actual barcode
            self.update_preview_with_barcode(barcode_path)

        except Exception as e:
            show_error(f"Failed to generate labels: {e}", self)
            logger.error(f"Label generation error: {e}")

    def update_preview_with_barcode(self, barcode_path: str):
        """Update preview with the actual generated barcode."""
        if not barcode_path or not os.path.exists(barcode_path):
            return

        pixmap = QPixmap(barcode_path)
        if pixmap.isNull():
            return

        # Scale to fit preview while maintaining aspect ratio
        self.preview.update_preview(pixmap)

    def print_label(self):
        """Print the last generated label."""
        if not self.last_barcode_png or not os.path.exists(self.last_barcode_png):
            show_warning("No label generated yet.", self)
            return

        print_image_via_dialog(self, self.last_barcode_png)

    def bulk_entry(self):
        """Open bulk entry dialog and process items."""
        dialog = BulkEntryDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            items = dialog.get_items()
            if not items:
                return

            # Get current settings
            width_mm, height_mm = self.get_label_size()
            dpi = int(self.dpi_combo.currentText())
            template_name = self.current_template["name"] if self.current_template else "Default"

            # Process each item
            for item in items:
                try:
                    # Generate barcode image
                    barcode_path = generate_barcode_image(
                        item["barcode"],
                        scale=self.current_template["barcode_scale"],
                        dpi=dpi,
                        font_path=None
                    )

                    # Generate PDF
                    pdf_filename = f"{item['item_name']} - {item['barcode']}.pdf"
                    pdf_path = os.path.join(self.output_folders["pdf"], pdf_filename)

                    details = {
                        "item": item["item_name"],
                        "sale": item["sale_price"],
                        "purchase": item["purchase_price"],
                        "company": item["company"],
                        "opening_stock": item["opening_stock"],
                        "count": 1  # One label per item in bulk mode
                    }

                    generate_label_pdf(pdf_path, width_mm, height_mm, barcode_path, details, self.current_template, dpi)

                    # Update database
                    gen_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    self.db.add_barcode_history(
                        self.current_user, item["item_name"], item["barcode"],
                        float(item["sale_price"]), float(item["purchase_price"]),
                        item["company"], gen_date, int(item["opening_stock"]), template_name
                    )
                    self.db.add_or_update_barcode(
                        item["item_name"], item["barcode"], float(item["sale_price"]),
                        float(item["purchase_price"]), item["company"],
                        1, int(item["opening_stock"]), template_name
                    )

                    # Update Excel
                    update_excel(
                        item["barcode"], item["item_name"], item["sale_price"],
                        item["purchase_price"], item["company"], gen_date,
                        item["opening_stock"]
                    )

                except Exception as e:
                    logger.error(f"Failed to process bulk item {item['item_name']}: {e}")
                    continue

            show_info(f"Processed {len(items)} items in bulk.", self)
            self.refresh_history()

    def refresh_history(self):
        """Refresh the history table with current filter."""
        search_text = self.search_field.text().strip()
        history = self.db.fetch_history(search_text)

        self.history_table.setRowCount(len(history))
        for row, record in enumerate(history):
            for col, value in enumerate(record):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignCenter)
                self.history_table.setItem(row, col, item)

        self.history_table.resizeRowsToContents()

    def clear_history(self):
        """Clear the entire barcode history."""
        if self.current_role != "admin":
            show_warning("Only administrators can clear history.", self)
            return

        reply = QMessageBox.question(
            self, "Clear History",
            "Are you sure you want to clear the entire barcode generation history?",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                self.db.clear_history()
                self.refresh_history()
                show_info("Barcode history cleared.", self)
                self.status_bar.showMessage("History cleared", 5000)
            except Exception as e:
                show_error(f"Failed to clear history: {e}", self)
                logger.error(f"Clear history error: {e}")

    def reprint_from_history(self, index):
        """Reprint a label from history selection."""
        barcode = self.history_table.item(index.row(), 3).text()
        if not barcode:
            return

        try:
            # Regenerate barcode image
            barcode_path = generate_barcode_image(
                barcode,
                scale=self.current_template["barcode_scale"],
                dpi=int(self.dpi_combo.currentText())
            )

            # Print it
            print_image_via_dialog(self, barcode_path)
        except Exception as e:
            show_error(f"Failed to reprint label: {e}", self)
            logger.error(f"Reprint error: {e}")

    def manage_templates(self):
        """Open template management dialog."""
        if self.current_role != "admin":
            show_warning("Only administrators can manage templates.", self)
            return

        dialog = TemplateDialog(self.db, self.current_user, self)
        if dialog.exec_() == QDialog.Accepted:
            # Update current template if one was selected
            selected_template = dialog.get_selected_template()
            if selected_template:
                self.current_template = selected_template
                self.load_templates()
                self.update_preview()

    def manage_users(self):
        """Open user management dialog."""
        if self.current_role != "admin":
            show_warning("Only administrators can manage users.", self)
            return

        dialog = UserManagementDialog(self.db, self)
        dialog.exec_()

    def set_output_folder(self):
        """Set custom output folder (admin only)."""
        if self.current_role != "admin":
            show_warning("Only administrators can change output folders.", self)
            return

        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if folder:
            global APP_DATA_PATH
            APP_DATA_PATH = folder
            self.output_folders = setup_output_folders(APP_DATA_PATH)
            show_info(f"Output folder set to:\n{APP_DATA_PATH}", self)


# ------------------------------
# Application Launcher
# ------------------------------
if __name__ == "__main__":
    multiprocessing.freeze_support()  # For PyInstaller support on Windows.

    if is_already_running():
        print("Another instance is already running. Exiting.")
        sys.exit(0)

    if platform.system() != "Windows":
        print("This application is designed to run on Windows.")
        sys.exit(1)

    # Enable High DPI scaling
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    app = QApplication(sys.argv)
    app.setStyle('Fusion')  # Modern style

    # Set application font
    font = QFont(DEFAULT_FONT, 10)
    app.setFont(font)

    # Show login dialog
    login = LoginDialog()
    if login.exec_() == QDialog.Accepted:
        window = BarcodeGeneratorUI(login.user, login.role)
        window.show()
        sys.exit(app.exec_())
    else:
        sys.exit(0)