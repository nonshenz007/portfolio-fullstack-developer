import sys
import ctypes
import multiprocessing
import subprocess
import platform
import logging
import os
import random
import datetime
import sqlite3

from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QFormLayout, QDialog,
    QPushButton, QMessageBox, QLineEdit, QFileDialog, QLabel,
    QComboBox, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView,
    QCheckBox, QStatusBar
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QImage, QPainter
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog

# ------------------------------
# Single Instance Check (Windows Only)
# ------------------------------
def is_already_running():
    if platform.system() != "Windows":
        return False
    mutex = ctypes.windll.kernel32.CreateMutexW(None, 1, "Global\\ProjectsByShenzC_FullSafeApp")
    return ctypes.GetLastError() == 183

# ------------------------------
# App Data Directory and Global Paths
# ------------------------------
def get_app_data_path():
    base_path = os.path.join(os.environ["APPDATA"], "BarcodeGenerator")
    os.makedirs(base_path, exist_ok=True)
    return base_path

DB_PATH = os.path.join(get_app_data_path(), "barcodes.db")
LOG_PATH = os.path.join(get_app_data_path(), "barcode_app.log")
output_dir = os.path.join(get_app_data_path(), "barcodes")
os.makedirs(output_dir, exist_ok=True)

# ------------------------------
# Logging Setup
# ------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(LOG_PATH), logging.StreamHandler()]
)
logger = logging.getLogger(__name__)

TITLE = "Projects \U0001F4C8 by Shenz C (v1 beta)"

LABEL_SIZES = {
    "Small (60x25 mm)": (60, 25),
    "Medium (100x25 mm)": (100, 25),
    "Large (150x50 mm)": (150, 50),
    "A4 Sheet (210x297 mm)": (210, 297),
    "2 Labels (50 * 25mm)": (100, 25),
    "Custom Size": None
}

# ------------------------------
# Database Manager
# ------------------------------
class DatabaseManager:
    def __init__(self, db_path=DB_PATH):
        self.db_path = db_path
        try:
            self.conn = sqlite3.connect(self.db_path)
            self.create_tables()
            logger.info("Database connected and tables created/verified.")
        except Exception as e:
            logger.error(f"Database connection failed: {e}")
            raise

    def create_tables(self):
        try:
            c = self.conn.cursor()
            # Barcode history with stock_count
            c.execute('''
                CREATE TABLE IF NOT EXISTS barcode_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    item_name TEXT,
                    barcode TEXT,
                    sale_price REAL,
                    purchase_price REAL,
                    company_name TEXT,
                    generation_date TEXT,
                    username TEXT DEFAULT 'admin',
                    stock_count INTEGER DEFAULT 0
                )
            ''')
            # Barcodes table with stock
            c.execute('''
                CREATE TABLE IF NOT EXISTS barcodes (
                    barcode TEXT PRIMARY KEY,
                    item_name TEXT,
                    sale_price REAL,
                    purchase_price REAL,
                    company_name TEXT,
                    quantity INTEGER,
                    stock INTEGER DEFAULT 0
                )
            ''')
            # Users table
            c.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    username TEXT PRIMARY KEY,
                    password TEXT,
                    role TEXT
                )
            ''')
            # Insert default admin if none exist.
            c.execute('SELECT COUNT(*) FROM users')
            if c.fetchone()[0] == 0:
                c.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)',
                          ("admin", "Kungfukenny", "admin"))
            self.conn.commit()
        except Exception as e:
            logger.error(f"Error creating tables: {e}")
            raise

    def add_barcode_history(self, username, item_name, barcode, sale_price, purchase_price, company_name,
                            generation_date, stock_count):
        try:
            c = self.conn.cursor()
            c.execute('''
                INSERT INTO barcode_history
                (username, item_name, barcode, sale_price, purchase_price, company_name, generation_date, stock_count)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (username, item_name, barcode, sale_price, purchase_price, company_name, generation_date, stock_count))
            self.conn.commit()
            logger.info("Barcode history added.")
        except Exception as e:
            logger.error(f"Failed to add barcode history: {e}")
            raise

    def add_or_update_barcode(self, item_name, barcode, sale_price, purchase_price, company_name, quantity, stock):
        try:
            c = self.conn.cursor()
            c.execute('SELECT quantity FROM barcodes WHERE barcode=?', (barcode,))
            result = c.fetchone()
            if result:
                new_quantity = result[0] + quantity
                c.execute('''
                    UPDATE barcodes
                    SET item_name=?, sale_price=?, purchase_price=?, company_name=?, quantity=?, stock=?
                    WHERE barcode=?
                ''', (item_name, sale_price, purchase_price, company_name, new_quantity, stock, barcode))
            else:
                c.execute('''
                    INSERT INTO barcodes
                    (barcode, item_name, sale_price, purchase_price, company_name, quantity, stock)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (barcode, item_name, sale_price, purchase_price, company_name, quantity, stock))
            self.conn.commit()
            logger.info("Barcode updated/added in barcodes table.")
        except Exception as e:
            logger.error(f"Failed to add or update barcode: {e}")
            raise

    def fetch_history(self, search_filter=""):
        try:
            c = self.conn.cursor()
            query = """
                SELECT id, username, item_name, barcode, sale_price, purchase_price, company_name, stock_count, generation_date
                FROM barcode_history
            """
            params = ()
            if search_filter:
                query += " WHERE item_name LIKE ? OR barcode LIKE ?"
                params = (f"%{search_filter}%", f"%{search_filter}%")
            query += " ORDER BY id DESC"
            c.execute(query, params)
            history = c.fetchall()
            logger.info("History fetched from database.")
            return history
        except Exception as e:
            logger.error(f"Failed to fetch history: {e}")
            return []

    def clear_history(self):
        try:
            c = self.conn.cursor()
            c.execute('DELETE FROM barcode_history')
            self.conn.commit()
            logger.info("Barcode history cleared.")
        except Exception as e:
            logger.error(f"Failed to clear history: {e}")
            raise

    def add_user(self, username, password, role="user"):
        try:
            c = self.conn.cursor()
            c.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)', (username, password, role))
            self.conn.commit()
            logger.info(f"User {username} added.")
        except Exception as e:
            logger.error(f"Failed to add user: {e}")
            raise

    def remove_user(self, username):
        try:
            c = self.conn.cursor()
            c.execute('DELETE FROM users WHERE username=?', (username,))
            self.conn.commit()
            logger.info(f"User {username} removed.")
        except Exception as e:
            logger.error(f"Failed to remove user: {e}")
            raise

    def fetch_users(self):
        try:
            c = self.conn.cursor()
            c.execute('SELECT username, role FROM users ORDER BY username ASC')
            users = c.fetchall()
            logger.info("Fetched users from database.")
            return users
        except Exception as e:
            logger.error(f"Failed to fetch users: {e}")
            return []

    def validate_user(self, username, password):
        try:
            c = self.conn.cursor()
            c.execute('SELECT username, role FROM users WHERE username=? AND password=?', (username, password))
            result = c.fetchone()
            logger.info(f"User {username} validation attempted.")
            return result
        except Exception as e:
            logger.error(f"User validation failed: {e}")
            return None

# ------------------------------
# Barcode Generation (Code128)
# ------------------------------
try:
    from barcode import get_barcode_class
    from barcode.writer import ImageWriter
except ImportError as e:
    logger.error("Missing dependency 'python-barcode'. Please install it before running this application.")
    raise e

def sanitize_barcode_input(barcode_str):
    # Code128 supports ASCII values 32 to 126.
    return "".join(ch for ch in barcode_str if 32 <= ord(ch) <= 126)

class BarcodeGeneratorUI(QWidget):
    def __init__(self, current_user, current_role):
        super().__init__()
        self.current_user = current_user
        self.current_role = current_role
        try:
            self.db = DatabaseManager()
        except Exception as e:
            QMessageBox.critical(self, "Error", "Failed to initialize the database.")
            sys.exit(1)
        self.last_barcode_png = None
        self.label_font = QFont("Arial", 8)
        self.initUI()

    def initUI(self):
        self.setWindowTitle(TITLE)
        self.setGeometry(250, 250, 1100, 600)
        self.setStyleSheet("""
            QWidget { background-color: #1c1c1c; color: #e0e0e0; font-family: 'Segoe UI', sans-serif; font-size: 14px; }
            QLineEdit, QComboBox { background-color: #2a2a2a; border: 1px solid #444444; border-radius: 6px; padding: 6px; color: #ffffff; }
            QPushButton { background-color: #333333; border: none; border-radius: 6px; padding: 10px 20px; color: #ffffff; font-weight: bold; }
            QLabel { color: #ffffff; font-size: 14px; }
            QTableWidget { background-color: #2a2a2a; border: 1px solid #444444; border-radius: 6px; }
        """)
        main_layout = QVBoxLayout()
        heading_label = QLabel(TITLE)
        heading_label.setAlignment(Qt.AlignCenter)
        heading_label.setStyleSheet("font-size: 18px; font-weight: 600; margin-bottom: 8px;")
        main_layout.addWidget(heading_label)
        # Row 1: Item Name, Purchase Price, Profit %, Sale Price, Calculator
        row1_layout = QHBoxLayout()
        row1_layout.addWidget(QLabel("Item Name:"))
        self.itemField = QLineEdit()
        row1_layout.addWidget(self.itemField)
        row1_layout.addWidget(QLabel("Purchase Price:"))
        self.purchaseField = QLineEdit()
        row1_layout.addWidget(self.purchaseField)
        row1_layout.addWidget(QLabel("Profit (%):"))
        self.profitField = QLineEdit()
        self.profitField.setPlaceholderText("optional")
        row1_layout.addWidget(self.profitField)
        row1_layout.addWidget(QLabel("Sale Price:"))
        self.saleField = QLineEdit()
        row1_layout.addWidget(self.saleField)
        calcBtn = QPushButton("Calc")
        calcBtn.clicked.connect(self.apply_profit_calculator)
        row1_layout.addWidget(calcBtn)
        main_layout.addLayout(row1_layout)
        # Row 2: Company, No. of Labels, Stock Count, Barcode
        row2_layout = QHBoxLayout()
        row2_layout.addWidget(QLabel("Company:"))
        self.companyField = QLineEdit()
        row2_layout.addWidget(self.companyField)
        row2_layout.addWidget(QLabel("No. of Labels:"))
        self.countField = QLineEdit()
        row2_layout.addWidget(self.countField)
        row2_layout.addWidget(QLabel("Stock Count:"))
        self.stockField = QLineEdit()
        row2_layout.addWidget(self.stockField)
        row2_layout.addWidget(QLabel("Barcode (optional):"))
        self.barcodeField = QLineEdit()
        self.barcodeField.setPlaceholderText("If empty, generate random 12-digit code")
        row2_layout.addWidget(self.barcodeField)
        main_layout.addLayout(row2_layout)
        # Row 3: Label Size, Custom Dimensions, Continuous Roll
        row3_layout = QHBoxLayout()
        row3_layout.addWidget(QLabel("Label Size:"))
        self.sizeBox = QComboBox()
        for label in LABEL_SIZES:
            self.sizeBox.addItem(label)
        row3_layout.addWidget(self.sizeBox)
        row3_layout.addWidget(QLabel("Custom W x H (mm):"))
        self.customWidth = QLineEdit()
        self.customHeight = QLineEdit()
        self.customWidth.setVisible(False)
        self.customHeight.setVisible(False)
        row3_layout.addWidget(self.customWidth)
        row3_layout.addWidget(self.customHeight)
        self.continuousRollCheck = QCheckBox("Continuous Roll")
        row3_layout.addWidget(self.continuousRollCheck)
        main_layout.addLayout(row3_layout)
        # Buttons: Generate PDF, Print (opens print dialog), Reprint, Clear History, Open Excel, Manage Users
        btn_layout = QHBoxLayout()
        genBtn = QPushButton("Generate Labels PDF")
        printNowBtn = QPushButton("Print Label (Select Printer)")
        reprintBtn = QPushButton("Reprint Selected")
        clearHistBtn = QPushButton("Clear History")
        openExcelBtn = QPushButton("Open Items Excel")
        genBtn.clicked.connect(self.generate_pdf)
        # Use a lambda to check for last barcode image and then call our print dialog
        printNowBtn.clicked.connect(lambda: print_image_via_dialog(self, self.last_barcode_png)
                                     if self.last_barcode_png else QMessageBox.warning(self, "Print Error", "Generate a barcode label first."))
        reprintBtn.clicked.connect(self.reprint_selected_barcode)
        clearHistBtn.clicked.connect(self.clear_history)
        openExcelBtn.clicked.connect(self.open_excel_file)
        btn_layout.addWidget(genBtn)
        btn_layout.addWidget(printNowBtn)
        btn_layout.addWidget(reprintBtn)
        btn_layout.addWidget(clearHistBtn)
        btn_layout.addWidget(openExcelBtn)
        if self.current_role == "admin":
            manageUsersBtn = QPushButton("Manage Users")
            manageUsersBtn.clicked.connect(self.manage_users)
        main_layout.addLayout(btn_layout)
        # Search History and History Table
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Search History:"))
        self.searchField = QLineEdit()
        self.searchField.textChanged.connect(self.refresh_history)
        search_layout.addWidget(self.searchField)
        main_layout.addLayout(search_layout)
        self.historyTable = QTableWidget()
        self.historyTable.setColumnCount(9)
        self.historyTable.setHorizontalHeaderLabels(
            ["ID", "User", "Item Name", "Barcode", "Sale Price", "Purchase Price", "Company", "Stock Count", "Date"]
        )
        self.historyTable.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        main_layout.addWidget(QLabel("Barcode Generation History:"))
        main_layout.addWidget(self.historyTable)
        self.statusBar = QStatusBar()
        main_layout.addWidget(self.statusBar)
        self.setLayout(main_layout)
        self.sizeBox.currentTextChanged.connect(self.toggleCustomSize)
        self.refresh_history()

    def toggleCustomSize(self):
        is_custom = (self.sizeBox.currentText() == "Custom Size")
        self.customWidth.setVisible(is_custom)
        self.customHeight.setVisible(is_custom)

    def apply_profit_calculator(self):
        p_txt = self.purchaseField.text().strip().replace(",", ".")
        profit_txt = self.profitField.text().strip().replace(",", ".")
        try:
            purchase_val = float(p_txt)
        except ValueError:
            QMessageBox.warning(self, "Error", "Enter a valid numeric Purchase Price.")
            return
        if not profit_txt:
            QMessageBox.warning(self, "Error", "Enter a valid numeric Profit (%) or type 0.")
            return
        try:
            profit_percent = float(profit_txt)
        except ValueError:
            QMessageBox.warning(self, "Error", "Profit (%) must be numeric.")
            return
        calc_sale = round(purchase_val + (purchase_val * (profit_percent / 100.0)))
        self.saleField.setText(str(calc_sale))

    def get_label_size(self):
        label = self.sizeBox.currentText()
        if label == "Custom Size":
            try:
                w = float(self.customWidth.text().strip().replace(",", "."))
                h = float(self.customHeight.text().strip().replace(",", "."))
                return w, h
            except ValueError:
                QMessageBox.warning(self, "Error", "Enter valid numeric custom dimensions.")
                logger.error("Invalid custom size.")
                return None
        return LABEL_SIZES[label]

    def generate_pdf(self):
        item_name = self.itemField.text().strip()
        purchase_txt = self.purchaseField.text().strip().replace(",", ".")
        sale_txt = self.saleField.text().strip().replace(",", ".")
        company = self.companyField.text().strip()
        count_txt = self.countField.text().strip()
        stock_txt = self.stockField.text().strip()
        barcode_in = self.barcodeField.text().strip()
        if not item_name or not purchase_txt or not sale_txt or not company or not count_txt or not stock_txt:
            QMessageBox.warning(self, "Error", "All fields except Barcode/Profit must be filled out.")
            return
        try:
            purchase_val = float(purchase_txt)
            sale_val = float(sale_txt)
            count_val = int(count_txt)
            stock_val = int(stock_txt)
        except ValueError:
            QMessageBox.warning(self, "Error", "Purchase, Sale, No. of Labels and Stock Count must be numeric.")
            return
        size = self.get_label_size()
        if not size:
            return
        w_mm, h_mm = size
        barcode_str = barcode_in if barcode_in else "".join(random.choices("0123456789", k=12))
        try:
            barcode_png = generate_barcode_image(barcode_str)
            self.last_barcode_png = barcode_png
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Barcode generation failed:\n{e}")
            return
        folder = QFileDialog.getExistingDirectory(self, "Select folder to save PDF")
        if not folder:
            return
        pdf_path = os.path.join(folder, f"labels_{barcode_str}.pdf")
        details = {"item": item_name, "sale": sale_txt, "purchase": purchase_txt, "company": company, "count": count_val}
        try:
            generate_label_pdf(pdf_path, w_mm, h_mm, self.last_barcode_png, details)
            QMessageBox.information(self, "Done", f"PDF saved:\n{pdf_path}")
            self.statusBar.showMessage("PDF generated successfully.", 5000)
        except Exception as ex:
            QMessageBox.warning(self, "Error", f"PDF generation failed:\n{ex}")
            logger.error(f"PDF generation error: {ex}")
            return
        gen_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            self.db.add_barcode_history(self.current_user, item_name, barcode_str,
                                        sale_val, purchase_val, company, gen_date, stock_val)
            self.db.add_or_update_barcode(item_name, barcode_str, sale_val, purchase_val, company, count_val, stock_val)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Database update failed:\n{e}")
            logger.error(f"Database update error: {e}")
        try:
            update_excel(
                item_code=barcode_str,
                item_name=item_name,
                sale_price=sale_txt,
                purchase_price=purchase_txt,
                company_name=company,
                generation_date=gen_date,
                stock_count=stock_val
            )
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Excel update failed:\n{e}")
            logger.error(f"Excel update error: {e}")
        self.refresh_history()

    def reprint_selected_barcode(self):
        selected_rows = self.historyTable.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.information(self, "Reprint", "Please select a barcode record to reprint.")
            return
        index = selected_rows[0].row()
        barcode_value = self.historyTable.item(index, 3).text()
        try:
            barcode_png = generate_barcode_image(barcode_value)
            print_image_via_dialog(self, barcode_png)
        except Exception as e:
            QMessageBox.warning(self, "Reprint Error", f"Failed to reprint barcode: {e}")
            logger.error(f"Reprint error: {e}")

    def refresh_history(self):
        search_text = self.searchField.text().strip()
        history = self.db.fetch_history(search_text)
        self.historyTable.setRowCount(len(history))
        for i, row in enumerate(history):
            for j, value in enumerate(row):
                cell_item = QTableWidgetItem(str(value))
                cell_item.setTextAlignment(Qt.AlignCenter)
                self.historyTable.setItem(i, j, cell_item)
        self.historyTable.resizeRowsToContents()

    def clear_history(self):
        reply = QMessageBox.question(
            self,
            "Clear History",
            "Are you sure you want to clear the entire barcode generation history?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            try:
                self.db.clear_history()
                self.refresh_history()
                QMessageBox.information(self, "Clear History", "Barcode generation history cleared.")
                self.statusBar.showMessage("History cleared.", 5000)
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to clear history: {e}")
                logger.error(f"Clear history error: {e}")

    def open_excel_file(self):
        excel_file = "Items.xlsx"
        if not os.path.exists(excel_file):
            QMessageBox.information(self, "Excel Not Found", "Items.xlsx does not exist yet.")
            return
        try:
            if platform.system() == "Windows":
                os.startfile(excel_file)
            else:
                subprocess.run(["open", excel_file])
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Could not open {excel_file}:\n{e}")
            logger.error(f"Open Excel error: {e}")

    def manage_users(self):
        dialog = UserManagementDialog(self)
        dialog.exec_()

# ------------------------------
# Application Launcher
# ------------------------------
if __name__ == "__main__":
    multiprocessing.freeze_support()  # For PyInstaller support on Windows.
    if is_already_running():
        print("Another instance is already running. Exiting.")
        sys.exit(0)
    if platform.system() != "Windows":
        print("This application is designed to run on Windows.")
        sys.exit(1)
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    app = QApplication(sys.argv)
    login = LoginDialog()
    if login.exec_() == QDialog.Accepted:
        window = BarcodeGeneratorUI(login.user, login.role)
        window.show()
        sys.exit(app.exec_())
    else:
        sys.exit(0)


# --- Barcode Generation Utility ---
def generate_barcode_image(barcode_str, output_path):
    safe_barcode = sanitize_barcode_input(barcode_str)
    if not safe_barcode:
        raise ValueError("No valid Code128 characters after sanitizing.")
    try:
        Code128 = get_barcode_class('code128')
        code128_obj = Code128(safe_barcode, writer=ImageWriter())
        options = {"font_path": r"C:\Windows\Fonts\arial.ttf", "module_height": 15.0, "module_width": 0.3, "font_size": 10}
        code128_obj.writer.set_options(options)
        saved_path = code128_obj.save(output_path[:-4])  # Remove .png if present
        if not os.path.isfile(saved_path):
            raise OSError(f"Barcode image not created at: {saved_path}")
        logger.info("Barcode image generated at %s", saved_path)
        return saved_path
    except Exception as e:
        logger.error(f"Barcode generation failed: {e}")
        raise

# --- PDF Generation Utility ---
def generate_label_pdf(pdf_path, w_mm, h_mm, barcode_path, details, font=None):
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    c = canvas.Canvas(pdf_path, pagesize=(w_mm * mm, h_mm * mm))
    count = details.get("count", 1)
    labels_per_row = 2 if w_mm >= 100 else 1
    label_width = w_mm / labels_per_row
    for i in range(0, count, labels_per_row):
        for j in range(labels_per_row):
            if i + j >= count:
                break
            x_offset = j * label_width
            draw_single_label(c, x_offset, 0, label_width, h_mm, barcode_path, details, font)
        c.showPage()
    c.save()

def draw_single_label(c, x_offset_mm, y_offset_mm, w_label_mm, h_label_mm, barcode_path, details, font=None):
    item = details.get("item", "")
    sale = details.get("sale", "")
    company = details.get("company", "")
    c.rect(x_offset_mm * mm, y_offset_mm * mm, w_label_mm * mm, h_label_mm * mm)
    if font:
        c.setFont(font.family(), font.pointSize())
    else:
        c.setFont("Helvetica", 8)
    c.drawCentredString((x_offset_mm + w_label_mm / 2) * mm, (y_offset_mm + h_label_mm - 5) * mm, item)
    c.drawImage(barcode_path, (x_offset_mm + w_label_mm / 2 - 20) * mm,
                (y_offset_mm + h_label_mm - 16) * mm, 40 * mm, 8 * mm)
    c.drawCentredString((x_offset_mm + w_label_mm / 2) * mm, (y_offset_mm + h_label_mm - 20) * mm, f"Sale: {sale} | {company}")

# --- Excel Export Utility ---
def update_excel(item_code, item_name, sale_price, purchase_price, company_name, generation_date, stock_count, excel_path=None):
    from openpyxl import Workbook, load_workbook
    file_name = excel_path or "Items.xlsx"
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Items"
        ws.append(["Item Code", "Item Name", "Sale Price", "Purchase Price", "Company", "Opening Stock", "Stock Count", "Generation Date"])
        wb.save(file_name)
    wb = load_workbook(file_name)
    ws = wb.active
    found = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == item_code:
            row[1].value = item_name
            row[2].value = sale_price
            row[3].value = purchase_price
            row[4].value = company_name
            row[5].value = stock_count  # Opening Stock
            row[6].value = stock_count
            row[7].value = generation_date
            found = True
            break
    if not found:
        ws.append([item_code, item_name, sale_price, purchase_price, company_name, stock_count, stock_count, generation_date])
    for attempt in range(3):
        try:
            wb.save(file_name)
            logger.info("Excel updated: item_code=%s (found=%s)", item_code, found)
            break
        except Exception as e:
            logger.error(f"Excel save failed: {e}")
            ret = QMessageBox.question(None, "Excel Locked",
                                       f"Unable to save {file_name} (maybe open/locked?).\nClose the file and click 'Retry' or 'Cancel' to skip updating.",
                                       QMessageBox.Retry | QMessageBox.Cancel)
            if ret == QMessageBox.Retry:
                continue
            else:
                break

# ------------------------------
# User Management Dialog
# ------------------------------
class UserManagementDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Manage Users")
        self.setFixedSize(400, 300)
        self.setStyleSheet("""
            QDialog { background-color: #1c1c1c; color: #e0e0e0; font-family: 'Segoe UI', sans-serif; }
            QLabel, QLineEdit, QComboBox, QPushButton { font-size: 14px; }
            QLineEdit, QComboBox { background-color: #2a2a2a; border: 1px solid #444444; padding: 4px; }
            QPushButton { background-color: #333333; border: none; padding: 6px 12px; }
        """)
        layout = QVBoxLayout()
        # Table for listing users
        self.userTable = QTableWidget()
        self.userTable.setColumnCount(2)
        self.userTable.setHorizontalHeaderLabels(["Username", "Role"])
        self.userTable.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.userTable)
        # Input fields for new user
        form_layout = QFormLayout()
        self.newUsername = QLineEdit()
        self.newPassword = QLineEdit()
        self.newPassword.setEchoMode(QLineEdit.Password)
        self.newRole = QComboBox()
        self.newRole.addItems(["user", "admin"])
        form_layout.addRow("Username:", self.newUsername)
        form_layout.addRow("Password:", self.newPassword)
        form_layout.addRow("Role:", self.newRole)
        layout.addLayout(form_layout)
        # Buttons to add and remove user
        btn_layout = QHBoxLayout()
        addBtn = QPushButton("Add User")
        removeBtn = QPushButton("Remove Selected")
        addBtn.clicked.connect(self.add_user)
        removeBtn.clicked.connect(self.remove_user)
        btn_layout.addWidget(addBtn)
        btn_layout.addWidget(removeBtn)
        layout.addLayout(btn_layout)
        self.setLayout(layout)
        self.db = DatabaseManager()
        self.refresh_users()

    def refresh_users(self):
        users = self.db.fetch_users()
        self.userTable.setRowCount(len(users))
        for row, user in enumerate(users):
            username_item = QTableWidgetItem(user[0])
            username_item.setTextAlignment(Qt.AlignCenter)
            role_item = QTableWidgetItem(user[1])
            role_item.setTextAlignment(Qt.AlignCenter)
            self.userTable.setItem(row, 0, username_item)
            self.userTable.setItem(row, 1, role_item)

    def add_user(self):
        username = self.newUsername.text().strip()
        password = self.newPassword.text().strip()
        role = self.newRole.currentText()
        if not username or not password:
            QMessageBox.warning(self, "Add User", "Please provide both username and password.")
            return
        try:
            self.db.add_user(username, password, role)
            QMessageBox.information(self, "Add User", f"User {username} added successfully.")
            self.refresh_users()
        except Exception as e:
            QMessageBox.warning(self, "Add User", f"Failed to add user: {e}")

    def remove_user(self):
        selected_rows = self.userTable.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.information(self, "Remove User", "Please select a user to remove.")
            return
        row = selected_rows[0].row()
        username = self.userTable.item(row, 0).text()
        if username == "admin":
            QMessageBox.warning(self, "Remove User", "Cannot remove the default admin user.")
            return
        try:
            self.db.remove_user(username)
            QMessageBox.information(self, "Remove User", f"User {username} removed successfully.")
            self.refresh_users()
        except Exception as e:
            QMessageBox.warning(self, "Remove User", f"Failed to remove user: {e}")

# ------------------------------
# Login Dialog (UI)
# ------------------------------
class LoginDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(TITLE)
        # Increase size to 500x300 to ensure full title is visible.
        self.setFixedSize(500, 300)
        self.setStyleSheet("""
            QDialog { background-color: #1c1c1c; color: #e0e0e0; font-family: 'Segoe UI', sans-serif; }
            QLabel { font-size: 16px; font-weight: 500; color: #ffffff; }
            QLineEdit { background-color: #2a2a2a; border: 1px solid #444444; border-radius: 6px; padding: 6px; color: #ffffff; }
            QPushButton { background-color: #333333; border: none; border-radius: 6px; padding: 8px 16px; color: #ffffff; font-weight: bold; font-size: 14px; }
        """)
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(10)
        title_label = QLabel(TITLE)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 18px; font-weight: 600; margin-bottom: 8px;")
        layout.addWidget(title_label)
        form_layout = QFormLayout()
        self.usernameField = QLineEdit()
        form_layout.addRow("Username:", self.usernameField)
        self.passwdField = QLineEdit()
        self.passwdField.setEchoMode(QLineEdit.Password)
        form_layout.addRow("Password:", self.passwdField)
        layout.addLayout(form_layout)
        btn = QPushButton("Login")
        btn.clicked.connect(self.check_credentials)
        layout.addWidget(btn, alignment=Qt.AlignCenter)
        self.setLayout(layout)
        self.user = None
        self.role = None

    def check_credentials(self):
        username = self.usernameField.text().strip()
        password = self.passwdField.text().strip()
        if not username or not password:
            QMessageBox.warning(self, "Access Denied", "Enter both username and password")
            return
        try:
            db = DatabaseManager()
            result = db.validate_user(username, password)
            if result:
                self.user, self.role = result
                logger.info(f"User {username} logged in successfully.")
                self.accept()
            else:
                QMessageBox.warning(self, "Access Denied", "Incorrect username or password")
                self.passwdField.clear()
        except Exception as e:
            QMessageBox.warning(self, "Error", "An error occurred during login.")
            logger.error(f"Login error: {e}")

# ------------------------------
# Printing via Print Dialog
# ------------------------------
def print_image_via_dialog(parent, image_path):
    try:
        printer = QPrinter(QPrinter.HighResolution)
        printer.setFullPage(True)
        dialog = QPrintDialog(printer, parent)
        if dialog.exec_() == QDialog.Accepted:
            painter = QPainter(printer)
            img = QImage(image_path)
            if img.isNull():
                raise IOError("Failed to open image for printing.")
            rect = painter.viewport()
            size = img.size()
            size.scale(rect.size(), Qt.KeepAspectRatio)
            painter.setViewport(rect.x(), rect.y(), size.width(), size.height())
            painter.setWindow(img.rect())
            painter.drawImage(0, 0, img)
            painter.end()
            QMessageBox.information(parent, "Printing", "Image successfully sent to printer.")
        else:
            QMessageBox.information(parent, "Printing", "Print job canceled.")
    except Exception as ex:
        QMessageBox.warning(parent, "Print Error", f"Unable to open printer:\n{ex}")
        logger.error(f"Printing error: {ex}")

# ------------------------------
# Main Application UI
# ------------------------------
class BarcodeGeneratorUI(QWidget):
    def __init__(self, current_user, current_role):
        super().__init__()
        self.current_user = current_user
        self.current_role = current_role
        try:
            self.db = DatabaseManager()
        except Exception as e:
            QMessageBox.critical(self, "Error", "Failed to initialize the database.")
            sys.exit(1)
        self.last_barcode_png = None
        self.label_font = QFont("Arial", 8)
        self.initUI()

    def initUI(self):
        self.setWindowTitle(TITLE)
        self.setGeometry(250, 250, 1100, 600)
        self.setStyleSheet("""
            QWidget { background-color: #1c1c1c; color: #e0e0e0; font-family: 'Segoe UI', sans-serif; font-size: 14px; }
            QLineEdit, QComboBox { background-color: #2a2a2a; border: 1px solid #444444; border-radius: 6px; padding: 6px; color: #ffffff; }
            QPushButton { background-color: #333333; border: none; border-radius: 6px; padding: 10px 20px; color: #ffffff; font-weight: bold; }
            QLabel { color: #ffffff; font-size: 14px; }
            QTableWidget { background-color: #2a2a2a; border: 1px solid #444444; border-radius: 6px; }
        """)
        main_layout = QVBoxLayout()
        heading_label = QLabel(TITLE)
        heading_label.setAlignment(Qt.AlignCenter)
        heading_label.setStyleSheet("font-size: 18px; font-weight: 600; margin-bottom: 8px;")
        main_layout.addWidget(heading_label)
        # Row 1: Item Name, Purchase Price, Profit %, Sale Price, Calculator
        row1_layout = QHBoxLayout()
        row1_layout.addWidget(QLabel("Item Name:"))
        self.itemField = QLineEdit()
        row1_layout.addWidget(self.itemField)
        row1_layout.addWidget(QLabel("Purchase Price:"))
        self.purchaseField = QLineEdit()
        row1_layout.addWidget(self.purchaseField)
        row1_layout.addWidget(QLabel("Profit (%):"))
        self.profitField = QLineEdit()
        self.profitField.setPlaceholderText("optional")
        row1_layout.addWidget(self.profitField)
        row1_layout.addWidget(QLabel("Sale Price:"))
        self.saleField = QLineEdit()
        row1_layout.addWidget(self.saleField)
        calcBtn = QPushButton("Calc")
        calcBtn.clicked.connect(self.apply_profit_calculator)
        row1_layout.addWidget(calcBtn)
        main_layout.addLayout(row1_layout)
        # Row 2: Company, No. of Labels, Stock Count, Barcode
        row2_layout = QHBoxLayout()
        row2_layout.addWidget(QLabel("Company:"))
        self.companyField = QLineEdit()
        row2_layout.addWidget(self.companyField)
        row2_layout.addWidget(QLabel("No. of Labels:"))
        self.countField = QLineEdit()
        row2_layout.addWidget(self.countField)
        row2_layout.addWidget(QLabel("Stock Count:"))
        self.stockField = QLineEdit()
        row2_layout.addWidget(self.stockField)
        row2_layout.addWidget(QLabel("Barcode (optional):"))
        self.barcodeField = QLineEdit()
        self.barcodeField.setPlaceholderText("If empty, generate random 12-digit code")
        row2_layout.addWidget(self.barcodeField)
        main_layout.addLayout(row2_layout)
        # Row 3: Label Size, Custom Dimensions, Continuous Roll
        row3_layout = QHBoxLayout()
        row3_layout.addWidget(QLabel("Label Size:"))
        self.sizeBox = QComboBox()
        for label in LABEL_SIZES:
            self.sizeBox.addItem(label)
        row3_layout.addWidget(self.sizeBox)
        row3_layout.addWidget(QLabel("Custom W x H (mm):"))
        self.customWidth = QLineEdit()
        self.customHeight = QLineEdit()
        self.customWidth.setVisible(False)
        self.customHeight.setVisible(False)
        row3_layout.addWidget(self.customWidth)
        row3_layout.addWidget(self.customHeight)
        self.continuousRollCheck = QCheckBox("Continuous Roll")
        row3_layout.addWidget(self.continuousRollCheck)
        main_layout.addLayout(row3_layout)
        # Buttons: Generate PDF, Print (opens print dialog), Reprint, Clear History, Open Excel, Manage Users
        btn_layout = QHBoxLayout()
        genBtn = QPushButton("Generate Labels PDF")
        printNowBtn = QPushButton("Print Label (Select Printer)")
        reprintBtn = QPushButton("Reprint Selected")
        clearHistBtn = QPushButton("Clear History")
        openExcelBtn = QPushButton("Open Items Excel")
        genBtn.clicked.connect(self.generate_pdf)
        # Use a lambda to check for last barcode image and then call our print dialog
        printNowBtn.clicked.connect(lambda: print_image_via_dialog(self, self.last_barcode_png)
                                     if self.last_barcode_png else QMessageBox.warning(self, "Print Error", "Generate a barcode label first."))
        reprintBtn.clicked.connect(self.reprint_selected_barcode)
        clearHistBtn.clicked.connect(self.clear_history)
        openExcelBtn.clicked.connect(self.open_excel_file)
        btn_layout.addWidget(genBtn)
        btn_layout.addWidget(printNowBtn)
        btn_layout.addWidget(reprintBtn)
        btn_layout.addWidget(clearHistBtn)
        btn_layout.addWidget(openExcelBtn)
        if self.current_role == "admin":
            manageUsersBtn = QPushButton("Manage Users")
            manageUsersBtn.clicked.connect(self.manage_users)
        main_layout.addLayout(btn_layout)
        # Search History and History Table
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Search History:"))
        self.searchField = QLineEdit()
        self.searchField.textChanged.connect(self.refresh_history)
        search_layout.addWidget(self.searchField)
        main_layout.addLayout(search_layout)
        self.historyTable = QTableWidget()
        self.historyTable.setColumnCount(9)
        self.historyTable.setHorizontalHeaderLabels(
            ["ID", "User", "Item Name", "Barcode", "Sale Price", "Purchase Price", "Company", "Stock Count", "Date"]
        )
        self.historyTable.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        main_layout.addWidget(QLabel("Barcode Generation History:"))
        main_layout.addWidget(self.historyTable)
        self.statusBar = QStatusBar()
        main_layout.addWidget(self.statusBar)
        self.setLayout(main_layout)
        self.sizeBox.currentTextChanged.connect(self.toggleCustomSize)
        self.refresh_history()

    def toggleCustomSize(self):
        is_custom = (self.sizeBox.currentText() == "Custom Size")
        self.customWidth.setVisible(is_custom)
        self.customHeight.setVisible(is_custom)

    def apply_profit_calculator(self):
        p_txt = self.purchaseField.text().strip().replace(",", ".")
        profit_txt = self.profitField.text().strip().replace(",", ".")
        try:
            purchase_val = float(p_txt)
        except ValueError:
            QMessageBox.warning(self, "Error", "Enter a valid numeric Purchase Price.")
            return
        if not profit_txt:
            QMessageBox.warning(self, "Error", "Enter a valid numeric Profit (%) or type 0.")
            return
        try:
            profit_percent = float(profit_txt)
        except ValueError:
            QMessageBox.warning(self, "Error", "Profit (%) must be numeric.")
            return
        calc_sale = round(purchase_val + (purchase_val * (profit_percent / 100.0)))
        self.saleField.setText(str(calc_sale))

    def get_label_size(self):
        label = self.sizeBox.currentText()
        if label == "Custom Size":
            try:
                w = float(self.customWidth.text().strip().replace(",", "."))
                h = float(self.customHeight.text().strip().replace(",", "."))
                return w, h
            except ValueError:
                QMessageBox.warning(self, "Error", "Enter valid numeric custom dimensions.")
                logger.error("Invalid custom size.")
                return None
        return LABEL_SIZES[label]

    def generate_pdf(self):
        item_name = self.itemField.text().strip()
        purchase_txt = self.purchaseField.text().strip().replace(",", ".")
        sale_txt = self.saleField.text().strip().replace(",", ".")
        company = self.companyField.text().strip()
        count_txt = self.countField.text().strip()
        stock_txt = self.stockField.text().strip()
        barcode_in = self.barcodeField.text().strip()
        if not item_name or not purchase_txt or not sale_txt or not company or not count_txt or not stock_txt:
            QMessageBox.warning(self, "Error", "All fields except Barcode/Profit must be filled out.")
            return
        try:
            purchase_val = float(purchase_txt)
            sale_val = float(sale_txt)
            count_val = int(count_txt)
            stock_val = int(stock_txt)
        except ValueError:
            QMessageBox.warning(self, "Error", "Purchase, Sale, No. of Labels and Stock Count must be numeric.")
            return
        size = self.get_label_size()
        if not size:
            return
        w_mm, h_mm = size
        barcode_str = barcode_in if barcode_in else "".join(random.choices("0123456789", k=12))
        try:
            barcode_png = generate_barcode_image(barcode_str)
            self.last_barcode_png = barcode_png
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Barcode generation failed:\n{e}")
            return
        folder = QFileDialog.getExistingDirectory(self, "Select folder to save PDF")
        if not folder:
            return
        pdf_path = os.path.join(folder, f"labels_{barcode_str}.pdf")
        details = {"item": item_name, "sale": sale_txt, "purchase": purchase_txt, "company": company, "count": count_val}
        try:
            generate_label_pdf(pdf_path, w_mm, h_mm, self.last_barcode_png, details)
            QMessageBox.information(self, "Done", f"PDF saved:\n{pdf_path}")
            self.statusBar.showMessage("PDF generated successfully.", 5000)
        except Exception as ex:
            QMessageBox.warning(self, "Error", f"PDF generation failed:\n{ex}")
            logger.error(f"PDF generation error: {ex}")
            return
        gen_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            self.db.add_barcode_history(self.current_user, item_name, barcode_str,
                                        sale_val, purchase_val, company, gen_date, stock_val)
            self.db.add_or_update_barcode(item_name, barcode_str, sale_val, purchase_val, company, count_val, stock_val)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Database update failed:\n{e}")
            logger.error(f"Database update error: {e}")
        try:
            update_excel(
                item_code=barcode_str,
                item_name=item_name,
                sale_price=sale_txt,
                purchase_price=purchase_txt,
                company_name=company,
                generation_date=gen_date,
                stock_count=stock_val
            )
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Excel update failed:\n{e}")
            logger.error(f"Excel update error: {e}")
        self.refresh_history()

    def reprint_selected_barcode(self):
        selected_rows = self.historyTable.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.information(self, "Reprint", "Please select a barcode record to reprint.")
            return
        index = selected_rows[0].row()
        barcode_value = self.historyTable.item(index, 3).text()
        try:
            barcode_png = generate_barcode_image(barcode_value)
            print_image_via_dialog(self, barcode_png)
        except Exception as e:
            QMessageBox.warning(self, "Reprint Error", f"Failed to reprint barcode: {e}")
            logger.error(f"Reprint error: {e}")

    def refresh_history(self):
        search_text = self.searchField.text().strip()
        history = self.db.fetch_history(search_text)
        self.historyTable.setRowCount(len(history))
        for i, row in enumerate(history):
            for j, value in enumerate(row):
                cell_item = QTableWidgetItem(str(value))
                cell_item.setTextAlignment(Qt.AlignCenter)
                self.historyTable.setItem(i, j, cell_item)
        self.historyTable.resizeRowsToContents()

    def clear_history(self):
        reply = QMessageBox.question(
            self,
            "Clear History",
            "Are you sure you want to clear the entire barcode generation history?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            try:
                self.db.clear_history()
                self.refresh_history()
                QMessageBox.information(self, "Clear History", "Barcode generation history cleared.")
                self.statusBar.showMessage("History cleared.", 5000)
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to clear history: {e}")
                logger.error(f"Clear history error: {e}")

    def open_excel_file(self):
        excel_file = "Items.xlsx"
        if not os.path.exists(excel_file):
            QMessageBox.information(self, "Excel Not Found", "Items.xlsx does not exist yet.")
            return
        try:
            if platform.system() == "Windows":
                os.startfile(excel_file)
            else:
                subprocess.run(["open", excel_file])
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Could not open {excel_file}:\n{e}")
            logger.error(f"Open Excel error: {e}")

    def manage_users(self):
        dialog = UserManagementDialog(self)
        dialog.exec_()

# ------------------------------
# Application Launcher
# ------------------------------
if __name__ == "__main__":
    multiprocessing.freeze_support()  # For PyInstaller support on Windows.
    if is_already_running():
        print("Another instance is already running. Exiting.")
        sys.exit(0)
    if platform.system() != "Windows":
        print("This application is designed to run on Windows.")
        sys.exit(1)
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    app = QApplication(sys.argv)
    login = LoginDialog()
    if login.exec_() == QDialog.Accepted:
        window = BarcodeGeneratorUI(login.user, login.role)
        window.show()
        sys.exit(app.exec_())
    else:
        sys.exit(0)


# --- Barcode Generation Utility ---
def generate_barcode_image(barcode_str, output_path):
    safe_barcode = sanitize_barcode_input(barcode_str)
    if not safe_barcode:
        raise ValueError("No valid Code128 characters after sanitizing.")
    try:
        Code128 = get_barcode_class('code128')
        code128_obj = Code128(safe_barcode, writer=ImageWriter())
        options = {"font_path": r"C:\Windows\Fonts\arial.ttf", "module_height": 15.0, "module_width": 0.3, "font_size": 10}
        code128_obj.writer.set_options(options)
        saved_path = code128_obj.save(output_path[:-4])  # Remove .png if present
        if not os.path.isfile(saved_path):
            raise OSError(f"Barcode image not created at: {saved_path}")
        logger.info("Barcode image generated at %s", saved_path)
        return saved_path
    except Exception as e:
        logger.error(f"Barcode generation failed: {e}")
        raise

# --- PDF Generation Utility ---
def generate_label_pdf(pdf_path, w_mm, h_mm, barcode_path, details, font=None):
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    c = canvas.Canvas(pdf_path, pagesize=(w_mm * mm, h_mm * mm))
    count = details.get("count", 1)
    labels_per_row = 2 if w_mm >= 100 else 1
    label_width = w_mm / labels_per_row
    for i in range(0, count, labels_per_row):
        for j in range(labels_per_row):
            if i + j >= count:
                break
            x_offset = j * label_width
            draw_single_label(c, x_offset, 0, label_width, h_mm, barcode_path, details, font)
        c.showPage()
    c.save()

def draw_single_label(c, x_offset_mm, y_offset_mm, w_label_mm, h_label_mm, barcode_path, details, font=None):
    item = details.get("item", "")
    sale = details.get("sale", "")
    company = details.get("company", "")
    c.rect(x_offset_mm * mm, y_offset_mm * mm, w_label_mm * mm, h_label_mm * mm)
    if font:
        c.setFont(font.family(), font.pointSize())
    else:
        c.setFont("Helvetica", 8)
    c.drawCentredString((x_offset_mm + w_label_mm / 2) * mm, (y_offset_mm + h_label_mm - 5) * mm, item)
    c.drawImage(barcode_path, (x_offset_mm + w_label_mm / 2 - 20) * mm,
                (y_offset_mm + h_label_mm - 16) * mm, 40 * mm, 8 * mm)
    c.drawCentredString((x_offset_mm + w_label_mm / 2) * mm, (y_offset_mm + h_label_mm - 20) * mm, f"Sale: {sale} | {company}")

# --- Excel Export Utility ---
def update_excel(item_code, item_name, sale_price, purchase_price, company_name, generation_date, stock_count, excel_path=None):
    from openpyxl import Workbook, load_workbook
    file_name = excel_path or "Items.xlsx"
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Items"
        ws.append(["Item Code", "Item Name", "Sale Price", "Purchase Price", "Company", "Opening Stock", "Stock Count", "Generation Date"])
        wb.save(file_name)
    wb = load_workbook(file_name)
    ws = wb.active
    found = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == item_code:
            row[1].value = item_name
            row[2].value = sale_price
            row[3].value = purchase_price
            row[4].value = company_name
            row[5].value = stock_count  # Opening Stock
            row[6].value = stock_count
            row[7].value = generation_date
            found = True
            break
    if not found:
        ws.append([item_code, item_name, sale_price, purchase_price, company_name, stock_count, stock_count, generation_date])
    for attempt in range(3):
        try:
            wb.save(file_name)
            logger.info("Excel updated: item_code=%s (found=%s)", item_code, found)
            break
        except Exception as e:
            logger.error(f"Excel save failed: {e}")
            ret = QMessageBox.question(None, "Excel Locked",
                                       f"Unable to save {file_name} (maybe open/locked?).\nClose the file and click 'Retry' or 'Cancel' to skip updating.",
                                       QMessageBox.Retry | QMessageBox.Cancel)
            if ret == QMessageBox.Retry:
                continue
            else:
                break