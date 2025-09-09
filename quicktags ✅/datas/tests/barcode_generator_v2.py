import sys
import ctypes
import multiprocessing
import subprocess
import platform
import logging
import os
import random
import datetime
import sqlite3
import csv
from typing import Dict, List, Tuple, Optional

from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QFormLayout, QDialog, QDialogButtonBox,
    QPushButton, QMessageBox, QLineEdit, QFileDialog, QLabel,
    QComboBox, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView,
    QCheckBox, QStatusBar, QGroupBox, QScrollArea, QTextEdit, QSpinBox,
    QDoubleSpinBox, QTabWidget, QSizePolicy, QGraphicsView, QGraphicsScene,
    QGraphicsPixmapItem, QSplitter, QToolButton, QMenu, QAction, QInputDialog
)
from PyQt5.QtCore import Qt, QSize, QRectF, pyqtSignal
from PyQt5.QtGui import (
    QFont, QImage, QPainter, QPixmap, QColor, QIcon,
    QTextCursor, QFontDatabase
)
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog

# ------------------------------
# Constants and Configurations
# ------------------------------
TITLE = "Barcode Generator v2"
LOGO_PATH = "logo.ico"  # Assumed to be in the same directory
DEFAULT_FONT = "Arial"
DEFAULT_FONT_SIZE = 10
DPI_OPTIONS = [150, 200, 300, 600]
DEFAULT_DPI = 300

# Label sizes in mm (width, height)
LABEL_SIZES = {
    "Small (60x25 mm)": (60, 25),
    "Medium (100x25 mm)": (100, 25),
    "Large (150x50 mm)": (150, 50),
    "A4 Sheet (210x297 mm)": (210, 297),
    "2 Labels (50x25 mm)": (100, 25),  # Will be split into two 50x25 labels
    "Custom Size": None
}


# ------------------------------
# Single Instance Check (Windows Only)
# ------------------------------
def is_already_running():
    if platform.system() != "Windows":
        return False
    mutex = ctypes.windll.kernel32.CreateMutexW(None, 1, "Global\\BarcodeGeneratorV2")
    return ctypes.GetLastError() == 183


# ------------------------------
# App Data Directory and Global Paths
# ------------------------------
def get_app_data_path():
    base_path = os.path.join(os.environ["APPDATA"], "BarcodeGeneratorV2")
    os.makedirs(base_path, exist_ok=True)
    return base_path


def get_today_folder(base_path: str) -> str:
    today = datetime.datetime.now().strftime("%d-%m-%y")
    today_path = os.path.join(base_path, today)
    os.makedirs(today_path, exist_ok=True)
    return today_path


def setup_output_folders(base_path: str) -> Dict[str, str]:
    today_path = get_today_folder(base_path)
    # In setup_output_folders()
    # Update folder structure
    folders = {
        "pdf": os.path.join(base_path, date_str, "PDF"),
        "png": os.path.join(base_path, date_str, "PNG"),
        "bmp": os.path.join(base_path, date_str, "BMP"),
        "excel": os.path.join(base_path, date_str, "Excel")
    }
    for folder in folders.values():
        os.makedirs(folder, exist_ok=True)
    return folders


APP_DATA_PATH = get_app_data_path()
DB_PATH = os.path.join(APP_DATA_PATH, "barcodes.db")
LOG_PATH = os.path.join(APP_DATA_PATH, "app.log")
TEMPLATES_PATH = os.path.join(APP_DATA_PATH, "templates")
os.makedirs(TEMPLATES_PATH, exist_ok=True)

# ------------------------------
# Logging Setup
# ------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def show_error(message: str, parent=None):
    """Show error message both in UI and log it."""
    logger.error(message)
    QMessageBox.critical(parent, "Error", message)


def show_warning(message: str, parent=None):
    """Show warning message both in UI and log it."""
    logger.warning(message)
    QMessageBox.warning(parent, "Warning", message)


def show_info(message: str, parent=None):
    """Show info message both in UI and log it."""
    logger.info(message)
    QMessageBox.information(parent, "Information", message)


# ------------------------------
# Database Manager
# ------------------------------
class DatabaseManager:
    def __init__(self, db_path=DB_PATH):
        self.db_path = db_path
        try:
            self.conn = sqlite3.connect(self.db_path)
            self.create_tables()
            self.create_default_templates()
            logger.info("Database connected and tables created/verified.")
        except Exception as e:
            show_error(f"Database connection failed: {e}")
            raise

    def create_tables(self):
        try:
            c = self.conn.cursor()
            # Barcode history with stock_count
            c.execute('''
                CREATE TABLE IF NOT EXISTS barcode_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    item_name TEXT,
                    barcode TEXT,
                    sale_price REAL,
                    purchase_price REAL,
                    company_name TEXT,
                    generation_date TEXT,
                    username TEXT DEFAULT 'admin',
                    opening_stock INTEGER DEFAULT 0,
                    template_used TEXT
                )
            ''')
            # Barcodes table with stock
            c.execute('''
                CREATE TABLE IF NOT EXISTS barcodes (
                    barcode TEXT PRIMARY KEY,
                    item_name TEXT,
                    sale_price REAL,
                    purchase_price REAL,
                    company_name TEXT,
                    quantity INTEGER,
                    opening_stock INTEGER DEFAULT 0,
                    template_used TEXT
                )
            ''')
            # Users table
            c.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    username TEXT PRIMARY KEY,
                    password TEXT,
                    role TEXT
                )
            ''')
            # Templates table
            c.execute('''
                CREATE TABLE IF NOT EXISTS templates (
                    name TEXT PRIMARY KEY,
                    font_family TEXT,
                    font_size INTEGER,
                    barcode_scale REAL,
                    show_price BOOLEAN,
                    show_company BOOLEAN,
                    show_stock BOOLEAN,
                    created_by TEXT,
                    created_date TEXT
                )
            ''')
            # Insert default admin if none exist
            c.execute('SELECT COUNT(*) FROM users')
            if c.fetchone()[0] == 0:
                c.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)',
                          ("admin", "admin123", "admin"))
            self.conn.commit()
        except Exception as e:
            show_error(f"Error creating tables: {e}")
            raise

    def create_default_templates(self):
        try:
            c = self.conn.cursor()
            c.execute('SELECT COUNT(*) FROM templates')
            if c.fetchone()[0] == 0:
                # Create a default template
                default_template = (
                    "Default", "Arial", 10, 1.0, True, True, False,
                    "system", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                )
                c.execute('''
                    INSERT INTO templates 
                    (name, font_family, font_size, barcode_scale, show_price, show_company, show_stock, created_by, created_date)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', default_template)
                self.conn.commit()
        except Exception as e:
            show_error(f"Error creating default templates: {e}")

    def add_barcode_history(self, username: str, item_name: str, barcode: str, sale_price: float,
                            purchase_price: float, company_name: str, generation_date: str,
                            opening_stock: int, template_used: str):
        try:
            c = self.conn.cursor()
            c.execute('''
                INSERT INTO barcode_history
                (username, item_name, barcode, sale_price, purchase_price, company_name, 
                 generation_date, opening_stock, template_used)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (username, item_name, barcode, sale_price, purchase_price,
                  company_name, generation_date, opening_stock, template_used))
            self.conn.commit()
            logger.info("Barcode history added.")
        except Exception as e:
            show_error(f"Failed to add barcode history: {e}")
            raise

    def add_or_update_barcode(self, item_name: str, barcode: str, sale_price: float,
                              purchase_price: float, company_name: str, quantity: int,
                              opening_stock: int, template_used: str):
        try:
            c = self.conn.cursor()
            c.execute('SELECT quantity FROM barcodes WHERE barcode=?', (barcode,))
            result = c.fetchone()
            if result:
                new_quantity = result[0] + quantity
                c.execute('''
                    UPDATE barcodes
                    SET item_name=?, sale_price=?, purchase_price=?, company_name=?, 
                        quantity=?, opening_stock=?, template_used=?
                    WHERE barcode=?
                ''', (item_name, sale_price, purchase_price, company_name,
                      new_quantity, opening_stock, template_used, barcode))
            else:
                c.execute('''
                    INSERT INTO barcodes
                    (barcode, item_name, sale_price, purchase_price, company_name, 
                     quantity, opening_stock, template_used)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (barcode, item_name, sale_price, purchase_price, company_name,
                      quantity, opening_stock, template_used))
            self.conn.commit()
            logger.info("Barcode updated/added in barcodes table.")
        except Exception as e:
            show_error(f"Failed to add or update barcode: {e}")
            raise

    def fetch_history(self, search_filter: str = "") -> List[Tuple]:
        try:
            c = self.conn.cursor()
            query = """
                SELECT id, username, item_name, barcode, sale_price, purchase_price, 
                       company_name, opening_stock, generation_date, template_used
                FROM barcode_history
            """
            params = ()
            if search_filter:
                query += " WHERE item_name LIKE ? OR barcode LIKE ?"
                params = (f"%{search_filter}%", f"%{search_filter}%")
            query += " ORDER BY id DESC"
            c.execute(query, params)
            history = c.fetchall()
            logger.info("History fetched from database.")
            return history
        except Exception as e:
            show_error(f"Failed to fetch history: {e}")
            return []

    def clear_history(self):
        try:
            c = self.conn.cursor()
            c.execute('DELETE FROM barcode_history')
            self.conn.commit()
            logger.info("Barcode history cleared.")
        except Exception as e:
            show_error(f"Failed to clear history: {e}")
            raise

    def add_user(self, username: str, password: str, role: str = "user"):
        try:
            c = self.conn.cursor()
            c.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)',
                      (username, password, role))
            self.conn.commit()
            logger.info(f"User {username} added.")
        except Exception as e:
            show_error(f"Failed to add user: {e}")
            raise

    def remove_user(self, username: str):
        try:
            c = self.conn.cursor()
            c.execute('DELETE FROM users WHERE username=?', (username,))
            self.conn.commit()
            logger.info(f"User {username} removed.")
        except Exception as e:
            show_error(f"Failed to remove user: {e}")
            raise

    def fetch_users(self) -> List[Tuple]:
        try:
            c = self.conn.cursor()
            c.execute('SELECT username, role FROM users ORDER BY username ASC')
            users = c.fetchall()
            logger.info("Fetched users from database.")
            return users
        except Exception as e:
            show_error(f"Failed to fetch users: {e}")
            return []

    def validate_user(self, username: str, password: str) -> Optional[Tuple]:
        try:
            c = self.conn.cursor()
            c.execute('SELECT username, role FROM users WHERE username=? AND password=?',
                      (username, password))
            result = c.fetchone()
            logger.info(f"User {username} validation attempted.")
            return result
        except Exception as e:
            show_error(f"User validation failed: {e}")
            return None

    def add_template(self, name: str, font_family: str, font_size: int, barcode_scale: float,
                     show_price: bool, show_company: bool, show_stock: bool, created_by: str):
        try:
            c = self.conn.cursor()
            created_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            c.execute('''
                INSERT INTO templates 
                (name, font_family, font_size, barcode_scale, show_price, show_company, show_stock, created_by, created_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (name, font_family, font_size, barcode_scale, show_price, show_company, show_stock, created_by,
                  created_date))
            self.conn.commit()
            logger.info(f"Template {name} added by {created_by}.")
            return True
        except sqlite3.IntegrityError:
            show_warning(f"Template '{name}' already exists.")
            return False
        except Exception as e:
            show_error(f"Failed to add template: {e}")
            return False

    def update_template(self, name: str, font_family: str, font_size: int, barcode_scale: float,
                        show_price: bool, show_company: bool, show_stock: bool):
        try:
            c = self.conn.cursor()
            c.execute('''
                UPDATE templates 
                SET font_family=?, font_size=?, barcode_scale=?, show_price=?, show_company=?, show_stock=?
                WHERE name=?
            ''', (font_family, font_size, barcode_scale, show_price, show_company, show_stock, name))
            self.conn.commit()
            logger.info(f"Template {name} updated.")
            return True
        except Exception as e:
            show_error(f"Failed to update template: {e}")
            return False

    def delete_template(self, name: str):
        try:
            c = self.conn.cursor()
            c.execute('DELETE FROM templates WHERE name=?', (name,))
            self.conn.commit()
            logger.info(f"Template {name} deleted.")
            return True
        except Exception as e:
            show_error(f"Failed to delete template: {e}")
            return False

    def get_template(self, name: str) -> Optional[Dict]:
        try:
            c = self.conn.cursor()
            c.execute('''
                SELECT name, font_family, font_size, barcode_scale, show_price, show_company, show_stock
                FROM templates WHERE name=?
            ''', (name,))
            result = c.fetchone()
            if result:
                return {
                    "name": result[0],
                    "font_family": result[1],
                    "font_size": result[2],
                    "barcode_scale": result[3],
                    "show_price": bool(result[4]),
                    "show_company": bool(result[5]),
                    "show_stock": bool(result[6])
                }
            return None
        except Exception as e:
            show_error(f"Failed to get template: {e}")
            return None

    def get_all_templates(self) -> List[Tuple]:
        try:
            c = self.conn.cursor()
            c.execute('SELECT name, created_by, created_date FROM templates ORDER BY name ASC')
            return c.fetchall()
        except Exception as e:
            show_error(f"Failed to get templates: {e}")
            return []


# ------------------------------
# Barcode Generation (Code128)
# ------------------------------
try:
    from barcode import get_barcode_class
    from barcode.writer import ImageWriter
except ImportError as e:
    show_error("Missing dependency 'python-barcode'. Please install it before running this application.")
    raise e


def sanitize_barcode_input(barcode_str: str) -> str:
    """Ensure barcode only contains valid Code128 characters."""
    return "".join(ch for ch in barcode_str if 32 <= ord(ch) <= 126)


def generate_barcode_image(barcode_str: str, scale: float = 1.0, dpi: int = 300,
                           font_path: str = None) -> str:
    """
    Generate a barcode image with the given specifications.
    Returns path to the generated image.
    """
    safe_barcode = sanitize_barcode_input(barcode_str)
    if not safe_barcode:
        raise ValueError("No valid Code128 characters after sanitizing.")

    try:
        Code128 = get_barcode_class('code128')
        output_folders = setup_output_folders(APP_DATA_PATH)
        base_filename = os.path.join(output_folders["png"], safe_barcode)

        # Custom writer with specified DPI and scale
        class CustomWriter(ImageWriter):
            def __init__(self):
                super().__init__()
                self.dpi = dpi
                self.set_options({
                    "module_width": 0.2 * scale,
                    "module_height": 15.0 * scale,
                    "font_size": 10,
                    "text_distance": 2
                })
                if font_path:
                    self.set_options({"font_path": font_path})

        writer = CustomWriter()
        code128_obj = Code128(safe_barcode, writer=writer)

        # Save to PNG first (highest quality)
        png_path = code128_obj.save(base_filename)

        # Convert to BMP if needed (for thermal printers)
        img = QImage(png_path)
        bmp_path = os.path.join(output_folders["bmp"], f"{safe_barcode}.bmp")
        img.save(bmp_path, "BMP")

        logger.info(f"Barcode images generated at {png_path} and {bmp_path}")
        return png_path
    except Exception as e:
        show_error(f"Barcode generation failed: {e}")
        raise


# ------------------------------
# Unified Label Generation System
# ------------------------------
# In generate_label_image()
# Add unified layout function
def generate_label_image(barcode_str: str, details: Dict, template: Dict,
                        width_mm: float, height_mm: float, dpi: int = 300) -> Tuple[str, str]:
    """
    Generate both PNG and BMP label images with identical layout to PDF
    Returns paths to the generated images (png_path, bmp_path)
    """
    try:
        output_folders = setup_output_folders(APP_DATA_PATH)
        base_filename = os.path.join(output_folders["png"], f"{details['item']} - {barcode_str}")
        
        # Create image with white background
        mm_to_px = dpi / 25.4
        width_px = int(width_mm * mm_to_px)
        height_px = int(height_mm * mm_to_px)
        
        image = QImage(width_px, height_px, QImage.Format_RGB32)
        image.fill(Qt.white)
        
        painter = QPainter(image)
        try:
            # Set font
            font = QFont(template["font_family"], template["font_size"])
            painter.setFont(font)
            
            # Calculate positions (in pixels)
            center_x = width_px / 2
            top_y = height_px - 5  # 5px from top
            
            # Draw item name (top center)
            painter.drawText(QRectF(0, top_y - 20, width_px, 20), 
                           Qt.AlignCenter, details["item"])
            
            # Generate barcode image
            barcode_img = generate_barcode_only(barcode_str, 
                                               scale=template["barcode_scale"], 
                                               dpi=dpi)
            
            # Draw barcode (center)
            barcode_pixmap = QPixmap(barcode_img)
            barcode_width = barcode_pixmap.width()
            barcode_height = barcode_pixmap.height()
            barcode_x = center_x - (barcode_width / 2)
            barcode_y = top_y + 20  # 20px below item name
            painter.drawPixmap(barcode_x, barcode_y, barcode_pixmap)
            
            # Draw additional info (bottom)
            info_lines = []
            if template["show_price"]:
                info_lines.append(f"Sale: {details['sale']}")
            if template["show_company"]:
                info_lines.append(details["company"])
            if template["show_stock"]:
                info_lines.append(f"Stock: {details['opening_stock']}")

            info_text = " | ".join(info_lines)
            if info_text:
                painter.drawText(QRectF(0, height_px - 30, width_px, 20), 
                               Qt.AlignCenter, info_text)
                               
        finally:
            painter.end()
        
        # Save PNG
        png_path = f"{base_filename}.png"
        image.save(png_path, "PNG")
        
        # Save BMP
        bmp_path = os.path.join(output_folders["bmp"], f"{details['item']} - {barcode_str}.bmp")
        image.save(bmp_path, "BMP")
        
        logger.info(f"Label images generated at {png_path} and {bmp_path}")
        return png_path, bmp_path
        
    except Exception as e:
        show_error(f"Label image generation failed: {e}")
        raise

def generate_barcode_only(barcode_str: str, scale: float = 1.0, dpi: int = 300) -> str:
    """Generate just the barcode portion without text"""
    safe_barcode = sanitize_barcode_input(barcode_str)
    if not safe_barcode:
        raise ValueError("No valid Code128 characters after sanitizing.")

    try:
        Code128 = get_barcode_class('code128')
        temp_dir = os.path.join(APP_DATA_PATH, "temp")
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, f"temp_{safe_barcode}.png")

        writer = ImageWriter()
        writer.set_options({
            "module_width": 0.2 * scale,
            "module_height": 15.0 * scale,
            "font_size": 0,  # No text
            "quiet_zone": 1,
            "dpi": dpi
        })

        code128 = Code128(safe_barcode, writer=writer)
        code128.save(temp_path)
        
        return temp_path
    except Exception as e:
        show_error(f"Barcode generation failed: {e}")
        raise


# ------------------------------
# PDF Generation for Labels (using ReportLab)
# ------------------------------
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import Paragraph
except ImportError as e:
    show_error("Missing dependency 'reportlab'. Please install it before running this application.")
    raise e


def draw_label(c: canvas.Canvas, x_offset: float, y_offset: float,
               width: float, height: float, barcode_path: str,
               details: Dict, template: Dict, dpi: int = 300):
    """
    Draw a single label with the given specifications.
    """
    # Set up font
    font_name = template["font_family"]
    try:
        pdfmetrics.registerFont(TTFont(font_name, f"{font_name}.ttf"))
        pdfmetrics.registerFont(TTFont(f"{font_name}-Bold", f"{font_name}.ttf"))
    except:
        font_name = "Helvetica"  # Fallback font

    # Item name (top center, bold)
    c.setFont(f"{font_name}-Bold", template["font_size"])
    c.drawCentredString((x_offset + width / 2) * mm, (y_offset + height - 5) * mm,
                        details["item"])

    # Barcode (center)
    barcode_width = 40 * template["barcode_scale"]
    barcode_height = 15
    c.drawImage(barcode_path,
                (x_offset + (width - barcode_width) / 2) * mm,
                (y_offset + (height - barcode_height) / 2) * mm,
                barcode_width * mm,
                barcode_height * mm,
                preserveAspectRatio=True)

    # Additional info (bottom)
    info_lines = []
    if template["show_price"]:
        price = details["sale"]
        if "." in price:
            price = price.rstrip("0").rstrip(".") if "." in price else price
        info_lines.append(f"Sale: {price}")
    if template["show_company"]:
        info_lines.append(details["company"])
    if template["show_stock"]:
        info_lines.append(f"Stock: {details['opening_stock']}")

    info_text = " | ".join(info_lines)
    if info_text:
        c.setFont(font_name, template["font_size"] - 2)
        c.drawCentredString((x_offset + width / 2) * mm, (y_offset + 5) * mm, info_text)


def generate_label_pdf(pdf_path: str, width_mm: float, height_mm: float,
                       barcode_path: str, details: Dict, template: Dict,
                       dpi: int = 300) -> bool:
    """
    Generate a PDF with labels according to specifications.
    Returns True if successful, False otherwise.
    """
    try:
        count = details.get("count", 1)
        is_dual = (width_mm, height_mm) == (100, 25)  # Special case for 2 labels per row

        # Create canvas with appropriate page size
        if is_dual:
            c = canvas.Canvas(pdf_path, pagesize=(width_mm * mm, height_mm * mm))
        else:
            c = canvas.Canvas(pdf_path,
                              pagesize=A4 if (width_mm, height_mm) == (210, 297) else (width_mm * mm, height_mm * mm))

        # For dual labels (2 per row)
        if is_dual:
            labels_per_page = 2
            labels_remaining = count

            while labels_remaining > 0:
                # First label in row
                draw_label(c, 0, 0, 50, 25, barcode_path, details, template, dpi)
                labels_remaining -= 1

                # Second label in row if needed
                if labels_remaining > 0:
                    draw_label(c, 50, 0, 50, 25, barcode_path, details, template, dpi)
                    labels_remaining -= 1

                c.showPage()
        else:
            # Single label per page
            for _ in range(count):
                draw_label(c, 0, 0, width_mm, height_mm, barcode_path, details, template, dpi)
                c.showPage()

        c.save()
        logger.info(f"PDF generated successfully at {pdf_path}")
        return True
    except Exception as e:
        show_error(f"PDF generation failed: {e}")
        return False


# ------------------------------
# Excel Export with Stock Management
# ------------------------------
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment
except ImportError as e:
    show_error("Missing dependency 'openpyxl'. Please install it before running this application.")
    raise e


def update_excel(item_code: str, item_name: str, sale_price: str,
                 purchase_price: str, company_name: str, generation_date: str,
                 opening_stock: int) -> bool:
    """
    Update or create Excel file with item data.
    Returns True if successful, False otherwise.
    """
    try:
        output_folders = setup_output_folders(APP_DATA_PATH)
        today = datetime.datetime.now().strftime("%d-%m-%y")
        file_name = os.path.join(output_folders["excel"], f"Stock {today}.xlsx")

        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "Items"
            # Header row with styling
            headers = ["Item Code", "Item Name", "Sale Price", "Purchase Price",
                       "Company", "Opening Stock", "Generation Date"]
            ws.append(headers)

            # Style header row
            bold_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = bold_font
                cell.alignment = Alignment(horizontal='center')

            wb.save(file_name)

        wb = load_workbook(file_name)
        ws = wb.active
        found = False

        # Check if item already exists
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
            if row[0].value == item_code:
                # Update existing row
                row[1].value = item_name
                row[2].value = sale_price
                row[3].value = purchase_price
                row[4].value = company_name
                row[5].value = opening_stock
                row[6].value = generation_date
                found = True
                break

        if not found:
            # Add new row
            ws.append([item_code, item_name, sale_price, purchase_price,
                       company_name, opening_stock, generation_date])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save with retry logic
        for attempt in range(3):
            try:
                wb.save(file_name)
                logger.info(f"Excel updated: item_code={item_code} (found={found})")
                return True
            except Exception as e:
                if attempt == 2:
                    show_error(f"Failed to save Excel after 3 attempts: {e}")
                    return False

                ret = QMessageBox.question(None, "Excel Locked",
                                           f"Unable to save {file_name} (maybe open/locked?).\n"
                                           "Close the file and click 'Retry' or 'Cancel' to skip updating.",
                                           QMessageBox.Retry | QMessageBox.Cancel)
                if ret == QMessageBox.Retry:
                    continue
                else:
                    logger.warning("User canceled Excel update.")
                    return False
    except Exception as e:
        show_error(f"Excel update failed: {e}")
        return False


# ------------------------------
# Preview Widget for Label
# ------------------------------
class LabelPreview(QGraphicsView):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.scene = QGraphicsScene(self)
        self.setScene(self.scene)
        self.setRenderHint(QPainter.Antialiasing)
        self.setRenderHint(QPainter.SmoothPixmapTransform)
        self.setRenderHint(QPainter.TextAntialiasing)
        self.setBackgroundBrush(QColor(240, 240, 240))
        self.setMinimumSize(400, 200)

    def update_preview(self, pixmap: QPixmap):
        """Update the preview with a new pixmap."""
        self.scene.clear()
        if pixmap.isNull():
            return

        # Scale the pixmap to fit the view while maintaining aspect ratio
        scaled_pixmap = pixmap.scaled(self.width(), self.height(),
                                      Qt.KeepAspectRatio, Qt.SmoothTransformation)
        item = QGraphicsPixmapItem(scaled_pixmap)
        self.scene.addItem(item)
        self.fitInView(item, Qt.KeepAspectRatio)
# In BarcodeGeneratorUI.init_ui()
# Update UI styling
self.setStyleSheet("""
    QWidget {
        font-family: 'Segoe UI', sans-serif;
    }
    QPushButton {
        border-radius: 4px;
        padding: 8px;
    }
    #generateBtn {
        background-color: #27ae60;
    }
    #printBtn {
        background-color: #3498db;
    }
""")
# In LabelPreview.update_preview()
# Add zoom/pan functionality
if pixmap.height() > self.max_height:
    scale = self.max_height / pixmap.height()
    self.scale(scale, scale)
    self.scale_factor = scale



